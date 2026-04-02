# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          محرك حساب الكميات الإماراتي — الإصدار الخامس عشر                  ║
║          UAE Master QTO Engine — Version 15                                  ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  كل ما في V14 + الجديد:                                                      ║
║  ✅ طبقة التحقق من المنطق — تحذيرات واضحة للمهندس                           ║
║  ✅ عدّ الفتحات تلقائياً من المسقط (D1، W1، MD1...)                         ║
║  ✅ دعم مجلد كامل (مخططات متعددة: إنشائي + معماري)                          ║
║  ✅ واجهة مستخدم ويب كاملة (رفع ملفات + إدخال + نتائج + Excel)             ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import re

# Ensure UTF-8 encoding for stdout/stderr to prevent encoding crashes on Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='backslashreplace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='backslashreplace')
import math
import base64
import logging
import json
import glob
import requests
import shutil
import tempfile
import time
from datetime import datetime
from typing import List, Dict, Optional, Tuple, Any, Set, Union, cast
from collections import defaultdict, Counter


def _safe_utf8_text(value: Any) -> str:
    text = value if isinstance(value, str) else str(value)
    return text.encode('utf-8', 'backslashreplace').decode('utf-8')

import ezdxf
import fitz  # PyMuPDF
import numpy as np
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from scipy.spatial import KDTree
from shapely.geometry import LineString, Polygon, Point, MultiLineString
from shapely.ops import polygonize, unary_union, snap

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_متاح = True
except ImportError:
    EXCEL_متاح = False
    مسجل_مؤقت = logging.getLogger("QTO_V15")
    مسجل_مؤقت.warning("openpyxl غير مثبت — تصدير Excel غير متاح")

logging.basicConfig(level=logging.INFO)
مسجل = logging.getLogger("QTO_V15")
logger = مسجل  # لتوافق نقاط النهاية الجديدة

التطبيق = FastAPI(title="محرك حساب الكميات الإماراتي v15.0")

# ─────────────────────────────────────────────────────────────────────────────
# نماذج توافق QS Hub (Legacy Compatibility Models)
# ─────────────────────────────────────────────────────────────────────────────

class FoundationInput(BaseModel):
    label: str
    length: float
    width: float
    depth: float
    pcc_thickness: float = 0.10
    quantity: int = 1

class NeckColumnInput(BaseModel):
    label: str
    length: float
    width: float
    foundation_depth: float
    quantity: int = 1

class TieBeamInput(BaseModel):
    label: str
    length: float
    width: float
    depth: float

class StrapBeamInput(BaseModel):
    label: str
    length: float
    width: float
    depth: float

class ColumnSuperInput(BaseModel):
    label: str
    length: float
    width: float
    quantity: int = 1

class BeamInput(BaseModel):
    label: str
    length: float
    width: float
    depth: float

class OpeningInput(BaseModel):
    symbol: str
    length: float
    height: float
    quantity: int = 1

class SolidBlockSubInput(BaseModel):
    label: str
    length: float
    quantity: int = 1

class QTOConstants(BaseModel):
    floor_height: float
    excavation_depth: float
    gfsl_level: float
    tb_depth: float
    no_of_floors: int
    pcc_thickness: float = 0.10
    slab_thickness: float = 0.20
    road_base_exists: bool = False
    road_base_thickness: float = 0.0
    staircase_concrete: float = 0.0

class QTORequest(BaseModel):
    file_path: str
    project_id: str
    unit: str = "mm"
    constants: QTOConstants
    pdf_drawing_scale: float = 100.0
    gemini_api_key: Optional[str] = None
    
    foundations:    List[FoundationInput]    = []
    neck_columns:   List[NeckColumnInput]    = []
    tie_beams:      List[TieBeamInput]       = []
    strap_beams:    List[StrapBeamInput]     = []
    solid_blocks:   List[SolidBlockSubInput] = []
    super_columns:  List[ColumnSuperInput]   = []
    beams:          List[BeamInput]          = []
    openings:       List[OpeningInput]       = []

# ─────────────────────────────────────────────────────────────────────────────
# الثوابت العالمية
# ─────────────────────────────────────────────────────────────────────────────

# طبقات يجب تجاهلها
طبقات_مُهملة = [
    "FURNITURE", "A-FURN", "A-ANNO", "DEFPOINTS", "GRID", "S-GRID",
    "DIMENSIONS", "A-FLOR-PATT", "HATCH", "A-HATCH", "TITLEBLOCK", "BORDER",
    "TEXT", "NOTES", "REVISION", "STAMP", "VP01", "VP02", "XREF", "DETL"
]

# طبقات الكمرات والكمرات الإنشائية
طبقات_الكمرات = ["BEAM", "BEAMS", "C-BEAMS", "S-BEAM", "GB", "GROUND BEAM",
                  "TIE BEAM", "TIE-BEAM", "TB", "STB", "STRAP", "STRAP BEAM"]

# طبقات الجدران المعمارية
طبقات_الجدران = ["A-WALL", "WALL1", "WALL", "A-WALL-PATT", "WALLS"]

# طبقات الأسس
طبقات_الأسس = ["FOU", "FOOTING", "FOUNDATION", "F-S", "PAD", "FOOTING-PLAN"]

# طبقات الأعمدة
طبقات_الأعمدة = ["COL", "COLUMN", "COLUMNS", "C-RFT", "NECK", "N.C", "NC"]

# طبقات البلاطات
طبقات_البلاطات = ["SLAB", "S-SLAB", "0 SLAB", "SOG", "S.O.G"]

# كلمات/عبارات الغرف الحقيقية فقط.
# لا نستخدم كلمات عامة مثل ROOF / FLOOR / HALL كبحث جزئي مباشر
# لأن ذلك كان يلتقط نصوصاً إنشائية مثل SHALL أو TOP ROOF PLAN.
عبارات_الغرف_القياسية = {
    "MASTER BED ROOM": ["MASTER BED ROOM", "MASTER BEDROOM"],
    "BED ROOM": ["BED ROOM", "BEDROOM"],
    "KITCHEN": ["KITCHEN"],
    "BATH": ["BATH", "BATH ROOM", "BATHROOM", "TOILET", "WC"],
    "LIVING": ["LIVING", "LIVING ROOM"],
    "CORRIDOR": ["CORRIDOR", "PASSAGE"],
    "BALCONY": ["BALCONY", "TERRACE", "OUTDOOR"],
    "LOBBY": ["LOBBY"],
    "PANTRY": ["PANTRY"],
    "LAUNDRY": ["LAUNDRY"],
    "DINING": ["DINING", "DINING ROOM"],
    "STORE": ["STORE"],
    "MAJLIS": ["MAJLIS"],
    "MAID ROOM": ["MAID ROOM", "MAID"],
    "DRIVER ROOM": ["DRIVER ROOM", "DRIVER"],
    "GARAGE": ["GARAGE"],
    "STUDY": ["STUDY"],
    "OFFICE": ["OFFICE"],
    "PRAYER": ["PRAYER", "PRAYER ROOM"],
    "HALL": ["HALL"],
    "DRESSING": ["DRESSING", "DRESSING ROOM", "CLOSET", "WALK-IN", "WARDROBE"],
    "FAMILY": ["FAMILY", "FAMILY ROOM", "FAMILY LIVING"],
    "ENTRANCE": ["ENTRANCE", "FOYER", "ENTRY"],
    "STAIRCASE": ["STAIRCASE", "STAIR", "STAIRS"],
    "LANDING": ["LANDING"],
    "GUARD ROOM": ["GUARD ROOM", "GUARD", "SECURITY"],
    "UTILITY": ["UTILITY", "UTILITY ROOM"],
    "غرفة": ["غرفة", "غرفة نوم"],
    "مطبخ": ["مطبخ"],
    "حمام": ["حمام", "دورة مياه"],
    "معيشة": ["معيشة", "صالة", "غرفة معيشة"],
    "ممر": ["ممر"],
    "بلكونة": ["بلكونة", "تراس"],
    "مجلس": ["مجلس"],
    "خادمة": ["خادمة", "غرفة خادمة"],
    "مرآب": ["مرآب", "كراج"],
    "مكتب": ["مكتب"],
    "مصلى": ["مصلى"],
    "مخزن": ["مخزن"],
}
كلمات_الغرف = sorted(
    {عبارة for عبارات in عبارات_الغرف_القياسية.values() for عبارة in عبارات},
    key=len,
    reverse=True,
)

كلمات_استبعاد_الغرف = [
    "GENERAL NOTES", "NOTE :", "NOTE", "NOTES", "SCHEDULE", "SCALE",
    "PLAN", "LEVEL", "TOP LVL", "TOP OF", "F.F.L", "LVL", "ROOF SYSTEM",
    "SLAB", "BEAM", "COLUMN", "RFT", "STEEL", "STIRRUPS", "DRAWINGS",
    "SPECIFICATION", "WATER PROOF", "WATERPROOF", "WINDOWS", "DOORS",
    "CONTROL DEVICE", "EXCAVATION", "MID-BEAM", "N.G.LVL", "PROOFING",
    "SHALL", "PROVIDED", "CONSTRUCTION JOINTS", "OPENING SHOWN",
]

# كلمات المناطق المبللة
مناطق_مبللة_كلمات = ["BATH", "TOILET", "KITCHEN", "PANTRY", "LAUNDRY", "MAID", "WC",
                      "حمام", "دورة", "مطبخ", "خادمة"]

# كلمات البلكونة
كلمات_بلكونة = ["BALCONY", "TERRACE", "OUTDOOR", "بلكونة", "تراس"]


def نظف_نص_غرفة(نص: str) -> str:
    قيمة = re.sub(r'%%[A-Z]', '', str(نص or '').upper())
    for قديم, جديد in {
        "MASTER BEDROOM": "MASTER BED ROOM",
        "BEDROOM": "BED ROOM",
        "BATHROOM": "BATH ROOM",
        "LIVINGROOM": "LIVING ROOM",
        "DININGROOM": "DINING ROOM",
        "MAIDROOM": "MAID ROOM",
        "DRIVERROOM": "DRIVER ROOM",
        "W.C": "WC",
    }.items():
        قيمة = قيمة.replace(قديم, جديد)
    قيمة = قيمة.replace("&", " ")
    قيمة = re.sub(r'[\{\}\[\]\(\),;:_\-]+', ' ', قيمة)
    قيمة = re.sub(r'\s+', ' ', قيمة).strip()
    return قيمة


def استخرج_نص_طابق_موحد(نص: str) -> str:
    """يحول أسماء الطوابق إلى رموز موحدة."""
    نص = str(نص or "").upper()
    if any(k in نص for k in ["GROUND", "GF"]): return "GF"
    if any(k in نص for k in ["FIRST", "1F"]): return "1F"
    if any(k in نص for k in ["SECOND", "2F"]): return "2F"
    if any(k in نص for k in ["ROOF", "RF"]) and "TOP" not in نص: return "RF"
    if any(k in نص for k in ["TOP ROOF", "PENTHOUSE", "TRF"]): return "TRF"
    if any(k in نص for k in ["BASEMENT", "B1"]): return "B1"
    if any(k in نص for k in ["MEZZANINE", "MZ"]): return "MZ"
    if any(k in نص for k in ["KITCHEN", "SERVICE", "KB"]): return "KB"
    return ""


def استخرج_اسم_غرفة_صالح(نص: str) -> Optional[str]:
    قيمة = نظف_نص_غرفة(نص)
    if not قيمة:
        return None
    if len(قيمة) > 35 or len(قيمة.split()) > 5:
        return None
    if re.match(r'^[\*\d]', قيمة):
        return None
    if any(كلمة in قيمة for كلمة in كلمات_استبعاد_الغرف):
        return None

    حدود = r'A-Z\u0600-\u06FF'
    for اسم_قياسي, عبارات in عبارات_الغرف_القياسية.items():
        for عبارة in sorted(عبارات, key=len, reverse=True):
            نمط = rf'(?<![{حدود}]){re.escape(عبارة)}(?![{حدود}])'
            if re.search(نمط, قيمة):
                return اسم_قياسي
    return None

# رؤوس جداول الجدول الإنشائي
رؤوس_الجداول = {
    "أساس":        ["FOOTING SCHEDULE", "FOOTINGS SCHEDULE", "SCHEDULE OF FOOTING",
                     "SCHEDULE OF FOOTINGS", "FOUNDATION SCHEDULE", "FOUNDATIONS SCHEDULE",
                     "SCHEDULE OF FOUNDATION", "SCHEDULE OF FOUNDATIONS", "FOOTING LAYOUT", "FOUNDATION LAYOUT",
                     "الأساسات", "القواعد", "جدول الأسس", "FOOTING SCH", "FTG SCH"],
    "عمود_عنق":    ["NECK COLUMN", "NECK COL", "NECK COLUMN SCHEDULE", "NECK COL SCHEDULE",
                     "بدن العمود", "عنق العمود", "أعمدة العنق"],
    "كمرة_ربط":    ["TIE BEAM", "TIE-BEAM", "TIEBEAM", "TIE BEAMS", "SCHEDULE OF TIE BEAM",
                     "SCHEDULE OF TIE BEAMS", "TIE BEAMS SCHEDULE", "B.W TIE BEAMS SCHEDULE",
                     "TIE BEAM LAYOUT", "TIE BEAM SCH", "GROUND BEAM", "GROUND BEAMS",
                     "جداول الميدات", "الميدات", "جدول الميدة"],
    "كمرة_ستراب":  ["STRAP BEAM", "STRAP BEAMS", "SCHEDULE OF STRAP BEAM", "SCHEDULE OF STRAP BEAMS",
                     "STB SCHEDULE", "S.B SCH", "STB SCH"],
    "كمرة":        ["BEAM SCHEDULE", "SCHEDULE OF BEAM", "ROOF BEAM", "SLAB BEAMS",
                     "BEAM LAYOUT", "FIRST FLOOR SLAB LAYOUT", "ROOF SLAB LAYOUT",
                     "جسور السقف", "جدول الكمرات"],
    "عمود":        ["COLUMN SCHEDULE", "COLUMNS SCHEDULE", "SCHEDULE OF COLUMN", "SCHEDULE OF COLUMNS",
                     "COLUMN LAYOUT", "GROUND COLUMN LAYOUT", "FIRST FLOOR COLUMN LAYOUT",
                     "COLUMN LAYOUT(FROM FOUNDATION TO FIRST FLOOR)",
                     "COLUMN LAYOUT(FROM FIRST FLOOR TO ROOF FLOOR)",
                     "COL LAYOUT", "C.S", "جدول الأعمدة", "الأعمدة"],
    "فتحة":        ["DOOR SCHEDULE", "WINDOW SCHEDULE", "OPENING SCHEDULE", "FINISH SCHEDULE"],
}

# حدود سماكة الجدران
حد_جدار_20_أدنى, حد_جدار_20_أعلى = 0.15, 0.25
حد_جدار_10_أدنى, حد_جدار_10_أعلى = 0.07, 0.14

# نطاق الربط بين الكمرة وتسميتها (متر)
نطاق_ربط_الكمرة = 15.0


# ─────────────────────────────────────────────────────────────────────────────
# نماذج البيانات
# ─────────────────────────────────────────────────────────────────────────────

class بيانات_الأساس(BaseModel):
    الرمز: str
    الطول: float       # من الجدول
    العرض: float       # من الجدول
    العمق: float       # من الجدول
    الكمية: int = 1
    سماكة_PCC: float = 0.10

class بيانات_عمود_العنق(BaseModel):
    الرمز: str
    الطول: float       # من الجدول
    العرض: float       # من الجدول
    عمق_الأساس: float  # يُحل تلقائياً من جدول الأسس
    الكمية: int = 1

class بيانات_كمرة_ربط(BaseModel):
    الرمز: str
    الطول: float       # ← من الرسم (يُقاس من الخطوط)
    العرض: float       # ← من الجدول
    العمق: float       # ← من الجدول

class بيانات_كمرة_ستراب(BaseModel):
    الرمز: str
    الطول: float       # ← من الرسم
    العرض: float       # ← من الجدول
    العمق: float       # ← من الجدول

class بيانات_كمرة(BaseModel):
    الرمز: str
    الطول: float       # ← من الرسم
    العرض: float       # ← من الجدول
    العمق: float       # ← من الجدول

class بيانات_بلوك_تحت_الأرض(BaseModel):
    الرمز: str = "SB"
    الطول: float
    الكمية: int = 1

class بيانات_العمود(BaseModel):
    الرمز: str
    الطول: float       # من الجدول
    العرض: float       # من الجدول
    الكمية: int = 1
    المستوى: str = ""

class بيانات_فتحة(BaseModel):
    الرمز: str
    الطول: float       # من الجدول
    الارتفاع: float    # من الجدول
    الكمية: int = 1    # من المسقط
    مصدر_الكمية: str = ""
    مصدر_الأبعاد: str = ""
    ملف_الأبعاد: str = ""


def رمز_فتحة_موحد(رمز: str) -> str:
    return re.sub(r'[\s\-_]', '', str(رمز or '').upper())

class بيانات_مناطق_مبللة(BaseModel):
    الدور: str         # "أرضي" أو "أول"
    المساحة: float
    المحيط: float

class ثوابت_المشروع(BaseModel):
    ارتفاع_الدور: float = 0.0
    ارتفاع_الدور_الأرضي: float = 0.0
    ارتفاع_الدور_الأول: float = 0.0
    ارتفاع_الدور_الثاني: float = 0.0
    ارتفاع_دور_السطح: float = 0.0
    ارتفاع_التصوينة: float = 1.0 # متر
    ارتفاع_مبنى_الخدمة: float = 0.0
    عمق_الحفر: float
    مستوى_بلاطة_الأرضي: float = 0.0
    عمق_كمرة_الربط: float = 0.0
    عدد_الأدوار: int = 0
    سماكة_PCC: float = 0.10
    سماكة_البلاطة: float = 0.0
    ارتفاع_صافي_التشطيب_القياسي: float = 0.0
    يوجد_رصيف_طرق: bool = False
    سماكة_رصيف_الطرق: float = 0.0
    خرسانة_الدرج: float = 0.0
    استخراج_الدرج_آليا: bool = True
    نوع_الحفر: str = "صافي"          # صافي / كتلي
    مساحة_حفر_يدوية: float = 0.0     # لو أراد المهندس فرض مساحة الحفر مباشرة
    ارتفاع_بلاط_الجدران_المبللة: float = 2.40
    طول_العتبة_لكل_باب: float = 1.00
    مفتاح_gemini: str = ""   # Gemini API Key لقراءة PDF المسح الضوئي
    استخراج_صارم_فقط: bool = True  # لا يستخدم أي نتائج يدوية/Gemini/افتراضية في الحساب

class طلب_حساب_الكميات(BaseModel):
    مسار_الملف: str = ""           # ملف واحد DXF أو PDF
    مسار_المجلد: str = ""          # مجلد كامل يحتوي مخططات متعددة
    رقم_المشروع: str
    وحدة_القياس: str = "mm"
    الثوابت: ثوابت_المشروع

    الأسس:             List[بيانات_الأساس]          = []
    أعمدة_العنق:       List[بيانات_عمود_العنق]      = []
    كمرات_الربط:       List[بيانات_كمرة_ربط]        = []
    كمرات_الستراب:     List[بيانات_كمرة_ستراب]      = []
    بلوكات_تحت_الأرض:  List[بيانات_بلوك_تحت_الأرض] = []
    الأعمدة:           List[بيانات_العمود]           = []
    الكمرات:           List[بيانات_كمرة]             = []
    الفتحات:           List[بيانات_فتحة]             = []
    مناطق_مبللة_بالدور: List[بيانات_مناطق_مبللة]    = []

    المحيط_الخارجي:              Optional[float] = None
    طول_جدران_بلوك_20_داخلي:    Optional[float] = None
    طول_جدران_بلوك_10_داخلي:    Optional[float] = None


# ─────────────────────────────────────────────────────────────────────────────
# أداة تحويل الأبعاد
# ─────────────────────────────────────────────────────────────────────────────

def إلى_متر(قيمة: str, وحدة_الرسم: str = "mm") -> Optional[float]:
    """
    يحول نصوص الأبعاد إلى متر بذكاء شديد.
    يتعامل مع: '300', '0.30', '300mm', '30cm', '0.3m', '300X400', 'D=300'
    """
    if not قيمة:
        return None
    # تنظيف النص: إزالة المسافات، تحويل للفواصل العشرية، إزالة الحروف غير الرقمية بجانب الرقم
    نص_منظف = str(قيمة).strip().upper().replace(",", ".")
    
    # البحث عن الرقم داخل النص (قد يكون 300MM أو [300])
    تطابق = re.search(r'([-+]?\d*\.\d+|\d+)', نص_منظف)
    if not تطابق:
        return None
        
    رقم = float(تطابق.group())
    if رقم == 0: return None
    
    # تحديد الوحدة بذكاء بناءً على القرائن
    وحدة_مكتشفة = وحدة_الرسم
    if 'MM' in نص_منظف: وحدة_مكتشفة = "mm"
    elif 'CM' in نص_منظف: وحدة_مكتشفة = "cm"
    elif 'M' in نص_منظف and 'MM' not in نص_منظف: وحدة_مكتشفة = "m"
    
    # التحويل لمتر
    # Rule 6: ALWAYS apply the unit conversion factor regardless of value size.
    if وحدة_مكتشفة == "mm":
        return round(float(رقم * 0.001), 4)
    elif وحدة_مكتشفة == "cm":
        return round(float(رقم * 0.01), 4)
    else: # متر
        return round(float(رقم), 4)

def إلى_صحيح(قيمة: str) -> Optional[int]:
    if not قيمة: return None
    # البحث عن أول عدد صحيح
    تطابق = re.search(r'\d+', str(قيمة).strip())
    return int(تطابق.group()) if تطابق else None


def استخراج_أبعاد_من_نص(قيمة: str, عدد_الأبعاد: int, وحدة_الرسم: str = "mm") -> Tuple[Optional[float], ...]:
    """
    يلتقط أبعاداً من نص واحد مثل:
    130x170x40   أو   20 x 50   أو   300X400
    ويحوّلها إلى المتر حسب وحدة الرسم.
    """
    if not قيمة:
        return tuple([None] * عدد_الأبعاد)

    نص = str(قيمة).upper().replace(",", ".")
    if عدد_الأبعاد == 3:
        مطابق_مباشر = re.search(r'(\d+(?:\.\d+)?)\s*[X]\s*(\d+(?:\.\d+)?)\s*[X]\s*(\d+(?:\.\d+)?)', نص)
    else:
        مطابق_مباشر = re.search(r'(\d+(?:\.\d+)?)', نص)

    if مطابق_مباشر:
        مطابقات: List[str] = [str(g) for g in مطابق_مباشر.groups()]
    else:
        مطابقات: List[str] = re.findall(r'(\d+(?:\.\d+)?)', نص)
    أبعاد = []
    # Use explicit loop to avoid slice indexing issues with complex type inference
    مطابقات_قائمة = cast(List[str], مطابقات)
    for i in range(min(len(مطابقات_قائمة), عدد_الأبعاد)):
        رقم = مطابقات_قائمة[i]
        قيمة_رقمية = float(رقم)
        if "MM" in نص:
            وحدة_مستنتجة = "mm"
        elif "CM" in نص:
            وحدة_مستنتجة = "cm"
        elif "M" in نص and "MM" not in نص:
            وحدة_مستنتجة = "m"
        else:
            وحدة_مستنتجة = وحدة_الرسم
            if وحدة_الرسم.lower() == "mm" and 100 <= قيمة_رقمية <= 2000:
                pass # Context confirms mm
            else:
                مسجل.warning(f"Ambiguous unit for value {قيمة_رقمية}, using drawing unit {وحدة_الرسم}")
        بعد = إلى_متر(رقم, وحدة_مستنتجة)
        أبعاد.append(بعد)
    # Pad to requested dimension count to prevent unpacking errors
    while len(أبعاد) < عدد_الأبعاد:
        أبعاد.append(None)
    return tuple(أبعاد)


# ─────────────────────────────────────────────────────────────────────────────
# قارئ جداول الجدول الإنشائي — مُحسَّن
# ─────────────────────────────────────────────────────────────────────────────

class قارئ_الجداول:
    """
    يقرأ جداول الجدول الإنشائي من DXF أو PDF.
    
    الإصلاحات عن V13:
    - يستخدم KDTree لتجميع النصوص (أسرع)
    - يفرق بين الجداول المتجاورة بشكل صحيح
    - يستخرج العرض والعمق لكمرات TB/STB/B (الطول يأتي من الرسم)
    - يحل عمق الأساس لأعمدة العنق تلقائياً
    - V15.1: يدعم PDF عبر تمرير نصوص جاهزة (pdf_texts)
    """

    def __init__(self, msp, مقياس: float, وحدة: str = "mm", pdf_texts: Optional[List[Dict]] = None):
        self.msp    = msp
        self.مقياس  = مقياس
        self.وحدة   = وحدة
        self._pdf_texts = pdf_texts  # If provided, skip msp query and use these texts
        self.كل_النصوص: List[Dict] = []
        self.عناوين_المساقط: List[Dict] = []
        self.سماكات_البلاطات: Dict[str, float] = {}
        self.منسوب_كمرة_الربط: float = 0.0
        self.سماكة_بلاطة_الأرضي: float = 0.0
        self.ارتفاع_الدور_المستخرج: float = 0.0

        # النتائج
        self.الأسس:         List[بيانات_الأساس]     = []
        self.أعمدة_العنق:   List[بيانات_عمود_العنق] = []
        self.جدول_كمرات_الربط:   Dict[str, Dict]    = {}  # رمز → {عرض، عمق}
        self.جدول_كمرات_الستراب: Dict[str, Dict]    = {}  # رمز → {عرض، عمق}
        self.جدول_الكمرات:        Dict[str, Dict]    = {}  # رمز → {عرض، عمق}
        self.الأعمدة:        List[بيانات_العمود]     = []
        self.الفتحات:        List[بيانات_فتحة]       = []

    @staticmethod
    def _رمز_فتحة_من_عنوان_تفصيلي(نص: str) -> str:
        قيمة = " ".join(str(نص or "").upper().split())
        قيمة = قيمة.replace("WINDOES", "WINDOWS")
        مطابق = re.search(r'WINDOWS?\s*#\s*0*(\d+)', قيمة)
        if مطابق:
            return f"W{int(مطابق.group(1))}"
        مطابق = re.search(r'DOORS?\s*#\s*0*(\d+)', قيمة)
        if مطابق:
            return f"D{int(مطابق.group(1))}"
        return ""

    @staticmethod
    def _اختيار_بعد_من_مرشحات(مرشحات: List[float], نوع: str, محور: str) -> float:
        قيم = sorted({
            round(float(ق or 0.0), 3)
            for ق in مرشحات
            if 0.35 <= float(ق or 0.0) <= 5.0
        })
        if not قيم:
            return 0.0

        if نوع == "door":
            if محور == "x":
                مفضلة = [ق for ق in قيم if 0.70 <= ق <= 2.50]
                if مفضلة:
                    return مفضلة[0]
                بديلة = [ق for ق in قيم if 0.45 <= ق <= 3.50]
                return بديلة[0] if بديلة else 0.0
            مفضلة = [ق for ق in قيم if 1.90 <= ق <= 3.20]
            if مفضلة:
                return مفضلة[0]
            بديلة = [ق for ق in قيم if 1.50 <= ق <= 3.80]
            return بديلة[0] if بديلة else 0.0

        if محور == "x":
            مفضلة = [ق for ق in قيم if 0.75 <= ق <= 3.00]
            if مفضلة:
                return مفضلة[0]
            ضيقة = [ق for ق in قيم if 0.45 <= ق < 0.75]
            return ضيقة[-1] if ضيقة else 0.0

        مفضلة = [ق for ق in قيم if 0.80 <= ق <= 2.40]
        if مفضلة:
            return مفضلة[0]
        بديلة = [ق for ق in قيم if 2.40 < ق <= 3.50]
        return بديلة[0] if بديلة else 0.0

    def _قياسات_الأبعاد_داخل_صندوق(self, xmin: float, xmax: float, ymin: float, ymax: float) -> Tuple[List[float], List[float]]:
        أفقية: List[float] = []
        رأسية: List[float] = []
        if self.msp is None:
            return أفقية, رأسية
        for كيان in self.msp.query('DIMENSION'):
            try:
                نقطة = getattr(كيان.dxf, "text_midpoint", None) or كيان.dxf.defpoint
                x = float(نقطة.x) * self.مقياس
                y = float(نقطة.y) * self.مقياس
                if not (xmin <= x <= xmax and ymin <= y <= ymax):
                    continue
                قياس = float(كيان.get_measurement())
                if not (0.35 <= قياس <= 5.0):
                    continue
                زاوية = float(getattr(كيان.dxf, "angle", 0.0) or 0.0)
                if abs(زاوية - 90.0) <= 0.5:
                    رأسية.append(قياس)
                else:
                    أفقية.append(قياس)
            except Exception:
                continue
        return أفقية, رأسية

    def _استخرج_الفتحات_من_تفاصيل_الأبواب_والشبابيك(self):
        عناوين = []
        for نص in self.كل_النصوص:
            رمز = self._رمز_فتحة_من_عنوان_تفصيلي(نص.get("نص", ""))
            if رمز:
                عناوين.append({
                    "الرمز": رمز,
                    "x": float(نص["x"]),
                    "y": float(نص["y"]),
                    "النص": نص["نص"],
                })

        if not عناوين:
            return

        صفوف: List[Dict[str, Any]] = []
        for عنوان in sorted(عناوين, key=lambda ع: (-ع["y"], ع["x"])):
            أضيف = False
            for صف in صفوف:
                if abs(صف["y"] - عنوان["y"]) <= 1.0:
                    صف["عناصر"].append(عنوان)
                    أضيف = True
                    break
            if not أضيف:
                صفوف.append({"y": عنوان["y"], "عناصر": [عنوان]})

        موجود = {رمز_فتحة_موحد(ف.الرمز) for ف in self.الفتحات if ف.الطول > 0 and ف.الارتفاع > 0}
        نتائج_تفصيلية: Dict[str, بيانات_فتحة] = {}

        for صف in صفوف:
            عناصر = sorted(صف["عناصر"], key=lambda ع: ع["x"])
            if not عناصر:
                continue
            xs = [ع["x"] for ع in عناصر]

            for idx, عنوان in enumerate(عناصر):
                رمز = رمز_فتحة_موحد(عنوان["الرمز"])
                if رمز in موجود:
                    continue

                # صندوقان مختلفان لأن موضع عنوان التفصيلة ليس ثابتًا دائمًا
                يسار_مجاور = xs[idx - 1] if idx > 0 else عنوان["x"] - 3.0
                يمين_مجاور = xs[idx + 1] if idx + 1 < len(xs) else عنوان["x"] + 4.5
                صناديق = [
                    (
                        عنوان["x"] - 0.5,
                        (يمين_مجاور - 0.5) if idx + 1 < len(xs) else عنوان["x"] + 4.5,
                        عنوان["y"] - 1.5,
                        عنوان["y"] + 20.0,
                    ),
                    (
                        ((يسار_مجاور + عنوان["x"]) / 2.0) if idx > 0 else عنوان["x"] - 2.0,
                        ((عنوان["x"] + يمين_مجاور) / 2.0) if idx + 1 < len(xs) else عنوان["x"] + 2.5,
                        عنوان["y"] - 8.0,
                        عنوان["y"] + 10.0,
                    ),
                ]

                أفقية: List[float] = []
                رأسية: List[float] = []
                for xmin, xmax, ymin, ymax in صناديق:
                    أف, رأ = self._قياسات_الأبعاد_داخل_صندوق(xmin, xmax, ymin, ymax)
                    أفقية.extend(أف)
                    رأسية.extend(رأ)

                نوع = "door" if رمز.startswith("D") else "window"
                طول = self._اختيار_بعد_من_مرشحات(أفقية, نوع, "x")
                ارتفاع = self._اختيار_بعد_من_مرشحات(رأسية, نوع, "y")
                if طول > 0 and ارتفاع > 0:
                    نتائج_تفصيلية[رمز] = بيانات_فتحة(
                        الرمز=عنوان["الرمز"],
                        الطول=طول,
                        الارتفاع=ارتفاع,
                        الكمية=0,
                        مصدر_الأبعاد="DETAIL",
                    )

        if نتائج_تفصيلية:
            for فتحة in sorted(نتائج_تفصيلية.values(), key=lambda ف: رمز_فتحة_موحد(ف.الرمز)):
                if رمز_فتحة_موحد(فتحة.الرمز) not in موجود:
                    self.الفتحات.append(فتحة)
            مسجل.info(f"تم استخراج {len(نتائج_تفصيلية)} فتحة من تفاصيل الأبواب/الشبابيك")

    @staticmethod
    def _نص_موحد(نص: str) -> str:
        return " ".join(str(نص).upper().split())

    def _استخراج_النصوص(self):
        # V15.1: If PDF texts were pre-extracted, use them directly
        if self._pdf_texts:
            for نص_جاهز in self._pdf_texts:
                قيمة = str(نص_جاهز.get("نص", "")).strip()
                if not قيمة:
                    continue
                كبير = self._نص_موحد(قيمة)
                self.كل_النصوص.append({
                    "نص": قيمة,
                    "x": float(نص_جاهز.get("x", 0)),
                    "y": float(نص_جاهز.get("y", 0)),
                    "كبير": كبير,
                    "الطبقة": نص_جاهز.get("الطبقة", "PDF"),
                })
                if "LAYOUT" in كبير:
                    self.عناوين_المساقط.append({
                        "نص": قيمة,
                        "كبير": كبير,
                        "x": float(نص_جاهز.get("x", 0)),
                        "y": float(نص_جاهز.get("y", 0)),
                    })
            مسجل.info(f"قارئ الجداول (PDF): {len(self.كل_النصوص)} نص")
            return

        for t in self.msp.query('TEXT MTEXT'):
            try:
                if t.dxftype() == 'TEXT':
                    قيمة = t.dxf.text.strip()
                else:
                    قيمة = t.plain_text().strip()
                
                # تنظيف أكواد أوتوكاد (مثل %%U و %%D)
                قيمة = re.sub(r'%%[pPduDCU]', '', قيمة)
                قيمة = قيمة.replace('%%U', '').replace('%%u', '').strip()
                
                x = t.dxf.insert.x * self.مقياس
                y = t.dxf.insert.y * self.مقياس
                if قيمة:
                    كبير = self._نص_موحد(قيمة)
                    self.كل_النصوص.append({
                        "نص": قيمة, "x": x, "y": y,
                        "كبير": كبير,
                        "الطبقة": t.dxf.layer.upper(),
                    })
                    if "LAYOUT" in كبير:
                        self.عناوين_المساقط.append({
                            "نص": قيمة,
                            "كبير": كبير,
                            "x": x,
                            "y": y,
                        })
            except Exception:
                pass
        مسجل.info(f"قارئ الجداول: {len(self.كل_النصوص)} نص")

    def _استخرج_السماكات_والمناسيب(self):
        مستويات_صريحة: List[float] = []
        مستويات_lvl: List[float] = []
        مستويات_ثانوية: List[float] = []
        مستويات_top: List[float] = []

        def _التقط_مناسيب(نص_كبير: str) -> List[float]:
            نتائج: List[float] = []
            for مطابق in re.finditer(r'(?:\(\s*([+-])\s*\)|([+-]))\s*(\d+(?:\.\d+)?)\s*M', نص_كبير):
                إشارة = مطابق.group(1) or مطابق.group(2) or "+"
                قيمة = float(مطابق.group(3))
                if إشارة == "-":
                    قيمة *= -1.0
                نتائج.append(قيمة)
            for مطابق in re.finditer(r'(\d+(?:\.\d+)?)\s*LVL', نص_كبير):
                نتائج.append(float(مطابق.group(1)))
            return نتائج

        def _اختر_فرق_دور(فروقات: List[float]) -> float:
            فروقات_موجبة = sorted({round(float(ف), 2) for ف in فروقات if ف > 0})
            for حد_أدنى, حد_أعلى in ((3.6, 4.4), (3.4, 4.6)):
                مرشحة = [ف for ف in فروقات_موجبة if حد_أدنى <= ف <= حد_أعلى]
                if مرشحة:
                    return round(min(مرشحة), 3)
            return 0.0

        for عنوان in self.عناوين_المساقط:
            نص_عنوان = self._نص_موحد(عنوان["نص"])
            if "SLAB LAYOUT" not in نص_عنوان:
                continue
            مرشحات = [
                ت for ت in self.كل_النصوص
                if abs(ت["x"] - عنوان["x"]) <= 18.0
                and عنوان["y"] <= ت["y"] <= (عنوان["y"] + 26.0)
                and ("SLAB THICKNESS" in ت["كبير"] or "THK" in ت["كبير"])
            ]
            for نص in مرشحات:
                مطابق = re.search(r'(\d+(?:\.\d+)?)\s*CM', نص["كبير"])
                if not مطابق:
                    مطابق = re.search(r'(\d+(?:\.\d+)?)\s*THK', نص["كبير"])
                if مطابق:
                    قيمة = float(مطابق.group(1))
                    self.سماكات_البلاطات[نص_عنوان] = round((قيمة * 0.01) if قيمة > 1.0 else قيمة, 3)
                    break

        for نص in self.كل_النصوص:
            كبير = نص["كبير"]
            if "GRADE SLAB" in كبير and "THK" in كبير:
                مطابق = re.search(r'GRADE SLAB\s+(\d+(?:\.\d+)?)\s*THK', كبير)
                if مطابق:
                    قيمة = float(مطابق.group(1))
                    self.سماكة_بلاطة_الأرضي = round((قيمة * 0.01) if قيمة > 1.0 else قيمة, 3)
            if "TIE BEAM" in كبير and "LEVEL" in كبير:
                مطابق = re.search(r'LEVEL\s*\+?(-?\d+(?:\.\d+)?)', كبير)
                if مطابق:
                    self.منسوب_كمرة_الربط = max(self.منسوب_كمرة_الربط, float(مطابق.group(1)))

            for قيمة_مستوى in _التقط_مناسيب(كبير):
                if قيمة_مستوى < 0.0:
                    continue
                if "TIE BEAM" in كبير:
                    continue
                مستويات_صريحة.append(قيمة_مستوى)
                if "LVL(" in كبير and "TOP LVL" not in كبير:
                    مستويات_lvl.append(قيمة_مستوى)
                elif "TOP LVL" in كبير:
                    مستويات_top.append(قيمة_مستوى)
                elif "TOP OF SLAB" in كبير:
                    مستويات_ثانوية.append(قيمة_مستوى)

        مستويات_lvl_مرتبة = sorted(set(round(ق, 2) for ق in مستويات_lvl if ق >= 0.0))
        فروقات_lvl = [
            round(مستويات_lvl_مرتبة[i + 1] - مستويات_lvl_مرتبة[i], 2)
            for i in range(len(مستويات_lvl_مرتبة) - 1)
        ]
        مستويات_ثانوية_مرتبة = sorted(set(round(ق, 2) for ق in مستويات_ثانوية if ق >= 0.0))
        فروقات_ثانوية = [
            round(مستويات_ثانوية_مرتبة[i + 1] - مستويات_ثانوية_مرتبة[i], 2)
            for i in range(len(مستويات_ثانوية_مرتبة) - 1)
        ]

        مستويات_top_مرتبة = sorted(set(round(ق, 2) for ق in مستويات_top if ق >= 0.0))
        فروقات_top = [
            round(مستويات_top_مرتبة[i + 1] - مستويات_top_مرتبة[i], 2)
            for i in range(len(مستويات_top_مرتبة) - 1)
        ]

        مستويات_مرتبة = sorted(set(round(ق, 2) for ق in مستويات_صريحة if ق >= 0.0))
        فروقات = [round(مستويات_مرتبة[i + 1] - مستويات_مرتبة[i], 2) for i in range(len(مستويات_مرتبة) - 1)]
        مرشحات_ارتفاع = [
            _اختر_فرق_دور(فروقات_lvl),
            _اختر_فرق_دور(فروقات_ثانوية),
            _اختر_فرق_دور(فروقات),
        ]
        مرشحات_ارتفاع = [ف for ف in مرشحات_ارتفاع if ف > 0]
        if مرشحات_ارتفاع:
            self.ارتفاع_الدور_المستخرج = round(min(مرشحات_ارتفاع), 3)

    def _تجميع_النصوص_في_جداول(self) -> List[List[Dict]]:
        """
        يستخدم KDTree للتجميع — أسرع من O(n²).
        كل جدول هو مجموعة نصوص ضمن مسافة 2 متر من بعضها.
        """
        if not self.كل_النصوص:
            return []

        نقاط = np.array([[t["x"], t["y"]] for t in self.كل_النصوص])
        شجرة = KDTree(نقاط)

        # V15.4: PDF schedule reader coords are pre-scaled to DXF-like range
        # (villa-meters: 20-30m). Use 2.0m to bridge split text block rows.
        # For DXF, 0.6m is sufficient.
        if self._pdf_texts:
            نصف_قطر_التجميع = 2.0
        else:
            نصف_قطر_التجميع = 0.6
        مجموعات_الجيران = شجرة.query_ball_point(نقاط, r=نصف_قطر_التجميع)

        # Union-Find لتجميع النصوص المترابطة
        أب = list(range(len(self.كل_النصوص)))

        def ابحث(x):
            while أب[x] != x:
                أب[x] = أب[أب[x]]
                x = أب[x]
            return x

        def ادمج(x, y):
            أب[ابحث(x)] = ابحث(y)

        for i, جيران in enumerate(مجموعات_الجيران):
            for j in جيران:
                ادمج(i, j)

        # تجميع حسب الأب
        مجموعات: Dict[int, List[Dict]] = defaultdict(list)
        for i, t in enumerate(self.كل_النصوص):
            مجموعات[ابحث(i)].append(t)

        # فلترة: جدول حقيقي يحتاج على الأقل 4 نصوص
        الجداول = [م for م in مجموعات.values() if len(م) >= 4]
        مسجل.info(f"اكتُشف {len(الجداول)} جدول محتمل")
        return الجداول

    @staticmethod
    def _تحديد_نوع(مجموعة: List[Dict]) -> Optional[str]:
        كل_النص = " ".join(t["كبير"] for t in مجموعة)
        # Deep Logging to file to avoid terminal garbling
        with open("potential_tables.txt", "a", encoding="utf-8", errors="backslashreplace") as f:
            f.write(f"\n--- TABLE ({len(مجموعة)} texts) ---\n")
            f.write(_safe_utf8_text(كل_النص[:1000]) + "\n")

        # 1. Try explicit header keywords first
        for نوع, كلمات in رؤوس_الجداول.items():
            if any(ك in كل_النص for ك in كلمات):
                return نوع

        if re.search(r'(?:SCHEDULE\s+OF\s+FOOT|FOOT\w*\s+SCHEDULE|FOUNDATION\s+LAYOUT)', كل_النص):
            return "أساس"
        if re.search(r'(?:SCHEDULE\s+OF\s+TIE\s*BEAMS?|TIE\s*BEAMS?\s+SCHEDULE|B\.W\s+TIE\s+BEAMS\s+SCHEDULE)', كل_النص):
            return "كمرة_ربط"
        if re.search(r'(?:SCHEDULE\s+OF\s+STRAP\s*BEAMS?|STRAP\s*BEAMS?\s+SCHEDULE)', كل_النص):
            return "كمرة_ستراب"
        if re.search(r'(?:SCHEDULE\s+OF\s+COLUMNS?|COLUMN\s+LAYOUT)', كل_النص):
            return "عمود"

        # 2. Heuristic: small groups (4-15 texts) containing a column mark + two dimension values
        #    → treat as inline column dimension annotation (e.g. "C3A 0.60 0.20")
        نصوص = [t["كبير"] for t in مجموعة]
        if 4 <= len(مجموعة) <= 120:
            رموز_أعمدة = [n for n in نصوص if re.search(r'(^|[^A-Z0-9])(C\d+[A-Z]?|C\d+(?:/[A-Z0-9]+)?|DC\d*|NC\d*)(?=$|[^A-Z0-9])', n.strip())]
            أرقام = [n for n in نصوص if re.match(r'^\d+\.?\d*$', n.strip()) and 0.1 <= float(n) <= 2000.0]
            if رموز_أعمدة and len(أرقام) >= 2 and (
                len(مجموعة) <= 24 or
                any(k in كل_النص for k in ["COLUMN", "COL", "TYPE", "SIZE", "REINFORCEMENT", "STIRRUPS", "LONG"])
            ):
                return "عمود"

        # 3. Heuristic: footing mark (F\d+) + dimension numbers
        if 4 <= len(مجموعة) <= 160:
            رموز_أسس = [n for n in نصوص if re.search(r'(^|[^A-Z0-9])((?:CF|F|WF|FTG)[\-_]?\d+[A-Z]?)(?=$|[^A-Z0-9])', n.strip())]
            أرقام = [n for n in نصوص if re.match(r'^\d+\.?\d*$', n.strip()) and 0.2 <= float(n.replace(',','')) <= 2000.0]
            if رموز_أسس and len(أرقام) >= 2 and (
                len(مجموعة) <= 36 or
                any(k in كل_النص for k in ["FOOTING", "FOUNDATION", "SHORT", "LONG", "WIDTH", "THICK", "BOTT", "TOP STEEL"])
            ):
                return "أساس"

        # 4. Heuristic: TB/STB mark + dimension numbers → tie beam
        if 4 <= len(مجموعة) <= 160:
            رموز_كمرات = [n for n in نصوص if re.search(r'(^|[^A-Z0-9])((?:CTB|STB|TB|GB|B)[\-_]?\d+[A-Z]?)(?=$|[^A-Z0-9])', n.strip())]
            أرقام = [n for n in نصوص if re.match(r'^\d+\.?\d*$', n.strip()) and 0.1 <= float(n) <= 2000.0]
            if رموز_كمرات and len(أرقام) >= 2 and (
                len(مجموعة) <= 36 or
                any(k in كل_النص for k in ["BEAM", "DEPTH", "WIDTH", "TOP STEEL", "BOTTOM STEEL", "STIRRUPS", "SIZE"])
            ):
                if any('STB' in n for n in رموز_كمرات):
                    return "كمرة_ستراب"
                return "كمرة_ربط"

        return None

    @staticmethod
    def _تجميع_بصفوف(مجموعة: List[Dict], تسامح: float = 0.08) -> List[List[Dict]]:
        # V15.4: PDF coords now DXF-scaled. Use 0.30m tolerance to handle
        # split text blocks that may have slight Y offsets within the same row.
        pdf_mode = any(t.get("الطبقة") == "PDF" for t in مجموعة[:20])
        if pdf_mode:
            تسامح = 0.30
        مرتب = sorted(مجموعة, key=lambda t: -t["y"])
        صفوف, صف_حالي = [], [مرتب[0]]
        for t in مرتب[1:]:
            if abs(t["y"] - صف_حالي[-1]["y"]) <= تسامح:
                صف_حالي.append(t)
            else:
                صفوف.append(sorted(صف_حالي, key=lambda x: x["x"]))
                صف_حالي = [t]
        صفوف.append(sorted(صف_حالي, key=lambda x: x["x"]))
        return صفوف

    @staticmethod
    def _ربط_الأعمدة(صف_الرأس: List[Dict]) -> Dict[int, str]:
        ربط = {}
        for i, خلية in enumerate(صف_الرأس):
            نص = خلية["كبير"]
            # Flexible mapping for various structural standards
            if any(k in نص for k in ["MARK", "TYPE", "SYM", "REF", "NO.", "رمز", "نوع", "بند", "COL TYPE", "FOOTING"]):
                ربط[i] = "رمز"
            elif "SIZE" in نص or "(B X D)" in نص or "(b x d)" in خلية["نص"]:
                ربط[i] = "أبعاد"
            elif any(k in نص for k in ["LENGTH", "SPAN", "طول", "L (MM)", "L (M)", "L "]):
                ربط[i] = "طول"
            elif any(k in نص for k in ["WIDTH", "BREADTH", "عرض", "W (MM)", "B (MM)", "W (M)", "B "]):
                ربط[i] = "عرض"
            elif any(k in نص for k in ["DEPTH", "HEIGHT", "THICK", "THK", "عمق", "ارتفاع", "سماكة", "D (MM)", "D (M)", "D "]):
                ربط[i] = "عمق"
            elif any(k in نص for k in ["QTY", "NOS", "COUNT", "NUM", "كمية", "عدد", "NUMBER"]):
                ربط[i] = "كمية"
        return ربط

    def _عناصر_حول_الرأس(self, الرأس: Dict, امتداد_يسار: float, امتداد_يمين: float,
                      امتداد_أسفل: float, امتداد_أعلى: float = 1.50) -> List[Dict]:
        return [
            t for t in self.كل_النصوص
            if (الرأس["x"] - امتداد_يسار) <= t["x"] <= (الرأس["x"] + امتداد_يمين)
            and (الرأس["y"] - امتداد_أسفل) <= t["y"] <= (الرأس["y"] + امتداد_أعلى)
        ]

    @staticmethod
    def _داخل_صندوق(x: float, y: float, صندوق: Dict) -> bool:
        return صندوق["xmin"] <= x <= صندوق["xmax"] and صندوق["ymin"] <= y <= صندوق["ymax"]

    def _صندوق_حول_عنوان(self, عبارة: str, هامش_يسار_x: float = 48.0, هامش_يمين_x: float = 12.0,
                      صعود_y: float = 32.0, نزول_y: float = -8.0) -> Optional[Dict]:
        مطابقات = [ع for ع in self.عناوين_المساقط if عبارة in ع["كبير"]]
        if not مطابقات:
            return None
        عنوان = sorted(مطابقات, key=lambda ع: (-ع["y"], ع["x"]))[0]
        return {
            "العنوان": عنوان["نص"],
            "x_title": عنوان["x"],
            "xmin": عنوان["x"] - هامش_يسار_x,
            "xmax": عنوان["x"] + هامش_يمين_x,
            "ymin": عنوان["y"] + نزول_y,
            "ymax": عنوان["y"] + صعود_y,
        }

    def _عد_الأسس_من_الهندسة(self):
        if not self.الأسس:
            return

        صندوق = self._صندوق_حول_عنوان("FOOTING LAYOUT")
        if not صندوق:
            return

        مساحات_الجدول = {
            أ.الرمز.upper(): round(أ.الطول * أ.العرض, 3)
            for أ in self.الأسس
            if أ.الطول and أ.العرض
        }
        if not مساحات_الجدول:
            return

        عدادات_هندسية = Counter()
        if self.msp is not None:
          for كيان in self.msp.query('LWPOLYLINE'):
            try:
                طبقة = كيان.dxf.layer.upper()
                if not getattr(كيان, "closed", False):
                    continue
                if not any(ك in طبقة for ك in ["S-BEAM", "FOOTING", "FOUND"]):
                    continue
                نقاط = [(ن[0] * self.مقياس, ن[1] * self.مقياس) for ن in كيان.get_points(format='xy')]
                if len(نقاط) < 3:
                    continue
                xs = [ن[0] for ن in نقاط]
                ys = [ن[1] for ن in نقاط]
                مركز_x = (min(xs) + max(xs)) / 2.0
                مركز_y = (min(ys) + max(ys)) / 2.0
                if not self._داخل_صندوق(مركز_x, مركز_y, صندوق):
                    continue
                مساحة = abs(sum(
                    نقاط[i][0] * نقاط[(i + 1) % len(نقاط)][1] -
                    نقاط[(i + 1) % len(نقاط)][0] * نقاط[i][1]
                    for i in range(len(نقاط))
                ) / 2.0)
                if مساحة < 0.50:
                    continue
                رمز, مساحة_مرجعية = min(
                    مساحات_الجدول.items(),
                    key=lambda زوج: abs(زوج[1] - مساحة)
                )
                سماحية = max(0.20, مساحة_مرجعية * 0.18)
                if abs(مساحة_مرجعية - مساحة) <= سماحية:
                    عدادات_هندسية[رمز] += 1
            except Exception:
                pass

        عدادات_نصية_هيكلية = Counter()
        عدادات_نصية_عامة = Counter()
        for نص in self.كل_النصوص:
            if not self._داخل_صندوق(نص["x"], نص["y"], صندوق):
                continue
            مطابق = re.fullmatch(r'(?:CF|F)\d+\*?', re.sub(r'[\s\-_]+', '', نص["كبير"]))
            if not مطابق:
                continue
            رمز = مطابق.group(0).replace("*", "")
            if نص.get("الطبقة", "").upper() != "TEXT":
                عدادات_نصية_هيكلية[رمز] += 1
            elif نص["x"] <= (صندوق.get("x_title", نص["x"]) - 8.0):
                عدادات_نصية_عامة[رمز] += 1

        if عدادات_نصية_هيكلية:
            عدادات = عدادات_نصية_هيكلية
        elif sum(عدادات_نصية_عامة.values()) > sum(عدادات_هندسية.values()):
            عدادات = عدادات_نصية_عامة
        else:
            عدادات = عدادات_هندسية

        if not عدادات:
            return

        for أساس in self.الأسس:
            if أساس.الرمز.upper() in عدادات:
                أساس.الكمية = عدادات[أساس.الرمز.upper()]

    def _خرائط_صفوف_الجدول(self, مجموعة: List[Dict], تسامح: float = 0.18) -> List[Dict[str, str]]:
        if not مجموعة:
            return []

        صفوف = self._تجميع_بصفوف(مجموعة, تسامح=تسامح)
        if len(صفوف) < 2:
            return []

        مرشحون = صفوف[: min(6, len(صفوف))]
        فهرس_الرأس = 0
        أكثر_ربط = 0
        for idx, صف in enumerate(مرشحون):
            عدد = len(self._ربط_الأعمدة(صف))
            if عدد > أكثر_ربط:
                أكثر_ربط = عدد
                فهرس_الرأس = idx

        صف_الرأس = صفوف[فهرس_الرأس]
        ربط_الأعمدة = self._ربط_الأعمدة(صف_الرأس)
        if not ربط_الأعمدة:
            return []

        مراسي = []
        for idx, اسم in ربط_الأعمدة.items():
            if idx < len(صف_الرأس):
                مراسي.append((صف_الرأس[idx]["x"], اسم))

        خرائط = []
        سماحية_x = 8.0 if self._pdf_texts else 4.0
        for صف in صفوف[فهرس_الرأس + 1:]:
            س = defaultdict(list)
            for خلية in صف:
                أقرب_x, أقرب_اسم = min(مراسي, key=lambda م: abs(خلية["x"] - م[0]))
            if abs(خلية["x"] - أقرب_x) <= سماحية_x:
                    س[أقرب_اسم].append(خلية["نص"])
            if س:
                خرائط.append({اسم: " ".join(قيم).strip() for اسم, قيم in س.items()})
        return خرائط

    def _استخرج_رمز_من_صف(self, صف: Dict[str, str], نمط: str) -> str:
        for قيمة in صف.values():
            مطابق = re.search(نمط, str(قيمة).upper())
            if مطابق:
                return مطابق.group(0).strip()
        return ""

    def _حوّل_بعد_هيكلي_إلى_متر(self, قيمة: str, حد_أدنى: float, حد_أعلى: float) -> Optional[float]:
        if not قيمة:
            return None
        نص = str(قيمة).strip().upper().replace(",", ".")
        مطابق = re.search(r'([-+]?\d*\.\d+|\d+)', نص)
        if not مطابق:
            return None
        رقم = float(مطابق.group(1))
        if رقم <= 0:
            return None

        مرشحون: List[float] = []
        if "MM" in نص:
            مرشحون.append(round(رقم * 0.001, 4))
        elif "CM" in نص:
            مرشحون.append(round(رقم * 0.01, 4))
        elif re.search(r'(?<!M)\bM\b', نص):
            مرشحون.append(round(رقم, 4))
        else:
            مرشحون.extend([
                round(رقم, 4),
                round(رقم * 0.01, 4),
                round(رقم * 0.001, 4),
            ])

        for مرشح in مرشحون:
            if حد_أدنى <= مرشح <= حد_أعلى:
                return مرشح
        return None

    def _أبعاد_من_نصوص_متتابعة(self, نصوص: List[str], حدود: List[Tuple[float, float]]) -> Optional[Tuple[float, ...]]:
        if len(نصوص) < len(حدود):
            return None
        أبعاد: List[float] = []
        for قيمة, (حد_أدنى, حد_أعلى) in zip(نصوص, حدود):
            بعد = self._حوّل_بعد_هيكلي_إلى_متر(قيمة, حد_أدنى, حد_أعلى)
            if بعد is None:
                return None
            أبعاد.append(بعد)
        return tuple(أبعاد)

    @staticmethod
    def _التقط_قيم_صريحة(نص: str) -> List[str]:
        return re.findall(r'(?<![A-Z0-9@])\d+(?:\.\d+)?(?![A-Z0-9])', نص.upper())

    def _استخرج_أبعاد_مرشحة(self, قيم: List[str], حدود: List[Tuple[float, float]], من_النهاية: bool = False) -> Optional[Tuple[float, ...]]:
        if not قيم:
            return None
        if من_النهاية:
            حدود_العمل = list(reversed(حدود))
            قيم_العمل = list(reversed(قيم))
        else:
            حدود_العمل = حدود
            قيم_العمل = قيم

        أبعاد: List[float] = []
        فهرس_القيمة = 0
        for حد_أدنى, حد_أعلى in حدود_العمل:
            بعد = None
            while فهرس_القيمة < len(قيم_العمل):
                قيمة = قيم_العمل[فهرس_القيمة]
                فهرس_القيمة += 1
                مرشح = self._حوّل_بعد_هيكلي_إلى_متر(قيمة, حد_أدنى, حد_أعلى)
                if مرشح is not None:
                    بعد = مرشح
                    break
            if بعد is None:
                return None
            أبعاد.append(بعد)

        if من_النهاية:
            أبعاد.reverse()
        return tuple(أبعاد)

    def _استخرج_أبعاد_حول_رمز(self, نص: str, بداية: int, نهاية: int, حدود: List[Tuple[float, float]], نافذة: int = 120) -> Optional[Tuple[float, ...]]:
        بعد = self._استخرج_أبعاد_مرشحة(
            self._التقط_قيم_صريحة(نص[نهاية: نهاية + نافذة]),
            حدود,
            من_النهاية=False,
        )
        if بعد:
            return بعد

        قبل = self._استخرج_أبعاد_مرشحة(
            self._التقط_قيم_صريحة(نص[max(0, بداية - نافذة): بداية]),
            حدود,
            من_النهاية=True,
        )
        if قبل:
            return قبل
        return None

    @staticmethod
    def _نص_كتلة_مرتبة(مجموعة: List[Dict]) -> str:
        return " ".join(
            t["كبير"]
            for t in sorted(مجموعة, key=lambda س: (-س["y"], س["x"]))
        )

    def _سجل_أساس_مباشر(self, عنصر: بيانات_الأساس):
        for idx, قائم in enumerate(self.الأسس):
            if قائم.الرمز.upper() != عنصر.الرمز.upper():
                continue
            if (عنصر.الطول * عنصر.العرض) > (قائم.الطول * قائم.العرض):
                self.الأسس[idx] = عنصر
            elif عنصر.الكمية > قائم.الكمية:
                قائم.الكمية = عنصر.الكمية
            return
        self.الأسس.append(عنصر)

    def _سجل_عمود_مباشر(self, عنصر: بيانات_العمود):
        for idx, قائم in enumerate(self.الأعمدة):
            if قائم.الرمز.upper() != عنصر.الرمز.upper():
                continue
            if (عنصر.الطول * عنصر.العرض) > (قائم.الطول * قائم.العرض):
                self.الأعمدة[idx] = عنصر
            elif عنصر.الكمية > قائم.الكمية:
                قائم.الكمية = عنصر.الكمية
            return
        self.الأعمدة.append(عنصر)

    @staticmethod
    def _سجل_كمرة_مباشر(جدول: Dict[str, Dict], رمز: str, عرض: float, عمق: float):
        قائم = جدول.get(رمز.upper())
        if قائم is None or (عرض * عمق) > (قائم["عرض"] * قائم["عمق"]):
            جدول[رمز.upper()] = {"عرض": عرض, "عمق": عمق}

    def _تحليل_مباشر_من_كتلة_نصية(self, مجموعة: List[Dict], نوع: str):
        نص = self._نص_كتلة_مرتبة(مجموعة)

        if نوع == "أساس":
            for مطابق in re.finditer(r'(?:CF|F|WF|FTG)[\-_]?\d+[A-Z]?\*?', نص):
                رمز = مطابق.group(0).replace("*", "").strip()
                أبعاد = self._استخرج_أبعاد_حول_رمز(نص, مطابق.start(), مطابق.end(), [(0.4, 20.0), (0.4, 20.0), (0.1, 3.0)])
                if not أبعاد:
                    continue
                self._سجل_أساس_مباشر(بيانات_الأساس(
                    الرمز=رمز,
                    الطول=أبعاد[0],
                    العرض=أبعاد[1],
                    العمق=أبعاد[2],
                    الكمية=1,
                ))

            for نمط, عكس in [
                (r'((?:CF|F|WF|FTG)[\-_]?\d+[A-Z]?\*?)((?:\s+(?:\d+(?:\.\d+)?)){3})', False),
                (r'((?:\d+(?:\.\d+)?\s+){3})((?:CF|F|WF|FTG)[\-_]?\d+[A-Z]?\*?)', True),
            ]:
                for مطابق in re.finditer(نمط, نص):
                    if عكس:
                        رمز = مطابق.group(2).replace("*", "").strip()
                        أرقام = re.findall(r'\d+(?:\.\d+)?', مطابق.group(1))[-3:]
                    else:
                        رمز = مطابق.group(1).replace("*", "").strip()
                        أرقام = re.findall(r'\d+(?:\.\d+)?', مطابق.group(2))[:3]
                    أبعاد = self._أبعاد_من_نصوص_متتابعة(أرقام, [(0.4, 20.0), (0.4, 20.0), (0.1, 3.0)])
                    if not أبعاد:
                        continue
                    self._سجل_أساس_مباشر(بيانات_الأساس(
                        الرمز=رمز,
                        الطول=أبعاد[0],
                        العرض=أبعاد[1],
                        العمق=أبعاد[2],
                        الكمية=1,
                    ))

        elif نوع in {"كمرة_ربط", "كمرة_ستراب", "كمرة"}:
            نمط_رمز = r'(?:CTB|STB|TB|GB|B)[\-_]?\d+[A-Z]?'
            for مطابق in re.finditer(نمط_رمز, نص):
                رمز = مطابق.group(0).strip()
                if رمز.upper() == "B":
                    continue
                أبعاد = self._استخرج_أبعاد_حول_رمز(نص, مطابق.start(), مطابق.end(), [(0.1, 2.0), (0.1, 2.0)])
                if not أبعاد:
                    continue
                if "STB" in رمز.upper():
                    self._سجل_كمرة_مباشر(self.جدول_كمرات_الستراب, رمز, أبعاد[0], أبعاد[1])
                elif any(بادئة in رمز.upper() for بادئة in ["TB", "CTB", "GB"]):
                    self._سجل_كمرة_مباشر(self.جدول_كمرات_الربط, رمز, أبعاد[0], أبعاد[1])
                else:
                    self._سجل_كمرة_مباشر(self.جدول_الكمرات, رمز, أبعاد[0], أبعاد[1])

            for مطابق in re.finditer(rf'((?:B[-_])?({نمط_رمز}))\s*\(?\s*(\d+(?:\.\d+)?)\s*[Xx]\s*(\d+(?:\.\d+)?)\s*\)?', نص):
                رمز = مطابق.group(2).strip()
                عرض = self._حوّل_بعد_هيكلي_إلى_متر(مطابق.group(3), 0.1, 2.0)
                عمق = self._حوّل_بعد_هيكلي_إلى_متر(مطابق.group(4), 0.1, 2.0)
                if not عرض or not عمق:
                    continue
                if "STB" in رمز.upper():
                    self._سجل_كمرة_مباشر(self.جدول_كمرات_الستراب, رمز, عرض, عمق)
                elif any(بادئة in رمز.upper() for بادئة in ["TB", "CTB", "GB"]):
                    self._سجل_كمرة_مباشر(self.جدول_كمرات_الربط, رمز, عرض, عمق)
                else:
                    self._سجل_كمرة_مباشر(self.جدول_الكمرات, رمز, عرض, عمق)

            for نمط, عكس in [
                (rf'(({نمط_رمز}))((?:\s+(?:\d+(?:\.\d+)?)){2})', False),
                (rf'((?:\d+(?:\.\d+)?\s+){{2}})(({نمط_رمز}))', True),
            ]:
                for مطابق in re.finditer(نمط, نص):
                    if عكس:
                        رمز = مطابق.group(2).strip()
                        أرقام = re.findall(r'\d+(?:\.\d+)?', مطابق.group(1))[-2:]
                    else:
                        رمز = مطابق.group(2).strip()
                        أرقام = re.findall(r'\d+(?:\.\d+)?', مطابق.group(3))[:2]
                    أبعاد = self._أبعاد_من_نصوص_متتابعة(أرقام, [(0.1, 2.0), (0.1, 2.0)])
                    if not أبعاد:
                        continue
                    if "STB" in رمز.upper():
                        self._سجل_كمرة_مباشر(self.جدول_كمرات_الستراب, رمز, أبعاد[0], أبعاد[1])
                    elif any(بادئة in رمز.upper() for بادئة in ["TB", "CTB", "GB"]):
                        self._سجل_كمرة_مباشر(self.جدول_كمرات_الربط, رمز, أبعاد[0], أبعاد[1])
                    else:
                        self._سجل_كمرة_مباشر(self.جدول_الكمرات, رمز, أبعاد[0], أبعاد[1])

        elif نوع == "عمود":
            for مطابق in re.finditer(r'(?:C\d+[A-Z]?|C\d+(?:/[A-Z0-9]+)?|C5/DC|DC\d*|NC\d*)', نص):
                رمز = مطابق.group(0).strip()
                أبعاد = self._استخرج_أبعاد_حول_رمز(نص, مطابق.start(), مطابق.end(), [(0.15, 1.5), (0.15, 1.5)])
                if not أبعاد:
                    continue
                self._سجل_عمود_مباشر(بيانات_العمود(
                    الرمز=رمز,
                    الطول=أبعاد[0],
                    العرض=أبعاد[1],
                    الكمية=1,
                ))

            for نمط, عكس in [
                (r'((?:C\d+[A-Z]?|C\d+(?:/[A-Z0-9]+)?|C5/DC|DC\d*|NC\d*))((?:\s+(?:\d+(?:\.\d+)?)){2})', False),
                (r'((?:\d+(?:\.\d+)?\s+){2})((?:C\d+[A-Z]?|C\d+(?:/[A-Z0-9]+)?|C5/DC|DC\d*|NC\d*))', True),
            ]:
                for مطابق in re.finditer(نمط, نص):
                    if عكس:
                        رمز = مطابق.group(2).strip()
                        أرقام = re.findall(r'\d+(?:\.\d+)?', مطابق.group(1))[-2:]
                    else:
                        رمز = مطابق.group(1).strip()
                        أرقام = re.findall(r'\d+(?:\.\d+)?', مطابق.group(2))[:2]
                    أبعاد = self._أبعاد_من_نصوص_متتابعة(أرقام, [(0.15, 1.5), (0.15, 1.5)])
                    if not أبعاد:
                        continue
                    self._سجل_عمود_مباشر(بيانات_العمود(
                        الرمز=رمز,
                        الطول=أبعاد[0],
                        العرض=أبعاد[1],
                        الكمية=1,
                    ))

    def _مسح_شامل_للنصوص_الإنشائية(self):
        نص = self._نص_كتلة_مرتبة(self.كل_النصوص)

        if not self.الأعمدة:
            for مطابق in re.finditer(r'(?<![A-Z0-9])(C\d+[A-Z]?|DC\d*|NC\d*)(?![A-Z0-9])\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)', نص):
                سياق = نص[max(0, مطابق.start() - 80): min(len(نص), مطابق.end() + 80)]
                if not any(k in سياق for k in ["COLUMN", "TYPE", "SIZE", "REINFORCEMENT", "STIRRUPS", "REMARKS"]):
                    continue
                طول = self._حوّل_بعد_هيكلي_إلى_متر(مطابق.group(2), 0.15, 1.5)
                عرض = self._حوّل_بعد_هيكلي_إلى_متر(مطابق.group(3), 0.15, 1.5)
                if طول and عرض:
                    self._سجل_عمود_مباشر(بيانات_العمود(
                        الرمز=مطابق.group(1),
                        الطول=max(طول, عرض),
                        العرض=min(طول, عرض),
                        الكمية=1,
                    ))

        if not self.جدول_كمرات_الربط and not self.جدول_كمرات_الستراب:
            for مطابق in re.finditer(r'(?<![A-Z0-9])((?:CTB|STB|TB|GB)\d+[A-Z]?)(?![A-Z0-9])\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)', نص):
                سياق = نص[max(0, مطابق.start() - 80): min(len(نص), مطابق.end() + 80)]
                if not any(k in سياق for k in ["BEAM", "DEPTH", "WIDTH", "TOP", "BOTTOM", "STIRRUPS", "SCHEDULE", "SIZE"]):
                    continue
                عرض = self._حوّل_بعد_هيكلي_إلى_متر(مطابق.group(2), 0.1, 2.0)
                عمق = self._حوّل_بعد_هيكلي_إلى_متر(مطابق.group(3), 0.1, 2.0)
                if not عرض or not عمق:
                    continue
                رمز = مطابق.group(1)
                if "STB" in رمز.upper():
                    self._سجل_كمرة_مباشر(self.جدول_كمرات_الستراب, رمز, عرض, عمق)
                else:
                    self._سجل_كمرة_مباشر(self.جدول_كمرات_الربط, رمز, عرض, عمق)

            for مطابق in re.finditer(r'(?:B[-_])?((?:CTB|STB|TB|GB)\d+[A-Z]?)\s*\(?\s*(\d+(?:\.\d+)?)\s*[Xx]\s*(\d+(?:\.\d+)?)\s*\)?', نص):
                عرض = self._حوّل_بعد_هيكلي_إلى_متر(مطابق.group(2), 0.1, 2.0)
                عمق = self._حوّل_بعد_هيكلي_إلى_متر(مطابق.group(3), 0.1, 2.0)
                if not عرض or not عمق:
                    continue
                رمز = مطابق.group(1)
                if "STB" in رمز.upper():
                    self._سجل_كمرة_مباشر(self.جدول_كمرات_الستراب, رمز, عرض, عمق)
                else:
                    self._سجل_كمرة_مباشر(self.جدول_كمرات_الربط, رمز, عرض, عمق)

    def _تحليل_أسس_من_خرائط(self, خرائط: List[Dict[str, str]]) -> List[بيانات_الأساس]:
        نتائج = []
        for صف in خرائط:
            رمز_خام = صف.get("رمز", "").strip() or self._استخرج_رمز_من_صف(صف, r'(?:CF|F|WF|FTG)[\-_]?\d+[A-Z]?\*?')
            مطابق = re.search(r'(?:CF|F|WF|FTG)[\-_]?\d+[A-Z]?\*?', رمز_خام.upper())
            رمز = مطابق.group(0) if مطابق else رمز_خام
            نص_أبعاد = صف.get("أبعاد", "")
            if not نص_أبعاد:
                نص_أبعاد = " ".join(صف.values())
            طول, عرض, عمق = استخراج_أبعاد_من_نص(نص_أبعاد, 3, self.وحدة)
            كمية = إلى_صحيح(صف.get("كمية", "1")) or 1
            if رمز and طول and عرض and عمق:
                نتائج.append(بيانات_الأساس(
                    الرمز=رمز, الطول=طول, العرض=عرض, العمق=عمق, الكمية=كمية
                ))
        return نتائج

    def _تحليل_كمرات_من_خرائط(self, خرائط: List[Dict[str, str]], نوع: str) -> Dict[str, Dict]:
        نتائج: Dict[str, Dict] = {}
        for صف in خرائط:
            if نوع == "كمرة_ستراب":
                نمط = r'(?:STB)[\-_]?\d+[A-Z]?\*?'
            elif نوع == "كمرة_ربط":
                نمط = r'(?:CTB|TB|GB)[\-_]?\d+[A-Z]?\*?'
            else:
                نمط = r'(?:B)[\-_]?\d+[A-Z]?\*?'

            رمز_خام = صف.get("رمز", "").strip() or self._استخرج_رمز_من_صف(صف, نمط)
            مطابق = re.search(نمط, رمز_خام.upper())
            رمز = مطابق.group(0) if مطابق else رمز_خام
            نص_أبعاد = صف.get("أبعاد", "")
            if not نص_أبعاد:
                نص_أبعاد = " ".join(صف.values())
            عرض, عمق = استخراج_أبعاد_من_نص(نص_أبعاد, 2, self.وحدة)
            if رمز and عرض and عمق:
                if رمز not in نتائج:
                    نتائج[رمز] = {"عرض": عرض, "عمق": عمق}
                else:
                    # لو تكرر الرمز بين أكثر من جدول، نحتفظ بالقطاع الأكبر
                    مساحة_قديم = نتائج[رمز]["عرض"] * نتائج[رمز]["عمق"]
                    مساحة_جديد = عرض * عمق
                    if مساحة_جديد > مساحة_قديم:
                        نتائج[رمز] = {"عرض": عرض, "عمق": عمق}
        return نتائج

    def _تحليل_أعمدة_من_خرائط(self, خرائط: List[Dict[str, str]]) -> List[بيانات_العمود]:
        نتائج: Dict[str, بيانات_العمود] = {}
        for صف in خرائط:
            رمز_خام = صف.get("رمز", "").strip() or self._استخرج_رمز_من_صف(صف, r'(?:C\d+[A-Z]?|C\d+(?:/[A-Z0-9]+)?|C5/DC|DC\d*|NC\d*)')
            مطابق = re.search(r'(?:C\d+[A-Z]?|C\d+(?:/[A-Z0-9]+)?|C5/DC|DC\d*|NC\d*)', رمز_خام.upper())
            رمز = مطابق.group(0) if مطابق else رمز_خام
            نص_أبعاد = صف.get("أبعاد", "")
            if not نص_أبعاد:
                نص_أبعاد = " ".join(صف.values())
            طول, عرض = استخراج_أبعاد_من_نص(نص_أبعاد, 2, self.وحدة)
            كمية = إلى_صحيح(صف.get("كمية", "1")) or 1
            if رمز and طول and عرض:
                مرشح = بيانات_العمود(الرمز=رمز, الطول=طول, العرض=عرض, الكمية=كمية)
                قائم = نتائج.get(رمز.upper())
                if قائم is None or (مرشح.الطول * مرشح.العرض) > (قائم.الطول * قائم.العرض):
                    نتائج[رمز.upper()] = مرشح
        return list(نتائج.values())

    def _قراءة_موجهة_بالرؤوس(self):
        مرشحات_أسس: List[List[بيانات_الأساس]] = []
        pdf_mode = bool(self._pdf_texts)

        foot_left = 20.0 if pdf_mode else 8.0
        foot_right = 42.0 if pdf_mode else 12.0
        foot_down = 28.0 if pdf_mode else 10.5
        beam_left = 20.0 if pdf_mode else 8.0
        beam_right = 42.0 if pdf_mode else 16.0
        beam_down = 24.0 if pdf_mode else 6.5
        col_left = 18.0 if pdf_mode else 8.0
        col_right = 32.0 if pdf_mode else 12.0
        col_down = 24.0 if pdf_mode else 10.0

        for رأس in self.كل_النصوص:
            كبير = رأس["كبير"]

            if "SCHEDULE OF FOOTINGS" in كبير or "SCHEDULE OF FOOTING" in كبير:
                نافذة = self._عناصر_حول_الرأس(رأس, foot_left, foot_right, foot_down)
                خرائط = self._خرائط_صفوف_الجدول(نافذة)
                عناصر = self._تحليل_أسس_من_خرائط(خرائط)
                if عناصر:
                    مرشحات_أسس.append(عناصر)

            elif "SCHEDULE OF STRAP BEAMS" in كبير or "STRAP BEAMS SCHEDULE" in كبير:
                نافذة = self._عناصر_حول_الرأس(رأس, beam_left, beam_right, beam_down)
                خرائط = self._خرائط_صفوف_الجدول(نافذة)
                self.جدول_كمرات_الستراب.update(self._تحليل_كمرات_من_خرائط(خرائط, "كمرة_ستراب"))

            elif "SCHEDULE OF TIE BEAM" in كبير or "SCHEDULE OF TIE BEAMS" in كبير or "TIE BEAMS SCHEDULE" in كبير:
                نافذة = self._عناصر_حول_الرأس(رأس, beam_left, beam_right, beam_down)
                خرائط = self._خرائط_صفوف_الجدول(نافذة)
                self.جدول_كمرات_الربط.update(self._تحليل_كمرات_من_خرائط(خرائط, "كمرة_ربط"))
                self.جدول_الكمرات.update(self._تحليل_كمرات_من_خرائط(خرائط, "كمرة"))

            elif "SLAB BEAMS" in كبير or "BEAM SCHEDULE" in كبير or "SCHEDULE OF BEAM" in كبير:
                نافذة = self._عناصر_حول_الرأس(رأس, beam_left, beam_right, beam_down)
                خرائط = self._خرائط_صفوف_الجدول(نافذة)
                self.جدول_الكمرات.update(self._تحليل_كمرات_من_خرائط(خرائط, "كمرة"))

            elif "SCHEDULE OF COLUMNS" in كبير or "COLUMN SCHEDULE" in كبير:
                نافذة = self._عناصر_حول_الرأس(رأس, col_left, col_right, col_down)
                خرائط = self._خرائط_صفوف_الجدول(نافذة)
                for عنصر in self._تحليل_أعمدة_من_خرائط(خرائط):
                    if not any(قائم.الرمز.upper() == عنصر.الرمز.upper() for قائم in self.الأعمدة):
                        self.الأعمدة.append(عنصر)

        if مرشحات_أسس:
            أفضل = max(
                مرشحات_أسس,
                key=lambda قائمة: (
                    len({ع.الرمز for ع in قائمة}),
                    sum(ع.الطول * ع.العرض * ع.الكمية for ع in قائمة),
                ),
            )
            موجود = {ع.الرمز.upper() for ع in self.الأسس}
            for عنصر in أفضل:
                if عنصر.الرمز.upper() not in موجود:
                    self.الأسس.append(عنصر)

    def _تحليل_أسس(self, صفوف):
        ربط = self._ربط_الأعمدة(صفوف[0])
        for صف in صفوف[1:]:
            # Clean and combine text in each cell
            ب = {}
            for i, خ in enumerate(صف):
                ك_اسم = ربط.get(i)
                if ك_اسم:
                    ب[ك_اسم] = ب.get(ك_اسم, "") + " " + خ["نص"]
            
            try:
                رمز = ب.get("رمز", "").strip()
                if not رمز: continue
                # Log the raw data found for this mark
                مسجل.info(f"  > تحليل صف الأساس [{رمز}]: {ب}")
                نص_أبعاد = ب.get("أبعاد", "")
                if نص_أبعاد:
                    ط, ع, عم = استخراج_أبعاد_من_نص(نص_أبعاد, 3, self.وحدة)
                else:
                    ط_نص = ب.get("طول", "0").upper()
                    if "CONT" in ط_نص or "مستمر" in ط_نص or "WF" in رمز.upper():
                        # لو كان أساس حائطي مستمر، نضع طولاً افتراضياً 1.0م 
                        # أو يتم حسابه لاحقاً من المسقط لو توفرت البيانات
                        ط = 1.0 
                        مسجل.info(f"    - تم رصد أساس حائطي (WF) أو مستمر: {رمز}")
                    else:
                        ط = إلى_متر(ط_نص, self.وحدة)
                    ع = إلى_متر(ب.get("عرض", "0"), self.وحدة)
                    عم = إلى_متر(ب.get("عمق", "0"), self.وحدة)
                
                كم = إلى_صحيح(ب.get("كمية", "1")) or 1
                if ط and ع and عم:
                    self.الأسس.append(بيانات_الأساس(
                        الرمز=رمز, الطول=ط, العرض=ع, العمق=عم, الكمية=كم
                    ))
            except Exception as خطأ:
                مسجل.warning(f"خطأ أساس: {خطأ}")

    def _تحليل_أعمدة_عنق(self, صفوف):
        ربط = self._ربط_الأعمدة(صفوف[0])
        for صف in صفوف[1:]:
            ب = {}
            for i, خ in enumerate(صف):
                ك_اسم = ربط.get(i)
                if ك_اسم:
                    ب[ك_اسم] = ب.get(ك_اسم, "") + " " + خ["نص"]
            try:
                رمز = ب.get("رمز", "").strip()
                if not رمز: continue
                ط = إلى_متر(ب.get("طول", "0"), self.وحدة)
                ع = إلى_متر(ب.get("عرض", "0"), self.وحدة)
                كم = إلى_صحيح(ب.get("كمية", "1")) or 1
                if ط and ع:
                    عمق_أ = self._ابحث_عمق_الأساس(رمز)
                    self.أعمدة_العنق.append(بيانات_عمود_العنق(
                        الرمز=رمز, الطول=ط, العرض=ع,
                        عمق_الأساس=عمق_أ, الكمية=كم
                    ))
            except Exception as خطأ:
                مسجل.warning(f"خطأ عمود عنق: {خطأ}")

    def _ابحث_عمق_الأساس(self, رمز_العنق: str) -> float:
        """يبحث عن عمق الأساس المرتبط بعمود العنق."""
        # رمز NC1 يرتبط بـ F1، NC2 بـ F2 وهكذا
        رقم = re.search(r'\d+', رمز_العنق)
        if رقم:
            رمز_أساس = f"F{رقم.group()}"
            for أ in self.الأسس:
                if أ.الرمز.upper() == رمز_أساس.upper():
                    return أ.العمق
        # لو لم يجد، يأخذ متوسط الأسس
        if self.الأسس:
            return sum(أ.العمق for أ in self.الأسس) / len(self.الأسس)
        return 0.0

    def _تحليل_كمرات(self, صفوف, نوع: str):
        ربط = self._ربط_الأعمدة(صفوف[0])
        for صف in صفوف[1:]:
            ب = {}
            for i, خ in enumerate(صف):
                ك_اسم = ربط.get(i)
                if ك_اسم:
                    ب[ك_اسم] = ب.get(ك_اسم, "") + " " + خ["نص"]
            try:
                رمز = ب.get("رمز", "").strip()
                if not رمز:
                    continue
                مسجل.info(f"  > تحليل صف كمرة [{رمز}]: {ب}")
                نص_أبعاد = ب.get("أبعاد", "")
                if نص_أبعاد:
                    ع, عم = استخراج_أبعاد_من_نص(نص_أبعاد, 2, self.وحدة)
                else:
                    ع = إلى_متر(ب.get("عرض", "0"), self.وحدة)
                    عم = إلى_متر(ب.get("عمق", "0"), self.وحدة)
                if ع and عم:
                    if نوع == "كمرة_ربط":
                        self.جدول_كمرات_الربط[رمز.upper()] = {"عرض": ع, "عمق": عم}
                    elif نوع == "كمرة_ستراب":
                        self.جدول_كمرات_الستراب[رمز.upper()] = {"عرض": ع, "عمق": عم}
                    else:
                        self.جدول_الكمرات[رمز.upper()] = {"عرض": ع, "عمق": عم}
            except Exception as خطأ:
                مسجل.warning(f"خطأ كمرة: {خطأ}")

    def _تحليل_أعمدة(self, صفوف):
        ربط = self._ربط_الأعمدة(صفوف[0])
        for صف in صفوف[1:]:
            ب = {}
            for i, خ in enumerate(صف):
                ك_اسم = ربط.get(i)
                if ك_اسم:
                    ب[ك_اسم] = ب.get(ك_اسم, "") + " " + خ["نص"]
            try:
                رمز = ب.get("رمز", "").strip()
                if not رمز: continue
                مسجل.info(f"  > تحليل صف العمود [{رمز}]: {ب}")
                ط = إلى_متر(ب.get("طول", "0"), self.وحدة)
                ع = إلى_متر(ب.get("عرض", "0"), self.وحدة)
                كم = إلى_صحيح(ب.get("كمية", "1")) or 1
                if ط and ع:
                    self.الأعمدة.append(بيانات_العمود(
                        الرمز=رمز, الطول=ط, العرض=ع, الكمية=كم
                    ))
            except Exception as خطأ:
                مسجل.warning(f"خطأ عمود: {خطأ}")

    def _تحليل_فتحات(self, صفوف):
        ربط = self._ربط_الأعمدة(صفوف[0])
        for صف in صفوف[1:]:
            ب = {}
            for i, خ in enumerate(صف):
                ك_اسم = ربط.get(i)
                if ك_اسم:
                    ب[ك_اسم] = ب.get(ك_اسم, "") + " " + خ["نص"]
            try:
                رمز = ب.get("رمز", "").strip()
                if not رمز: continue
                مسجل.info(f"  > تحليل صف الفتحة [{رمز}]: {ب}")
                ط = إلى_متر(ب.get("طول", "0"), self.وحدة)
                ار = إلى_متر(ب.get("عمق", "0"), self.وحدة)
                كم = إلى_صحيح(ب.get("كمية", "1")) or 1
                if ط and ار:
                    self.الفتحات.append(بيانات_فتحة(
                        الرمز=رمز, الطول=ط, الارتفاع=ار, الكمية=كم,
                        مصدر_الكمية="SCHEDULE",
                        مصدر_الأبعاد="SCHEDULE",
                    ))
            except Exception as خطأ:
                مسجل.warning(f"خطأ فتحة: {خطأ}")

    def قراءة(self):
        self._استخراج_النصوص()
        جداول = self._تجميع_النصوص_في_جداول()
        for مجموعة in جداول:
            نوع = self._تحديد_نوع(مجموعة)
            if not نوع:
                continue
            self._تحليل_مباشر_من_كتلة_نصية(مجموعة, نوع)
            صفوف = self._تجميع_بصفوف(مجموعة)
            if len(صفوف) < 2:
                continue
            
            # Use the most descriptive header row (the one with the most mapped columns)
            header_row_idx = 0
            max_mapped = 0
            for r_idx in range(min(3, len(صفوف))):
                mapped_count = len(self._ربط_الأعمدة(صفوف[r_idx]))
                if mapped_count > max_mapped:
                    max_mapped = mapped_count
                    header_row_idx = r_idx
            
            مسجل.info(f"تحليل جدول [{نوع}] — {len(صفوف)} صف (Header Row: {header_row_idx})")
            rows_to_parse = صفوف[header_row_idx+1:]
            header_row = صفوف[header_row_idx]
            
            if نوع == "أساس":         self._تحليل_أسس([header_row] + rows_to_parse)
            elif نوع == "عمود_عنق":   self._تحليل_أعمدة_عنق([header_row] + rows_to_parse)
            elif نوع == "كمرة_ربط":   self._تحليل_كمرات([header_row] + rows_to_parse, "كمرة_ربط")
            elif نوع == "كمرة_ستراب": self._تحليل_كمرات([header_row] + rows_to_parse, "كمرة_ستراب")
            elif نوع == "كمرة":       self._تحليل_كمرات([header_row] + rows_to_parse, "كمرة")
            elif نوع == "عمود":       self._تحليل_أعمدة([header_row] + rows_to_parse)
            elif نوع == "فتحة":       self._تحليل_فتحات([header_row] + rows_to_parse)

        self._استخرج_السماكات_والمناسيب()
        self._مسح_شامل_للنصوص_الإنشائية()

        # Fallback/upgrade path:
        # بعض ملفات DXF لا تتجمع كجداول كاملة بالمسافة فقط، لذلك نقرأها من رؤوسها مباشرة.
        self._قراءة_موجهة_بالرؤوس()
        self._عد_الأسس_من_الهندسة()
        self._استخرج_الفتحات_من_تفاصيل_الأبواب_والشبابيك()

        مسجل.info(
            f"الجداول: {len(self.الأسس)} أساس | "
            f"{len(self.أعمدة_العنق)} عنق | "
            f"{len(self.جدول_كمرات_الربط)} نوع TB | "
            f"{len(self.جدول_كمرات_الستراب)} نوع STB | "
            f"{len(self.جدول_الكمرات)} نوع B | "
            f"{len(self.الأعمدة)} عمود | "
            f"{len(self.الفتحات)} فتحة"
        )


# ─────────────────────────────────────────────────────────────────────────────
# قارئ الكمرات من الرسم (الطول من الخطوط + ربط بالتسمية)
# ─────────────────────────────────────────────────────────────────────────────

class قارئ_كمرات_الرسم:
    """
    يمشي على المسقط الرئيسي ويقيس طول كل كمرة.
    يربطها بأقرب تسمية (TB1، STB1، B1...) في نطاق 15 متر.
    ثم يجمع العرض والعمق من جدول الكمرات.

    الإصلاح الرئيسي من الكود السابق:
    - يفلتر المسقط الرئيسي فقط (يتجاهل الجداول والتفاصيل)
    - يستخدم KDTree للبحث عن التسميات
    """

    def __init__(self, msp, مقياس: float,
                 جدول_tb: Dict, جدول_stb: Dict, جدول_b: Dict):
        self.msp      = msp
        self.مقياس    = مقياس
        self.جدول_tb  = جدول_tb
        self.جدول_stb = جدول_stb
        self.جدول_b   = جدول_b

        self.تسميات_الكمرات: List[Dict] = []  # {نص، x، y}
        self.نصوص_مرجعية: List[Dict] = []
        self.عناوين_المساقط: List[Dict] = []
        self.نقاط_استبعاد: List[Tuple[float, float]] = []
        self.صناديق_المساقط: List[Dict] = []
        self.كمرات_الربط:   List[بيانات_كمرة_ربط]   = []
        self.كمرات_الستراب: List[بيانات_كمرة_ستراب]  = []
        self.الكمرات:        List[بيانات_كمرة]         = []

    @staticmethod
    def _مساحة_شوليس(نقاط: List[Tuple]) -> float:
        """حساب المساحة بصيغة شوليس — دقيق رياضياً."""
        مساحة = 0.0
        ن = len(نقاط)
        for i in range(ن):
            j = (i + 1) % ن
            مساحة += نقاط[i][0] * نقاط[j][1]
            مساحة -= نقاط[j][0] * نقاط[i][1]
        return abs(مساحة) / 2.0

    @staticmethod
    def _نص_موحد(نص: str) -> str:
        return " ".join(str(نص).upper().split())

    @staticmethod
    def _داخل_صندوق(x: float, y: float, صندوق: Dict) -> bool:
        return صندوق["xmin"] <= x <= صندوق["xmax"] and صندوق["ymin"] <= y <= صندوق["ymax"]

    def _مرشحو_المقياس_الهندسي(self) -> List[float]:
        مرشحون: List[float] = []
        for قيمة in (self.مقياس, 1.0, 0.1, 0.01, 0.001):
            if قيمة > 0 and قيمة not in مرشحون:
                مرشحون.append(قيمة)
        return مرشحون

    def _اختبر_مقياس_هندسي(self, مقياس_مرشح: float) -> int:
        عدد_صالح = 0
        for كيان in self.msp.query('LWPOLYLINE LINE'):
            try:
                طبقة = كيان.dxf.layer.upper()
                if not any(ط in طبقة for ط in طبقات_الكمرات):
                    continue

                if كيان.dxftype() == 'LWPOLYLINE':
                    نقاط = [(ن[0] * مقياس_مرشح, ن[1] * مقياس_مرشح) for ن in كيان.get_points(format='xy')]
                else:
                    نقاط = [
                        (كيان.dxf.start.x * مقياس_مرشح, كيان.dxf.start.y * مقياس_مرشح),
                        (كيان.dxf.end.x * مقياس_مرشح,   كيان.dxf.end.y * مقياس_مرشح)
                    ]

                if len(نقاط) < 2 or not self._هل_في_المسقط_الرئيسي(نقاط):
                    continue

                طول, _, _, _ = self._قياس_عنصر_كمرة(نقاط)
                if طول >= 0.40:
                    عدد_صالح += 1
            except Exception:
                pass
        return عدد_صالح

    def _اعتمد_مقياس_هندسي_إن_لزم(self):
        المقياس_الحالي = self.مقياس
        الدرجة_الحالية = self._اختبر_مقياس_هندسي(المقياس_الحالي)
        أفضل_مقياس = المقياس_الحالي
        أفضل_درجة = الدرجة_الحالية

        for مقياس_مرشح in self._مرشحو_المقياس_الهندسي():
            درجة = self._اختبر_مقياس_هندسي(مقياس_مرشح)
            if درجة > أفضل_درجة:
                أفضل_درجة = درجة
                أفضل_مقياس = مقياس_مرشح

        if أفضل_مقياس != المقياس_الحالي and أفضل_درجة >= max(25, الدرجة_الحالية * 3):
            self.مقياس = أفضل_مقياس
            مسجل.info(
                f"تم تعديل مقياس قراءة الكمرات من {المقياس_الحالي:g} إلى {أفضل_مقياس:g} "
                f"اعتماداً على {أفضل_درجة} عنصر كمرة صالح هندسياً"
            )

    def _هل_في_المسقط_الرئيسي(self, نقاط: List) -> bool:
        """
        يفلتر المسقط الرئيسي ويتجاهل الجداول والتفاصيل.
        المسقط الرئيسي عادةً في X < 0 أو في المنطقة الأكبر.
        """
        if not نقاط:
            return False
        # الطريقة: تجاهل العناصر الصغيرة جداً (تفاصيل) أو البعيدة جداً (جداول)
        xs = [ن[0] for ن in نقاط]
        ys = [ن[1] for ن in نقاط]
        # مركز العنصر
        مركز_x = sum(xs) / len(xs)
        مركز_y = sum(ys) / len(ys)
        # تجاهل العناصر في منطقة X إيجابية كبيرة (عادةً جداول)
        # هذا يعتمد على تصميم المكتب — نستخدم الإحداثي السلبي كمؤشر
        امتداد_x = max(xs) - min(xs)
        امتداد_y = max(ys) - min(ys)
        قطر = math.hypot(امتداد_x, امتداد_y)
        if قطر < 0.40:
            return False
        if قطر > 90.0:
            return False
        return True

    # عبارات تدل على مساقط تحت الأرض (substructure)
    _عبارات_مجال_sub = frozenset([
        "TIE BEAM LAYOUT", "TIE BEAM PLAN",
        "FOUNDATION LAYOUT", "FOUNDATION PLAN",
        "GROUND BEAM LAYOUT", "GROUND BEAM PLAN",
        "STRAP BEAM LAYOUT", "STRAP BEAM PLAN",
        "SUBSTRUCTURE LAYOUT",
    ])
    # عبارات تدل على مساقط فوق الأرض (superstructure)
    _عبارات_مجال_super = frozenset([
        "SLAB LAYOUT", "SLAB PLAN",
        "ROOF SLAB LAYOUT", "ROOF SLAB PLAN",
        "BEAM LAYOUT", "BEAM PLAN",
        "FIRST FLOOR LAYOUT", "SECOND FLOOR LAYOUT",
        "ROOF BEAM LAYOUT", "ROOF BEAM PLAN",
        "SUPERSTRUCTURE LAYOUT",
    ])

    def _مجال_المسقط(self, نص: str) -> str:
        كبير = self._نص_موحد(نص)
        if "COLUMN LAYOUT" in كبير:
            return ""
        for عبارة in self._عبارات_مجال_sub:
            if عبارة in كبير:
                return "SUB"
        for عبارة in self._عبارات_مجال_super:
            if عبارة in كبير:
                return "SUPER"
        return ""

    def _صندوق_حول_عنوان(self, عنوان: Dict, هامش_يسار_x: float = 48.0, هامش_يمين_x: float = 12.0,
                     صعود_y: float = 32.0, نزول_y: float = -8.0) -> Dict:
        return {
            "xmin": عنوان["x"] - هامش_يسار_x,
            "xmax": عنوان["x"] + هامش_يمين_x,
            "ymin": عنوان["y"] + نزول_y,
            "ymax": عنوان["y"] + صعود_y,
            "العنوان": عنوان["نص"],
            "المجال": عنوان.get("المجال", ""),
        }

    @staticmethod
    def _زاوج_محاور(عناصر: List[Tuple[float, float, float, str]]) -> List[Dict]:
        مستخدم = set()
        سنترلانات: List[Dict] = []
        for i, أول in enumerate(عناصر):
            if i in مستخدم:
                continue
            موضع1, بداية1, نهاية1, طبقة1 = أول
            for j, ثان in enumerate(عناصر[i + 1:], start=i + 1):
                if j in مستخدم:
                    continue
                موضع2, بداية2, نهاية2, طبقة2 = ثان
                if abs(بداية1 - بداية2) <= 0.08 and abs(نهاية1 - نهاية2) <= 0.08 and 0.08 <= abs(موضع2 - موضع1) <= 0.40:
                    مستخدم.add(i)
                    مستخدم.add(j)
                    سنترلانات.append({
                        "pos": (موضع1 + موضع2) / 2.0,
                        "start": min(بداية1, بداية2),
                        "end": max(نهاية1, نهاية2),
                        "source": f"{طبقة1}+{طبقة2}",
                    })
                    break
        return سنترلانات

    @staticmethod
    def _ادمج_سنترلانات(سنترلانات: List[Dict]) -> List[Dict]:
        مدمجة: List[Dict] = []
        مرتبة = sorted(سنترلانات, key=lambda ع: (round(ع["pos"], 3), ع["start"], ع["end"]))
        for عنصر in مرتبة:
            if not مدمجة:
                مدمجة.append({**عنصر, "sources": {عنصر["source"]}})
                continue
            سابق = مدمجة[-1]
            if abs(سابق["pos"] - عنصر["pos"]) <= 0.06 and عنصر["start"] - سابق["end"] <= 0.25:
                سابق["end"] = max(سابق["end"], عنصر["end"])
                سابق["sources"].add(عنصر["source"])
            else:
                مدمجة.append({**عنصر, "sources": {عنصر["source"]}})
        for عنصر in مدمجة:
            عنصر["source"] = ",".join(sorted(عنصر.pop("sources")))
        return مدمجة

    def _استخرج_سنترلانات_tb_احتياطي(self) -> Tuple[Dict[str, float], Dict[str, float]]:
        أطوال_tb: Dict[str, float] = defaultdict(float)
        أطوال_stb: Dict[str, float] = defaultdict(float)
        صناديق_tb_خام = [
            ص for ص in self.صناديق_المساقط
            if "TIE BEAM" in self._نص_موحد(ص.get("العنوان", ""))
        ]
        صناديق_tb: List[Dict] = []
        for صندوق in sorted(صناديق_tb_خام, key=lambda ص: ص.get("عدد_التسميات", 0), reverse=True):
            مركز_x = (صندوق["xmin"] + صندوق["xmax"]) / 2.0
            مركز_y = (صندوق["ymin"] + صندوق["ymax"]) / 2.0
            if any(
                abs(((م["xmin"] + م["xmax"]) / 2.0) - مركز_x) <= 6.0 and
                abs(((م["ymin"] + م["ymax"]) / 2.0) - مركز_y) <= 6.0
                for م in صناديق_tb
            ):
                continue
            صناديق_tb.append(صندوق)
        if not صناديق_tb:
            return أطوال_tb, أطوال_stb

        for صندوق in صناديق_tb:
            تسميات = [
                ت for ت in self.تسميات_الكمرات
                if self._داخل_صندوق(ت["x"], ت["y"], صندوق)
                and (ت["نص"].startswith("TB") or ت["نص"].startswith("STB"))
            ]
            if not تسميات:
                continue

            أفقية: List[Tuple[float, float, float, str]] = []
            رأسية: List[Tuple[float, float, float, str]] = []
            for كيان in self.msp.query('LINE'):
                try:
                    طبقة = كيان.dxf.layer.upper()
                    if طبقة not in {"WALL", "S-BEAM"}:
                        continue
                    x1 = كيان.dxf.start.x * self.مقياس
                    y1 = كيان.dxf.start.y * self.مقياس
                    x2 = كيان.dxf.end.x * self.مقياس
                    y2 = كيان.dxf.end.y * self.مقياس
                    cx = (x1 + x2) / 2.0
                    cy = (y1 + y2) / 2.0
                    if not self._داخل_صندوق(cx, cy, صندوق):
                        continue
                    if abs(y1 - y2) <= 0.02 and abs(x2 - x1) >= 0.80:
                        بداية, نهاية = sorted((x1, x2))
                        أفقية.append((round((y1 + y2) / 2.0, 3), round(بداية, 3), round(نهاية, 3), طبقة))
                    elif abs(x1 - x2) <= 0.02 and abs(y2 - y1) >= 0.80:
                        بداية, نهاية = sorted((y1, y2))
                        رأسية.append((round((x1 + x2) / 2.0, 3), round(بداية, 3), round(نهاية, 3), طبقة))
                except Exception:
                    pass

            سنترلانات_أفقية = self._ادمج_سنترلانات(self._زاوج_محاور(أفقية))
            for عنصر in سنترلانات_أفقية:
                عنصر["orientation"] = "H"
            سنترلانات_رأسية = self._ادمج_سنترلانات(self._زاوج_محاور(رأسية))
            for عنصر in سنترلانات_رأسية:
                عنصر["orientation"] = "V"
            سنترلانات = سنترلانات_أفقية + سنترلانات_رأسية
            for عنصر in سنترلانات:
                طول = عنصر["end"] - عنصر["start"]
                if طول < 0.50 or طول > 60.0:
                    continue
                اتجاه = عنصر.get("orientation", "H")
                if any(abs(عنصر["pos"] - نقطة[0]) < 0.01 and abs(((عنصر["start"] + عنصر["end"]) / 2.0) - نقطة[1]) < 0.01
                       for نقطة in self.نقاط_استبعاد):
                    continue

                if اتجاه == "H":
                    منتصف_x = (عنصر["start"] + عنصر["end"]) / 2.0
                    منتصف_y = عنصر["pos"]
                else:
                    منتصف_x = عنصر["pos"]
                    منتصف_y = (عنصر["start"] + عنصر["end"]) / 2.0

                مرشحون = [
                    (math.hypot(ت["x"] - منتصف_x, ت["y"] - منتصف_y), ت["نص"])
                    for ت in تسميات
                ]
                if not مرشحون:
                    continue
                مسافة, رمز = min(مرشحون, key=lambda زوج: زوج[0])
                if مسافة > 4.5 or رمز.startswith("WF"):
                    continue
                if رمز.startswith("STB"):
                    أطوال_stb[رمز] += طول
                else:
                    أطوال_tb[رمز] += طول
        return أطوال_tb, أطوال_stb

    def _ينتمي_للمنطقة(self, رمز: str, مجال: str) -> bool:
        نوع = self._تصنيف_الكمرة(رمز)
        if مجال == "SUB":
            return نوع in {"TB", "STB"}
        if مجال == "SUPER":
            return نوع == "B"
        return False

    def _جمع_مراكز_العناصر(self) -> List[Dict]:
        عناصر: List[Dict] = []
        for كيان in self.msp.query('LWPOLYLINE LINE'):
            try:
                طبقة = كيان.dxf.layer.upper()
                if not any(ط in طبقة for ط in طبقات_الكمرات):
                    continue

                if كيان.dxftype() == 'LWPOLYLINE':
                    نقاط = [(ن[0] * self.مقياس, ن[1] * self.مقياس)
                             for ن in كيان.get_points(format='xy')]
                else:
                    نقاط = [
                        (كيان.dxf.start.x * self.مقياس, كيان.dxf.start.y * self.مقياس),
                        (كيان.dxf.end.x * self.مقياس,   كيان.dxf.end.y * self.مقياس)
                    ]

                if len(نقاط) < 2 or not self._هل_في_المسقط_الرئيسي(نقاط):
                    continue

                طول, اتجاه, امتداد_x, امتداد_y = self._قياس_عنصر_كمرة(نقاط)
                if طول < 0.40:
                    continue

                عناصر.append({
                    "x": sum(ن[0] for ن in نقاط) / len(نقاط),
                    "y": sum(ن[1] for ن in نقاط) / len(نقاط),
                    "الطول": طول,
                    "الاتجاه": اتجاه,
                    "امتداد_x": امتداد_x,
                    "امتداد_y": امتداد_y,
                    "الطبقة": طبقة,
                })
            except Exception:
                pass
        return عناصر

    def _قياس_عنصر_كمرة(self, نقاط: List[Tuple[float, float]]) -> Tuple[float, str, float, float]:
        امتداد_x = max(ن[0] for ن in نقاط) - min(ن[0] for ن in نقاط)
        امتداد_y = max(ن[1] for ن in نقاط) - min(ن[1] for ن in نقاط)
        اتجاه = "H" if امتداد_x >= امتداد_y else "V"
        طول_المسار = sum(
            math.sqrt((نقاط[i + 1][0] - نقاط[i][0]) ** 2 +
                      (نقاط[i + 1][1] - نقاط[i][1]) ** 2)
            for i in range(len(نقاط) - 1)
        )

        # كثير من الكمرات تُرسم كحدود مستطيلة رفيعة، لذلك نأخذ المحور الرئيسي لا محيط الشكل.
        if len(نقاط) >= 4 and min(امتداد_x, امتداد_y) <= 0.35 and max(امتداد_x, امتداد_y) >= 0.60:
            طول = max(امتداد_x, امتداد_y)
        else:
            طول = طول_المسار
        return طول, اتجاه, امتداد_x, امتداد_y

    def _ابحث_عن_تسمية_محلية(self, مركز_x: float, مركز_y: float,
                           اتجاه: str, صندوق: Dict) -> Optional[str]:
        تسميات = صندوق.get("التسميات") or []
        if not تسميات:
            return None

        مرشحات = []
        if اتجاه == "V":
            if صندوق.get("المجال") == "SUB":
                مرشحات_stb = [
                    ت for ت in تسميات
                    if ت["نص"].startswith("STB")
                    and abs(ت["x"] - مركز_x) <= 2.5
                    and مركز_y <= ت["y"] <= (مركز_y + 12.0)
                ]
                if مرشحات_stb:
                    قائمة_البحث = مرشحات_stb
                else:
                    قائمة_البحث = [
                        ت for ت in تسميات
                        if abs(ت["x"] - مركز_x) <= 2.5 and مركز_y <= ت["y"] <= (مركز_y + 12.0)
                    ]
            else:
                قائمة_البحث = [
                    ت for ت in تسميات
                    if abs(ت["x"] - مركز_x) <= 2.5 and مركز_y <= ت["y"] <= (مركز_y + 12.0)
                ]
        else:
            قائمة_البحث = [
                ت for ت in تسميات
                if abs(ت["y"] - مركز_y) <= 4.0
            ]

        قائمة_البحث = قائمة_البحث or تسميات
        مسافة, رمز = min(
            ((math.hypot(ت["x"] - مركز_x, ت["y"] - مركز_y), ت["نص"]) for ت in قائمة_البحث),
            key=lambda زوج: زوج[0],
            default=(1e9, None)
        )
        if مسافة <= 11.0:
            return رمز
        return None

    def _أقرب_عنوان_للمجموعة(self, مركز_x: float, مركز_y: float) -> Optional[Dict]:
        عناوين = [ع for ع in self.عناوين_المساقط if ع.get("المجال")]
        if not عناوين:
            return None
        مسافة, عنوان = min(
            ((math.hypot(ع["x"] - مركز_x, ع["y"] - مركز_y), ع) for ع in عناوين),
            key=lambda زوج: زوج[0]
        )
        if مسافة > 25.0:
            return None
        return عنوان

    def _كوّن_صناديق_من_التسميات(self):
        if not self.تسميات_الكمرات:
            return

        نقاط = [(ت["x"], ت["y"]) for ت in self.تسميات_الكمرات]
        for فهارس in self._عنقدة_نقاط(نقاط, 8.0):
            عناصر = [self.تسميات_الكمرات[i] for i in فهارس]
            if len(عناصر) < 3:
                continue

            xs = [ع["x"] for ع in عناصر]
            ys = [ع["y"] for ع in عناصر]
            مركز_x = sum(xs) / len(xs)
            مركز_y = sum(ys) / len(ys)
            عرض = max(xs) - min(xs)
            ارتفاع = max(ys) - min(ys)

            نصوص_قريبة = [
                ن for ن in self.نصوص_مرجعية
                if math.hypot(ن["x"] - مركز_x, ن["y"] - مركز_y) <= 22.0
            ]
            يوجد_تخطيط = any(
                any(ك in ن["نص"] for ك in ["LAYOUT", "FOUNDATION", "SLAB LAYOUT", "ROOF SLAB"])
                for ن in نصوص_قريبة
            )
            يوجد_جدول = any(
                any(ك in ن["نص"] for ك in ["SCHEDULE", "SECTION", "DETAIL", "DESCRIPTION"])
                for ن in نصوص_قريبة
            )

            if يوجد_جدول and not يوجد_تخطيط:
                continue
            if not يوجد_تخطيط and (عرض < 8.0 or ارتفاع < 8.0):
                continue

            عائلات: Dict[str, List[Dict]] = defaultdict(list)
            for ع in عناصر:
                عائلات[self._تصنيف_الكمرة(ع["نص"])].append(ع)

            for نوع_عائلة, عناصر_العائلة in عائلات.items():
                if len(عناصر_العائلة) < 2:
                    continue
                xs2 = [ع["x"] for ع in عناصر_العائلة]
                ys2 = [ع["y"] for ع in عناصر_العائلة]
                رمز_افتراضي = "TB1" if نوع_عائلة == "TB" else "STB1" if نوع_عائلة == "STB" else "B1"
                هامش = 3.0 if نوع_عائلة == "B" else 4.0
                self.صناديق_المساقط.append({
                    "xmin": min(xs2) - هامش,
                    "ymin": min(ys2) - هامش,
                    "xmax": max(xs2) + هامش,
                    "ymax": max(ys2) + هامش,
                    "النوع_الافتراضي": رمز_افتراضي,
                    "عدد_التسميات": len(عناصر_العائلة),
                    "التسميات": عناصر_العائلة,
                    "المجال": "SUB" if نوع_عائلة in {"TB", "STB"} else "SUPER",
                })

    @staticmethod
    def _عنقدة_نقاط(نقاط: List[Tuple[float, float]], مسافة: float = 10.0) -> List[List[int]]:
        if not نقاط:
            return []
        مزارة = [False] * len(نقاط)
        عناقيد: List[List[int]] = []
        for فهرس, نقطة in enumerate(نقاط):
            if مزارة[فهرس]:
                continue
            مزارة[فهرس] = True
            مكدس = [فهرس]
            عنقود: List[int] = []
            while مكدس:
                i = مكدس.pop()
                عنقود.append(i)
                x1: float = float(نقاط[i][0])
                y1: float = float(نقاط[i][1])
                for j, (x2_raw, y2_raw) in enumerate(نقاط):
                    if مزارة[j]:
                        continue
                    x2: float = float(x2_raw)
                    y2: float = float(y2_raw)
                    if math.hypot(x2 - x1, y2 - y1) <= مسافة:
                        مزارة[j] = True
                        مكدس.append(j)
            عناقيد.append(عنقود)
        return عناقيد

    def _كوّن_صناديق_المساقط(self):
        self.صناديق_المساقط = []
        عناوين_مستخدمة = []
        # جمع كل ألفاظ دالة على مساقط sub أو super
        كل_عبارات_المسقط = self._عبارات_مجال_sub | self._عبارات_مجال_super
        for عنوان in self.عناوين_المساقط:
            نص_موحد = self._نص_موحد(عنوان["نص"])
            # قبول أي عنوان يحتوي إحدى العبارات المعروفة
            if not any(ع in نص_موحد for ع in كل_عبارات_المسقط):
                continue
            صندوق = self._صندوق_حول_عنوان(عنوان)
            تسميات_محلية = [
                ت for ت in self.تسميات_الكمرات
                if self._داخل_صندوق(ت["x"], ت["y"], صندوق)
                and self._ينتمي_للمنطقة(ت["نص"], صندوق["المجال"])
            ]
            if not تسميات_محلية:
                continue
            # تحديد الرمز الافتراضي: ckeck for any sub-structure beam keyword
            مجال = صندوق.get("المجال", "")
            هو_كمرة_ربط = (
                "TIE BEAM" in نص_موحد or
                "GROUND BEAM" in نص_موحد or
                "STRAP BEAM" in نص_موحد or
                "FOUNDATION" in نص_موحد or
                مجال == "SUB"
            )
            if هو_كمرة_ربط:
                رمز_افتراضي = "STB1" if any(ت["نص"].startswith("STB") for ت in تسميات_محلية) else "TB1"
            else:
                رمز_افتراضي = "B1"
            self.صناديق_المساقط.append({
                **صندوق,
                "النوع_الافتراضي": رمز_افتراضي,
                "عدد_التسميات": len(تسميات_محلية),
                "التسميات": تسميات_محلية,
            })
            عناوين_مستخدمة.append(نص_موحد)

        عناصر = self._جمع_مراكز_العناصر()
        if self.عناوين_المساقط and عناصر:
            نقاط = [(ع["x"], ع["y"]) for ع in عناصر]
            for فهارس in self._عنقدة_نقاط(نقاط, 5.5):
                if len(فهارس) < 3:
                    continue
                عناصر_العنقود = [عناصر[i] for i in فهارس]
                xs = [ع["x"] for ع in عناصر_العنقود]
                ys = [ع["y"] for ع in عناصر_العنقود]
                مركز_x = sum(xs) / len(xs)
                مركز_y = sum(ys) / len(ys)
                عنوان = self._أقرب_عنوان_للمجموعة(مركز_x, مركز_y)
                if not عنوان:
                    continue

                مجال = عنوان["المجال"]
                تسميات_محلية = []
                for تسمية in self.تسميات_الكمرات:
                    if not self._ينتمي_للمنطقة(تسمية["نص"], مجال):
                        continue
                    if not (min(xs) - 8.0 <= تسمية["x"] <= max(xs) + 8.0 and
                            min(ys) - 10.0 <= تسمية["y"] <= max(ys) + 10.0):
                        continue
                    أقرب_عنصر = min(
                        math.hypot(تسمية["x"] - ع["x"], تسمية["y"] - ع["y"])
                        for ع in عناصر_العنقود
                    )
                    if أقرب_عنصر <= 11.0:
                        تسميات_محلية.append(تسمية)

                رموز_محلية = [ت["نص"] for ت in تسميات_محلية]
                if مجال == "SUB":
                    if any(r.startswith("STB") for r in رموز_محلية):
                        رمز_افتراضي = "STB1"
                    else:
                        رمز_افتراضي = "TB1"
                else:
                    رمز_افتراضي = "B1"

                self.صناديق_المساقط.append({
                    "xmin": min(xs) - 2.0,
                    "ymin": min(ys) - 2.0,
                    "xmax": max(xs) + 2.0,
                    "ymax": max(ys) + 2.0,
                    "النوع_الافتراضي": رمز_افتراضي,
                    "عدد_التسميات": len(تسميات_محلية),
                    "التسميات": تسميات_محلية,
                    "المجال": مجال,
                    "العنوان": عنوان["نص"],
                })

        if not self.صناديق_المساقط:
            self._كوّن_صناديق_من_التسميات()

        مسجل.info(f"صناديق مساقط الكمرات المعتمدة: {len(self.صناديق_المساقط)}")

    def _صندوق_للنقطة(self, x: float, y: float) -> Optional[Dict]:
        for صندوق in self.صناديق_المساقط:
            if صندوق["xmin"] <= x <= صندوق["xmax"] and صندوق["ymin"] <= y <= صندوق["ymax"]:
                return صندوق
        return None

    def استخراج_تسميات_الكمرات(self):
        """يستخرج كل نصوص TB، STB، B من الرسم."""
        for t in self.msp.query('TEXT MTEXT'):
            try:
                خام = (t.dxf.text if t.dxftype()=='TEXT' else t.plain_text()).strip().upper()
                قيمة = re.sub(r'[\s\-_]+', '', خام)
                ins = t.dxf.insert
                نقطة = (ins.x * self.مقياس, ins.y * self.مقياس)
                if any(k in خام for k in [
                    "LAYOUT", "FOUNDATION", "SLAB LAYOUT", "ROOF SLAB",
                    "SCHEDULE", "SECTION", "DETAIL", "DESCRIPTION"
                ]):
                    self.نصوص_مرجعية.append({"نص": خام, "x": نقطة[0], "y": نقطة[1]})
                    مجال = self._مجال_المسقط(خام)
                    if مجال:
                        self.عناوين_المساقط.append({
                            "نص": خام, "x": نقطة[0], "y": نقطة[1], "المجال": مجال
                        })
                if any(k in خام for k in [
                    "SCHEDULE", "SECTION", "DETAIL", "REINFORCEMENT", "NOTES",
                    "TYPE", "SIZE", "TOP STEEL", "BOTTOM STEEL", "MIDSPAN",
                    "SUPPORT", "STIRRUP"
                ]):
                    self.نقاط_استبعاد.append(نقطة)
                if re.match(r'^(TB|STB|GB|B)\d+\*?$', قيمة):
                    self.تسميات_الكمرات.append({
                        "نص": قيمة,
                        "x": نقطة[0],
                        "y": نقطة[1]
                    })
            except Exception:
                pass
        مسجل.info(f"وُجد {len(self.تسميات_الكمرات)} تسمية كمرة في الرسم")

    def _ابحث_عن_أقرب_تسمية(self, مركز_x: float, مركز_y: float,
                              شجرة: Optional[KDTree],
                              نقاط_التسميات: List) -> Optional[str]:
        """يبحث عن أقرب تسمية للكمرة في نطاق محافظ لتقليل سحب التفاصيل المجاورة."""
        if شجرة is None or not نقاط_التسميات:
            return None
        مسافة, فهرس = شجرة.query((مركز_x, مركز_y))
        if مسافة <= min(8.0, نطاق_ربط_الكمرة):
            return self.تسميات_الكمرات[فهرس]["نص"]
        return None

    def _تصنيف_الكمرة(self, رمز: str) -> str:
        """يصنف الكمرة: TB أو STB أو B."""
        if رمز.startswith("STB") or "STRAP" in رمز:
            return "STB"
        elif رمز.startswith("TB") or "TIE" in رمز or "GB" in رمز:
            return "TB"
        else:
            return "B"

    def قراءة(self):
        self._اعتمد_مقياس_هندسي_إن_لزم()
        self.استخراج_تسميات_الكمرات()
        self._كوّن_صناديق_المساقط()

        # بناء KDTree من التسميات
        شجرة = None
        نقاط_ت = []
        if self.تسميات_الكمرات:
            نقاط_ت = [(ت["x"], ت["y"]) for ت in self.تسميات_الكمرات]
            شجرة = KDTree(نقاط_ت)
        شجرة_استبعاد = KDTree(self.نقاط_استبعاد) if self.نقاط_استبعاد else None

        أطوال_TB: Any = cast(Any, defaultdict(float))
        أطوال_STB: Any = cast(Any, defaultdict(float))
        أطوال_B: Any = cast(Any, defaultdict(float))

        for كيان in self.msp.query('LWPOLYLINE LINE'):
            try:
                طبقة = كيان.dxf.layer.upper()
                # فقط طبقات الكمرات
                if not any(ط in طبقة for ط in طبقات_الكمرات):
                    continue

                if كيان.dxftype() == 'LWPOLYLINE':
                    نقاط = [(ن[0]*self.مقياس, ن[1]*self.مقياس)
                             for ن in كيان.get_points(format='xy')]
                else:  # LINE
                    نقاط = [
                        (كيان.dxf.start.x*self.مقياس, كيان.dxf.start.y*self.مقياس),
                        (كيان.dxf.end.x*self.مقياس,   كيان.dxf.end.y*self.مقياس)
                    ]

                if len(نقاط) < 2:
                    continue

                # فلترة المسقط الرئيسي
                if not self._هل_في_المسقط_الرئيسي(نقاط):
                    continue

                # حساب الطول
                طول, اتجاه, _, _ = self._قياس_عنصر_كمرة(نقاط)
                if طول < 0.1:  # تجاهل الخطوط القصيرة جداً
                    continue

                # مركز الكمرة
                مركز_x = sum(ن[0] for ن in نقاط) / len(نقاط)
                مركز_y = sum(ن[1] for ن in نقاط) / len(نقاط)

                صندوق = self._صندوق_للنقطة(مركز_x, مركز_y)
                if صندوق is None and self.صناديق_المساقط:
                    continue

                if شجرة_استبعاد is not None:
                    مسافة_استبعاد, _ = شجرة_استبعاد.query((مركز_x, مركز_y))
                    if مسافة_استبعاد <= 8.0:
                        continue

                # البحث عن أقرب تسمية
                رمز = None
                if صندوق is not None:
                    رمز = self._ابحث_عن_تسمية_محلية(مركز_x, مركز_y, اتجاه, صندوق)
                if not رمز:
                    رمز = self._ابحث_عن_أقرب_تسمية(مركز_x, مركز_y, شجرة, نقاط_ت)
                if not رمز:
                    # محاولة التصنيف من اسم الطبقة
                    if صندوق is not None:
                        رمز = صندوق["النوع_الافتراضي"]
                    elif "STB" in طبقة or "STRAP" in طبقة:
                        رمز = "STB1"
                    elif "TB" in طبقة or "TIE" in طبقة or "GB" in طبقة:
                        رمز = "TB1"
                    else:
                        رمز = "B1"

                نوع = self._تصنيف_الكمرة(رمز)
                if نوع == "TB":
                    أطوال_TB[رمز] += طول
                elif نوع == "STB":
                    أطوال_STB[رمز] += طول
                else:
                    أطوال_B[رمز] += طول

            except Exception as خطأ:
                مسجل.warning(f"خطأ في قراءة الكمرة: {خطأ}")

        احتياطي_tb, احتياطي_stb = self._استخرج_سنترلانات_tb_احتياطي()
        for رمز, طول in احتياطي_tb.items():
            if طول > (أطوال_TB.get(رمز, 0.0) * 1.10):
                أطوال_TB[رمز] = طول
            elif رمز not in أطوال_TB:
                أطوال_TB[رمز] = طول
        for رمز, طول in احتياطي_stb.items():
            if طول > (أطوال_STB.get(رمز, 0.0) * 1.10):
                أطوال_STB[رمز] = طول
            elif رمز not in أطوال_STB:
                أطوال_STB[رمز] = طول

        # بناء القوائم النهائية مع ربط الجدول
        for رمز_raw, طول_raw in أطوال_TB.items():
            رمز: str = str(رمز_raw)
            طول_ف: float = float(طول_raw)
            بيانات_جدول = self.جدول_tb.get(رمز, self.جدول_tb.get("TB1", {}))
            self.كمرات_الربط.append(بيانات_كمرة_ربط(
                الرمز=رمز,
                الطول=round(طول_ف, 3),
                العرض=float(بيانات_جدول.get("عرض", 0.20)),
                العمق=float(بيانات_جدول.get("عمق", 0.50))
            ))

        for رمز_raw, طول_raw in أطوال_STB.items():
            رمز: str = str(رمز_raw)
            طول_ف: float = float(طول_raw)
            بيانات_جدول = self.جدول_stb.get(رمز, self.جدول_stb.get("STB1", {}))
            self.كمرات_الستراب.append(بيانات_كمرة_ستراب(
                الرمز=رمز,
                الطول=round(طول_ف, 3),
                العرض=float(بيانات_جدول.get("عرض", 0.20)),
                العمق=float(بيانات_جدول.get("عمق", 0.50))
            ))

        for رمز_raw, طول_raw in أطوال_B.items():
            رمز: str = str(رمز_raw)
            طول_ف: float = float(طول_raw)
            بيانات_جدول = self.جدول_b.get(رمز, self.جدول_b.get("B1", {}))
            self.الكمرات.append(بيانات_كمرة(
                الرمز=رمز,
                الطول=round(طول_ف, 3),
                العرض=float(بيانات_جدول.get("عرض", 0.25)),
                العمق=float(بيانات_جدول.get("عمق", 0.60))
            ))

        مسجل.info(
            f"الكمرات من الرسم: "
            f"{len(self.كمرات_الربط)} TB | "
            f"{len(self.كمرات_الستراب)} STB | "
            f"{len(self.الكمرات)} B"
        )


class قارئ_الأعمدة_من_الرسم:
    """
    يقرأ Column Layout من الرسم مباشرةً:
    - يحدد عناقيد C1/C2/C3... القريبة من عنوان COLUMN LAYOUT
    - ينظف صف الجدول الظاهر داخل نفس المنطقة
    - يحدّث كميات الأعمدة إن كانت الأبعاد مقروءة من جدول الأعمدة
    """

    def __init__(self, msp, مقياس: float, أعمدة_الجدول: Optional[List[بيانات_العمود]] = None):
        self.msp = msp
        self.مقياس = مقياس
        self.أعمدة_الجدول = أعمدة_الجدول or []
        self.تسميات_الأعمدة: List[Dict] = []
        self.نصوص_مرجعية: List[Dict] = []
        self.كل_النصوص: List[Dict] = []
        self.إحصاء_حسب_المستوى: Dict[str, Dict[str, int]] = {}
        self.مراكز_المستويات: Dict[str, Tuple[float, float]] = {}
        self.مقاطع_محلية_حسب_المستوى: Dict[str, Counter] = {}
        self.مقاطع_حسب_المستوى_والرمز: Dict[Tuple[str, str], Counter] = {}
        self.خريطة_أبعاد_مستدلة: Dict[str, Tuple[float, float]] = {}
        self.صناديق_المساقط: List[Dict] = []
        self.الأعمدة: List[بيانات_العمود] = []

    @staticmethod
    def _عنقدة_نقاط(نقاط: List[Tuple[float, float]], مسافة: float = 10.0) -> List[List[int]]:
        if not نقاط:
            return []
        مزارة = [False] * len(نقاط)
        عناقيد: List[List[int]] = []
        for فهرس in range(len(نقاط)):
            if مزارة[فهرس]:
                continue
            مزارة[فهرس] = True
            مكدس = [فهرس]
            عنقود: List[int] = []
            while مكدس:
                i = مكدس.pop()
                عنقود.append(i)
                x1, y1 = نقاط[i]
                for j, (x2, y2) in enumerate(نقاط):
                    if مزارة[j]:
                        continue
                    if math.hypot(x2 - x1, y2 - y1) <= مسافة:
                        مزارة[j] = True
                        مكدس.append(j)
            عناقيد.append(عنقود)
        return عناقيد

    @staticmethod
    def _طبع_رمز(رمز: str) -> str:
        رمز = re.sub(r'[\s\-_]+', '', رمز.upper().strip())
        رمز = رمز.replace("&PC", "").replace("*", "")
        if رمز.startswith("BC") and re.match(r'^BC\d+[A-Z]?$', رمز):
            رمز = رمز[1:]
        if رمز.startswith("B-"):
            رمز = رمز[2:]
        if رمز in {"C5/DC", "C5DC"}:
            return "C5"
        if رمز == "NC":
            return "NC"
        مطابقة = re.match(r'^(C\d+[A-Z]?)', رمز)
        if مطابقة:
            return مطابقة.group(1)
        مطابقة = re.match(r'^(C\d+(?:/[A-Z0-9]+)?)', رمز)
        if مطابقة:
            return مطابقة.group(1)
        مطابقة = re.match(r'^(DC\d*|NC\d*)', رمز)
        if مطابقة:
            return مطابقة.group(1)
        return رمز

    @staticmethod
    def _نص_موحد(نص: str) -> str:
        return " ".join(str(نص).upper().split())

    @staticmethod
    def _داخل_صندوق(x: float, y: float, صندوق: Dict) -> bool:
        return صندوق["xmin"] <= x <= صندوق["xmax"] and صندوق["ymin"] <= y <= صندوق["ymax"]

    @staticmethod
    def _تقريب_5سم(قيمة: float) -> float:
        return round(round(قيمة / 0.05) * 0.05, 2)

    def _مرشحو_المقياس_الهندسي(self) -> List[float]:
        مرشحون: List[float] = []
        for قيمة in (self.مقياس, 1.0, 0.1, 0.01, 0.001):
            if قيمة > 0 and قيمة not in مرشحون:
                مرشحون.append(قيمة)
        return مرشحون

    def _اختبر_مقياس_هندسي(self, مقياس_مرشح: float) -> int:
        عدد_صالح = 0
        for e in self.msp.query('LWPOLYLINE'):
            try:
                طبقة = e.dxf.layer.upper()
                if طبقة not in {"COLUMN", "COL", "COLUMNS", "0 COL", "HIDDEN", "CONC"}:
                    continue
                نقاط = [(ن[0] * مقياس_مرشح, ن[1] * مقياس_مرشح) for ن in e.get_points(format='xy')]
                if len(نقاط) < 2:
                    continue
                xs = [ن[0] for ن in نقاط]
                ys = [ن[1] for ن in نقاط]
                أ = self._تقريب_5سم(min(max(xs) - min(xs), max(ys) - min(ys)))
                ب = self._تقريب_5سم(max(max(xs) - min(xs), max(ys) - min(ys)))
                if 0.15 <= أ <= 0.60 and 0.15 <= ب <= 1.20:
                    عدد_صالح += 1
            except Exception:
                pass
        return عدد_صالح

    def _اعتمد_مقياس_هندسي_إن_لزم(self):
        المقياس_الحالي = self.مقياس
        الدرجة_الحالية = self._اختبر_مقياس_هندسي(المقياس_الحالي)
        أفضل_مقياس = المقياس_الحالي
        أفضل_درجة = الدرجة_الحالية

        for مقياس_مرشح in self._مرشحو_المقياس_الهندسي():
            درجة = self._اختبر_مقياس_هندسي(مقياس_مرشح)
            if درجة > أفضل_درجة:
                أفضل_درجة = درجة
                أفضل_مقياس = مقياس_مرشح

        if أفضل_مقياس != المقياس_الحالي and أفضل_درجة >= max(6, الدرجة_الحالية * 3):
            self.مقياس = أفضل_مقياس
            مسجل.info(
                f"تم تعديل مقياس قراءة الأعمدة من {المقياس_الحالي:g} إلى {أفضل_مقياس:g} "
                f"اعتماداً على {أفضل_درجة} مقطع عمود مرشح"
            )

    def _استخرج_التسميات(self):
        for t in self.msp.query('TEXT MTEXT'):
            try:
                خام = (t.dxf.text if t.dxftype() == 'TEXT' else t.plain_text()).strip().upper()
                قيمة = re.sub(r'[\s\-_]+', '', خام)
                ins = t.dxf.insert
                نقطة = (ins.x * self.مقياس, ins.y * self.مقياس)
                طبقة = t.dxf.layer.upper()
                self.كل_النصوص.append({"نص": خام, "x": نقطة[0], "y": نقطة[1], "الطبقة": طبقة})
                if "COLUMN LAYOUT" in خام or "SCHEDULE OF COLUMNS" in خام or "COLUMN SCHEDULE" in خام or "COL LAYOUT" in خام:
                    self.نصوص_مرجعية.append({"نص": خام, "x": نقطة[0], "y": نقطة[1], "الطبقة": طبقة})
                # Extended column mark regex: C1, C2, C3A, C3B, C4/DC, DC, NC
                if re.match(r'^(?:B)?(C\d+[A-Z]?\*?|C\d+(?:/[A-Z0-9]+)?|C\d+&PC|C5/DC|DC\d*|NC\d*)$', قيمة):
                    self.تسميات_الأعمدة.append({"نص": قيمة, "x": نقطة[0], "y": نقطة[1], "الطبقة": طبقة})
            except Exception:
                pass

    def _صناديق_مستويات_من_العناوين(self) -> List[Dict]:
        عناوين = [ن for ن in self.نصوص_مرجعية if "COLUMN LAYOUT" in self._نص_موحد(ن["نص"])]
        if not عناوين:
            return []
        صناديق: List[Dict] = []
        عناوين_مرتبة = sorted(عناوين, key=lambda ن: (-ن["y"], ن["x"]))
        for عنوان in عناوين_مرتبة:
            نص = self._نص_موحد(عنوان["نص"])
            صناديق.append({
                "المستوى": استخرج_نص_طابق_موحد(نص),
                "العنوان": عنوان["نص"],
                "xmin": عنوان["x"] - 48.0,
                "xmax": عنوان["x"] + 12.0,
                "ymin": عنوان["y"] - 8.0,
                "ymax": عنوان["y"] + 32.0,
            })
        return صناديق

    def _حدد_المستوى(self, عناصر: List[Dict]) -> str:
        if not عناصر:
            return ""
        مركز_x = sum(ع["x"] for ع in عناصر) / len(عناصر)
        مركز_y = sum(ع["y"] for ع in عناصر) / len(عناصر)
        أقرب = ""
        أقرب_مسافة = 1e9
        for ن in self.نصوص_مرجعية:
            مسافة = math.hypot(ن["x"] - مركز_x, ن["y"] - مركز_y)
            if مسافة < أقرب_مسافة:
                أقرب_مسافة = مسافة
                أقرب = ن["نص"]
        if not أقرب or أقرب_مسافة > 35.0:
            return ""
        return استخرج_نص_طابق_موحد(أقرب)

    def _نظف_عنقود_الأعمدة(self, عناصر: List[Dict]) -> List[Dict]:
        if len(عناصر) < 8:
            return عناصر
        xs = [ع["x"] for ع in عناصر]
        if max(xs) - min(xs) < 6.0:
            return عناصر
        عتبة_يمين = max(xs) - 2.5
        منقى = [ع for ع in عناصر if ع["x"] < عتبة_يمين]
        return منقى if len(منقى) >= max(1, len(عناصر) - 6) else عناصر

    def _استخرج_مقاطع_مرشحة(self, مركز: Optional[Tuple[float, float]] = None,
                             نصف_قطر: float = 15.0) -> Counter:
        مرشحات: Counter = Counter()
        for e in self.msp.query('LWPOLYLINE'):
            try:
                طبقة = e.dxf.layer.upper()
                if طبقة not in {"COLUMN", "COL", "COLUMNS", "0 COL", "HIDDEN", "CONC"}:
                    continue
                نقاط = [(ن[0] * self.مقياس, ن[1] * self.مقياس) for ن in e.get_points(format='xy')]
                if len(نقاط) < 2:
                    continue
                xs = [ن[0] for ن in نقاط]
                ys = [ن[1] for ن in نقاط]
                مركز_x = (min(xs) + max(xs)) / 2.0
                مركز_y = (min(ys) + max(ys)) / 2.0
                if مركز is not None and math.hypot(مركز_x - مركز[0], مركز_y - مركز[1]) > نصف_قطر:
                    continue
                أ = self._تقريب_5سم(min(max(xs) - min(xs), max(ys) - min(ys)))
                ب = self._تقريب_5سم(max(max(xs) - min(xs), max(ys) - min(ys)))
                if مركز is None:
                    صالح = 0.15 <= أ <= 0.60 and 0.15 <= ب <= 1.20
                else:
                    صالح = 0.15 <= أ <= 0.60 and 0.15 <= ب <= 1.20
                if صالح:
                    مرشحات[(أ, ب)] += 1
            except Exception:
                pass
        return مرشحات

    def _استخرج_درجات_التسليح(self) -> Dict[str, float]:
        درجات: Dict[str, float] = {}
        for رأس in self.نصوص_مرجعية:
            if "SCHEDULE OF COLUMNS" not in رأس["نص"] and "COLUMN SCHEDULE" not in رأس["نص"]:
                continue
            صفوف = [
                ن for ن in self.كل_النصوص
                if abs(ن["x"] - رأس["x"]) <= 3.5
                and رأس["y"] - 22.0 <= ن["y"] <= رأس["y"] - 2.0
                and re.match(r'^(C\d+[A-Z]?|C\d+(?:/[A-Z0-9]+)?|C5/DC|DC\d*|NC\d*)$', re.sub(r'[\s\-_]+', '', ن["نص"]))
            ]
            for صف in صفوف:
                رمز = self._طبع_رمز(re.sub(r'[\s\-_]+', '', صف["نص"]))
                أفضل = 0.0
                for ن in self.كل_النصوص:
                    if abs(ن["y"] - صف["y"]) > 1.0:
                        continue
                    if not (0.0 < ن["x"] - صف["x"] < 8.5):
                        continue
                    تطابق = re.search(r'(\d+)\s*T(\d+)', ن["نص"])
                    if تطابق:
                        عدد = int(تطابق.group(1))
                        قطر = int(تطابق.group(2))
                        أفضل = max(أفضل, عدد * (قطر ** 2))
                if أفضل > 0:
                    درجات[رمز] = max(درجات.get(رمز, 0.0), أفضل)
        return درجات

    def _ابن_خريطة_الأبعاد_الاستدلالية(self) -> Dict[str, Tuple[float, float]]:
        رموز = sorted({
            رمز for عدادات in self.إحصاء_حسب_المستوى.values()
            for رمز in عدادات.keys()
        })
        if not رموز:
            return {}

        درجات = self._استخرج_درجات_التسليح()
        مرشحات = self._استخرج_مقاطع_مرشحة()
        أحجام = [حجم for حجم, _ in sorted(مرشحات.items(), key=lambda kv: (kv[0][0] * kv[0][1], kv[1]))]
        if not أحجام:
            return {}

        # رموز أصغر تسليحاً ← مقاطع أصغر. هذا مسار احتياطي حين يغيب نص الأبعاد.
        ترتيب_الرموز = sorted(رموز, key=lambda ر: (درجات.get(ر, 0.0), ر))
        خريطة: Dict[str, Tuple[float, float]] = {}
        مقام = max(1, len(ترتيب_الرموز) - 1)
        for i, رمز in enumerate(ترتيب_الرموز):
            فهرس = round(i * (len(أحجام) - 1) / مقام)
            خريطة[رمز] = أحجام[فهرس]
        return خريطة

    def _ابن_أعمدة_بأبعاد_الجدول(self):
        if not self.إحصاء_حسب_المستوى:
            return
        خريطة_جدول = {ع.الرمز.upper(): (ع.الطول, ع.العرض) for ع in self.أعمدة_الجدول if ع.الرمز and ع.الطول and ع.العرض}
        خريطة_أساس_للجدول: Dict[str, Tuple[float, float]] = {}
        for رمز, أبعاد in خريطة_جدول.items():
            مطابقة_أساس = re.match(r'^(C\d+|DC\d*|NC\d*)', رمز)
            if not مطابقة_أساس:
                continue
            رمز_أساس = مطابقة_أساس.group(1)
            قائم = خريطة_أساس_للجدول.get(رمز_أساس)
            if قائم is None or (أبعاد[0] * أبعاد[1]) > (قائم[0] * قائم[1]):
                خريطة_أساس_للجدول[رمز_أساس] = أبعاد
        self.خريطة_أبعاد_مستدلة = self._ابن_خريطة_الأبعاد_الاستدلالية() if not خريطة_جدول else {}
        for مستوى, عدادات in self.إحصاء_حسب_المستوى.items():
            for رمز, كمية in عدادات.items():
                أبعاد = خريطة_جدول.get(رمز.upper())
                if not أبعاد:
                    مطابقة_أساس = re.match(r'^(C\d+|DC\d*|NC\d*)', رمز.upper())
                    if مطابقة_أساس:
                        أبعاد = خريطة_أساس_للجدول.get(مطابقة_أساس.group(1))
                if not أبعاد:
                    محلية_حسب_الرمز = self.مقاطع_حسب_المستوى_والرمز.get((مستوى, رمز.upper()), Counter())
                    if محلية_حسب_الرمز:
                        أبعاد = محلية_حسب_الرمز.most_common(1)[0][0]
                if not أبعاد and len(عدادات) == 1:
                    محلية = self.مقاطع_محلية_حسب_المستوى.get(مستوى, Counter())
                    if محلية:
                        أبعاد = محلية.most_common(1)[0][0]
                if not أبعاد:
                    أبعاد = self.خريطة_أبعاد_مستدلة.get(رمز.upper())
                if not أبعاد:
                    continue
                self.الأعمدة.append(بيانات_العمود(
                    الرمز=رمز.upper(),
                    الطول=max(أبعاد),
                    العرض=min(أبعاد),
                    الكمية=كمية,
                    المستوى=مستوى,
                ))

    def قراءة(self):
        self._اعتمد_مقياس_هندسي_إن_لزم()
        self._استخرج_التسميات()
        self.صناديق_المساقط = self._صناديق_مستويات_من_العناوين()
        for صندوق in self.صناديق_المساقط:
            عناصر_هيكلية = []
            عناصر_نصية = []
            for ت in self.تسميات_الأعمدة:
                if not self._داخل_صندوق(ت["x"], ت["y"], صندوق):
                    continue
                if ت.get("الطبقة") == "TEXT":
                    مقطع_قريب = self._استخرج_مقاطع_مرشحة((ت["x"], ت["y"]), 1.5)
                    if not مقطع_قريب:
                        continue
                    عناصر_نصية.append(ت)
                else:
                    عناصر_هيكلية.append(ت)
            عناصر = عناصر_هيكلية or عناصر_نصية
            if len(عناصر) < 3:
                continue
            عدادات = Counter(self._طبع_رمز(ع["نص"]) for ع in عناصر)
            if not عدادات:
                continue
            مستوى = صندوق["المستوى"]
            عدادات_مستوى = Counter(self.إحصاء_حسب_المستوى.get(مستوى, {}))
            عدادات_مستوى.update(عدادات)
            self.إحصاء_حسب_المستوى[مستوى] = dict(عدادات_مستوى)
            مركز_x = sum(ع["x"] for ع in عناصر) / len(عناصر)
            مركز_y = sum(ع["y"] for ع in عناصر) / len(عناصر)
            self.مراكز_المستويات[مستوى] = (مركز_x, مركز_y)
            مقاطع_محلية = self._استخرج_مقاطع_مرشحة((مركز_x, مركز_y), 15.0)
            if مستوى not in self.مقاطع_محلية_حسب_المستوى:
                self.مقاطع_محلية_حسب_المستوى[مستوى] = Counter()
            self.مقاطع_محلية_حسب_المستوى[مستوى].update(مقاطع_محلية)
            for عنصر in عناصر:
                رمز = self._طبع_رمز(عنصر["نص"])
                مقاطع_قريبة = self._استخرج_مقاطع_مرشحة((عنصر["x"], عنصر["y"]), 1.5)
                if مقاطع_قريبة:
                    مفتاح = (مستوى, رمز.upper())
                    if مفتاح not in self.مقاطع_حسب_المستوى_والرمز:
                        self.مقاطع_حسب_المستوى_والرمز[مفتاح] = Counter()
                    self.مقاطع_حسب_المستوى_والرمز[مفتاح].update(مقاطع_قريبة)

        if not self.إحصاء_حسب_المستوى:
            نقاط = [(ت["x"], ت["y"]) for ت in self.تسميات_الأعمدة]
            for فهارس in sorted(self._عنقدة_نقاط(نقاط, 10.0), key=len, reverse=True):
                عناصر = [self.تسميات_الأعمدة[i] for i in فهارس]
                if len(عناصر) < 6:
                    continue
                مستوى = self._حدد_المستوى(عناصر)
                if not مستوى:
                    continue
                منقى = self._نظف_عنقود_الأعمدة(عناصر)
                عدادات = Counter(self._طبع_رمز(ع["نص"]) for ع in منقى)
                if عدادات:
                    self.إحصاء_حسب_المستوى[مستوى] = dict(عدادات)
                    مركز_x = sum(ع["x"] for ع in منقى) / len(منقى)
                    مركز_y = sum(ع["y"] for ع in منقى) / len(منقى)
                    self.مراكز_المستويات[مستوى] = (مركز_x, مركز_y)
                    self.مقاطع_محلية_حسب_المستوى[مستوى] = self._استخرج_مقاطع_مرشحة((مركز_x, مركز_y), 15.0)
        self._ابن_أعمدة_بأبعاد_الجدول()
        if self.إحصاء_حسب_المستوى and not self.الأعمدة:
            مسجل.warning("تم عدّ الأعمدة من المسقط، لكن أبعاد القطاعات لم تُقرأ من جدول الأعمدة بعد.")
        elif self.خريطة_أبعاد_مستدلة:
            مسجل.warning(f"تم استخدام أبعاد أعمدة استدلالية لرموز: {sorted(self.خريطة_أبعاد_مستدلة.keys())}")
        مسجل.info(
            f"قارئ الأعمدة من الرسم: "
            f"{sum(sum(v.values()) for v in self.إحصاء_حسب_المستوى.values())} تسمية صالحة | "
            f"{len(self.الأعمدة)} عمود مدمج مع الأبعاد"
        )


# ─────────────────────────────────────────────────────────────────────────────
# كاشف الجدران — مُحسَّن (إصلاح O(n²))
# ─────────────────────────────────────────────────────────────────────────────

class كاشف_الجدران:
    """
    يكتشف أزواج الخطوط المتوازية لتصنيف سماكة الجدار.
    V15.2: حدود مرتفعة (15000 خط، 150 جار) + KDTree مكاني للبحث عن الأزواج القريبة.
    """

    def __init__(self, خطوط: List[LineString]):
        self.خطوط    = خطوط
        self.جدران_20: List[LineString] = []
        self.جدران_10: List[LineString] = []

    @staticmethod
    def _زاوية(خط: LineString) -> float:
        إ = list(خط.coords)
        return math.atan2(إ[-1][1]-إ[0][1], إ[-1][0]-إ[0][0]) % math.pi

    @staticmethod
    def _متوازيان(ز1: float, ز2: float, تسامح: float = 0.05) -> bool:
        فرق = abs(ز1 - ز2)
        return فرق < تسامح or abs(فرق - math.pi) < تسامح

    @staticmethod
    def _مسافة_عمودية(خ1: LineString, خ2: LineString) -> float:
        return خ2.distance(خ1.interpolate(0.5, normalized=True))

    @staticmethod
    def _خط_مركز(خ1: LineString, خ2: LineString) -> LineString:
        إ1, إ2 = list(خ1.coords), list(خ2.coords)
        return LineString([
            ((إ1[0][0]+إ2[0][0])/2, (إ1[0][1]+إ2[0][1])/2),
            ((إ1[-1][0]+إ2[-1][0])/2, (إ1[-1][1]+إ2[-1][1])/2)
        ])

    def اكتشاف(self):
        """
        V15.2: تحسينات جذرية:
        1. حد أعلى 15000 خط بدلاً من 5000
        2. فحص 150 جار بدلاً من 50
        3. KDTree مكاني لتسريع البحث عن الأزواج القريبة ضمن كل مجموعة زاوية
        """
        if len(self.خطوط) > 15000:
            # ترتيب حسب الطول (الأطول أهم) ثم أخذ أفضل 15000
            خطوط_مرتبة = sorted(self.خطوط, key=lambda خ: خ.length, reverse=True)
            خطوط_فحص = خطوط_مرتبة[:15000]
            مسجل.warning(f"عدد الخطوط كبير ({len(self.خطوط)}) — تحليل أطول 15000 خط")
        else:
            خطوط_فحص = self.خطوط

        # فلترة الخطوط القصيرة جداً (أقل من 0.3 متر) — ضوضاء لا تمثل جدراناً
        خطوط_فحص = [خ for خ in خطوط_فحص if خ.length >= 0.3]

        # تجميع بالزاوية (كل 5 درجات)
        مجموعات: Dict[int, List[int]] = defaultdict(list)
        زوايا = []
        مراكز = []
        for i, خط in enumerate(خطوط_فحص):
            ز = self._زاوية(خط)
            زوايا.append(ز)
            مركز = خط.interpolate(0.5, normalized=True)
            مراكز.append((مركز.x, مركز.y))
            مجموعات[int(math.degrees(ز) // 5)].append(i)

        مُعالَج = set()
        for مجموعة in مجموعات.values():
            if len(مجموعة) < 2:
                continue

            # بناء KDTree مكاني لمراكز خطوط هذه المجموعة — للبحث السريع عن اقتراب
            إحداثيات_المجموعة = np.array([مراكز[i] for i in مجموعة])
            شجرة_المجموعة = KDTree(إحداثيات_المجموعة)

            for idx_في_المجموعة, إ in enumerate(مجموعة):
                # البحث عن أقرب 150 خط ضمن مسافة 0.5 متر (أقصى سماكة جدار ممكنة)
                مسافات, فهارس = شجرة_المجموعة.query(
                    إحداثيات_المجموعة[idx_في_المجموعة], k=min(150, len(مجموعة))
                )
                if isinstance(مسافات, (int, float)):
                    مسافات = [مسافات]
                    فهارس = [فهارس]
                for مسافة_تقريبية, idx_جار in zip(مسافات, فهارس):
                    ج = مجموعة[idx_جار]
                    if إ >= ج:
                        continue
                    # تصفية أولية — المسافة التقريبية بين المراكز > 1.5 متر → لا يمكن أن يكون جدار
                    if مسافة_تقريبية > 1.5:
                        continue
                    if (إ, ج) in مُعالَج:
                        continue
                    مُعالَج.add((إ, ج))
                    if not self._متوازيان(زوايا[إ], زوايا[ج]):
                        continue
                    مسافة = self._مسافة_عمودية(خطوط_فحص[إ], خطوط_فحص[ج])
                    if حد_جدار_20_أدنى <= مسافة <= حد_جدار_20_أعلى:
                        self.جدران_20.append(self._خط_مركز(خطوط_فحص[إ], خطوط_فحص[ج]))
                    elif حد_جدار_10_أدنى <= مسافة <= حد_جدار_10_أعلى:
                        self.جدران_10.append(self._خط_مركز(خطوط_فحص[إ], خطوط_فحص[ج]))

    @property
    def طول_20(self) -> float:
        return round(sum(ج.length for ج in self.جدران_20), 2)

    @property
    def طول_10(self) -> float:
        return round(sum(ج.length for ج in self.جدران_10), 2)


# ─────────────────────────────────────────────────────────────────────────────
# Gemini Vision — لقراءة PDF المسح الضوئي
# ─────────────────────────────────────────────────────────────────────────────

class قارئ_gemini_vision:
    """
    يستخدم Gemini Vision API لاستخراج البيانات من PDF المسح الضوئي.
    يُستدعى فقط عند فشل الاستخراج المتجه.
    """

    def __init__(self, مفتاح_api: str):
        self.مفتاح = مفتاح_api
        # Updated to Gemini 2.0 Flash for superior vision and speed
        self.رابط  = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={مفتاح_api}"

    def استخراج_من_صفحة(self, صورة_bytes: bytes) -> Dict:
        """يرسل صورة الصفحة لـ Gemini ويطلب استخراج بيانات QTO."""
        صورة_base64 = base64.b64encode(صورة_bytes).decode('utf-8')

        # Bilingual Prompt for Global Drawing Standards
        الطلب = {
            "contents": [{
                "parts": [
                    {
                        "inline_data": {
                            "mime_type": "image/png",
                            "data": صورة_base64
                        }
                    },
                    {
                        "text": """
You are a UAE Quantity Surveying expert. Analyze this structural or architectural drawing.
أنت خبير حساب كميات إماراتي متخصص في قراءة المخططات الإنشائية والمعمارية.

Extract the following data in JSON format ONLY. Do not add any text, backticks or markdown.
استخرج البيانات التالية بتنسيق JSON فقط بدون أي نص إضافي أو علامات.

{
  "الأسس": [{"الرمز": "F1", "الطول": 1.5, "العرض": 1.5, "العمق": 0.6, "الكمية": 4}],
  "أعمدة_العنق": [{"الرمز": "NC1", "الطول": 0.3, "العرض": 0.3, "الكمية": 4}],
  "كمرات_الربط": [{"الرمز": "TB1", "الطول": 5.0, "العرض": 0.2, "العمق": 0.5}],
  "كمرات_الستراب": [{"الرمز": "STB1", "الطول": 4.0, "العرض": 0.2, "العمق": 0.5}],
  "الأعمدة": [{"الرمز": "C1", "الطول": 0.3, "العرض": 0.3, "الارتفاع": 3.0, "الكمية": 4}],
  "الكمرات": [{"الرمز": "B1", "الطول": 5.0, "العرض": 0.25, "العمق": 0.6}],
  "الفتحات": [{"الرمز": "D1", "الطول": 0.9, "الارتفاع": 2.1, "الكمية": 2}],
  "المحيط_الخارجي": 0,
  "المساحة_الكلية": 0
}

Rules:
- All dimensions must be in METERS (convert from mm/cm if necessary: divide mm by 1000, divide cm by 100).
- If an element is not found, return an empty list [].
- Strictly return JSON only.
- Handle English drawing labels (e.g. F, C, TB, B, D, W) and match them to the keys above.
- For الأعمدة (columns): الكمية is the TOTAL count of that column type across ALL floors in the entire building (not per floor). الارتفاع is the clear height of one column in meters from schedule or section; use 0 if not found.
- For كمرات_الربط (tie beams) and الكمرات (beams): الطول is the TOTAL LENGTH of ALL beams of that type combined (sum from schedule or estimate from plan). Minimum 1.0 m.
- For المحيط_الخارجي: measure or estimate the outer perimeter of the building in meters from the plan.
- For المساحة_الكلية: measure or estimate the total floor area in square meters.
"""
                    }
                ]
            }]
        }

        try:
            استجابة = requests.post(self.رابط, json=الطلب, timeout=30)
            استجابة.raise_for_status()
            نص = استجابة.json()["candidates"][0]["content"]["parts"][0]["text"]
            # تنظيف الاستجابة
            نص = re.sub(r'```json|```', '', نص).strip()
            بيانات = json.loads(نص) if نص else {}
            
            # خريطة الترجمة لضمان عمل المحرك العربي مع مفاتيح إنجليزية إن وجدت
            خريطة = {
                "foundations": "الأسس", "neck_columns": "أعمدة_العنق",
                "tie_beams": "كمرات_الربط", "strap_beams": "كمرات_الستراب",
                "columns": "الأعمدة", "beams": "الكمرات", "openings": "الفتحات",
                "perimeter": "المحيط_الخارجي", "total_area": "المساحة_الكلية"
            }
            نتيجة_منقحة = {}
            for k, v in بيانات.items():
                مفتاح_عربي = خريطة.get(k.lower(), k)
                نتيجة_منقحة[مفتاح_عربي] = v
            return نتيجة_منقحة
        except Exception as خطأ:
            مسجل.error(f"خطأ Gemini: {خطأ}")
            return {}

    def استخراج_من_pdf(self, مسار_pdf: str) -> Dict:
        """يحول كل صفحة PDF لصورة ويرسلها لـ Gemini."""
        import json
        نتيجة_مجمعة = {
            "الأسس": [], "أعمدة_العنق": [],
            "كمرات_الربط": [], "كمرات_الستراب": [],
            "الأعمدة": [], "الكمرات": [], "الفتحات": [],
            "المحيط_الخارجي": 0, "المساحة_الكلية": 0
        }
        وثيقة = fitz.open(مسار_pdf)
        for صفحة in وثيقة:
            صورة = صفحة.get_pixmap(dpi=150)
            بيانات = self.استخراج_من_صفحة(صورة.tobytes("png"))
            # دمج النتائج
            for مفتاح in ["الأسس", "أعمدة_العنق", "كمرات_الربط",
                           "كمرات_الستراب", "الأعمدة", "الكمرات", "الفتحات"]:
                نتيجة_مجمعة[مفتاح].extend(بيانات.get(مفتاح, []))
            if بيانات.get("المحيط_الخارجي", 0) > نتيجة_مجمعة["المحيط_الخارجي"]:
                نتيجة_مجمعة["المحيط_الخارجي"] = بيانات["المحيط_الخارجي"]
            if بيانات.get("المساحة_الكلية", 0) > نتيجة_مجمعة["المساحة_الكلية"]:
                نتيجة_مجمعة["المساحة_الكلية"] = بيانات["المساحة_الكلية"]
        return نتيجة_مجمعة


# ─────────────────────────────────────────────────────────────────────────────
# محرك الفضاء الدلالي — V14
# ─────────────────────────────────────────────────────────────────────────────

class محرك_الفضاء:

    def __init__(self, مسار: str, وحدة: str, مفتاح_gemini: str = "", pdf_drawing_scale: float = 0.0):
        self.مسار          = مسار
        self.وحدة          = وحدة
        self.مقياس         = {"m": 1.0, "cm": 0.01, "mm": 0.001}.get(وحدة.lower(), 0.001)
        self.مفتاح_gemini  = مفتاح_gemini
        self.pdf_drawing_scale = pdf_drawing_scale  # e.g. 100 means 1:100
        self.وثيقة_dxf: Optional[ezdxf.document.Drawing] = None
        self.خطوط:  List[LineString] = []
        self.نصوص:  List[Dict]       = []
        self.نصوص_خام: List[Dict]    = []
        self.مضلعات: List[Polygon]   = []
        self.الغرف:  List[Dict]      = []
        self.مساحات_البلاطات: List[Dict] = []

        # نتائج القراءة
        self.قارئ: Optional[قارئ_الجداول]      = None
        self.كمرات: Optional[قارئ_كمرات_الرسم] = None
        self.قارئ_أعمدة_الرسم: Optional[قارئ_الأعمدة_من_الرسم] = None
        self.بيانات_gemini: Dict                = {}
        self.عداد_الفتحات: Dict[str, int]      = {}
        self.عداد_الفتحات_حسب_المستوى: Dict[str, Dict[str, int]] = {}
        self.صناديق_المستويات: List[Dict]      = []
        self.ملف_المشروع_المستخرج: Dict        = {}
        self.الطبقات_المكتشفة: List[str]     = []

    def _طبقة_صالحة(self, طبقة: str) -> bool:
        return not any(م in طبقة.upper() for م in طبقات_مُهملة)

    def _ابن_صناديق_المستويات(self, msp) -> List[Dict]:
        عناوين: List[Dict] = []
        for نص in msp.query('TEXT MTEXT'):
            try:
                خام = (نص.dxf.text if نص.dxftype() == 'TEXT' else نص.plain_text()).strip().upper()
                if "GROUND FLOOR PLAN" in خام:
                    مستوى = "GF"
                elif "FIRST FLOOR PLAN" in خام:
                    مستوى = "1F"
                elif "SECOND FLOOR PLAN" in خام or "2ND FLOOR PLAN" in خام:
                    مستوى = "2F"
                elif "TOP OF ROOF PLAN" in خام or "TOP ROOF PLAN" in خام:
                    مستوى = "TRF"
                elif "ROOF PLAN" in خام and "TOP" not in خام:
                    مستوى = "RF"
                else:
                    continue
                عناوين.append({
                    "المستوى": مستوى,
                    "x": نص.dxf.insert.x * self.مقياس,
                    "y": نص.dxf.insert.y * self.مقياس,
                })
            except Exception:
                pass

        if not عناوين:
            return []

        عناوين_فريدة: List[Dict] = []
        for عنوان in sorted(عناوين, key=lambda ع: (ع["المستوى"], ع["x"], -ع["y"])):
            مرشح = next(
                (
                    م for م in عناوين_فريدة
                    if م["المستوى"] == عنوان["المستوى"] and abs(م["x"] - عنوان["x"]) <= 8.0
                ),
                None,
            )
            if مرشح:
                مرشح["x"] = (مرشح["x"] + عنوان["x"]) / 2.0
                مرشح["ys"].append(عنوان["y"])
            else:
                عناوين_فريدة.append({
                    "المستوى": عنوان["المستوى"],
                    "x": عنوان["x"],
                    "ys": [عنوان["y"]],
                })

        مراكز = sorted(
            [{"المستوى": ع["المستوى"], "x": round(ع["x"], 3), "y": max(ع["ys"])} for ع in عناوين_فريدة],
            key=lambda ع: ع["x"]
        )
        gaps = [round(مراكز[i + 1]["x"] - مراكز[i]["x"], 3) for i in range(len(مراكز) - 1) if (مراكز[i + 1]["x"] - مراكز[i]["x"]) > 0]
        if gaps:
            gaps_sorted = sorted(gaps)
            نصف_افتراضي = max(25.0, min(45.0, gaps_sorted[len(gaps_sorted) // 2] / 2.0))
        else:
            نصف_افتراضي = 35.0

        ymin = min(ع["y"] for ع in مراكز) - 130.0
        ymax = max(ع["y"] for ع in مراكز) + 35.0

        صناديق: List[Dict] = []
        for i, عنوان in enumerate(مراكز):
            if i == 0:
                xmin = عنوان["x"] - نصف_افتراضي
            else:
                xmin = (مراكز[i - 1]["x"] + عنوان["x"]) / 2.0

            if i == len(مراكز) - 1:
                xmax = عنوان["x"] + نصف_افتراضي
            else:
                xmax = (عنوان["x"] + مراكز[i + 1]["x"]) / 2.0

            صناديق.append({
                "المستوى": عنوان["المستوى"],
                "xmin": xmin,
                "xmax": xmax,
                "ymin": ymin,
                "ymax": ymax,
                "center_x": عنوان["x"],
            })
        return صناديق

    def _استخرج_ملف_المشروع(self, msp) -> Dict:
        مستويات_صريحة = set()
        يوجد_مبنى_خدمة = False

        for نص in msp.query('TEXT MTEXT'):
            try:
                خام = (نص.dxf.text if نص.dxftype() == 'TEXT' else نص.plain_text()).strip().upper()
                كبير = " ".join(str(خام).split())
                if "GROUND FLOOR" in كبير:
                    مستويات_صريحة.add("GF")
                if "FIRST FLOOR" in كبير or "1ST FLOOR" in كبير or "FIRST SLAB LAYOUT" in كبير:
                    مستويات_صريحة.add("1F")
                if "SECOND FLOOR" in كبير or "2ND FLOOR" in كبير or "SECOND SLAB LAYOUT" in كبير:
                    مستويات_صريحة.add("2F")
                if ("ROOF PLAN" in كبير or "ROOF SLAB LAYOUT" in كبير or "ROOF FLOOR" in كبير) and "TOP" not in كبير:
                    مستويات_صريحة.add("RF")
                if "TOP ROOF" in كبير:
                    مستويات_صريحة.add("TRF")
                if any(كلمة in كبير for كلمة in ["KITCHEN BLOCK", "SERVICE BLOCK", "SERVANT BLOCK"]):
                    مستويات_صريحة.add("KB")
                    يوجد_مبنى_خدمة = True
            except Exception:
                pass

        مستويات_إنشائية = [م for م in ["GF", "1F", "2F"] if م in مستويات_صريحة]
        if مستويات_إنشائية == ["GF"]:
            نوع_المشروع = "GROUND ONLY"
        elif مستويات_إنشائية:
            نوع_المشروع = f"G+{len(مستويات_إنشائية) - 1}"
        else:
            نوع_المشروع = "UNKNOWN"

        ترتيب = {"GF": 0, "1F": 1, "2F": 2, "RF": 3, "TRF": 4, "KB": 5}
        return {
            "المستويات_الصريحة": sorted(مستويات_صريحة, key=lambda م: ترتيب.get(م, 99)),
            "المستويات_الإنشائية": مستويات_إنشائية,
            "نوع_المشروع": نوع_المشروع,
            "يوجد_مبنى_خدمة": يوجد_مبنى_خدمة,
        }

    def _مستوى_من_نقطة(self, x: float, y: float) -> str:
        if not self.صناديق_المستويات:
            return ""
        for صندوق in self.صناديق_المستويات:
            if صندوق["xmin"] <= x <= صندوق["xmax"] and صندوق["ymin"] <= y <= صندوق["ymax"]:
                return صندوق["المستوى"]
        مرشحة = [ص for ص in self.صناديق_المستويات if ص["ymin"] <= y <= ص["ymax"]]
        if مرشحة:
            xmin_عام = min(ص["xmin"] for ص in مرشحة)
            xmax_عام = max(ص["xmax"] for ص in مرشحة)
            if x < (xmin_عام - 8.0) or x > (xmax_عام + 8.0):
                return ""

            الأقرب = min(مرشحة, key=lambda ص: abs(ص["center_x"] - x))
            نصف_عرض_الصندوق = max(6.0, ((الأقرب["xmax"] - الأقرب["xmin"]) / 2.0) + 4.0)
            if abs(الأقرب["center_x"] - x) <= نصف_عرض_الصندوق:
                return الأقرب["المستوى"]
        return ""

    # ── DXF ──────────────────────────────────────────────────────────────────

    def استخراج_dxf(self):
        """يستخرج الخطوط والنصوص من ملف DXF/DWG مع دعم مساحات الورق والنموذج."""
        وثيقة = ezdxf.readfile(self.مسار)
        self.وثيقة_dxf = وثيقة
        self.الطبقات_المكتشفة = [layer.dxf.name for layer in وثيقة.layers]
        msp = وثيقة.modelspace()
        
        # 1. قراءة الجداول من مساحة النموذج (غالباً ما تكون الجداول والرموز هنا)
        self.قارئ = قارئ_الجداول(msp, self.مقياس, self.وحدة)
        self.قارئ.قراءة()

        # 1.25 عدّ الفتحات من المسقط مباشرة، ثم تحديث جدول الفتحات إن وجد
        عداد = عدّاد_الفتحات(msp, self.مقياس)
        self.عداد_الفتحات = عداد.عدّ()
        self.عداد_الفتحات_حسب_المستوى = {
            مستوى: dict(sorted(عدادات.items()))
            for مستوى, عدادات in sorted(عداد.العداد_حسب_المستوى.items())
        }
        self.صناديق_المستويات = self._ابن_صناديق_المستويات(msp)
        self.ملف_المشروع_المستخرج = self._استخرج_ملف_المشروع(msp)
        if self.قارئ.الفتحات:
            self.قارئ.الفتحات = عداد.تحديث_الفتحات(self.قارئ.الفتحات)
        elif self.عداد_الفتحات:
            self.قارئ.الفتحات = عداد.إنشاء_فتحات_من_العد()

        # 1.5 قراءة الأعمدة من المسقط لالتقاط الكميات حسب المستوى
        self.قارئ_أعمدة_الرسم = قارئ_الأعمدة_من_الرسم(
            msp, self.مقياس, self.قارئ.الأعمدة
        )
        self.قارئ_أعمدة_الرسم.قراءة()

        # 2. قراءة الكمرات من الرسم
        self.كمرات = قارئ_كمرات_الرسم(
            msp, self.مقياس,
            self.قارئ.جدول_كمرات_الربط,
            self.قارئ.جدول_كمرات_الستراب,
            self.قارئ.جدول_الكمرات
        )
        self.كمرات.قراءة()

        # 3. استخراج الهندسة والنصوص من جميع مساحات العمل
        سجل_هندسة = []

        def أضف_للسجل(هندسة, طبقة):
            سجل_هندسة.append((هندسة, طبقة))

        # نستخدم الـ ModelSpace فقط إذا كان مليئاً فعلياً بالهندسة، لتجنب
        # مضاعفة الغرف والمساقط من الـ paper space.
        msp_entities_count = len(msp.query("*"))
        layout_info = [(l.name, len(l.query("*"))) for l in وثيقة.layouts]
        مسجل.info(f"إحصائيات الملف: ModelSpace ({msp_entities_count} كيان) | Layouts: {layout_info}")
        
        فضاءات = [msp]
        # V15.2: Raised threshold from 2000 to 5000 — many real STR drawings have 3000+ entities
        استخدام_اللايوات = msp_entities_count < 5000
        if استخدام_اللايوات:
            for layout in list(وثيقة.layouts):
                # نفضل الكلمات التي تدل على مساقط
                name_upper = layout.name.upper()
                if any(k in name_upper for k in ["GROUND", "FIRST", "FLOOR", "FLOR", "PLAN", "AR-", "A10"]):
                    فضاءات.append(layout)
                elif layout.name.lower() != "model" and len(فضاءات) < 7:
                    فضاءات.append(layout)
        else:
            مسجل.info("تم تجاهل Layouts لأن ModelSpace يحتوي هندسة كافية للاستخراج الصارم")
        
        مسجل.info(f"جاري البحث في {len(فضاءات)} مساحة مختارة... ({[l.name for l in فضاءات]})")

        for فضاء in فضاءات:
            for كيان in فضاء.query('INSERT'):
                ليير = كيان.dxf.layer
                if not self._طبقة_صالحة(ليير): continue
                for ع in كيان.virtual_entities():
                    if ع.dxftype() == 'LINE':
                        ب = (ع.dxf.start.x*self.مقياس, ع.dxf.start.y*self.مقياس)
                        ن = (ع.dxf.end.x*self.مقياس, ع.dxf.end.y*self.مقياس)
                        if ب != ن:
                            أضف_للسجل(LineString([ب, ن]), ليير)

            # V15.2: Added CIRCLE and SPLINE support
            for كيان in فضاء.query('LINE LWPOLYLINE POLYLINE ARC CIRCLE SPLINE'):
                ليير = كيان.dxf.layer
                if not self._طبقة_صالحة(ليير): continue
                if كيان.dxftype() == 'LINE':
                    ب = (كيان.dxf.start.x*self.مقياس, كيان.dxf.start.y*self.مقياس)
                    ن = (كيان.dxf.end.x*self.مقياس, كيان.dxf.end.y*self.مقياس)
                    if ب != ن:
                        أضف_للسجل(LineString([ب, ن]), ليير)
                elif كيان.dxftype() in ('LWPOLYLINE', 'POLYLINE'):
                    نقاط = [(ن[0]*self.مقياس, ن[1]*self.مقياس)
                             for ن in كيان.get_points(format='xy')]
                    if len(نقاط) >= 2:
                        أضف_للسجل(LineString(نقاط), ليير)
                elif كيان.dxftype() == 'ARC':
                    زاوية_بداية = int(كيان.dxf.start_angle)
                    زاوية_نهاية = int(كيان.dxf.end_angle)
                    if زاوية_نهاية <= زاوية_بداية:
                        زاوية_نهاية += 360
                    نقاط = [
                        (كيان.dxf.center.x + كيان.dxf.radius*math.cos(math.radians(ز)),
                         كيان.dxf.center.y + كيان.dxf.radius*math.sin(math.radians(ز)))
                        for ز in range(زاوية_بداية, زاوية_نهاية+1, 5)
                    ]
                    if len(نقاط) >= 2:
                        أضف_للسجل(LineString([(ن[0]*self.مقياس, ن[1]*self.مقياس) for ن in نقاط]), ليير)
                elif كيان.dxftype() == 'CIRCLE':
                    # V15.2: Circles (columns, posts) → 36-sided polygon approximation
                    cx, cy, r = كيان.dxf.center.x, كيان.dxf.center.y, كيان.dxf.radius
                    نقاط = [
                        ((cx + r*math.cos(math.radians(ز)))*self.مقياس,
                         (cy + r*math.sin(math.radians(ز)))*self.مقياس)
                        for ز in range(0, 361, 10)
                    ]
                    if len(نقاط) >= 3:
                        أضف_للسجل(LineString(نقاط), ليير)
                elif كيان.dxftype() == 'SPLINE':
                    # V15.2: Spline → flatten to polyline
                    try:
                        نقاط_spline = list(كيان.flattening(0.1))
                        if len(نقاط_spline) >= 2:
                            نقاط = [(ن.x*self.مقياس, ن.y*self.مقياس) for ن in نقاط_spline]
                            أضف_للسجل(LineString(نقاط), ليير)
                    except Exception:
                        pass

            # استخراج النصوص
            for نص in فضاء.query('TEXT MTEXT'):
                try:
                    قيمة_خام = (نص.dxf.text if نص.dxftype()=='TEXT'
                                else نص.plain_text()).strip()
                    ins = نص.dxf.insert
                    x = ins.x*self.مقياس
                    y = ins.y*self.مقياس
                    self.نصوص_خام.append({
                        "قيمة": قيمة_خام,
                        "خام": قيمة_خام,
                        "المستوى": self._مستوى_من_نقطة(x, y),
                        "نقطة": Point(x, y)
                    })
                    اسم_غرفة = استخرج_اسم_غرفة_صالح(قيمة_خام)
                    if اسم_غرفة:
                        self.نصوص.append({
                            "قيمة": اسم_غرفة,
                            "خام": قيمة_خام,
                            "المستوى": self._مستوى_من_نقطة(x, y),
                            "نقطة": Point(x, y)
                        })
                except Exception:
                    pass

        # فلترة ذكية للجداول الضخمة
        if len(سجل_هندسة) > 10000:
            structural_layers = طبقات_الجدران + طبقات_الكمرات + طبقات_الأعمدة + ["DOOR", "GLAZ", "WINDOW"]
            مهم = [h for h in سجل_هندسة if h[1] and any(s.upper() in h[1].upper() for s in structural_layers)]
            if len(مهم) > 500:
                مسجل.info(f"مخطط ضخم ({len(سجل_هندسة)} خط): تم الإبقاء على {len(مهم)} خط كحدود هيكلية")
                سجل_هندسة = مهم
        
        self.خطوط = [h[0] for h in سجل_هندسة]
        مسجل.info(f"DXF: {len(self.خطوط)} خط | {len(self.نصوص)} تسمية")

        # استخراج مساحات البلاطات من الطبقات الإنشائية مباشرة.
        self.مساحات_البلاطات = self._استخراج_مساحات_البلاطات(msp)

    def _أضف_خط(self, بداية, نهاية, طبقة="0"):
        ب = (بداية.x*self.مقياس, بداية.y*self.مقياس)
        ن = (نهاية.x*self.مقياس,  نهاية.y*self.مقياس)
        if ب != ن:
            ls = LineString([ب, ن])
            self.خطوط.append(ls)

    def _استخراج_مساحات_البلاطات(self, msp) -> List[Dict]:
        المرشحات = []

        صناديق_العناوين: List[Dict] = []
        if self.قارئ and getattr(self.قارئ, "عناوين_المساقط", None):
            for عنوان in self.قارئ.عناوين_المساقط:
                نص = " ".join(str(عنوان.get("نص", "")).upper().split())
                # قبول أي عنوان يدل على مسقط slab/beam
                عبارات_slab = [
                    "SLAB LAYOUT", "SLAB PLAN",
                    "BEAM LAYOUT", "BEAM PLAN",
                    "ROOF SLAB", "FLOOR SLAB",
                ]
                if not any(ع in نص for ع in عبارات_slab):
                    continue
                if "FIRST" in نص:
                    مفتاح = "FIRST SLAB LAYOUT"
                elif "SECOND" in نص or "2ND" in نص:
                    مفتاح = "SECOND SLAB LAYOUT"
                elif "TOP ROOF" in نص:
                    مفتاح = "TOP ROOF SLAB LAYOUT"
                elif "ROOF" in نص:
                    مفتاح = "ROOF SLAB LAYOUT"
                else:
                    مفتاح = نص
                صناديق_العناوين.append({
                    "العنوان": مفتاح,
                    "xmin": عنوان["x"] - 30.0,
                    "xmax": عنوان["x"] + 30.0,
                    "ymin": عنوان["y"] - 2.0,
                    "ymax": عنوان["y"] + 34.0,
                })

        def _سجل(طبقة: str, مساحة: float, عنوان: str = ""):
            if 3.0 < مساحة < 2000.0:
                المرشحات.append({
                    "الطبقة": طبقة,
                    "المساحة": round(float(مساحة), 3),
                    "العنوان": عنوان,
                })

        for كيان in msp.query('LWPOLYLINE POLYLINE'):
            طبقة = كيان.dxf.layer.upper()
            if not any(ط in طبقة for ط in طبقات_البلاطات):
                continue
            try:
                نقاط = [(ن[0]*self.مقياس, ن[1]*self.مقياس) for ن in كيان.get_points(format='xy')]
                مغلق = getattr(كيان, "closed", False) or getattr(كيان, "is_closed", False)
                if len(نقاط) >= 3 and (مغلق or نقاط[0] == نقاط[-1]):
                    مساحة = abs(sum(
                        نقاط[i][0] * نقاط[(i + 1) % len(نقاط)][1] -
                        نقاط[(i + 1) % len(نقاط)][0] * نقاط[i][1]
                        for i in range(len(نقاط))
                    ) / 2.0)
                    عنوان = ""
                    if صناديق_العناوين:
                        xs = [ن[0] for ن in نقاط]
                        ys = [ن[1] for ن in نقاط]
                        مركز_x = (min(xs) + max(xs)) / 2.0
                        مركز_y = (min(ys) + max(ys)) / 2.0
                        for صندوق in صناديق_العناوين:
                            if صندوق["xmin"] <= مركز_x <= صندوق["xmax"] and صندوق["ymin"] <= مركز_y <= صندوق["ymax"]:
                                عنوان = صندوق["العنوان"]
                                break
                    _سجل(طبقة, مساحة, عنوان)
            except Exception:
                pass

        مجموعات: Dict[Tuple[str, float], Dict] = {}
        for عنصر in المرشحات:
            مفتاح = (عنصر.get("العنوان", ""), round(عنصر["المساحة"], 1))
            if مفتاح not in مجموعات:
                مجموعات[مفتاح] = {
                    "المساحة": round(float(عنصر["المساحة"]), 3),
                    "التكرار": 0,
                    "الطبقات": set(),
                    "العنوان": عنصر.get("العنوان", ""),
                }
            مجموعات[مفتاح]["التكرار"] += 1
            مجموعات[مفتاح]["الطبقات"].add(عنصر["الطبقة"])

        نتائج = []
        for قيمة in sorted(مجموعات.values(), key=lambda ع: (-ع["المساحة"], -ع["التكرار"])):
            نتائج.append({
                "المساحة": قيمة["المساحة"],
                "التكرار": قيمة["التكرار"],
                "الطبقات": sorted(قيمة["الطبقات"]),
                "العنوان": قيمة.get("العنوان", ""),
            })

        if نتائج:
            مسجل.info(f"مساحات البلاطات من الرسم: {نتائج}")
        return نتائج

    # ── PDF ──────────────────────────────────────────────────────────────────

    def استخراج_pdf(self):
        وثيقة = fitz.open(self.مسار)
        عدد_المتجهات = 0
        كل_نصوص_pdf: List[Dict] = []  # ALL text blocks for schedule reading
        عدد_صفحات = len(وثيقة)

        # V15.3: PDF coordinates are in points (1pt = 1/72 inch = 0.000353m).
        # Override مقياس for PDF — the user-selected "unit" applies to DXF only.
        مقياس_pdf = 0.0254 / 72.0  # points → meters (≈ 0.000353)

        for idx_صفحة, صفحة in enumerate(وثيقة):
            # V15.3: Multi-page offset — shift each page's X to prevent coordinate collision
            إزاحة_x = idx_صفحة * (صفحة.rect.width + 50) * مقياس_pdf if عدد_صفحات > 1 else 0.0
            ارتفاع_صفحة = صفحة.rect.height

            مسارات = صفحة.get_drawings()
            for مسار in مسارات:
                for عنصر in مسار.get("items", []):
                    نوع = عنصر[0]
                    if نوع == "l":
                        ن1, ن2 = عنصر[1], عنصر[2]
                        self.خطوط.append(LineString([
                            (ن1.x*مقياس_pdf + إزاحة_x, (ارتفاع_صفحة - ن1.y)*مقياس_pdf),
                            (ن2.x*مقياس_pdf + إزاحة_x, (ارتفاع_صفحة - ن2.y)*مقياس_pdf)
                        ]))
                        عدد_المتجهات += 1
                    elif نوع == "re":
                        # Rectangle: 4 corner points → 4 line segments
                        مستطيل = عنصر[1]  # fitz.Rect
                        x0 = مستطيل.x0*مقياس_pdf + إزاحة_x
                        y0 = (ارتفاع_صفحة - مستطيل.y0)*مقياس_pdf
                        x1 = مستطيل.x1*مقياس_pdf + إزاحة_x
                        y1 = (ارتفاع_صفحة - مستطيل.y1)*مقياس_pdf
                        self.خطوط.append(LineString([(x0,y0),(x1,y0)]))
                        self.خطوط.append(LineString([(x1,y0),(x1,y1)]))
                        self.خطوط.append(LineString([(x1,y1),(x0,y1)]))
                        self.خطوط.append(LineString([(x0,y1),(x0,y0)]))
                        عدد_المتجهات += 4
                    elif نوع == "c":
                        # Cubic Bézier curve → linearize into 12 segments
                        نقاط_خام = [ن for ن in عنصر[1:] if hasattr(ن, 'x') and hasattr(ن, 'y')]
                        if len(نقاط_خام) < 4:
                            if len(نقاط_خام) >= 2:
                                self.خطوط.append(LineString([
                                    (ن.x*مقياس_pdf + إزاحة_x, (ارتفاع_صفحة - ن.y)*مقياس_pdf) for ن in نقاط_خام
                                ]))
                                عدد_المتجهات += 1
                            continue
                        ن1, ن2, ن3, ن4 = نقاط_خام[:4]
                        نقاط = []
                        for خطوة in range(13):
                            t = خطوة / 12.0
                            mt = 1.0 - t
                            x = mt**3*ن1.x + 3*mt**2*t*ن2.x + 3*mt*t**2*ن3.x + t**3*ن4.x
                            y = mt**3*ن1.y + 3*mt**2*t*ن2.y + 3*mt*t**2*ن3.y + t**3*ن4.y
                            نقاط.append((x*مقياس_pdf + إزاحة_x, (ارتفاع_صفحة - y)*مقياس_pdf))
                        if len(نقاط) >= 2:
                            self.خطوط.append(LineString(نقاط))
                            عدد_المتجهات += 1
                    elif نوع == "qu":
                        # Quadratic Bézier curve → linearize into 8 segments
                        نقاط_خام = [ن for ن in عنصر[1:] if hasattr(ن, 'x') and hasattr(ن, 'y')]
                        if len(نقاط_خام) < 3:
                            if len(نقاط_خام) >= 2:
                                self.خطوط.append(LineString([
                                    (ن.x*مقياس_pdf + إزاحة_x, (ارتفاع_صفحة - ن.y)*مقياس_pdf) for ن in نقاط_خام
                                ]))
                                عدد_المتجهات += 1
                            continue
                        ن1, ن2, ن3 = نقاط_خام[:3]
                        نقاط = []
                        for خطوة in range(9):
                            t = خطوة / 8.0
                            mt = 1.0 - t
                            x = mt**2*ن1.x + 2*mt*t*ن2.x + t**2*ن3.x
                            y = mt**2*ن1.y + 2*mt*t*ن2.y + t**2*ن3.y
                            نقاط.append((x*مقياس_pdf + إزاحة_x, (ارتفاع_صفحة - y)*مقياس_pdf))
                        if len(نقاط) >= 2:
                            self.خطوط.append(LineString(نقاط))
                            عدد_المتجهات += 1

            # V15.4: Scale factor for schedule reader — so schedule coords match
            # the DXF-tuned window sizes (8m, 12m) in _عناصر_حول_الرأس
            مقياس_جدول = self.pdf_drawing_scale if self.pdf_drawing_scale > 0 else 100.0

            # Extract ALL text blocks — both for room names AND schedule parsing
            for كتلة in صفحة.get_text("blocks"):
                if len(كتلة) > 6 and كتلة[6] != 0:  # skip image blocks
                    continue
                نص_كامل = كتلة[4].strip()
                if not نص_كامل:
                    continue

                bx0, by0, bx1, by1 = كتلة[0], كتلة[1], كتلة[2], كتلة[3]
                # Block center in small PDF-meter coords (for geometry texts)
                مركز_x = (bx0 + bx1) / 2 * مقياس_pdf + إزاحة_x
                مركز_y = (ارتفاع_صفحة - (by0 + by1) / 2) * مقياس_pdf

                # V15.4: Split multi-line blocks into per-line entries for schedule
                # reader so each table row gets its own text with correct Y position.
                سطور = نص_كامل.split('\n')
                سطور_فعلية = [(i, s.strip()) for i, s in enumerate(سطور) if s.strip()]
                ارتفاع_سطر = (by1 - by0) / max(len(سطور), 1)  # in PDF points

                for idx_سطر, نص_سطر in سطور_فعلية:
                    # Per-line center in PDF points
                    y_pts = by0 + (idx_سطر + 0.5) * ارتفاع_سطر
                    x_pts = (bx0 + bx1) / 2

                    # Small coords (for raw/room texts)
                    sx = x_pts * مقياس_pdf + إزاحة_x
                    sy = (ارتفاع_صفحة - y_pts) * مقياس_pdf

                    # Schedule coords: scaled up to DXF-like villa meters
                    كل_نصوص_pdf.append({
                        "نص": نص_سطر,
                        "x": sx * مقياس_جدول,
                        "y": sy * مقياس_جدول,
                        "كبير": " ".join(نص_سطر.upper().split()),
                        "الطبقة": "PDF",
                        "الصفحة": idx_صفحة,
                    })

                # Keep original block-level entries for room names / raw texts
                self.نصوص_خام.append({
                    "قيمة": نص_كامل,
                    "خام": نص_كامل,
                    "المستوى": self._مستوى_من_نقطة(مركز_x, مركز_y),
                    "نقطة": Point(مركز_x, مركز_y),
                    "الطبقة": "PDF",
                })

                # Also check for room names (existing behavior)
                اسم_غرفة = استخرج_اسم_غرفة_صالح(نص_كامل)
                if اسم_غرفة:
                    self.نصوص.append({
                        "قيمة": اسم_غرفة,
                        "خام": نص_كامل,
                        "المستوى": self._مستوى_من_نقطة(مركز_x, مركز_y),
                        "نقطة": Point(مركز_x, مركز_y)
                    })

        مسجل.info(f"PDF: {len(self.خطوط)} خط | {len(self.نصوص)} تسمية غرف | {len(كل_نصوص_pdf)} نص كلي")

        # ── 1. Create schedule reader from PDF texts ─────────────────────────
        if كل_نصوص_pdf:
            self.قارئ = قارئ_الجداول(
                msp=None, مقياس=مقياس_pdf, وحدة=self.وحدة,
                pdf_texts=كل_نصوص_pdf
            )
            self.قارئ.قراءة()
            مسجل.info(
                f"PDF Schedule Reader: {len(self.قارئ.الأسس)} أساس | "
                f"{len(self.قارئ.أعمدة_العنق)} عنق | "
                f"{len(self.قارئ.جدول_كمرات_الربط)} TB | "
                f"{len(self.قارئ.جدول_كمرات_الستراب)} STB | "
                f"{len(self.قارئ.جدول_الكمرات)} B | "
                f"{len(self.قارئ.الأعمدة)} عمود | "
                f"{len(self.قارئ.الفتحات)} فتحة"
            )

        # ── 2. Build level boxes from PDF texts ──────────────────────────────
        self._ابن_صناديق_المستويات_من_نصوص(كل_نصوص_pdf)
        self._استخرج_ملف_المشروع_من_نصوص(كل_نصوص_pdf)

        # ── 3. Gemini Vision fallback — use it for STR PDFs with few/no schedules
        جدول_فارغ = (
            not self.قارئ or
            (len(self.قارئ.الأسس) == 0 and len(self.قارئ.الأعمدة) == 0
             and len(self.قارئ.جدول_كمرات_الربط) == 0)
        )

        if عدد_المتجهات < 100 and self.مفتاح_gemini:
            # Scanned PDF — always use Gemini
            مسجل.info("PDF يبدو مسح ضوئي — تفعيل Gemini Vision...")
            قارئ_vision = قارئ_gemini_vision(self.مفتاح_gemini)
            self.بيانات_gemini = قارئ_vision.استخراج_من_pdf(self.مسار)
            مسجل.info(f"Gemini استخرج: {self.بيانات_gemini}")
        elif عدد_المتجهات < 100:
            مسجل.warning("PDF مسح ضوئي ولا يوجد مفتاح Gemini — الدقة ستكون منخفضة")
            self._pdf_opencv_fallback()
        elif جدول_فارغ and self.مفتاح_gemini:
            # Vector PDF but schedule reading failed — use Gemini as upgrade
            مسجل.info("PDF متجه لكن الجداول فارغة — تفعيل Gemini Vision للجداول الإنشائية...")
            قارئ_vision = قارئ_gemini_vision(self.مفتاح_gemini)
            self.بيانات_gemini = قارئ_vision.استخراج_من_pdf(self.مسار)
            مسجل.info(f"Gemini استخرج (fallback): {self.بيانات_gemini}")

    def _ابن_صناديق_المستويات_من_نصوص(self, نصوص: List[Dict]):
        """Builds level boxes from PDF text blocks (same logic as DXF _ابن_صناديق_المستويات)."""
        عناوين: List[Dict] = []
        for نص in نصوص:
            خام = نص.get("كبير", "")
            مستوى = ""
            if "GROUND FLOOR PLAN" in خام:
                مستوى = "GF"
            elif "FIRST FLOOR PLAN" in خام:
                مستوى = "1F"
            elif "SECOND FLOOR PLAN" in خام or "2ND FLOOR PLAN" in خام:
                مستوى = "2F"
            elif "TOP OF ROOF PLAN" in خام or "TOP ROOF PLAN" in خام:
                مستوى = "TRF"
            elif "ROOF PLAN" in خام and "TOP" not in خام:
                مستوى = "RF"
            if مستوى:
                عناوين.append({
                    "المستوى": مستوى,
                    "x": نص["x"],
                    "y": نص["y"],
                })

        if not عناوين:
            return

        عناوين_فريدة: List[Dict] = []
        for عنوان in sorted(عناوين, key=lambda ع: (ع["المستوى"], ع["x"], -ع["y"])):
            مرشح = next(
                (م for م in عناوين_فريدة
                 if م["المستوى"] == عنوان["المستوى"] and abs(م["x"] - عنوان["x"]) <= 8.0),
                None,
            )
            if مرشح:
                مرشح["x"] = (مرشح["x"] + عنوان["x"]) / 2.0
                مرشح["ys"].append(عنوان["y"])
            else:
                عناوين_فريدة.append({
                    "المستوى": عنوان["المستوى"],
                    "x": عنوان["x"],
                    "ys": [عنوان["y"]],
                })

        مراكز = sorted(
            [{"المستوى": ع["المستوى"], "x": round(ع["x"], 3), "y": max(ع["ys"])} for ع in عناوين_فريدة],
            key=lambda ع: ع["x"]
        )
        gaps = [round(مراكز[i + 1]["x"] - مراكز[i]["x"], 3) for i in range(len(مراكز) - 1) if (مراكز[i + 1]["x"] - مراكز[i]["x"]) > 0]
        if gaps:
            gaps_sorted = sorted(gaps)
            نصف_افتراضي = max(25.0, min(45.0, gaps_sorted[len(gaps_sorted) // 2] / 2.0))
        else:
            نصف_افتراضي = 35.0

        ymin = min(ع["y"] for ع in مراكز) - 130.0
        ymax = max(ع["y"] for ع in مراكز) + 35.0

        صناديق: List[Dict] = []
        for i, عنوان in enumerate(مراكز):
            if i == 0:
                xmin = عنوان["x"] - نصف_افتراضي
            else:
                xmin = (مراكز[i - 1]["x"] + عنوان["x"]) / 2.0
            if i == len(مراكز) - 1:
                xmax = عنوان["x"] + نصف_افتراضي
            else:
                xmax = (عنوان["x"] + مراكز[i + 1]["x"]) / 2.0
            صناديق.append({
                "المستوى": عنوان["المستوى"],
                "xmin": xmin, "xmax": xmax,
                "ymin": ymin, "ymax": ymax,
                "center_x": عنوان["x"],
            })
        self.صناديق_المستويات = صناديق

    def _استخرج_ملف_المشروع_من_نصوص(self, نصوص: List[Dict]):
        """Extract project file metadata from PDF text blocks."""
        مستويات_صريحة = set()
        يوجد_مبنى_خدمة = False

        for نص in نصوص:
            كبير = نص.get("كبير", "")
            if "GROUND FLOOR" in كبير:
                مستويات_صريحة.add("GF")
            if "FIRST FLOOR" in كبير or "1ST FLOOR" in كبير:
                مستويات_صريحة.add("1F")
            if "SECOND FLOOR" in كبير or "2ND FLOOR" in كبير:
                مستويات_صريحة.add("2F")
            if ("ROOF PLAN" in كبير or "ROOF SLAB" in كبير or "ROOF FLOOR" in كبير) and "TOP" not in كبير:
                مستويات_صريحة.add("RF")
            if "TOP ROOF" in كبير:
                مستويات_صريحة.add("TRF")
            if any(كلمة in كبير for كلمة in ["KITCHEN BLOCK", "SERVICE BLOCK", "SERVANT BLOCK"]):
                مستويات_صريحة.add("KB")
                يوجد_مبنى_خدمة = True

        self.ملف_المشروع_المستخرج = {
            "المستويات_الصريحة": list(مستويات_صريحة),
            "يوجد_مبنى_خدمة": يوجد_مبنى_خدمة,
        }
        if مستويات_صريحة:
            مسجل.info(f"PDF مستويات صريحة: {مستويات_صريحة}")

    def _pdf_opencv_fallback(self):
        """استخراج احتياطي بـ OpenCV إن لم يوجد Gemini."""
        try:
            import cv2
            وثيقة = fitz.open(self.مسار)
            for صفحة in وثيقة:
                صورة_خام = صفحة.get_pixmap(dpi=150)
                صورة = np.frombuffer(صورة_خام.samples, dtype=np.uint8).reshape(
                    صورة_خام.h, صورة_خام.w, صورة_خام.n)
                رمادي = cv2.cvtColor(صورة, cv2.COLOR_RGB2GRAY) if صورة_خام.n >= 3 else صورة
                حواف  = cv2.Canny(رمادي, 50, 150, apertureSize=3)
                خطوط  = cv2.HoughLinesP(حواف, 1, np.pi/180, 80,
                                         minLineLength=30, maxLineGap=5)
                if خطوط is not None:
                    # تحديد المقياس بناءً على حجم الصفحة الفعلي
                    عرض_الصفحة_m = صفحة.rect.width / 72 * 0.0254  # بوصة → متر
                    مقياس_px = عرض_الصفحة_m / صورة_خام.w
                    for خط in خطوط:
                        x1,y1,x2,y2 = خط[0]
                        self.خطوط.append(LineString([
                            (x1*مقياس_px, y1*مقياس_px),
                            (x2*مقياس_px, y2*مقياس_px)
                        ]))
        except ImportError:
            مسجل.warning("OpenCV غير مثبت")

    # ── الهندسة والمضلعات ─────────────────────────────────────────────────────

    def بناء_الهندسة(self):
        if not self.خطوط:
            return

        # V15.5: For PDFs with massive line count, aggressively filter short lines
        # to remove hatching, annotation marks, furniture details, etc.
        if len(self.خطوط) > 50000:
            # Keep only lines ≥ 0.5m (after scale correction) — short lines are noise
            خطوط_طويلة = [خ for خ in self.خطوط if خ.length >= 0.5]
            if len(خطوط_طويلة) > 100:
                مسجل.info(f"فلترة الخطوط القصيرة: {len(self.خطوط)} → {len(خطوط_طويلة)} (≥0.5م)")
                self.خطوط = خطوط_طويلة

        # Fast union approach without O(n^2) buffer snapping
        مسجل.info(f"جاري معالجة ودمج {len(self.خطوط)} خط بتكنولوجيا سريعة...")
        مدمج = unary_union(self.خطوط)
        # Label-Centric Filtering: Only keep lines near Floor Plan/Room labels
        # This prevents grabbing the massive site border/framepolygons
        focus_labels = [n for n in self.نصوص if any(k in n["قيمة"] for k in ["PLAN", "FLOOR", "ROOF", "SCHEDULE", "LAYOUT"])]
        if focus_labels:
            focus_points = [n["نقطة"] for n in focus_labels]
            focus_tree = KDTree([(p.x, p.y) for p in focus_points])
            filtered_lines = []
            for line in self.خطوط:
                dist, _ = focus_tree.query((line.centroid.x, line.centroid.y))
                if dist < 60.0: # Keep entities within 60 meters of a label
                    filtered_lines.append(line)
            if filtered_lines:
                self.خطوط = filtered_lines
                مدمج = unary_union(self.خطوط)

        # Snap all vertices to a 0.10m (10cm) grid to close CAD gaps
        # V15.2: reduced from 0.20m to 0.10m for better room boundary accuracy
        try:
            from shapely import set_precision
            مدمج = set_precision(مدمج, 0.10)
        except ImportError:
            مسجل.warning("إصدار Shapely قديم، لم يتم تطبيق دمج الشبكة (set_precision)")
        
        # V15.2: Wider area filter — catch small utility rooms (≥0.3m²) and large open plans (≤800m²)
        self.مضلعات = [م for م in polygonize(مدمج) if 0.3 < م.area < 800.0]
        
        if not self.مضلعات:
             all_polys = list(polygonize(مدمج))
             areas = [p.area for p in all_polys[:10]]
             مسجل.info(f"DEBUG: 0 polygons passed filter. Found {len(all_polys)} total empty spaces. Top 10 areas: {areas}.")
        مسجل.info(f"تكوّن {len(self.مضلعات)} مضلع (غرفة)")

    def ربط_الفضاء(self):
        """
        V15.2: ربط محسّن — KDTree للبحث السريع + تسامح أوسع (1.0م بدل 0.35م)
        + حد أدنى مساحة 0.5م² بدل 1.0م².
        """
        if not self.مضلعات:
            return
        self.الغرف = []
        نصوص_صالحة = [ن for ن in self.نصوص if ن.get("قيمة") and ن.get("نقطة")]
        if not نصوص_صالحة:
            return

        # V15.2: بناء KDTree من نقاط النصوص للبحث السريع
        إحداثيات_النصوص = np.array([(ن["نقطة"].x, ن["نقطة"].y) for ن in نصوص_صالحة])
        شجرة_النصوص = KDTree(إحداثيات_النصوص)

        مفاتيح_فريدة = set()

        for مضلع in self.مضلعات:
            if مضلع.area < 0.5:
                continue

            مركز = مضلع.centroid
            المستوى = self._مستوى_من_نقطة(مركز.x, مركز.y)
            أفضل_نص = None
            أفضل_مفتاح = None

            # V15.2: بحث KDTree عن النصوص القريبة بدلاً من فحص كل النصوص
            # نصف قطر البحث = max(أبعاد المضلع) + 1.5م هامش
            حدود_المضلع = مضلع.bounds  # (minx, miny, maxx, maxy)
            بُعد_أقصى = max(حدود_المضلع[2] - حدود_المضلع[0], حدود_المضلع[3] - حدود_المضلع[1])
            نصف_قطر = بُعد_أقصى / 2.0 + 1.5
            فهارس_قريبة = شجرة_النصوص.query_ball_point([مركز.x, مركز.y], نصف_قطر)

            for فهرس in فهارس_قريبة:
                نص = نصوص_صالحة[فهرس]
                مستوى_النص = نص.get("المستوى", "")
                if المستوى and مستوى_النص and مستوى_النص != المستوى:
                    continue

                داخل = مضلع.contains(نص["نقطة"])
                مسافة_حد = مضلع.distance(نص["نقطة"])
                # V15.2: تسامح أوسع (1.0م) — PDF text placement imprecise
                if not داخل and مسافة_حد > 1.0:
                    continue

                مسافة_مركز = نص["نقطة"].distance(مركز)
                مفتاح_مفاضلة = (
                    0 if داخل else 1,
                    0 if (المستوى and مستوى_النص == المستوى) else (1 if not مستوى_النص else 2),
                    round(مسافة_حد, 4),
                    round(مسافة_مركز, 4),
                    -len(str(نص.get("قيمة", ""))),
                )
                if أفضل_مفتاح is None or مفتاح_مفاضلة < أفضل_مفتاح:
                    أفضل_مفتاح = مفتاح_مفاضلة
                    أفضل_نص = نص

            if not أفضل_نص:
                continue

            # V15.5: منع ربط مضلعات كبيرة بتسميات غرف صغيرة (حمام 71م² = خطأ)
            اسم_مطابق = str(أفضل_نص.get("قيمة", "")).strip().upper()
            _أسماء_غرف_صغيرة = {"BATH", "BATH ROOM", "BATHROOM", "TOILET", "WC", "W.C"}
            if مضلع.area > 15.0 and (اسم_مطابق in _أسماء_غرف_صغيرة or any(k in اسم_مطابق for k in ["TOILET", "WC"])):
                continue

            سجل = {
                "الاسم": أفضل_نص["قيمة"],
                "المساحة": round(مضلع.area, 4),
                "المحيط": round(مضلع.length, 4),
                "المستوى": المستوى or أفضل_نص.get("المستوى", ""),
                "النص_الخام": أفضل_نص.get("خام", ""),
                "مركز_x": round(مركز.x, 3),
                "مركز_y": round(مركز.y, 3),
            }
            مفتاح_فريد = (
                سجل["الاسم"],
                سجل["المستوى"],
                round(سجل["المساحة"], 2),
                round(سجل["مركز_x"], 1),
                round(سجل["مركز_y"], 1),
            )
            if مفتاح_فريد in مفاتيح_فريدة:
                continue
            مفاتيح_فريدة.add(مفتاح_فريد)
            self.الغرف.append(سجل)

    def اكتشاف_المحيط_الخارجي(self) -> float:
        """
        يأخذ محيط أكبر مضلع — لكن يتجاهل المضلعات الكبيرة جداً
        (قد تكون إطار ورقة وليس الفيلا).
        """
        if not self.مضلعات:
            return 0.0
        # فلترة: مساحة معقولة للفيلا (50 م² إلى 800 م²)
        مضلعات_معقولة = [م for م in self.مضلعات if 50 < م.area < 800]
        if مضلعات_معقولة:
            return round(max(مضلعات_معقولة, key=lambda م: م.area).length, 2)
        return round(max(self.مضلعات, key=lambda م: م.area).length, 2)

    def اكتشاف_مساحة_الحفر_المستخرجة(self) -> float:
        """
        يعيد مساحة حفر مستخرجة من أكبر مضلع معقول في الرسم.
        V15.5-fix: نقسّم مساحة الغرف على عدد الأدوار المكتشفة لتمثّل بصمة دور واحد فقط.
        """
        مساحة_الغرف_كلية = round(sum(غ["المساحة"] for غ in self.الغرف if "FLOOR" not in غ["الاسم"]), 2) if self.الغرف else 0.0
        # عدد الأدوار المكتشفة من المستويات الفريدة للغرف
        _مستويات = set()
        for غ in (self.الغرف or []):
            مست = str((غ or {}).get("المستوى", "") or "").strip().upper()
            if مست in ("GF", "1F", "2F"):
                _مستويات.add(مست)
        عدد_أدوار_غرف_محلي = max(len(_مستويات), 1)
        مساحة_الغرف = round(مساحة_الغرف_كلية / عدد_أدوار_غرف_محلي, 2)
        if not self.مضلعات:
            return مساحة_الغرف
        مضلعات_معقولة = [م for م in self.مضلعات if 50 < م.area < 5000]
        if مضلعات_معقولة:
            مساحة_مستخرجة = round(max(مضلعات_معقولة, key=lambda م: م.area).area, 2)
            if self.قارئ and getattr(self.قارئ, "سماكة_بلاطة_الأرضي", 0.0) > 0 and مساحة_الغرف > 0:
                if مساحة_الغرف > مساحة_مستخرجة * 1.35:
                    return round(مساحة_مستخرجة * 1.15, 2)
                return max(مساحة_مستخرجة, مساحة_الغرف)
            return مساحة_مستخرجة
        مساحة_مستخرجة = round(max(self.مضلعات, key=lambda م: م.area).area, 2)
        if self.قارئ and getattr(self.قارئ, "سماكة_بلاطة_الأرضي", 0.0) > 0 and مساحة_الغرف > 0:
            if مساحة_الغرف > مساحة_مستخرجة * 1.35:
                return round(مساحة_مستخرجة * 1.15, 2)
            return max(مساحة_مستخرجة, مساحة_الغرف)
        return مساحة_مستخرجة

    def اكتشاف_إطار_الفيلا(self) -> Tuple[float, float]:
        """
        يكتشف أبعاد الفيلا من أكبر مضلع معقول.
        الإصلاح: لا يأخذ أبعاد كل الخطوط (قد تشمل الجداول).
        """
        مضلعات_معقولة = [م for م in self.مضلعات if 50 < م.area < 800]
        if مضلعات_معقولة:
            أكبر = max(مضلعات_معقولة, key=lambda م: م.area)
            حدود = أكبر.bounds  # (minx, miny, maxx, maxy)
            return حدود[2]-حدود[0], حدود[3]-حدود[1]
        # fallback
        if self.خطوط:
            إحداثيات = [إ for خ in self.خطوط[:1000] for إ in خ.coords]
            xs = [إ[0] for إ in إحداثيات]
            ys = [إ[1] for إ in إحداثيات]
            return max(xs)-min(xs), max(ys)-min(ys)
        return 10.0, 10.0

    def درجة_الثقة(self) -> Dict:
        كل = len(self.الغرف)
        مسمّاة = sum(1 for غ in self.الغرف if غ["الاسم"] != "فضاء_غير_معروف")
        if كل == 0:
            return {"النتيجة": 0, "الدرجة": "F"}
        نسبة  = مسمّاة / كل
        نتيجة = round(نسبة * 100)
        درجة  = "ممتاز" if نتيجة>=85 else "جيد جداً" if نتيجة>=70 else "جيد" if نتيجة>=50 else "ضعيف"
        return {"النتيجة": نتيجة, "الدرجة": درجة, "مسمّاة": مسمّاة, "الكل": كل}

    def _ensure_dxf(self, file_path: str) -> str:
        import os
        import subprocess
        
        if file_path.lower().endswith('.dxf'):
            return file_path
        if not file_path.lower().endswith('.dwg'):
            raise ValueError(f"Unsupported format: {file_path}")

        مسجل.info(f"Native DWG detected. Converting to DXF... ({file_path})")
        accore_path = r"C:\Program Files\Autodesk\AutoCAD 2022\accoreconsole.exe"
        if not os.path.exists(accore_path):
             accore_path = r"C:\Program Files\Autodesk\AutoCAD 2023\accoreconsole.exe"
             if not os.path.exists(accore_path):
                 accore_path = r"C:\Program Files\Autodesk\AutoCAD 2024\accoreconsole.exe"
                 if not os.path.exists(accore_path):
                     raise FileNotFoundError("AutoCAD Core Console not found.")

        out_dir = os.environ.get("TEMP", os.path.dirname(os.path.abspath(file_path)))
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        temp_dxf = os.path.join(out_dir, f"temp_convert_{base_name}.dxf")
        if os.path.exists(temp_dxf):
            try: os.remove(temp_dxf)
            except: pass

        script_path = os.path.join(out_dir, "export_dxf.scr")
        with open(script_path, "w", encoding="utf-8") as f:
            f.write("FILEDIA\n0\n")
            f.write(f"DXFOUT\n\"{temp_dxf}\"\nV\n2018\n16\n\n")
            f.write("QUIT\nY\n")

        try:
            subprocess.run([accore_path, "/i", file_path, "/s", script_path], 
                           check=True, capture_output=True, text=True, input="\n"*50)
            if os.path.exists(temp_dxf):
                مسجل.info("DWG Conversion Successful!")
                return temp_dxf
            else:
                raise RuntimeError("Conversion completed but DXF file not found.")
        except Exception as e:
            مسجل.error(f"AutoCAD core console failure: {e}")
            raise RuntimeError(f"Failed to convert DWG to DXF: {e}")

    def _كشف_مقياس_من_النصوص(self) -> float:
        """
        V15.2: يبحث في النصوص عن عبارة المقياس مثل "SCALE 1:100" أو "1:50" أو "SCALE: 1/200"
        ويعيد المقياس المكتشف (مثلاً 100 لرسم 1:100).
        يعيد 0 إذا لم يجد شيئاً.
        """
        import re
        نمط_المقياس = re.compile(
            r'(?:SCALE|مقياس)\s*[=:]?\s*1\s*[:/]\s*(\d+)', re.IGNORECASE
        )
        نمط_مقياس_بسيط = re.compile(r'\b1\s*:\s*(\d{2,4})\b')
        for نص in self.نصوص_خام:
            خام = نص.get("خام", "") or str(نص.get("قيمة", ""))
            تطابق = نمط_المقياس.search(خام)
            if تطابق:
                مقياس = int(تطابق.group(1))
                if 10 <= مقياس <= 500:
                    مسجل.info(f"تم كشف مقياس الرسم من النصوص: 1:{مقياس} — النص: '{خام.strip()}'")
                    return float(مقياس)
            تطابق2 = نمط_مقياس_بسيط.search(خام)
            if تطابق2:
                مقياس = int(تطابق2.group(1))
                if 10 <= مقياس <= 500:
                    مسجل.info(f"تم كشف مقياس الرسم (نمط بسيط) من النصوص: 1:{مقياس} — النص: '{خام.strip()}'")
                    return float(مقياس)
        return 0.0

    def تصحيح_المقياس_تلقائيا(self):
        """
        Auto-corrects incorrect unit selections. e.g. DWG drawn in cm,
        but user selected mm. Checks bounding box. If the villa is 
        0.5 meters across, it was probably cm not mm.
        
        V15.2: Also scans text entities for "SCALE 1:xxx" annotations first.
        """
        if not self.خطوط: return

        # V15.2: Try detecting scale from text annotations before heuristics
        if self.pdf_drawing_scale <= 0:
            مقياس_نصي = self._كشف_مقياس_من_النصوص()
            if مقياس_نصي > 0:
                self.pdf_drawing_scale = مقياس_نصي
        
        # Test boundaries before polygon generation
        xs = [إ[0] for خ in self.خطوط[:1000] for إ in خ.coords]
        ys = [إ[1] for خ in self.خطوط[:1000] for إ in خ.coords]
        
        width = max(xs) - min(xs)
        height = max(ys) - min(ys)
        
        # In meters, a typical house is 5m to 50m wide.
        if width > 0 and height > 0:
            area_m2 = width * height
            مسجل.info(f"الأبعاد الأولية: {width:.2f}m x {height:.2f}m (المساحة: {area_m2:.2f} m²)")
            
            # If house is less than 2x2 meters, scale up heavily (mismatched cm vs mm)
            if width < 2.0 or height < 2.0:
                 biggest = max(width, height)
                 if self.pdf_drawing_scale > 0:
                     # Use the user-provided or text-detected drawing scale
                     scale_factor = self.pdf_drawing_scale
                     مسجل.info(f"استخدام مقياس PDF: {scale_factor}")
                 elif self.مسار.lower().endswith('.pdf'):
                     # PDF files: coordinates are in points (0.353mm/pt).
                     # Most architectural plans are 1:100 scale.
                     # A1 paper ≈ 0.84m after pt→m conversion → ×100 = 84m (correct)
                     scale_factor = 100.0
                     مسجل.info("PDF بدون نص مقياس → افتراضي 1:100")
                 elif biggest < 0.05:
                     scale_factor = 1000.0
                 elif biggest < 0.5:
                     scale_factor = 100.0
                 elif biggest < 2.0:
                     target = 20.0
                     scale_factor = round(target / biggest)
                     scale_factor = max(10.0, min(scale_factor, 100.0))
                 else:
                     scale_factor = 10.0
                 مسجل.warning(f"المخطط صغير جداً! أصحح المقياس... (x{scale_factor})")
                 from shapely.affinity import scale
                 self.خطوط = [scale( line, xfact=scale_factor, yfact=scale_factor, origin=(0,0)) for line in self.خطوط]
                 for t in self.نصوص:
                     pt = t["نقطة"]
                     t["نقطة"] = Point(pt.x * scale_factor, pt.y * scale_factor)
                 for t in self.نصوص_خام:
                     if "نقطة" in t:
                         pt = t["نقطة"]
                         t["نقطة"] = Point(pt.x * scale_factor, pt.y * scale_factor)
            elif width > 200.0 or height > 200.0:
                 # It's over 200 meters? Might be drawn in mm but processed as meters
                 scale_factor = 0.001 if width > 2000.0 else 0.01 if width > 500.0 else 0.1
                 مسجل.warning(f"المخطط ضخم جداً! أصحح المقياس... (x{scale_factor})")
                 from shapely.affinity import scale
                 self.خطوط = [scale( line, xfact=scale_factor, yfact=scale_factor, origin=(0,0)) for line in self.خطوط]
                 for t in self.نصوص:
                     pt = t["نقطة"]
                     t["نقطة"] = Point(pt.x * scale_factor, pt.y * scale_factor)
                 for t in self.نصوص_خام:
                     if "نقطة" in t:
                         pt = t["نقطة"]
                         t["نقطة"] = Point(pt.x * scale_factor, pt.y * scale_factor)

    # ── تنفيذ ─────────────────────────────────────────────────────────────────

    def تنفيذ(self) -> Dict:
        # Detect extension
        root, ext = os.path.splitext(self.مسار)
        ext = ext.lower()
        
        # If no extension (common in multer uploads), default to .dxf
        if not ext:
            مسجل.info(f"No extension found for {self.مسار}. Defaulting to DXF processing.")
            ext = '.dxf'
            
        if ext in ('.dxf', '.dwg'):
            if ext == '.dwg':
                self.مسار = self._ensure_dxf(self.مسار)
            self.استخراج_dxf()
        elif ext == '.pdf':
            self.استخراج_pdf()
        else:
            raise NotImplementedError(f"Format '{ext}' not supported for file: {self.مسار}. Use .dxf, .dwg or .pdf")

        # Auto correct the scale based on bounding box
        self.تصحيح_المقياس_تلقائيا()

        self.بناء_الهندسة()
        self.ربط_الفضاء()

        كاشف = كاشف_الجدران(self.خطوط)
        كاشف.اكتشاف()

        ط_فيلا, ع_فيلا = self.اكتشاف_إطار_الفيلا()

        ق = self.قارئ
        ك = self.كمرات
        ق_أ = self.قارئ_أعمدة_الرسم

        # أكبر مضلع في نطاق البصمة المعقول للفيلا (50-800 م²)
        _مضلعات_بصمة = [م for م in self.مضلعات if 50 < م.area < 800] if self.مضلعات else []
        مساحة_أكبر_مضلع = round(max(_مضلعات_بصمة, key=lambda م: م.area).area, 2) if _مضلعات_بصمة else 0.0

        return {
            "المساحة_الكلية":           round(sum(غ["المساحة"] for غ in self.الغرف if "FLOOR" not in غ["الاسم"]), 2),
            "المحيط_الخارجي":          self.اكتشاف_المحيط_الخارجي(),
            "مساحة_الحفر":             self.اكتشاف_مساحة_الحفر_المستخرجة(),
            "مساحة_أكبر_مضلع":        مساحة_أكبر_مضلع,
            "مساحات_البلاطات":         self.مساحات_البلاطات,
            "تكرار_عناوين_المساقط":    ({
                عنوان: sum(
                    1 for ع in (ق.عناوين_المساقط if ق else [])
                    if str(ع.get("كبير", "")).strip() == عنوان
                )
                for عنوان in {
                    str(ع.get("كبير", "")).strip()
                    for ع in (ق.عناوين_المساقط if ق else [])
                    if str(ع.get("كبير", "")).strip()
                }
            }),
            "سماكات_البلاطات_المستخرجة": (ق.سماكات_البلاطات if ق else {}),
            "منسوب_كمرة_الربط_المستخرج": (ق.منسوب_كمرة_الربط if ق else 0.0),
            "سماكة_بلاطة_الأرضي_المستخرجة": (ق.سماكة_بلاطة_الأرضي if ق else 0.0),
            "ارتفاع_الدور_المستخرج": (ق.ارتفاع_الدور_المستخرج if ق else 0.0),
            "ملف_المشروع_المستخرج": self.ملف_المشروع_المستخرج,
            "طول_جدران_بلوك_20":       كاشف.طول_20,
            "طول_جدران_بلوك_10":       كاشف.طول_10,
            "الغرف":                   self.الغرف,
            "درجة_الثقة":              self.درجة_الثقة(),
            "الجداول": {
                "الأسس":         [أ.__dict__ for أ in (ق.الأسس         if ق else [])],
                "أعمدة_العنق":   [ع.__dict__ for ع in (ق.أعمدة_العنق   if ق else [])],
                "الأعمدة":        [ع.__dict__ for ع in (ق.الأعمدة        if ق else [])],
                "الفتحات":        [ف.__dict__ for ف in (ق.الفتحات        if ق else [])],
            },
            "الأعمدة_من_الرسم": {
                "الأعمدة": [ع.__dict__ for ع in (ق_أ.الأعمدة if ق_أ else [])],
                "إحصاء_حسب_المستوى": (ق_أ.إحصاء_حسب_المستوى if ق_أ else {}),
            },
            "الكمرات_من_الرسم": {
                "كمرات_الربط":   [ك.__dict__ for ك in (ك.كمرات_الربط   if ك else [])],
                "كمرات_الستراب": [ك.__dict__ for ك in (ك.كمرات_الستراب if ك else [])],
                "الكمرات":        [ك.__dict__ for ك in (ك.الكمرات        if ك else [])],
            },
            "بيانات_gemini": self.بيانات_gemini,
            "عداد_الفتحات": self.عداد_الفتحات,
            "عداد_الفتحات_حسب_المستوى": self.عداد_الفتحات_حسب_المستوى,
        }


# ─────────────────────────────────────────────────────────────────────────────
# الكتاب المقدس لحساب الكميات — 31 بنداً
# ─────────────────────────────────────────────────────────────────────────────

class كتاب_الكميات:

    @staticmethod
    def احسب(بيانات_الغرف: List[Dict], الفضاء: Dict,
              الطلب: طلب_حساب_الكميات) -> Dict[str, List[Dict]]:

        ث = الطلب.الثوابت

        # ── دمج المصادر (الأولوية: يدوي > Gemini > تلقائي) ──────────────────
        جداول    = الفضاء.get("الجداول", {})
        كمرات_رسم = الفضاء.get("الكمرات_من_الرسم", {})
        جيميني   = الفضاء.get("بيانات_gemini", {})
        استخراج_صارم = getattr(ث, "استخراج_صارم_فقط", False)
        # ── إذا كان الوضع الصارم مفعّلاً لكن الجداول فارغة وجيميني يملك بيانات → نتراجع
        _جداول_فارغة = not any(جداول.get(k) for k in ("الأسس", "أعمدة_العنق", "كمرات_الربط", "كمرات_الستراب", "الأعمدة", "الكمرات"))
        _جيميني_موجود = any(جيميني.get(k) for k in ("الأسس", "أعمدة_العنق", "كمرات_الربط", "الأعمدة", "الكمرات"))
        if استخراج_صارم and _جداول_فارغة and _جيميني_موجود:
            استخراج_صارم = False
        سماكات_البلاطات_المستخرجة = الفضاء.get("سماكات_البلاطات_المستخرجة", {}) or {}
        منسوب_كمرة_الربط_المستخرج = float(الفضاء.get("منسوب_كمرة_الربط_المستخرج", 0.0) or 0.0)
        سماكة_بلاطة_الأرضي_المستخرجة = float(الفضاء.get("سماكة_بلاطة_الأرضي_المستخرجة", 0.0) or 0.0)
        ارتفاع_الدور_المستخرج = float(الفضاء.get("ارتفاع_الدور_المستخرج", 0.0) or 0.0)
        ملف_المشروع_المستخرج = الفضاء.get("ملف_المشروع_المستخرج", {}) or {}
        مستويات_المشروع_الصريحة = set(ملف_المشروع_المستخرج.get("المستويات_الصريحة", []) or [])
        مستويات_المشروع_الإنشائية = list(ملف_المشروع_المستخرج.get("المستويات_الإنشائية", []) or [])
        يوجد_مبنى_خدمة_صريح = bool(ملف_المشروع_المستخرج.get("يوجد_مبنى_خدمة", False))

        سماكات_غير_أرضية = [v for k, v in سماكات_البلاطات_المستخرجة.items() if "GRADE" not in k and v > 0]
        سماكة_بلاطة_افتراضية_مستخرجة = (sum(سماكات_غير_أرضية) / len(سماكات_غير_أرضية)) if سماكات_غير_أرضية else 0.0
        سماكة_مدخلة_للبلاطة = float(ث.سماكة_البلاطة or 0.0)
        ارتفاع_دور_مدخل = float(ث.ارتفاع_الدور or 0.0)
        ارتفاع_أرضي_مدخل = float(getattr(ث, "ارتفاع_الدور_الأرضي", 0.0) or 0.0)
        ارتفاع_أول_مدخل = float(getattr(ث, "ارتفاع_الدور_الأول", 0.0) or 0.0)
        ارتفاع_ثان_مدخل = float(getattr(ث, "ارتفاع_الدور_الثاني", 0.0) or 0.0)
        ارتفاع_سطح_مدخل = float(getattr(ث, "ارتفاع_دور_السطح", 0.0) or 0.0)
        ارتفاع_خدمة_مدخل = float(getattr(ث, "ارتفاع_مبنى_الخدمة", 0.0) or 0.0)
        ارتفاع_صافي_تشطيب_مدخل = float(getattr(ث, "ارتفاع_صافي_التشطيب_القياسي", 0.0) or 0.0)
        مستوى_أرضي_مدخل = float(ث.مستوى_بلاطة_الأرضي or 0.0)

        سماكة_بلاطة_فعلية = سماكة_بلاطة_افتراضية_مستخرجة if سماكة_بلاطة_افتراضية_مستخرجة > 0 else سماكة_مدخلة_للبلاطة
        سماكة_بلاطة_أرضي_فعلية = سماكة_بلاطة_الأرضي_المستخرجة if سماكة_بلاطة_الأرضي_المستخرجة > 0 else 0.0
        مستوى_بلاطة_الأرضي_فعلي = منسوب_كمرة_الربط_المستخرج if منسوب_كمرة_الربط_المستخرج > 0 else مستوى_أرضي_مدخل
        ارتفاع_الدور_فعلي = ارتفاع_الدور_المستخرج if ارتفاع_الدور_المستخرج > 0 else (
            ارتفاع_أرضي_مدخل or ارتفاع_أول_مدخل or ارتفاع_ثان_مدخل or ارتفاع_سطح_مدخل or ارتفاع_خدمة_مدخل or ارتفاع_دور_مدخل
        )
        عدد_الأدوار_فعلي = 0
        if استخراج_صارم:
            عدد_الأدوار_فعلي = len(مستويات_المشروع_الإنشائية)
            if عدد_الأدوار_فعلي <= 0:
                عدد_الأدوار_فعلي = sum(
                    1 for مفتاح in ["FIRST SLAB LAYOUT", "SECOND SLAB LAYOUT", "ROOF SLAB LAYOUT"]
                    if سماكات_البلاطات_المستخرجة.get(مفتاح, 0.0) > 0
                )
        عدد_الأدوار_مرجعي = عدد_الأدوار_فعلي if عدد_الأدوار_فعلي > 0 else int(ث.عدد_الأدوار or 0)
        ارتفاعات_يدوية_حسب_المستوى = {
            "GF": ارتفاع_أرضي_مدخل,
            "1F": ارتفاع_أول_مدخل,
            "2F": ارتفاع_ثان_مدخل,
            "RF": ارتفاع_سطح_مدخل,
            "TRF": ارتفاع_سطح_مدخل,
            "KB": ارتفاع_خدمة_مدخل,
        }

        def _نظّف_قاموس(ب: dict) -> dict:
            """Sanitise a Gemini dict before Pydantic conversion.
            - None/0 counts → 1  (default quantity)
            - Missing or None dimensions → 0.0  (so Pydantic doesn't reject the item)
            """
            نظيف = dict(ب)
            if نظيف.get("الكمية") in (None, 0):
                نظيف["الكمية"] = 1
            for مفتاح in ("الطول", "العرض", "العمق", "عمق_الأساس", "الارتفاع"):
                if مفتاح not in نظيف or نظيف[مفتاح] is None:
                    نظيف[مفتاح] = 0.0
            return نظيف

        def _حوّل(قائمة, فئة):
            نتيجة = []
            for ب in قائمة:
                try:
                    if isinstance(ب, dict):
                        نتيجة.append(فئة(**_نظّف_قاموس(ب)))
                    else:
                        نتيجة.append(ب)
                except Exception:
                    pass
            return نتيجة

        def _وحّد_العناصر(قائمة, مفتاح_حجم):
            """Dedupe by الرمز, keeping the entry with the largest dimension sum."""
            أفضل = {}
            for ع in قائمة:
                رمز = getattr(ع, "الرمز", "")
                حجم = sum(float(getattr(ع, م, 0) or 0) for م in مفتاح_حجم)
                if رمز not in أفضل or حجم > أفضل[رمز][1]:
                    أفضل[رمز] = (ع, حجم)
            return [v[0] for v in أفضل.values()]

        def _طبع_المستوى(مستوى: str) -> str:
            قيمة = str(مستوى or "").strip().upper()
            if قيمة in {"GF", "G", "GROUND", "GROUND FLOOR", "GROUND FLOOR LVL", "أرضي"}:
                return "GF"
            if قيمة in {"1F", "FF", "FIRST", "FIRST FLOOR", "FIRST FLOOR LVL", "1ST FLOOR", "1ST FLOOR LVL", "أول"}:
                return "1F"
            if قيمة in {"2F", "SF", "SECOND", "SECOND FLOOR", "SECOND FLOOR LVL", "2ND FLOOR", "2ND FLOOR LVL", "ثاني"}:
                return "2F"
            if قيمة in {"RF", "ROOF", "ROOF FLOOR", "ROOF LVL", "سطح"}:
                return "RF"
            if قيمة in {"TRF", "TOP ROOF", "TOP ROOF LVL", "TOP ROOF PLAN"}:
                return "TRF"
            if قيمة in {"KB", "KITCHEN BLOCK", "SERVICE", "SERVICE BLOCK"}:
                return "KB"
            return قيمة

        def _ارتفاع_الدور_للمستوى(مستوى: str) -> float:
            مستوى_مطبع = _طبع_المستوى(مستوى)
            ارتفاع_يدوي = float(ارتفاعات_يدوية_حسب_المستوى.get(مستوى_مطبع, 0.0) or 0.0)
            if ارتفاع_يدوي > 0:
                return ارتفاع_يدوي
            if ارتفاع_الدور_المستخرج > 0:
                return ارتفاع_الدور_المستخرج
            if ارتفاع_دور_مدخل > 0:
                return ارتفاع_دور_مدخل
            return 0.0

        # الأسس
        if استخراج_صارم:
            الأسس = _حوّل(جداول.get("الأسس", []), بيانات_الأساس)
        else:
            _أسس_يدوي = الطلب.الأسس
            _أسس_جداول = _حوّل(جداول.get("الأسس", []), بيانات_الأساس)
            _أسس_جيميني = _حوّل(جيميني.get("الأسس", []), بيانات_الأساس)
            الأسس = _أسس_يدوي or _أسس_جداول or _أسس_جيميني
        # إزالة تكرار الأسس: Gemini قد يرسل نفس الرمز من جداول مختلفة
        الأسس = _وحّد_العناصر(الأسس, ["الطول", "العرض", "العمق"])

        # أعمدة العنق
        if استخراج_صارم:
            أعمدة_العنق = _حوّل(جداول.get("أعمدة_العنق", []), بيانات_عمود_العنق)
        else:
            أعمدة_العنق = الطلب.أعمدة_العنق or \
                          _حوّل(جداول.get("أعمدة_العنق", []), بيانات_عمود_العنق) or \
                          _حوّل(جيميني.get("أعمدة_العنق", []), بيانات_عمود_العنق)

        # الكمرات — الطول من الرسم، العرض/العمق من الجدول
        if استخراج_صارم:
            كمرات_الربط = _حوّل(كمرات_رسم.get("كمرات_الربط", []), بيانات_كمرة_ربط)
        else:
            كمرات_الربط = الطلب.كمرات_الربط or \
                          _حوّل(كمرات_رسم.get("كمرات_الربط", []), بيانات_كمرة_ربط) or \
                          _حوّل(جيميني.get("كمرات_الربط", []), بيانات_كمرة_ربط)

        if استخراج_صارم:
            كمرات_الستراب = _حوّل(كمرات_رسم.get("كمرات_الستراب", []), بيانات_كمرة_ستراب)
        else:
            كمرات_الستراب = الطلب.كمرات_الستراب or \
                            _حوّل(كمرات_رسم.get("كمرات_الستراب", []), بيانات_كمرة_ستراب) or \
                            _حوّل(جيميني.get("كمرات_الستراب", []), بيانات_كمرة_ستراب)

        if استخراج_صارم:
            الكمرات = _حوّل(كمرات_رسم.get("الكمرات", []), بيانات_كمرة)
        else:
            الكمرات = الطلب.الكمرات or \
                      _حوّل(كمرات_رسم.get("الكمرات", []), بيانات_كمرة) or \
                      _حوّل(جيميني.get("الكمرات", []), بيانات_كمرة)

        أعماق_ربط_مستخرجة = [ك.العمق for ك in كمرات_الربط if getattr(ك, "العمق", 0) > 0]
        عمق_كمرة_الربط_فعلي = (sum(أعماق_ربط_مستخرجة) / len(أعماق_ربط_مستخرجة)) if أعماق_ربط_مستخرجة else (0.0 if استخراج_صارم else ث.عمق_كمرة_الربط)

        # ── إزالة التكرار وتقدير أطوال كمرات Gemini ──────────────────────────
        # Gemini قد يعطي نفس الرمز عدة مرات من صفحات/جداول مختلفة.
        # نحتفظ بالبُعد الأكبر (الأكثر واقعية عادةً) لكل رمز.
        كمرات_الربط = _وحّد_العناصر(كمرات_الربط, ["الطول", "العرض", "العمق"])
        كمرات_الستراب = _وحّد_العناصر(كمرات_الستراب, ["الطول", "العرض", "العمق"])
        الكمرات = _وحّد_العناصر(الكمرات, ["الطول", "العرض", "العمق"])

        # Gemini يعطي العرض والعمق فقط — الطول يأتي من الهندسة عادةً.
        # عند غيابه نقسم المحيط على عدد الأنواع كتقدير مبدئي.
        محيط_للتقدير = float(الفضاء.get("المحيط_الخارجي", 0) or جيميني.get("المحيط_الخارجي", 0) or 0)
        # If still 0, estimate perimeter from excavation/foundation footprint area
        if محيط_للتقدير <= 0:
            مساحة_حفر_تقدير = float(الفضاء.get("مساحة_الحفر", 0) or 0)
            if مساحة_حفر_تقدير <= 0:
                مساحة_حفر_تقدير = sum(
                    float(getattr(أ, "الطول", 0) or 0) * float(getattr(أ, "العرض", 0) or 0) * int(getattr(أ, "الكمية", 1) or 1)
                    for أ in الأسس
                )
            if مساحة_حفر_تقدير > 0:
                import math as _math
                # Foundations cover ~25% of building footprint; perimeter of rectangle with aspect 1.5
                محيط_للتقدير = round(5.0 * _math.sqrt(مساحة_حفر_تقدير * 4.0), 1)
        if محيط_للتقدير > 0:
            أنواع_كر_فريدة = {ك.الرمز for ك in كمرات_الربط if getattr(ك, "الطول", 0) <= 0}
            if أنواع_كر_فريدة:
                طول_لكل_نوع = محيط_للتقدير / len(أنواع_كر_فريدة)
                كمرات_ربط_جديدة = []
                for ك in كمرات_الربط:
                    if ك.الرمز in أنواع_كر_فريدة and getattr(ك, "الطول", 0) <= 0:
                        كمرات_ربط_جديدة.append(بيانات_كمرة_ربط(الرمز=ك.الرمز, الطول=طول_لكل_نوع, العرض=ك.العرض, العمق=ك.العمق))
                    else:
                        كمرات_ربط_جديدة.append(ك)
                كمرات_الربط = كمرات_ربط_جديدة

            أنواع_كم_فريدة = {ك.الرمز for ك in الكمرات if getattr(ك, "الطول", 0) <= 0}
            if أنواع_كم_فريدة:
                طول_كمرة = محيط_للتقدير / len(أنواع_كم_فريدة)
                كمرات_جديدة = []
                for ك in الكمرات:
                    if ك.الرمز in أنواع_كم_فريدة and getattr(ك, "الطول", 0) <= 0:
                        كمرات_جديدة.append(بيانات_كمرة(الرمز=ك.الرمز, الطول=طول_كمرة, العرض=ك.العرض, العمق=ك.العمق))
                    else:
                        كمرات_جديدة.append(ك)
                الكمرات = كمرات_جديدة

        # الأعمدة
        if استخراج_صارم:
            الأعمدة = _حوّل(الفضاء.get("الأعمدة_من_الرسم", {}).get("الأعمدة", []), بيانات_العمود) or \
                      _حوّل(جداول.get("الأعمدة", []), بيانات_العمود)
        else:
            الأعمدة = الطلب.الأعمدة or \
                      _حوّل(جداول.get("الأعمدة", []), بيانات_العمود) or \
                      _حوّل(الفضاء.get("الأعمدة_من_الرسم", {}).get("الأعمدة", []), بيانات_العمود) or \
                      _حوّل(جيميني.get("الأعمدة", []), بيانات_العمود)
        # إزالة تكرار الأعمدة
        الأعمدة = _وحّد_العناصر(الأعمدة, ["الطول", "العرض"])

        if استخراج_صارم and الأعمدة:
            أعماق_الأسس_المتاحة = [أ.العمق for أ in الأسس if getattr(أ, "العمق", 0) > 0]
            عمق_أساس_مرجعي = (sum(أعماق_الأسس_المتاحة) / len(أعماق_الأسس_المتاحة)) if أعماق_الأسس_المتاحة else 0.0
            أعمدة_عادية = []
            for ع in الأعمدة:
                if ع.الرمز.upper().startswith("NC"):
                    أعمدة_العنق.append(بيانات_عمود_العنق(
                        الرمز=ع.الرمز,
                        الطول=ع.الطول,
                        العرض=ع.العرض,
                        عمق_الأساس=عمق_أساس_مرجعي,
                        الكمية=ع.الكمية,
                    ))
                else:
                    أعمدة_عادية.append(ع)
            الأعمدة = أعمدة_عادية

            if not أعمدة_العنق and الأعمدة and عمق_أساس_مرجعي > 0:
                ترتيب_المستويات = {"GF": 0, "": 1, "1F": 2, "2F": 3, "RF": 4, "TRF": 5, "KB": 9}
                أقل_رتبة = min(
                    ترتيب_المستويات.get(_طبع_المستوى(getattr(ع, "المستوى", "") or ""), 99)
                    for ع in الأعمدة
                )
                أعمدة_مرشحة_للعنق = [
                    ع for ع in الأعمدة
                    if ترتيب_المستويات.get(_طبع_المستوى(getattr(ع, "المستوى", "") or ""), 99) == أقل_رتبة
                ]
                if أعمدة_مرشحة_للعنق:
                    for ع in أعمدة_مرشحة_للعنق:
                        أعمدة_العنق.append(بيانات_عمود_العنق(
                            الرمز=f"NC-{ع.الرمز}",
                            الطول=ع.الطول,
                            العرض=ع.العرض,
                            عمق_الأساس=عمق_أساس_مرجعي,
                            الكمية=ع.الكمية,
                        ))
                    مسجل.info(
                        f"تم اشتقاق {len(أعمدة_العنق)} نوع أعمدة عنق من أعمدة المستوى الأدنى "
                        f"بمرجع عمق أساس {عمق_أساس_مرجعي:.2f}م"
                    )

            if not يوجد_مبنى_خدمة_صريح:
                الأعمدة = [ع for ع in الأعمدة if _طبع_المستوى(getattr(ع, "المستوى", "")) != "KB"]

            مستويات_الأعمدة_المستخرجة = {
                _طبع_المستوى(getattr(ع, "المستوى", "") or "")
                for ع in الأعمدة
                if _طبع_المستوى(getattr(ع, "المستوى", "") or "") in {"GF", "1F", "2F"}
            }
            if not مستويات_المشروع_الإنشائية and مستويات_الأعمدة_المستخرجة:
                عدد_الأدوار_مرجعي = max(1, len(مستويات_الأعمدة_المستخرجة))

        # الفتحات
        if استخراج_صارم:
            الفتحات = _حوّل(جداول.get("الفتحات", []), بيانات_فتحة)
        else:
            الفتحات = الطلب.الفتحات or \
                      _حوّل(جداول.get("الفتحات", []), بيانات_فتحة) or \
                      _حوّل(جيميني.get("الفتحات", []), بيانات_فتحة)

        # ── الهندسة ───────────────────────────────────────────────────────────
        المساحة_الكلية = الفضاء["المساحة_الكلية"]
        if استخراج_صارم:
            المحيط_الخارجي = الفضاء["المحيط_الخارجي"]
            جدران_20 = الفضاء.get("طول_جدران_بلوك_20", 0.0)
            جدران_10 = الفضاء.get("طول_جدران_بلوك_10", 0.0)
        else:
            المحيط_الخارجي = الطلب.المحيط_الخارجي or الفضاء["المحيط_الخارجي"]
            جدران_20 = الطلب.طول_جدران_بلوك_20_داخلي \
                       if الطلب.طول_جدران_بلوك_20_داخلي is not None \
                       else الفضاء.get("طول_جدران_بلوك_20", 0.0)
            جدران_10 = الطلب.طول_جدران_بلوك_10_داخلي \
                       if الطلب.طول_جدران_بلوك_10_داخلي is not None \
                       else الفضاء.get("طول_جدران_بلوك_10", 0.0)

        جدران_20_خام = float(جدران_20 or 0.0)
        جدران_10_خام = float(جدران_10 or 0.0)

        # V15.5: Sanity-cap wall lengths and perimeter based on villa area.
        # A typical UAE villa has perimeter ≈ 4*sqrt(area) and total internal
        # wall length ≈ 1.5-3× perimeter.  PDF noise can inflate these 10-50×.
        # V15.5b: Use ARCH-detected floor count (not just config) for per-floor area
        مستويات_غرف_مبدئية = {
            _طبع_المستوى((غ or {}).get("المستوى", ""))
            for غ in (بيانات_الغرف or [])
            if _طبع_المستوى((غ or {}).get("المستوى", "")) in {"GF", "1F", "2F"}
        }
        عدد_أدوار_رئيسية = len({م for م in مستويات_المشروع_الصريحة if م in {"GF", "1F", "2F"}}) if مستويات_المشروع_الصريحة else 0
        عدد_أدوار_غرف = len(مستويات_غرف_مبدئية)
        عدد_أدوار_للقسمة = max(عدد_أدوار_رئيسية, عدد_أدوار_غرف, عدد_الأدوار_مرجعي, 1)
        if المساحة_الكلية > 10:
            # V15.5: استخدم مساحة الدور الواحد (لا المجموع) لتقدير المحيط
            مساحة_دور = المساحة_الكلية / عدد_أدوار_للقسمة
            محيط_مرجعي = 4.0 * math.sqrt(مساحة_دور) * 1.3  # generous
            if المحيط_الخارجي > محيط_مرجعي * 1.5:
                مسجل.warning(
                    f"تصحيح المحيط الخارجي ({المحيط_الخارجي:.1f}م) → {محيط_مرجعي:.1f}م (بناءً على المساحة {المساحة_الكلية:.1f}م²)"
                )
                المحيط_الخارجي = round(محيط_مرجعي, 2)
            حد_جدار_أقصى = المحيط_الخارجي * 1.5  # internal walls ≤ 1.5× perimeter
            حد_جدار_10_أقصى = المحيط_الخارجي * 0.5  # 10cm partitions much shorter
            if جدران_20_خام > حد_جدار_أقصى:
                مسجل.warning(
                    f"تصحيح طول بلوك 20 ({جدران_20_خام:.1f}م) → {حد_جدار_أقصى:.1f}م"
                )
                جدران_20 = حد_جدار_أقصى
            if جدران_10_خام > حد_جدار_10_أقصى:
                مسجل.warning(
                    f"تصحيح طول بلوك 10 ({جدران_10_خام:.1f}م) → {حد_جدار_10_أقصى:.1f}م"
                )
                جدران_10 = حد_جدار_10_أقصى
        elif المحيط_الخارجي > 0:
            حد_جدار_أقصى = المحيط_الخارجي * 1.5
            حد_جدار_10_أقصى = المحيط_الخارجي * 0.5
            if جدران_20_خام > حد_جدار_أقصى:
                جدران_20 = حد_جدار_أقصى
            if جدران_10_خام > حد_جدار_10_أقصى:
                جدران_10 = حد_جدار_10_أقصى

        # ── تصنيف الغرف ───────────────────────────────────────────────────────
        أسماء_مبللة_قياسية = {"BATH", "KITCHEN", "PANTRY", "LAUNDRY", "حمام", "مطبخ"}
        أسماء_بلكونة_قياسية = {"BALCONY", "بلكونة"}

        def _اسم_غرفة_قياسي(غرفة: Dict) -> str:
            return str(غرفة.get("الاسم", "") or "").strip().upper()

        def _مستوى_غرفة(غرفة: Dict) -> str:
            return _طبع_المستوى(غرفة.get("المستوى", ""))

        def _غرفة_مبللة(غرفة: Dict) -> bool:
            اسم = _اسم_غرفة_قياسي(غرفة)
            مبللة = اسم in أسماء_مبللة_قياسية or any(كلمة in اسم for كلمة in ["TOILET", "WC", "KITCHEN", "LAUNDRY", "PANTRY", "حمام", "مطبخ"])
            if not مبللة:
                return False
            # V15.5: حد أقصى لمساحة الغرف المبللة — حمام 71م² = تصنيف خاطئ حتماً
            مساحة = غرفة.get("المساحة", 0.0)
            if اسم in {"BATH", "BATH ROOM"} or any(k in اسم for k in ["TOILET", "WC"]):
                return مساحة <= 15.0
            if اسم == "KITCHEN":
                return مساحة <= 30.0
            return مساحة <= 20.0  # PANTRY, LAUNDRY, MAID

        def _غرفة_بلكونة(غرفة: Dict) -> bool:
            اسم = _اسم_غرفة_قياسي(غرفة)
            return اسم in أسماء_بلكونة_قياسية or any(كلمة in اسم for كلمة in كلمات_بلكونة)

        مستويات_الغرف_المستخرجة = {_مستوى_غرفة(غ) for غ in بيانات_الغرف if _مستوى_غرفة(غ)}
        مستويات_تشطيب_مستهدفة = {
            م for م in مستويات_المشروع_الصريحة
            if م in {"GF", "1F", "2F", "RF"}
        }
        if يوجد_مبنى_خدمة_صريح:
            مستويات_تشطيب_مستهدفة.add("KB")
        if استخراج_صارم and not مستويات_تشطيب_مستهدفة:
            مستويات_تشطيب_مستهدفة = {
                م for م in مستويات_الغرف_المستخرجة
                if م in {"GF", "1F", "2F", "RF"}
            }
            if يوجد_مبنى_خدمة_صريح and "KB" in مستويات_الغرف_المستخرجة:
                مستويات_تشطيب_مستهدفة.add("KB")
        if استخراج_صارم and مستويات_الغرف_المستخرجة:
            غرف_التشطيب = [غ for غ in بيانات_الغرف if _مستوى_غرفة(غ) in مستويات_تشطيب_مستهدفة]
            if not غرف_التشطيب:
                غرف_التشطيب = [
                    غ for غ in بيانات_الغرف
                    if _مستوى_غرفة(غ) != "TRF" and (_مستوى_غرفة(غ) != "KB" or يوجد_مبنى_خدمة_صريح)
                ]
        else:
            غرف_التشطيب = list(بيانات_الغرف)

        if (not استخراج_صارم) and الطلب.مناطق_مبللة_بالدور:
            م_أرضي  = next((د.المساحة  for د in الطلب.مناطق_مبللة_بالدور if د.الدور=="أرضي"), 0.0)
            م_أول   = next((د.المساحة  for د in الطلب.مناطق_مبللة_بالدور if د.الدور=="أول"),  0.0)
            م_ثان   = next((د.المساحة  for د in الطلب.مناطق_مبللة_بالدور if د.الدور=="ثاني"),  0.0)
            مح_أرضي = next((د.المحيط   for د in الطلب.مناطق_مبللة_بالدور if د.الدور=="أرضي"), 0.0)
            مح_أول  = next((د.المحيط   for د in الطلب.مناطق_مبللة_بالدور if د.الدور=="أول"),  0.0)
            مح_ثان  = next((د.المحيط   for د in الطلب.مناطق_مبللة_بالدور if د.الدور=="ثاني"),  0.0)
            م_سطح = مح_سطح = 0.0
            م_خدمة = مح_خدمة = 0.0
            غ_مبللة = [غ for غ in غرف_التشطيب if _غرفة_مبللة(غ)]
        else:
            غ_مبللة = [غ for غ in غرف_التشطيب if _غرفة_مبللة(غ)]
            if any(_مستوى_غرفة(غ) for غ in غ_مبللة):
                غ_مبللة_أرضي = [غ for غ in غ_مبللة if _مستوى_غرفة(غ) == "GF"]
                غ_مبللة_أول = [غ for غ in غ_مبللة if _مستوى_غرفة(غ) == "1F"]
                غ_مبللة_ثان = [غ for غ in غ_مبللة if _مستوى_غرفة(غ) == "2F"]
                غ_مبللة_سطح = [غ for غ in غ_مبللة if _مستوى_غرفة(غ) == "RF"]
                غ_مبللة_خدمة = [غ for غ in غ_مبللة if _مستوى_غرفة(غ) == "KB"] if يوجد_مبنى_خدمة_صريح else []
            else:
                غ_مبللة_أرضي = list(غ_مبللة)
                غ_مبللة_أول = []
                غ_مبللة_ثان = []
                غ_مبللة_سطح = []
                غ_مبللة_خدمة = []
            م_أرضي  = sum(غ["المساحة"] for غ in غ_مبللة_أرضي)
            م_أول   = sum(غ["المساحة"] for غ in غ_مبللة_أول)
            م_ثان   = sum(غ["المساحة"] for غ in غ_مبللة_ثان)
            م_سطح   = sum(غ["المساحة"] for غ in غ_مبللة_سطح)
            م_خدمة  = sum(غ["المساحة"] for غ in غ_مبللة_خدمة)
            مح_أرضي = sum(غ["المحيط"]  for غ in غ_مبللة_أرضي)
            مح_أول  = sum(غ["المحيط"]  for غ in غ_مبللة_أول)
            مح_ثان  = sum(غ["المحيط"]  for غ in غ_مبللة_ثان)
            مح_سطح  = sum(غ["المحيط"]  for غ in غ_مبللة_سطح)
            مح_خدمة = sum(غ["المحيط"]  for غ in غ_مبللة_خدمة)

        مح_مبلل = مح_أرضي + مح_أول + مح_ثان + مح_سطح + مح_خدمة
        م_مبلل_كلي = م_أرضي + م_أول + م_ثان + م_سطح + م_خدمة
        if استخراج_صارم:
            غ_بلكونة = [غ for غ in غرف_التشطيب if _غرفة_بلكونة(غ)]
        else:
            غ_بلكونة = [غ for غ in بيانات_الغرف if _غرفة_بلكونة(غ)]
        م_بلكونة_غرف = sum(غ["المساحة"] for غ in غ_بلكونة)
        م_بلكونة = م_بلكونة_غرف
        غ_جافة = [غ for غ in غرف_التشطيب if not _غرفة_مبللة(غ) and not _غرفة_بلكونة(غ)]
        م_جاف = sum(غ["المساحة"] for غ in غ_جافة)
        مح_جاف = sum(غ["المحيط"] for غ in غ_جافة)
        مستويات_واجهات_مغلقة = sorted({
            _مستوى_غرفة(غ)
            for غ in غرف_التشطيب
            if _مستوى_غرفة(غ) in مستويات_تشطيب_مستهدفة and not _غرفة_بلكونة(غ)
        })
        عدد_مستويات_واجهات = len(مستويات_واجهات_مغلقة) if مستويات_واجهات_مغلقة else عدد_الأدوار_مرجعي

        طول_جدران_20_من_الغرف = 0.0
        طول_جدران_10_من_الغرف = 0.0
        if استخراج_صارم and المحيط_الخارجي > 0:
            غرف_مغلقة = [غ for غ in غرف_التشطيب if not _غرفة_بلكونة(غ)]
            if غرف_مغلقة:
                مستويات_مغلقة = sorted({
                    _مستوى_غرفة(غ)
                    for غ in غرف_مغلقة
                    if _مستوى_غرفة(غ) in مستويات_تشطيب_مستهدفة
                })
                عدد_مستويات_رئيسية = len([م for م in مستويات_مغلقة if م in {"GF", "1F", "2F"}])
                عدد_مستويات_علوية = len([م for م in مستويات_مغلقة if م == "RF"])
                عدد_مستويات_خدمة = len([م for م in مستويات_مغلقة if م == "KB"])
                معامل_غلاف = max(
                    1.0,
                    float(عدد_مستويات_رئيسية) + (0.5 * float(عدد_مستويات_علوية + عدد_مستويات_خدمة))
                )
                محيط_الغرف_المغلقة = sum(float(غ.get("المحيط", 0.0) or 0.0) for غ in غرف_مغلقة)
                طول_جدران_داخلي_مقدر = max(0.0, (محيط_الغرف_المغلقة - (المحيط_الخارجي * معامل_غلاف)) / 2.0)
                مجموع_خام_داخلي = max(0.0, جدران_20_خام) + max(0.0, جدران_10_خام)
                if طول_جدران_داخلي_مقدر > 0:
                    if مجموع_خام_داخلي > 0:
                        نسبة_20 = max(0.0, جدران_20_خام) / مجموع_خام_داخلي
                        نسبة_10 = max(0.0, جدران_10_خام) / مجموع_خام_داخلي
                    else:
                        نسبة_20 = 0.0
                        نسبة_10 = 1.0
                    طول_جدران_20_من_الغرف = round(طول_جدران_داخلي_مقدر * نسبة_20, 3)
                    طول_جدران_10_من_الغرف = round(طول_جدران_داخلي_مقدر * نسبة_10, 3)

                    حد_20_منطقي = max(المحيط_الخارجي * 2.5, طول_جدران_20_من_الغرف * 1.8)
                    حد_10_منطقي = max(المحيط_الخارجي * 2.5, طول_جدران_10_من_الغرف * 1.8)

                    if جدران_20 <= 0 or جدران_20 > حد_20_منطقي:
                        if طول_جدران_20_من_الغرف > 0:
                            مسجل.info(
                                f"تم استبدال طول بلوك 20 الداخلي من الرسم الخام ({جدران_20:.2f}م) "
                                f"بطول مشتق من الغرف ({طول_جدران_20_من_الغرف:.2f}م)"
                            )
                            جدران_20 = round(طول_جدران_20_من_الغرف, 3)

                    if جدران_10 <= 0 or جدران_10 > حد_10_منطقي:
                        if طول_جدران_10_من_الغرف > 0:
                            مسجل.info(
                                f"تم استبدال طول بلوك 10 الداخلي من الرسم الخام ({جدران_10:.2f}م) "
                                f"بطول مشتق من الغرف ({طول_جدران_10_من_الغرف:.2f}م)"
                            )
                            جدران_10 = round(طول_جدران_10_من_الغرف, 3)

        # ── الدرج ───────────────────────────────────────────────────────────
        خرسانة_الدرج_مستخرجة = 0.0
        if ث.استخراج_الدرج_آليا:
            غرف_الدرج = [غ for غ in بيانات_الغرف if "STAIR" in str(غ.get("الاسم", "")).upper()]
            for غ in غرف_الدرج:
                # تقدير حجم الخرسانة: المساحة * متوسط سمك الدرج (15سم بلاطة + 15سم مثلثات / 2 = ~25سم)
                مساحة_الدرج = float(غ.get("المساحة", 0.0) or 0.0)
                if مساحة_الدرج > 0:
                    سمك_مقدر = 0.25
                    حجم_الدورة = مساحة_الدرج * سمك_مقدر
                    خرسانة_الدرج_مستخرجة += حجم_الدورة
            
            if خرسانة_الدرج_مستخرجة > 0:
                مسجل.info(f"تم استخراج خرسانة الدرج آلياً: {خرسانة_الدرج_مستخرجة:.2f} م³")

        # ── التصوينة (Parapet) ────────────────────────────────────────────────
        مساحة_بلوك_التصوينة = 0.0
        طول_التصوينة = 0.0
        if المحيط_الخارجي > 0:
            # التصوينة تكون حول السطح (RF) وأيضاً أعلى السطح (TRF)
            طول_التصوينة = المحيط_الخارجي
            مساحة_بلوك_التصوينة = طول_التصوينة * ث.ارتفاع_التصوينة
            مسجل.info(f"تم حساب كميات التصوينة بناءً على المحيط ({طول_التصوينة:.2f}م) وارتفاع {ث.ارتفاع_التصوينة}م")

        # ── الفتحات ───────────────────────────────────────────────────────────
        def _هو_باب(رمز: str) -> bool:
            رمز = رمز.upper()
            return رمز.startswith("D") or رمز.startswith("MD") or رمز.startswith("DR")

        def _هو_نافذة(رمز: str) -> bool:
            رمز = رمز.upper()
            return رمز.startswith("W") or رمز.startswith("WIN")

        م_أبواب = sum(ف.الطول * ف.الارتفاع * ف.الكمية for ف in الفتحات if _هو_باب(ف.الرمز))
        ع_أبواب = sum(ف.الطول * ف.الكمية for ف in الفتحات if _هو_باب(ف.الرمز))
        عدد_الأبواب = sum(ف.الكمية for ف in الفتحات if _هو_باب(ف.الرمز))
        م_نوافذ = sum(ف.الطول * ف.الارتفاع * ف.الكمية for ف in الفتحات if _هو_نافذة(ف.الرمز))
        م_باب_رئيسي = sum(ف.الطول * ف.الارتفاع * ف.الكمية for ف in الفتحات if ف.الرمز.upper().startswith("MD"))

        مساحات_البلاطات = sorted(
            الفضاء.get("مساحات_البلاطات", []),
            key=lambda ع: (
                str(ع.get("العنوان", "")),
                -ع.get("المساحة", 0),
                -ع.get("التكرار", 0)
            )
        )
        تكرار_عناوين_المساقط = {
            str(ك): max(1, int(v or 1))
            for ك, v in (الفضاء.get("تكرار_عناوين_المساقط", {}) or {}).items()
        }
        مجموعات_البلاطات_بالعنوان: Dict[str, List[Dict]] = {}
        for مجموعة_بلاطة in مساحات_البلاطات:
            عنوان = str(مجموعة_بلاطة.get("العنوان", "") or "").strip().upper()
            if not عنوان:
                عنوان = "UNTITLED"
            تكرار_خام = max(1, int(مجموعة_بلاطة.get("التكرار", 1) or 1))
            معامل_تكرار_العنوان = max(1, تكرار_عناوين_المساقط.get(عنوان, 1))
            تكرار_فعلي = max(1, int(round(تكرار_خام / معامل_تكرار_العنوان)))
            نسخة_مهيأة = dict(مجموعة_بلاطة)
            نسخة_مهيأة["العنوان"] = عنوان
            نسخة_مهيأة["التكرار_الفعلي"] = تكرار_فعلي
            مجموعات_البلاطات_بالعنوان.setdefault(عنوان, []).append(نسخة_مهيأة)

        def _أكبر_بلاطة_بعنوان(عنوان: str) -> float:
            مرشح = مجموعات_البلاطات_بالعنوان.get(عنوان, [])
            return max((float(م.get("المساحة", 0) or 0.0) for م in مرشح), default=0.0)

        def _مجموع_بلاطات_بعنوان(عنوان: str, حد_أدنى: float = 0.0) -> float:
            مرشح = مجموعات_البلاطات_بالعنوان.get(عنوان, [])
            مجموع = 0.0
            for م in مرشح:
                مساحة = float(م.get("المساحة", 0) or 0.0)
                if مساحة < حد_أدنى:
                    continue
                مجموع += مساحة * max(1, int(م.get("التكرار_الفعلي", 1) or 1))
            return مجموع

        مساحة_بلاطة_أول = _أكبر_بلاطة_بعنوان("FIRST SLAB LAYOUT")
        مساحة_بلاطة_ثان = _أكبر_بلاطة_بعنوان("SECOND SLAB LAYOUT")
        مساحة_بلاطة_سقف = _أكبر_بلاطة_بعنوان("ROOF SLAB LAYOUT")
        مساحة_بلاطات_أعلى_السطح = _مجموع_بلاطات_بعنوان("TOP ROOF SLAB LAYOUT", حد_أدنى=1.0)
        مساحة_بلاطة_رئيسية = max(مساحة_بلاطة_أول, مساحة_بلاطة_ثان, مساحة_بلاطة_سقف)
        بلاطات_غير_معنونة = مجموعات_البلاطات_بالعنوان.get("UNTITLED", [])
        أكبر_غير_معنونة_مرتبة = sorted([
            (
                float(م.get("المساحة", 0) or 0.0),
                max(1, int(م.get("التكرار_الفعلي", 1) or 1))
            )
            for م in بلاطات_غير_معنونة
            if float(م.get("المساحة", 0) or 0.0) > 1.0
        ], reverse=True)

        if مساحة_بلاطة_سقف <= 0 and أكبر_غير_معنونة_مرتبة:
            حد_الأول = مساحة_بلاطة_أول * 0.95 if مساحة_بلاطة_أول > 0 else float("inf")
            for مساحة_مرشحة, _ in أكبر_غير_معنونة_مرتبة:
                if مساحة_مرشحة < حد_الأول:
                    مساحة_بلاطة_سقف = مساحة_مرشحة
                    break
            if مساحة_بلاطة_سقف <= 0:
                مساحة_بلاطة_سقف = أكبر_غير_معنونة_مرتبة[0][0]

        if أكبر_غير_معنونة_مرتبة and مساحة_بلاطة_سقف > 0:
            مجموع_خدمة_احتياطي = sum(
                مساحة * تكرار for مساحة, تكرار in أكبر_غير_معنونة_مرتبة
                if مساحة < (مساحة_بلاطة_سقف * 0.9)
            )
            if مجموع_خدمة_احتياطي > 0 and (
                مساحة_بلاطات_أعلى_السطح <= 0 or مساحة_بلاطات_أعلى_السطح >= (مساحة_بلاطة_سقف * 0.9)
            ):
                مساحة_بلاطات_أعلى_السطح = مجموع_خدمة_احتياطي

        مساحة_بلاطة_رئيسية = max(مساحة_بلاطة_أول, مساحة_بلاطة_ثان, مساحة_بلاطة_سقف)

        if not any(عنوان != "UNTITLED" for عنوان in مجموعات_البلاطات_بالعنوان):
            مساحة_بلاطة_رئيسية = مساحات_البلاطات[0]["المساحة"] if مساحات_البلاطات else 0.0
            مساحة_بلاطات_أعلى_السطح = 0.0
            if مساحة_بلاطة_رئيسية:
                for مجموعة_بلاطة in مساحات_البلاطات[1:]:
                    if مجموعة_بلاطة.get("المساحة", 0) < (مساحة_بلاطة_رئيسية * 0.85):
                        تكرار = max(1, min(مجموعة_بلاطة.get("التكرار", 1), عدد_الأدوار_مرجعي))
                        مساحة_بلاطات_أعلى_السطح += مجموعة_بلاطة.get("المساحة", 0) * تكرار

        if استخراج_صارم and غ_بلكونة:
            مستويات_بلكونات = {_مستوى_غرفة(غ) for غ in غ_بلكونة if _مستوى_غرفة(غ)}
            مرشحات_بلكونة_بلاطات = sorted({
                round(float(م.get("المساحة", 0.0) or 0.0), 3)
                for عنوان, مجموعات in مجموعات_البلاطات_بالعنوان.items()
                for م in مجموعات
                if عنوان == "UNTITLED"
                and int(م.get("التكرار_الفعلي", 1) or 1) == 1
                and 8.0 <= float(م.get("المساحة", 0.0) or 0.0) <= 15.0
            })
            if مستويات_بلكونات and مستويات_بلكونات.issubset({"RF", "TRF"}) and مرشحات_بلكونة_بلاطات:
                مساحة_مرشحة = min(
                    مرشحات_بلكونة_بلاطات,
                    key=lambda قيمة: abs(قيمة - max(float(غ.get("المساحة", 0.0) or 0.0) for غ in غ_بلكونة))
                )
                if مساحة_مرشحة > 0 and م_بلكونة_غرف > (مساحة_مرشحة * 1.5):
                    مسجل.info(
                        f"تم استبدال مساحة البلكونة من الغرف ({م_بلكونة_غرف:.2f}م²) "
                        f"بمرشح بلاطة من الرسم ({مساحة_مرشحة:.2f}م²)"
                    )
                    م_بلكونة = مساحة_مرشحة

        سماكات_البلاطات_حسب_المستوى = {
            "GF": float(سماكات_البلاطات_المستخرجة.get("FIRST SLAB LAYOUT", 0.0) or 0.0),
            "1F": float(سماكات_البلاطات_المستخرجة.get("SECOND SLAB LAYOUT", 0.0) or سماكات_البلاطات_المستخرجة.get("ROOF SLAB LAYOUT", 0.0) or 0.0),
            "2F": float(سماكات_البلاطات_المستخرجة.get("ROOF SLAB LAYOUT", 0.0) or 0.0),
            "RF": float(سماكات_البلاطات_المستخرجة.get("TOP ROOF SLAB LAYOUT", 0.0) or 0.0),
            "TRF": float(سماكات_البلاطات_المستخرجة.get("TOP ROOF SLAB LAYOUT", 0.0) or 0.0),
        }
        if يوجد_مبنى_خدمة_صريح:
            سماكات_البلاطات_حسب_المستوى["KB"] = 0.0

        def _سماكة_بلاطة_للمستوى(مستوى: str) -> float:
            مستوى_مطبع = _طبع_المستوى(مستوى)
            سماكة = float(سماكات_البلاطات_حسب_المستوى.get(مستوى_مطبع, 0.0) or 0.0)
            if سماكة > 0:
                return سماكة
            return سماكة_بلاطة_فعلية

        def _ارتفاع_حائط_للمستوى(مستوى: str) -> float:
            ارتفاع_دور = _ارتفاع_الدور_للمستوى(مستوى)
            if ارتفاع_دور <= 0:
                return ارتفاع_صافي_تشطيب_مدخل if ارتفاع_صافي_تشطيب_مدخل > 0 else 0.0
            خصم_تشطيب_مستوى = max(
                _سماكة_بلاطة_للمستوى(مستوى),
                min(0.60, عمق_الجسر_فعلي) if عمق_الجسر_فعلي > 0 else (0.0 if استخراج_صارم else 0.50)
            )
            ارتفاع_حائط = max(0.0, ارتفاع_دور - خصم_تشطيب_مستوى)
            if ارتفاع_حائط <= 0 and ارتفاع_صافي_تشطيب_مدخل > 0:
                return ارتفاع_صافي_تشطيب_مدخل
            return ارتفاع_حائط

        النتائج: Dict[str, List[Dict]] = {
            "البنية_التحتية": [],
            "البنية_الفوقية": [],
            "المعمارية":      [],
            "التشطيبات":      [],
            "الفتحات":        [],
        }
        تح = النتائج["البنية_التحتية"]
        فوق = النتائج["البنية_الفوقية"]
        معم = النتائج["المعمارية"]
        تشط = النتائج["التشطيبات"]

        def بند(قسم, اسم, كمية, وحدة):
            قسم.append({"البند": اسم, "الكمية": round(float(كمية), 3), "الوحدة": وحدة})

        # ══════════════════════════════════════════════════════════════════════
        # أ. البنية التحتية
        # ══════════════════════════════════════════════════════════════════════

        # 1. الحفر
        مساحة_حفر_صافي = sum((أ.الطول * أ.العرض * أ.الكمية) for أ in الأسس) if الأسس else 0.0

        # V15.6: أولوية مساحة البصمة:
        # 1) أكبر مضلع معقول (50-800م²) — الأدق
        # 2) المحيط / 4.5 مربع
        # 3) مجموع الغرف / عدد الأدوار — fallback
        مساحة_أكبر_مضلع = float(الفضاء.get("مساحة_أكبر_مضلع", 0.0) or 0.0)
        if مساحة_أكبر_مضلع > 30:
            مساحة_بصمة = مساحة_أكبر_مضلع
            مسجل.info(f"مساحة البصمة من المضلع: {مساحة_بصمة:.1f}م²")
        elif المحيط_الخارجي > 10:
            مساحة_بصمة = round((المحيط_الخارجي / 4.5) ** 2 * 0.75, 2)
            مسجل.info(f"مساحة البصمة من المحيط: {مساحة_بصمة:.1f}م²")
        else:
            مساحة_بصمة = المساحة_الكلية / عدد_أدوار_للقسمة
            مسجل.info(f"مساحة البصمة من الغرف: {مساحة_بصمة:.1f}م²")

        مساحة_حفر_مستخرجة = float(الفضاء.get("مساحة_الحفر", 0.0) or 0.0)
        if استخراج_صارم:
            if سماكة_بلاطة_أرضي_فعلية > 0 and (الأسس or كمرات_الربط):
                if مساحة_بصمة > 10:
                    م_حفر = مساحة_بصمة
                elif مساحة_حفر_صافي > 0:
                    م_حفر = max(مساحة_حفر_صافي, مساحة_حفر_مستخرجة)
                else:
                    م_حفر = مساحة_حفر_مستخرجة or مساحة_بصمة
            elif مساحة_حفر_صافي > 0:
                م_حفر = max(مساحة_حفر_صافي, الفضاء.get("مساحة_الحفر", 0.0) or 0.0)
            else:
                م_حفر = الفضاء.get("مساحة_الحفر", مساحة_بصمة)
        elif ث.مساحة_حفر_يدوية > 0:
            م_حفر = ث.مساحة_حفر_يدوية
        elif str(ث.نوع_الحفر).strip().lower() in {"كتلي", "bulk", "mass"}:
            م_حفر = الفضاء.get("مساحة_الحفر", مساحة_بصمة)
        elif مساحة_حفر_صافي > 0:
            م_حفر = مساحة_حفر_صافي
        else:
            م_حفر = الفضاء.get("مساحة_الحفر", مساحة_بصمة)
        # V15.5: cap excavation area — الحفر الكتلي عادة 1.0-1.5× البصمة
        if مساحة_بصمة > 10 and م_حفر > مساحة_بصمة * 3.0:
            مسجل.warning(f"تصحيح مساحة الحفر ({م_حفر:.1f}م²) → {مساحة_بصمة * 2.0:.1f}م² (بصمة×2.0)")
            م_حفر = round(مساحة_بصمة * 2.0, 2)
        مساحة_بصمة_فعلي = مساحة_بصمة if مساحة_بصمة > 10 else (م_حفر if م_حفر > 0 else مساحة_حفر_صافي)
        بند(تح, "مساحة الحفر",  م_حفر,                    "م²")
        بند(تح, "حجم الحفر",    م_حفر * ث.عمق_الحفر,      "م³")

        # 2. الأسس — F1، F2... × الكمية ثم المجموع
        م_pcc_كلي = 0.0
        ح_أسس_كلي = 0.0
        بيتومين_إجمالي = 0.0          # V15.6: مجمّع كل أنواع البيتومين
        for أ in الأسس:
            م_أ = أ.العرض * أ.الطول * أ.الكمية
            ح_أ = م_أ * أ.العمق
            pcc_ط = أ.الطول + 0.20
            pcc_ع = أ.العرض + 0.20
            ح_pcc = pcc_ط * pcc_ع * أ.سماكة_PCC * أ.الكمية
            # مساحة البيتومين = (قاع PCC) + (4 جوانب بأبعاد PCC × عمق الأساس) × الكمية
            بيتو  = (pcc_ط * pcc_ع + 2*(pcc_ط+pcc_ع)*أ.العمق) * أ.الكمية
            م_pcc_كلي += pcc_ط * pcc_ع * أ.الكمية
            ح_أسس_كلي += ح_أ
            بيتومين_إجمالي += بيتو
            بند(تح, f"مساحة الأساس ({أ.الرمز})",   م_أ,   "م²")
            بند(تح, f"حجم الأساس ({أ.الرمز})",     ح_أ,   "م³")
            بند(تح, f"PCC الأساس ({أ.الرمز})",     ح_pcc, "م³")
            بند(تح, f"بيتومين الأساس ({أ.الرمز})", بيتو,  "م²")
        بند(تح, "إجمالي حجم الأسس", ح_أسس_كلي, "م³")

        # 3. أعمدة العنق
        ح_ع_عنق_كلي = 0.0
        for ع in أعمدة_العنق:
            ار = مستوى_بلاطة_الأرضي_فعلي + (ث.عمق_الحفر - ع.عمق_الأساس - 0.10)
            ار = max(0.0, ار)
            ح  = ع.الطول * ع.العرض * ار * ع.الكمية
            # البيتومين على محيط عمود العنق × ارتفاعه الفعلي فقط (ار)، لا كامل عمق الحفر
            بيتو = (2*(ع.الطول+ع.العرض)) * ار * ع.الكمية
            ح_ع_عنق_كلي += ح
            بيتومين_إجمالي += بيتو
            بند(تح, f"حجم عمود العنق ({ع.الرمز})",    ح,    "م³")
            بند(تح, f"بيتومين عمود العنق ({ع.الرمز})", بيتو, "م²")
        بند(تح, "إجمالي حجم أعمدة العنق", ح_ع_عنق_كلي, "م³")

        # 4. كمرات الربط — الطول من الرسم
        ح_كر_كلي = 0.0
        for ك in كمرات_الربط:
            ح  = ك.الطول * ك.العرض * ك.العمق
            ح_pcc = ك.الطول * (ك.العرض+0.20) * ث.سماكة_PCC
            # بيتومين كمرة الربط = وجهان جانبيان + الوجه السفلي
            بيتو  = ك.الطول * (2*ك.العمق + ك.العرض)
            م_pcc_كلي += ك.الطول * (ك.العرض+0.20)
            ح_كر_كلي  += ح
            بيتومين_إجمالي += بيتو
            بند(تح, f"حجم كمرة الربط ({ك.الرمز})",    ح,     "م³")
            بند(تح, f"PCC كمرة الربط ({ك.الرمز})",    ح_pcc, "م³")
            بند(تح, f"بيتومين كمرة الربط ({ك.الرمز})", بيتو, "م²")
        بند(تح, "إجمالي حجم كمرات الربط", ح_كر_كلي, "م³")

        # 5. كمرات الستراب — الطول من الرسم
        ح_ست_كلي = 0.0
        for س in كمرات_الستراب:
            ح = س.الطول * س.العرض * س.العمق
            ح_ست_كلي += ح
            بند(تح, f"حجم كمرة الستراب ({س.الرمز})", ح, "م³")
        بند(تح, "إجمالي حجم كمرات الستراب", ح_ست_كلي, "م³")

        # 6. البلوك تحت الأرض
        ح_بلوك_تحت = 0.0
        بلوكات_تحت_الأرض = [] if استخراج_صارم else الطلب.بلوكات_تحت_الأرض
        for ب in بلوكات_تحت_الأرض:
            ار = (مستوى_بلاطة_الأرضي_فعلي + ث.عمق_الحفر) - عمق_كمرة_الربط_فعلي - ث.سماكة_PCC
            م  = ب.الطول * ار * ب.الكمية
            بيتو = م * 2
            ح_بلوك_تحت += ب.الطول * ار * ب.الكمية * 0.20
            بند(تح, f"مساحة بلوك تحت الأرض ({ب.الرمز})",   م,    "م²")
            بند(تح, f"بيتومين بلوك تحت الأرض ({ب.الرمز})", بيتو, "م²")

        # 7. بلاطة على الأرض (V15.5: بصمة دور واحد)
        بند(تح, "مساحة بلاطة على الأرض",  مساحة_بصمة_فعلي,        "م²")
        بند(تح, "حجم بلاطة على الأرض",    مساحة_بصمة_فعلي * سماكة_بلاطة_أرضي_فعلية, "م³")

        # 8. الردم
        ح_pcc_إج = (sum((أ.الطول+0.20)*(أ.العرض+0.20)*أ.سماكة_PCC*أ.الكمية for أ in الأسس) +
                    sum(ك.الطول*(ك.العرض+0.20)*ث.سماكة_PCC for ك in كمرات_الربط))
        ح_ع_عنق  = sum(ع.الطول*ع.العرض*max(0,مستوى_بلاطة_الأرضي_فعلي+(ث.عمق_الحفر-ع.عمق_الأساس-0.10))*ع.الكمية for ع in أعمدة_العنق)
        ح_كر     = sum(ك.الطول*ك.العرض*ك.العمق for ك in كمرات_الربط)
        ح_ست     = sum(س.الطول*س.العرض*س.العمق for س in كمرات_الستراب)
        ح_أسس    = sum(أ.العرض*أ.الطول*أ.العمق*أ.الكمية for أ in الأسس)
        ح_إنشاء_كلي = ح_pcc_إج + ح_ع_عنق + ح_كر + ح_ست + ح_أسس + ح_بلوك_تحت
        حجم_حفر = م_حفر * (مستوى_بلاطة_الأرضي_فعلي + ث.عمق_الحفر)
        if ح_إنشاء_كلي > 0.1:
            ردم = حجم_حفر - ح_إنشاء_كلي
        else:
            # V15.5: بدون بيانات إنشائية، الردم ≈ 35% من حجم الحفر (تقدير)
            ردم = حجم_حفر * 0.35
        بند(تح, "حجم الردم", max(0, ردم), "م³")
        # V15.5: إجمالي PCC كبند مجمع
        if ح_pcc_إج > 0:
            بند(تح, "PCC للمؤسسات", ح_pcc_إج, "م³")
        # V15.6: إجمالي عزل البيتومين (مجمّع من جميع الأسس + العنق + كمرات الربط)
        بند(تح, "إجمالي عزل البيتومين", round(بيتومين_إجمالي, 2), "م²")

        # 9. مبيد النمل الأبيض (V15.5: بصمة دور واحد)
        بند(تح, "مبيد النمل الأبيض", مساحة_بصمة if استخراج_صارم else مساحة_بصمة * 1.15, "م²")

        # 10. نايلون أسود (V15.5: بصمة دور واحد)
        بند(تح, "نايلون أسود (بولي إيثيلين)", مساحة_بصمة if استخراج_صارم else مساحة_بصمة * 1.15, "م²")

        # 11. رصيف الطرق
        if (not استخراج_صارم) and ث.يوجد_رصيف_طرق:
            بند(تح, "مساحة رصيف الطرق",  م_حفر,                         "م²")
            بند(تح, "حجم رصيف الطرق",    م_حفر * ث.سماكة_رصيف_الطرق,   "م³")

        # ══════════════════════════════════════════════════════════════════════
        # ب. البنية الفوقية
        # ══════════════════════════════════════════════════════════════════════

        # 12. البلاطات (Net Area Deduction implemented)
        if مساحة_بلاطة_رئيسية > 0 or مساحة_بلاطات_أعلى_السطح > 0:
            سماكة_أول = سماكات_البلاطات_المستخرجة.get("FIRST SLAB LAYOUT", سماكة_بلاطة_فعلية)
            سماكة_ثان = سماكات_البلاطات_المستخرجة.get("SECOND SLAB LAYOUT", سماكة_بلاطة_فعلية)
            سماكة_سقف = سماكات_البلاطات_المستخرجة.get("ROOF SLAB LAYOUT", سماكة_بلاطة_فعلية)
            سماكة_أعلى_السطح = سماكات_البلاطات_المستخرجة.get("TOP ROOF SLAB LAYOUT", سماكة_بلاطة_فعلية)
            if مساحة_بلاطة_أول > 0:
                بند(فوق, "حجم بلاطة الدور الأول", مساحة_بلاطة_أول * سماكة_أول, "م³")
            if مساحة_بلاطة_ثان > 0:
                بند(فوق, "حجم بلاطة الدور الثاني", مساحة_بلاطة_ثان * سماكة_ثان, "م³")
            if مساحة_بلاطة_سقف > 0:
                بند(فوق, "حجم بلاطة السقف", مساحة_بلاطة_سقف * سماكة_سقف, "م³")
            if مساحة_بلاطات_أعلى_السطح > 0:
                بند(فوق, "حجم بلاطات أعلى السطح", مساحة_بلاطات_أعلى_السطح * سماكة_أعلى_السطح, "م³")
        elif not استخراج_صارم:
            مساحة_الخصم_افتراضي = 105.0 # Fallback only if slab layers are unavailable
            for دور in range(1, عدد_الأدوار_مرجعي + 1):
                بند(فوق, f"حجم البلاطة (الدور {دور})",
                    max(0, (المساحة_الكلية - مساحة_الخصم_افتراضي)) * سماكة_بلاطة_فعلية, "م³")

        # 13. الأعمدة (Net Height Deduction implemented)
        ح_أع_كلي = 0.0
        أعماق_كمرات_مستخرجة = [ك.العمق for ك in الكمرات if getattr(ك, "العمق", 0) > 0]
        if أعماق_كمرات_مستخرجة:
            عمق_الجسر_فعلي = sum(أعماق_كمرات_مستخرجة) / len(أعماق_كمرات_مستخرجة)
        else:
            عمق_الجسر_فعلي = 0.0 if استخراج_صارم else 0.60
        for ع in الأعمدة:
            مستوى = _طبع_المستوى(getattr(ع, "المستوى", "") or "")
            ارتفاع_مرجعي = _ارتفاع_الدور_للمستوى(مستوى) if مستوى else ارتفاع_الدور_فعلي
            # إذا استخرج Gemini ارتفاع العمود مباشرة، نستخدمه — وإلا نحسبه من ارتفاع الدور
            ارتفاع_مستخرج = getattr(ع, "الارتفاع", 0) or 0
            if ارتفاع_مستخرج > 0:
                ارتفاع_صافي = max(0.0, ارتفاع_مستخرج - عمق_الجسر_فعلي)
            else:
                ارتفاع_صافي = max(0.0, ارتفاع_مرجعي - عمق_الجسر_فعلي) if ارتفاع_مرجعي > 0 else 0.0
            # الكمية الآن كلية (كل الأدوار مجمعة) — لا ضرب في عدد الأدوار
            ح = ع.الطول * ع.العرض * ارتفاع_صافي * ع.الكمية
            ح_أع_كلي += ح
            اسم_البند = f"حجم العمود ({ع.الرمز})" if not مستوى else f"حجم العمود ({ع.الرمز} - {مستوى})"
            بند(فوق, اسم_البند, ح, "م³")
        بند(فوق, "إجمالي حجم الأعمدة", ح_أع_كلي, "م³")

        # 14. الكمرات — الطول من الرسم
        ح_كم_كلي = 0.0
        for ك in الكمرات:
            ح = ك.الطول * ك.العرض * max(0, ك.العمق - سماكة_بلاطة_فعلية)
            ح_كم_كلي += ح
            بند(فوق, f"حجم الكمرة ({ك.الرمز})", ح, "م³")
        بند(فوق, "إجمالي حجم الكمرات", ح_كم_كلي, "م³")

        ارتفاعات_واجهات_فعالة = [
            _ارتفاع_حائط_للمستوى(م)
            for م in مستويات_واجهات_مغلقة
            if _ارتفاع_حائط_للمستوى(م) > 0
        ]
        مجموع_ارتفاعات_الواجهات = sum(ارتفاعات_واجهات_فعالة)

        محيط_غرف_داخلي_مرجعي = sum(
            float(غ.get("المحيط", 0.0) or 0.0)
            for غ in غرف_التشطيب
            if not _غرفة_بلكونة(غ)
        )
        مساحة_حوائط_الغرف_المرجعية = sum(
            float(غ.get("المحيط", 0.0) or 0.0) * _ارتفاع_حائط_للمستوى(_مستوى_غرفة(غ))
            for غ in غرف_التشطيب
            if not _غرفة_بلكونة(غ) and _ارتفاع_حائط_للمستوى(_مستوى_غرفة(غ)) > 0
        )
        ارتفاع_حائط_داخلي_مرجح = (
            مساحة_حوائط_الغرف_المرجعية / محيط_غرف_داخلي_مرجعي
            if محيط_غرف_داخلي_مرجعي > 0 else 0.0
        )
        ارتفاع_حائط_فعلي = max(ارتفاعات_واجهات_فعالة, default=ارتفاع_حائط_داخلي_مرجح)

        if استخراج_صارم and ارتفاع_حائط_فعلي <= 0:
            مسجل.warning("تعذر استخراج ارتفاعات الأدوار من الرسم ولا يوجد إدخال يدوي بديل — البنود المعتمدة على الارتفاع ستخرج صفراً")
        elif (not استخراج_صارم) and ارتفاع_حائط_فعلي <= 0:
            ارتفاع_حائط_فعلي = ارتفاع_صافي_تشطيب_مدخل

        if مجموع_ارتفاعات_الواجهات <= 0 and ارتفاع_حائط_فعلي > 0 and عدد_مستويات_واجهات > 0:
            مجموع_ارتفاعات_الواجهات = ارتفاع_حائط_فعلي * عدد_مستويات_واجهات

        مساحة_واجهات_خارجية = max(0.0, (المحيط_الخارجي * مجموع_ارتفاعات_الواجهات) - (م_نوافذ + م_باب_رئيسي))
        مساحة_بلوك_20_داخلي = max(0.0, (جدران_20 * ارتفاع_حائط_داخلي_مرجح) - م_أبواب)
        مساحة_بلوك_10_داخلي = max(0.0, (جدران_10 * ارتفاع_حائط_داخلي_مرجح) - م_أبواب)

        # 15. الدرج
        خرسانة_الدرج_نهائية = خرسانة_الدرج_مستخرجة if خرسانة_الدرج_مستخرجة > 0 else ث.خرسانة_الدرج
        if خرسانة_الدرج_نهائية > 0:
            بند(فوق, "خرسانة الدرج", خرسانة_الدرج_نهائية, "م³")

        # ══════════════════════════════════════════════════════════════════════
        # ج. المعمارية
        # ══════════════════════════════════════════════════════════════════════

        # 16. بلوك 20 خارجي
        بند(معم, "بلوك 20 سم خارجي", مساحة_واجهات_خارجية, "م²")
        if مساحة_بلوك_التصوينة > 0:
            بند(معم, "بلوك 20 سم تصوينة السطح", مساحة_بلوك_التصوينة, "م²")

        # 17. بلوك 20 داخلي
        بند(معم, "بلوك 20 سم داخلي", مساحة_بلوك_20_داخلي, "م²")

        # 18. بلوك 10 داخلي
        بند(معم, "بلوك 10 سم داخلي", مساحة_بلوك_10_داخلي, "م²")

        # 19. لياسة داخلية
        if استخراج_صارم and مساحة_حوائط_الغرف_المرجعية > 0:
            مساحة_اللياسة_الاجمالية = max(0, مساحة_حوائط_الغرف_المرجعية - (م_أبواب + م_نوافذ))
        else:
            مساحة_اللياسة_الاجمالية = max(0, (((جدران_20 + جدران_10) * 2 + المحيط_الخارجي) * ارتفاع_حائط_فعلي) - (م_أبواب + م_نوافذ))
        
        # دهان الوجه الداخلي (من الأعلى) للتصوينة فقط — الوجه الخارجي مضمون في تشطيب الواجهة
        دهان_التصوينة = مساحة_بلوك_التصوينة * 1
        بند(معم, "لياسة داخلية", مساحة_اللياسة_الاجمالية, "م²")

        # 20. تشطيب الواجهة الخارجية
        # يشمل BOQ أحياناً: سور المجمع + مبنى الخدمة — نُعلِّم كـ REQUIRES_MANUAL_INPUT
        تشطيب_واجهة_خارجية = مساحة_واجهات_خارجية + مساحة_بلوك_التصوينة
        if not استخراج_صارم:
            تشطيب_واجهة_خارجية = max(0, تشطيب_واجهة_خارجية + 2.5)
        _حالة_خارجية = "REQUIRES_MANUAL_INPUT" if تشطيب_واجهة_خارجية <= 1.0 else ""
        بند_خارج = {"البند": "تشطيب الواجهة الخارجية", "الكمية": تشطيب_واجهة_خارجية, "الوحدة": "م²", "الحالة": _حالة_خارجية}
        معم.append(بند_خارج)

        # 21. عزل مائي
        عزل_مائي = max(0, م_مبلل_كلي + م_بلكونة)
        if not استخراج_صارم:
            عزل_مائي *= 1.1
        بند(معم, "عزل مائي", عزل_مائي, "م²")

        # 22. نظام السقف المركب — Roof_Slab_Area + Parapet_Upturn
        # + upstands التصوينة (محيط × ارتفاع التصوينة)
        _مساحة_سقف_للعزل = مساحة_بلاطة_سقف + مساحة_بلاطات_أعلى_السطح
        if _مساحة_سقف_للعزل <= 0:
            # fallback: بلاطة أولى → مساحة دور واحد (وليس المساحة الكلية لكل الأدوار)
            _مساحة_سقف_للعزل = مساحة_بلاطة_أول if مساحة_بلاطة_أول > 0 else (المساحة_الكلية / max(عدد_أدوار_للقسمة, 1))
        _upstands_التصوينة = طول_التصوينة * ث.ارتفاع_التصوينة
        نظام_السقف_المركب = _مساحة_سقف_للعزل + _upstands_التصوينة
        بند(معم, "نظام السقف المركب", نظام_السقف_المركب, "م²")

        # ══════════════════════════════════════════════════════════════════════
        # د. التشطيبات
        # ══════════════════════════════════════════════════════════════════════

        # 23. بلاط جاف
        بند(تشط, "بلاط المناطق الجافة",         م_جاف,                              "م²")
        # 24. سكرتة — Bible equation: محيط كل الغرف الجافة - عروض الأبواب
        سكرتة = max(0, مح_جاف - ع_أبواب)
        بند(تشط, "سكرتة",                        سكرتة,                              "م.ط")
        # 25. دهان (Internal Paint)
        # نحسب مساحة جدران الغرف الجافة (تدهن بالكامل)
        مساحة_جدران_جافة = sum(
            float(غ.get("المحيط", 0.0) or 0.0) * _ارتفاع_حائط_للمستوى(_مستوى_غرفة(غ))
            for غ in غرف_التشطيب
            if not _غرفة_مبللة(غ) and not _غرفة_بلكونة(غ) and _ارتفاع_حائط_للمستوى(_مستوى_غرفة(غ)) > 0
        )
        # نحسب الجزء المتبقي من الغرف المبللة (فوق البلاط)
        مساحة_جدران_مبللة_فوق_البلاط = sum(
            float(غ.get("المحيط", 0.0) or 0.0) * max(0, _ارتفاع_حائط_للمستوى(_مستوى_غرفة(غ)) - ث.ارتفاع_بلاط_الجدران_المبللة)
            for غ in غرف_التشطيب
            if _غرفة_مبللة(غ) and _ارتفاع_حائط_للمستوى(_مستوى_غرفة(غ)) > 0
        )
        # الخصم الكلي للفتحات (الأبواب والنوافذ) - نعتبرها موزعة
        مساحة_دهان_صافي = max(0, (مساحة_جدران_جافة + مساحة_جدران_مبللة_فوق_البلاط) - (م_أبواب + م_نوافذ))
        # دهان التصوينة = بند خارجي منفصل — لا يُضاف للدهان الداخلي
        بند(تشط, "دهان", مساحة_دهان_صافي, "م²")
        
        # 26. سقف جاف
        بند(تشط, "سقف المناطق الجافة",           م_جاف,                              "م²")
        # 27. سيراميك مبلل
        # إذا لم تُستخرج غرف مبللة كافية (< 3م² مجموع) → مشكوك في الاستخراج → MANUAL
        _حالة_مبلل = "REQUIRES_MANUAL_INPUT" if م_مبلل_كلي < 3.0 else ""
        بند_مبلل = {"البند": "سيراميك المناطق المبللة", "الكمية": م_مبلل_كلي if not _حالة_مبلل else None, "الوحدة": "م²", "الحالة": _حالة_مبلل}
        تشط.append(بند_مبلل)
        # 28. بلاط جدران — ارتفاع البلاط = ارتفاع الحائط الفعلي - 0.3م (تشطيب الأرضية)
        if غ_مبللة and any(_ارتفاع_حائط_للمستوى(_مستوى_غرفة(غ)) > 0 for غ in غ_مبللة):
            مساحة_بلاط_الجدران = sum(
                float(غ.get("المحيط", 0.0) or 0.0) * max(0.0, _ارتفاع_حائط_للمستوى(_مستوى_غرفة(غ)) - 0.3)
                for غ in غ_مبللة
                if _ارتفاع_حائط_للمستوى(_مستوى_غرفة(غ)) > 0
            )
            if مساحة_بلاط_الجدران <= 0:
                مساحة_بلاط_الجدران = مح_مبلل * ث.ارتفاع_بلاط_الجدران_المبللة  # fallback
        else:
            مساحة_بلاط_الجدران = مح_مبلل * ث.ارتفاع_بلاط_الجدران_المبللة
        بند(تشط, "بلاط الجدران",                 مساحة_بلاط_الجدران,                 "م²")
        # 29. سقف مبلل
        بند(تشط, "سقف المناطق المبللة",          م_مبلل_كلي,                         "م²")
        # 30. بلاط بلكونة
        بند(تشط, "بلاط البلكونة",                م_بلكونة,                           "م²")
        # 31. عتبات رخام
        بند(تشط, "عتبات رخام",                   عدد_الأبواب * ث.طول_العتبة_لكل_باب, "م.ط")

        # ══════════════════════════════════════════════════════════════════════
        # هـ. الفتحات
        # ══════════════════════════════════════════════════════════════════════
        for ف in الفتحات:
            النتائج["الفتحات"].append({
                "البند":    f"فتحة ({ف.الرمز})",
                "الكمية":   round(ف.الطول*ف.الارتفاع*ف.الكمية, 3),
                "الوحدة":   "م²",
                "العدد":    ف.الكمية,
                "الطول":    ف.الطول,
                "الارتفاع": ف.الارتفاع,
            })

        return النتائج


# ─────────────────────────────────────────────────────────────────────────────
# دالة مساعدة: تعزيز درجة الثقة بناءً على جودة الاستخراج الإنشائي
# ─────────────────────────────────────────────────────────────────────────────

def _حساب_درجة_ثقة_محسّنة(درجة_الغرف: dict, مسطح: list) -> dict:
    """
    تدمج درجة ثقة الغرف مع جودة الاستخراج الإنشائي.
    - 50%: نسبة الفضاءات المسمّاة (المقياس الأصلي)
    - 50%: نسبة البنود الإنشائية الأساسية ذات الكميات غير الصفرية
    """
    بنود_إنشائية_أساسية = {
        "حجم الحفر",
        "إجمالي حجم الأسس",
        "إجمالي حجم أعمدة العنق",
        "إجمالي حجم كمرات الربط",
        "إجمالي حجم الأعمدة",
        "إجمالي حجم الكمرات",
    }
    درجة_غرف = float(درجة_الغرف.get("النتيجة", 0))
    عدد_مستخرج = sum(
        1 for ب in مسطح
        if ب.get("البند") in بنود_إنشائية_أساسية and float(ب.get("الكمية", 0) or 0) > 0
    )
    درجة_إنشاء = min(100.0, round(عدد_مستخرج / len(بنود_إنشائية_أساسية) * 100))
    درجة_مدمجة = max(0, min(100, round(درجة_غرف * 0.50 + درجة_إنشاء * 0.50)))
    درجة_حرف = "ممتاز" if درجة_مدمجة >= 85 else "جيد جداً" if درجة_مدمجة >= 70 else "جيد" if درجة_مدمجة >= 50 else "ضعيف"
    return {
        **درجة_الغرف,
        "النتيجة":      درجة_مدمجة,
        "الدرجة":       درجة_حرف,
        "درجة_الغرف":   round(درجة_غرف),
        "درجة_الإنشاء": round(درجة_إنشاء),
    }


# ─────────────────────────────────────────────────────────────────────────────
# دالة مساعدة: تعليم البنود الإنشائية التي تحتاج إدخالاً يدوياً (نظام 70/30)
# ─────────────────────────────────────────────────────────────────────────────

def _علّم_بنود_str_يدوية(نتائج: dict, درجة_إنشاء: float) -> dict:
    """
    نظام 70/30:
      - 70% ARCH: تُستخرج تلقائياً من PDF (المعمارية + التشطيبات) — لا تغيير.
      - 30% STR:  عند درجة_إنشاء < 50%، يُعلَّم كل بند إنشائي بقيمة صفر
                  على أنه REQUIRES_MANUAL_INPUT بدلاً من إخراج أصفار مضللة.

    يُعدَّل النتائج في المكان (in-place) ويُعيد نفس المرجع.
    البنود ذات قيم حقيقية غير صفرية تبقى كما هي حتى عند confidence منخفض.
    """
    if درجة_إنشاء >= 50:
        return نتائج  # الاستخراج كافٍ — لا حاجة للتعليم

    # confidence < 50% → جميع بنود STR غير موثوقة بما فيها القيم غير الصفرية الخاطئة
    أقسام_إنشائية = {"البنية_التحتية", "البنية_الفوقية"}
    for قسم, بنود in نتائج.items():
        if قسم not in أقسام_إنشائية:
            continue
        for بند in بنود:
            بند["الكمية"]  = None
            بند["الحالة"] = "REQUIRES_MANUAL_INPUT"
    return نتائج


# ─────────────────────────────────────────────────────────────────────────────
# نظام معايير المتوسطات — Averages Sanity Clamp
# المصدر: EQUATION_ITEMS_AVG.txt — 159 مشروع — تاريخ: 2026-03-25
# ─────────────────────────────────────────────────────────────────────────────

_متوسطات_الكميات: Dict[str, Dict[str, float]] = {
    # ── G (5 مشاريع، أرضي فقط) ──────────────────────────────────────────────
    "G": {
        "إجمالي حجم الأسس":              120.9,
        "PCC للمؤسسات":                  30.7,
        "إجمالي حجم أعمدة العنق":         5.3,
        "إجمالي حجم كمرات الربط":        19.5,
        "حجم بلاطة على الأرض":           28.8,
        "حجم الحفر":                     344.5,
        "حجم الردم":                      85.3,
        "مبيد النمل الأبيض":             242.3,
        "نايلون أسود (بولي إيثيلين)":    210.7,
        "نايلون أسود":                   210.7,
        "إجمالي عزل البيتومين":          420.0,      # V15.6
        "حجم بلاطة الدور الأول":          65.0,       # V15.6 — بلاطة السقف G type ≈ 350m²×0.20m=70m³
        "بلوك 20 سم خارجي":              67.1,
        "بلوك 20 سم داخلي":              72.3,
        "بلوك 20 سم تصوينة السطح":      22.0,
        "بلوك 10 سم داخلي":              12.9,
        "نظام السقف المركب":             110.0,
        "بلاط المناطق الجافة":           515.7,
        "سقف المناطق الجافة":            515.7,
        "عتبات رخام":                     10.1,
    },
    # ── G+1 (101 مشروع) ───────────────────────────────────────────────────────
    "G+1": {
        "إجمالي حجم الأسس":               58.0,
        "PCC للمؤسسات":                   22.0,
        "إجمالي حجم أعمدة العنق":          4.8,
        "إجمالي حجم كمرات الربط":         23.4,
        "إجمالي حجم كمرات الستراب":       23.4,
        "مساحة بلوك تحت الأرض":           96.7,
        "حجم بلاطة على الأرض":            29.5,
        "حجم الحفر":                      626.2,
        "حجم الردم":                      614.5,
        "مبيد النمل الأبيض":              482.7,
        "نايلون أسود (بولي إيثيلين)":     419.8,
        "نايلون أسود":                    419.8,
        "إجمالي عزل البيتومين":           580.0,
        "حجم بلاطة الدور الأول":          140.0,
        "حجم البلاطة (الدور 1)":          140.0,
        "حجم بلاطة السقف":                55.0,
        "حجم البلاطة (الدور 2)":           55.0,
        "حجم بلاطة الدور الثاني":          55.0,
        "إجمالي حجم الكمرات":             16.0,
        "إجمالي حجم الأعمدة":             29.0,
        "خرسانة الدرج":                   10.3,
        "بلوك 20 سم خارجي":              389.7,
        "بلوك 20 سم تصوينة السطح":      25.0,
        "بلوك 20 سم داخلي":              292.0,
        "بلوك 10 سم داخلي":              210.0,
        "لياسة داخلية":                 1398.8,
        "تشطيب الواجهة الخارجية":         825.0,
        "عزل مائي":                        40.0,
        "نظام السقف المركب":              222.3,
        "بلاط المناطق الجافة":            318.1,
        "سكرتة":                          340.0,
        "دهان":                           920.0,
        "سقف المناطق الجافة":             318.1,
        "سيراميك المناطق المبللة":          60.0,
        "بلاط الجدران":                   285.0,
        "سقف المناطق المبللة":             60.0,
        "بلاط البلكونة":                   12.0,
        "عتبات رخام":                      21.0,
    },
    # ── G+2 (46 مشروع) ───────────────────────────────────────────────────────
    "G+2": {
        "إجمالي حجم الأسس":               52.6,
        "PCC للمؤسسات":                   24.0,
        "إجمالي حجم أعمدة العنق":          4.8,
        "إجمالي حجم كمرات الربط":         23.8,
        "إجمالي حجم كمرات الستراب":       23.8,
        "مساحة بلوك تحت الأرض":           96.7,
        "حجم بلاطة على الأرض":            23.8,
        "حجم الحفر":                      751.0,
        "حجم الردم":                      775.7,
        "مبيد النمل الأبيض":              512.9,
        "نايلون أسود (بولي إيثيلين)":     446.0,
        "نايلون أسود":                    446.0,
        "إجمالي عزل البيتومين":           750.0,
        "حجم بلاطة الدور الأول":           84.9,
        "حجم البلاطة (الدور 1)":           84.9,
        "حجم بلاطة السقف":                29.1,
        "حجم البلاطة (الدور 2)":           29.1,
        "حجم بلاطة الدور الثاني":          29.1,
        "إجمالي حجم الكمرات":             17.5,
        "إجمالي حجم الأعمدة":             27.3,
        "خرسانة الدرج":                    9.7,
        "بلوك 20 سم خارجي":              251.5,
        "بلوك 20 سم تصوينة السطح":      35.0,
        "بلوك 20 سم داخلي":              299.7,
        "بلوك 10 سم داخلي":              126.8,
        "لياسة داخلية":                 1604.1,
        "تشطيب الواجهة الخارجية":         756.1,
        "عزل مائي":                        52.0,
        "نظام السقف المركب":              248.1,
        "بلاط المناطق الجافة":            432.5,
        "سكرتة":                          588.0,
        "دهان":                          1510.5,
        "سقف المناطق الجافة":             432.5,
        "سيراميك المناطق المبللة":          87.4,
        "بلاط الجدران":                   143.6,
        "سقف المناطق المبللة":             87.4,
        "بلاط البلكونة":                   38.3,
        "عتبات رخام":                      33.0,
    },
}

# المساحة المرجعية (إجمالي المسطح المبني) لكل نوع أدوار
_مساحة_مرجعية_للأدوار: Dict[str, float] = {
    "G":   350.0,
    "G+1": 650.0,
    "G+2": 950.0,
}



def _اكتشف_نوع_المشروع(نتائج: dict) -> str:
    """اكتشف نوع المشروع (G / G+1 / G+2) من البلاطات المستخرجة."""
    def _قيمة(*أسماء):
        for قسم, بنود in نتائج.items():
            for بند in بنود:
                if بند.get("البند") in أسماء:
                    try:
                        return float(بند.get("الكمية") or 0)
                    except (TypeError, ValueError):
                        return 0.0
        return 0.0

    has_slab2 = _قيمة(
        "حجم بلاطة السقف", "حجم بلاطة الدور الثاني",
        "حجم البلاطة (الدور 2)", "حجم بلاطات أعلى السطح"
    ) > 0
    has_slab1 = _قيمة(
        "حجم بلاطة الدور الأول", "حجم البلاطة (الدور 1)"
    ) > 0
    return "G+2" if has_slab2 else ("G+1" if has_slab1 else "G")


# ─────────────────────────────────────────────────────────────────────────────
# 1/4 — فرض قواعد الكتاب المقدس (Bible Rules)
# ─────────────────────────────────────────────────────────────────────────────

def _فرض_قواعد_الكتاب(نتائج: dict, بيانات_فضاء: dict) -> None:
    """
    Bible Rules — قيود مطلقة لا تُنتهك أبداً.
    يُعدَّل النتائج in-place.
    """
    def _ابحث(اسم):
        for قسم, بنود in نتائج.items():
            for بند in بنود:
                if بند.get("البند") == اسم:
                    return بند
        return None

    # ── القاعدة 2: السقف = الأرضية (ABSOLUTE) ────────────────────────────
    for أرض, سقف in [
        ("بلاط المناطق الجافة", "سقف المناطق الجافة"),
        ("سيراميك المناطق المبللة", "سقف المناطق المبللة"),
    ]:
        بند_أرض = _ابحث(أرض)
        بند_سقف = _ابحث(سقف)
        if بند_أرض and بند_سقف:
            try:
                ق_أرض = float(بند_أرض.get("الكمية") or 0)
                ق_سقف = float(بند_سقف.get("الكمية") or 0)
            except (TypeError, ValueError):
                continue
            if ق_أرض > 0 or ق_سقف > 0:
                val = max(ق_أرض, ق_سقف)
                بند_أرض["الكمية"] = val
                بند_سقف["الكمية"] = val
                if بند_أرض.get("الحالة") == "REQUIRES_MANUAL_INPUT":
                    بند_أرض["الحالة"] = "BIBLE_ENFORCED"
                if بند_سقف.get("الحالة") == "REQUIRES_MANUAL_INPUT":
                    بند_سقف["الحالة"] = "BIBLE_ENFORCED"

    # ── القاعدة 1: رصيف الطرق يُخصم من الردم ────────────────────────────
    بند_ردم = _ابحث("حجم الردم")
    بند_رصيف = _ابحث("حجم رصيف الطرق")
    if بند_ردم and بند_رصيف:
        try:
            ردم = float(بند_ردم.get("الكمية") or 0)
            رصيف = float(بند_رصيف.get("الكمية") or 0)
            if رصيف > 0 and ردم > رصيف:
                بند_ردم["الكمية"] = round(ردم - رصيف, 1)
        except (TypeError, ValueError):
            pass


# ─────────────────────────────────────────────────────────────────────────────
# 2/4 — المتوسطات كشبكة أمان ذكية (Smart Averages)
# ─────────────────────────────────────────────────────────────────────────────

def _تصحيح_بالمتوسطات_الذكي(نتائج: dict, بيانات_فضاء: dict) -> dict:
    """
    المتوسطات كشبكة أمان ذكية — ليست بديلاً أعمى.

    ┌──────────────────────────────────────────────────────────────────┐
    │ Engine استخرج قيمة حقيقية ≤ 3× المتوسط              → أبقها   │
    │ Engine فشل (0 أو MANUAL)                              → avg    │
    │ Engine مبالغ (> 3× المتوسط)                           → avg    │
    │ بنود إنشائية بدون confidence + engine خاطئ            → avg    │
    └──────────────────────────────────────────────────────────────────┘
    """
    import math as _math

    نوع = _اكتشف_نوع_المشروع(نتائج)

    _بنود_محيط = {
        "بلوك 20 سم خارجي", "بلوك 20 سم تصوينة السطح",
        "بلوك 20 سم داخلي", "بلوك 10 سم داخلي",
        "تشطيب الواجهة الخارجية", "سكرتة",
        "مساحة بلوك تحت الأرض", "عتبات رخام",
    }

    # بنود يُخطئ فيها engine دائماً (البيانات الإنشائية غالباً ناقصة)
    _بنود_str_fallback = {
        "إجمالي حجم الأسس", "إجمالي حجم أعمدة العنق",
        "إجمالي حجم كمرات الربط", "إجمالي حجم كمرات الستراب",
        "خرسانة الدرج", "إجمالي حجم الأعمدة", "إجمالي حجم الكمرات",
        "إجمالي عزل البيتومين",    # V15.6: البيتومين الكلي
    }

    # بنود block و parapet — engine يبالغ بسبب wall detection
    _بنود_block_override = {
        "بلوك 20 سم خارجي", "بلوك 20 سم تصوينة السطح",
        "بلوك 20 سم داخلي", "بلوك 10 سم داخلي",
    }

    # بنود يُستبدل فيها engine بالمتوسط دائماً (engine يُخطئ بشكل بنيوي)
    _بنود_دائماً_avg = {
        "نظام السقف المركب", "عزل مائي",
    }

    مساحة_المشروع = float(بيانات_فضاء.get("المساحة_الكلية", 0) or 0)

    # ── حماية نوع المشروع: G غير منطقي إذا المساحة < 50 أو > 400 ──
    if نوع == "G" and (مساحة_المشروع < 50 or مساحة_المشروع > 400):
        نوع = "G+1"

    مساحة_مرجعية = _مساحة_مرجعية_للأدوار.get(نوع, 650.0)
    if مساحة_المشروع >= 50:
        معامل_خطي = max(0.25, min(4.0, مساحة_المشروع / مساحة_مرجعية))
        معامل_محيط = max(0.30, min(2.5, _math.sqrt(مساحة_المشروع / مساحة_مرجعية)))
    else:
        معامل_خطي = معامل_محيط = 1.0

    متوسطات = _متوسطات_الكميات.get(نوع, {})
    if not متوسطات:
        return نتائج

    for قسم, بنود in نتائج.items():
        for بند in بنود:
            اسم = بند.get("البند", "")
            if اسم not in متوسطات:
                continue
            معامل = معامل_محيط if اسم in _بنود_محيط else معامل_خطي
            avg_مقيّس = round(متوسطات[اسم] * معامل, 1)
            حالة = بند.get("الحالة", "") or ""
            try:
                كمية = float(بند.get("الكمية") or 0)
            except (TypeError, ValueError):
                كمية = 0.0

            # ── MANUAL أو صفر → avg fallback ──
            if كمية == 0.0 or "REQUIRES_MANUAL_INPUT" in حالة:
                بند["الكمية"] = avg_مقيّس
                بند["الحالة"] = "AVG_FALLBACK"

            # ── بنود دائماً بالمتوسط (engine يُخطئ بنيوياً) ──
            elif اسم in _بنود_دائماً_avg:
                بند["الكمية"] = avg_مقيّس
                بند["الحالة"] = "AVG_OVERRIDE"

            # ── Block items: engine wall-detection يبالغ دائماً → avg (تصحيح إذا > 1.4× avg) ──
            elif اسم in _بنود_block_override and كمية > avg_مقيّس * 1.4:
                بند["الكمية"] = avg_مقيّس
                بند["الحالة"] = "AVG_BLOCK_CORRECTED"

            # ── STR fallback items: بدون بيانات إنشائية موثوقة → avg ──
            elif اسم in _بنود_str_fallback and كمية > avg_مقيّس * 2.0:
                بند["الكمية"] = avg_مقيّس
                بند["الحالة"] = "AVG_STR_CORRECTED"

            # ── أي بند مبالغ > 3× avg → cap ──
            elif كمية > avg_مقيّس * 3.0 and avg_مقيّس > 0:
                بند["الكمية"] = avg_مقيّس
                بند["الحالة"] = "AVG_CAPPED"

            # ── بند أقل من 0.25× avg → engine أخطأ (مقياس خاطئ غالباً) ──
            elif كمية > 0 and avg_مقيّس > 0 and كمية < avg_مقيّس * 0.25:
                بند["الكمية"] = avg_مقيّس
                بند["الحالة"] = "AVG_FLOOR_CORRECTED"

            # ── ضمن المدى → keep engine value ──
            # else: pass

    # ── إضافة بنود مفقودة من المتوسطات (المحرك لم يُخرجها) ──────────────
    _الموجودة = {بند.get("البند", "") for قسم in نتائج.values() for بند in قسم}
    for اسم_avg, قيمة_avg in متوسطات.items():
        if اسم_avg in _الموجودة:
            continue
        _مع = معامل_محيط if اسم_avg in _بنود_محيط else معامل_خطي
        _avg = round(قيمة_avg * _مع, 1)
        if _avg <= 0:
            continue
        # تحديد القسم المناسب
        if any(k in اسم_avg for k in ("حفر", "ردم", "أسس", "عنق", "ربط", "ستراب", "بلوك تحت", "بلاطة على", "PCC", "نمل", "نايلون", "رصيف", "بيتومين")):
            _قسم = "البنية_التحتية"
        elif any(k in اسم_avg for k in ("أعمدة", "كمرات", "بلاطة الدور", "بلاطة السقف", "بلاطات", "درج")):
            _قسم = "البنية_الفوقية"
        elif "بلوك" in اسم_avg:
            _قسم = "المعمارية"
        else:
            _قسم = "التشطيبات"
        if _قسم not in نتائج:
            نتائج[_قسم] = []
        نتائج[_قسم].append({
            "البند": اسم_avg, "الكمية": _avg, "الوحدة": "م²", "الحالة": "AVG_INJECTED",
        })

    return نتائج


# ─────────────────────────────────────────────────────────────────────────────
# 3/4 — Gemini 2.5 Flash Supervisor بكل المعرفة
# ─────────────────────────────────────────────────────────────────────────────

# ── بيانات تدريب حقيقية من 289 مشروع فيلل إماراتية ──────────────────────────
# يُستخدم هذا القاموس في prompt مشرف Gemini فقط — لا يُعدّل _متوسطات_الكميات
# جميع القيم من avg_qty_result_305.json (مستخلصة من BOQ)
_نطاقات_تدريب_289 = {
    # item_code: { avg, min, max, count, unit, ar_name }
    "EXCAVATION":       {"avg": 269.72, "min": 2.0,   "max": 1170.0, "count": 97,  "unit": "m3",  "ar": "حجم الحفر"},
    "BACKFILL":         {"avg": 301.66, "min": 2.02,  "max": 1120.0, "count": 54,  "unit": "m3",  "ar": "حجم الردم"},
    "ANTI_TERMITE":     {"avg": 129.10, "min": 1.0,   "max": 723.96, "count": 46,  "unit": "m2",  "ar": "مبيد النمل الأبيض"},
    "DAMP_PROOF":       {"avg": 98.48,  "min": 1.0,   "max": 349.0,  "count": 31,  "unit": "m2",  "ar": "نايلون أسود"},
    "LEAN_CONCRETE":    {"avg": 4.41,   "min": 4.01,  "max": 5.07,   "count": 10,  "unit": "m3",  "ar": "PCC للمؤسسات"},
    "RAFT_CONCRETE":    {"avg": 37.58,  "min": 2.5,   "max": 254.0,  "count": 34,  "unit": "m3",  "ar": "خرسانة الحصيرة"},
    "FOOTING_CONCRETE": {"avg": 17.98,  "min": 3.01,  "max": 59.94,  "count": 113, "unit": "m3",  "ar": "إجمالي حجم الأسس"},
    "TIE_BEAM":         {"avg": 8.30,   "min": 3.06,  "max": 30.0,   "count": 101, "unit": "m3",  "ar": "إجمالي حجم كمرات الربط"},
    "COLUMN_CONCRETE":  {"avg": 7.88,   "min": 3.0,   "max": 28.35,  "count": 127, "unit": "m3",  "ar": "إجمالي حجم الأعمدة"},
    "BEAM_CONCRETE":    {"avg": 3.78,   "min": 1.0,   "max": 5.82,   "count": 101, "unit": "m3",  "ar": "إجمالي حجم الكمرات"},
    "SLAB_CONCRETE":    {"avg": 10.95,  "min": 2.2,   "max": 57.0,   "count": 129, "unit": "m3",  "ar": "حجم البلاطات"},
    "STAIR_CONCRETE":   {"avg": 10.64,  "min": 1.0,   "max": 40.0,   "count": 296, "unit": "m3",  "ar": "خرسانة الدرج"},
    "STEEL":            {"avg": 4.85,   "min": 4.08,  "max": 5.06,   "count": 17,  "unit": "ton", "ar": "حديد التسليح"},
    "BLOCK_WORK":       {"avg": 71.99,  "min": 4.0,   "max": 360.0,  "count": 110, "unit": "m2",  "ar": "بلوك (إجمالي)"},
    "PLASTER":          {"avg": 274.46, "min": 1.0,   "max": 1235.1, "count": 233, "unit": "m2",  "ar": "لياسة داخلية"},
    "WATERPROOF_FLOOR": {"avg": 31.41,  "min": 0.5,   "max": 300.0,  "count": 189, "unit": "m2",  "ar": "عزل مائي"},
    "FLOOR_TILES":      {"avg": 58.46,  "min": 7.47,  "max": 202.1,  "count": 82,  "unit": "m2",  "ar": "بلاط المناطق الجافة"},
    "WALL_TILES":       {"avg": 127.06, "min": 0.65,  "max": 340.0,  "count": 39,  "unit": "m2",  "ar": "بلاط الجدران"},
    "CEILING_AREA":     {"avg": 34.0,   "min": 1.0,   "max": 226.0,  "count": 262, "unit": "m2",  "ar": "سقف المناطق الجافة"},
    "PAINTING":         {"avg": 198.51, "min": 2.0,   "max": 995.0,  "count": 220, "unit": "m2",  "ar": "دهان"},
    "DOORS":            {"avg": 5.94,   "min": 1.0,   "max": 15.07,  "count": 92,  "unit": "no.", "ar": "أبواب"},
    "ROAD_BASE":        {"avg": 44.09,  "min": 2.04,  "max": 167.25, "count": 51,  "unit": "m2",  "ar": "رصيف الطرق"},
}

# إجمالي العينة: 289 مشروع صالح | G=15 | G+1=208 | G+2=24 | UNKNOWN villa-like=42
_إحصاء_مجموعة_التدريب = {
    "total_analyzed":      289,
    "ground_only":         15,
    "g_plus_1":            208,
    "g_plus_1_service":    0,
    "g_plus_2":            24,
    "unknown_villa_like":  42,
    "data_source":         "UAE villa-like BOQ Excel files with PDF companions, 2020-2025",
}


def _مراجعة_gemini_المتكاملة(نتائج: dict, بيانات_فضاء: dict,
                              مسار_المجلد: str, مفتاح_api: str) -> dict:
    """
    Gemini 2.5 Flash Supervisor — يراجع بكل المعرفة:
    ● Bible Rules: القواعد المطلقة
    ● Averages: المتوسطات المرجعية المقيّسة لهذا المشروع
    ● Training: أنماط 289 مشروع سابق
    ● Equations: معادلات حساب الكميات الإماراتية

    يقرأ كل صفحات PDF → يستخرج بيانات إنشائية حقيقية → يصحح
    """
    if not مفتاح_api or not مسار_المجلد:
        return نتائج

    # ── 1. جمع صور PDF ──────────────────────────────────────────
    صور_base64 = []
    ملفات_pdf = []
    for جذر, _, ملفات in os.walk(مسار_المجلد):
        for ملف in ملفات:
            if ملف.lower().endswith(".pdf"):
                ملفات_pdf.append(os.path.join(جذر, ملف))

    if not ملفات_pdf:
        مسجل.warning("مشرف Gemini المتكامل: لا توجد ملفات PDF")
        return نتائج

    for مسار_pdf in ملفات_pdf:
        try:
            وثيقة = fitz.open(مسار_pdf)
            for i, صفحة in enumerate(وثيقة):
                if i >= 10:  # حد أقصى 10 صفحات لكل PDF
                    break
                مصفوفة = fitz.Matrix(2.0, 2.0)  # 144 DPI
                صورة = صفحة.get_pixmap(matrix=مصفوفة)
                بايتات = صورة.tobytes("png")
                صور_base64.append(base64.b64encode(بايتات).decode("utf-8"))
            وثيقة.close()
        except Exception as خطأ:
            مسجل.warning(f"مشرف Gemini المتكامل: فشل قراءة {مسار_pdf}: {خطأ}")

    if not صور_base64:
        مسجل.warning("مشرف Gemini المتكامل: لم يتم استخراج أي صور")
        return نتائج

    if len(صور_base64) > 25:
        صور_base64 = صور_base64[:25]

    # ── 2. بناء جدول المتوسطات المقيّسة ──────────────────────────
    import math as _math
    نوع = _اكتشف_نوع_المشروع(نتائج)
    متوسطات = _متوسطات_الكميات.get(نوع, {})
    مساحة_المشروع = float(بيانات_فضاء.get("المساحة_الكلية", 0) or 0)

    # ── حماية نوع المشروع (نفس القاعدة) ──
    if نوع == "G" and (مساحة_المشروع < 50 or مساحة_المشروع > 400):
        نوع = "G+1"
        متوسطات = _متوسطات_الكميات.get(نوع, {})

    مساحة_مرجعية = _مساحة_مرجعية_للأدوار.get(نوع, 650.0)
    if مساحة_المشروع >= 50:
        معامل_خطي = max(0.25, min(4.0, مساحة_المشروع / مساحة_مرجعية))
        معامل_محيط = max(0.30, min(2.5, _math.sqrt(مساحة_المشروع / مساحة_مرجعية)))
    else:
        معامل_خطي = معامل_محيط = 1.0

    _بنود_محيط_ref = {
        "بلوك 20 سم خارجي", "بلوك 20 سم تصوينة السطح",
        "بلوك 20 سم داخلي", "بلوك 10 سم داخلي",
        "تشطيب الواجهة الخارجية", "سكرتة",
        "مساحة بلوك تحت الأرض", "عتبات رخام",
    }

    سطور_avg = ""
    for اسم, قيمة in متوسطات.items():
        if اسم in ("نايلون أسود",):
            continue  # skip aliases
        معامل = معامل_محيط if اسم in _بنود_محيط_ref else معامل_خطي
        مقيّس = round(قيمة * معامل, 1)
        سطور_avg += f"  - {اسم}: {مقيّس}\n"

    # ── بناء جدول نطاقات التدريب (min–max من 289 مشروع حقيقي) ───────────
    سطور_تدريب = ""
    for _كود, _d in _نطاقات_تدريب_289.items():
        سطور_تدريب += (
            f"  - {_d['ar']} ({_d['unit']}): "
            f"avg={_d['avg']:.1f}  |  range [{_d['min']:.1f} – {_d['max']:.1f}]"
            f"  (n={_d['count']} projects)\n"
        )

    # إحصاء corpus التدريب لهذا النوع
    _نوع_للعدد = {"G": "ground_only", "G+1": "g_plus_1", "G+2": "g_plus_2"}.get(نوع, "g_plus_1")
    _عدد_الكوربوس = _إحصاء_مجموعة_التدريب.get(_نوع_للعدد, 33)

    # ── 3. ملخص engine output ──────────────────────────────────
    ملخص_engine = {}
    for قسم, بنود in نتائج.items():
        for بند in بنود:
            اسم = بند.get("البند", "?")
            كمية = بند.get("الكمية")
            ملخص_engine[اسم] = {
                "qty": كمية,
                "unit": بند.get("الوحدة", ""),
                "status": بند.get("الحالة", ""),
            }

    ملخص_نص = json.dumps(ملخص_engine, ensure_ascii=False, indent=2)

    معلومات = (
        f"المساحة الكلية: {مساحة_المشروع} م²\n"
        f"المحيط الخارجي: {بيانات_فضاء.get('المحيط_الخارجي', 0)} م\n"
        f"عدد الغرف: {len(بيانات_فضاء.get('الغرف', []))}\n"
        f"نوع المشروع: {نوع}\n"
        f"درجة ثقة الإنشاء: {بيانات_فضاء.get('درجة_الثقة', {}).get('النتيجة', 0)}%\n"
    )

    # ── 4. بناء الطلب بكل المعرفة ────────────────────────────────
    prompt = f"""You are an EXPERT UAE Quantity Surveyor (QS) with 20+ years of villa quantification.
You are SUPERVISOR of a QTO engine. READ every drawing page → VERIFY quantities → CORRECT errors.

## BIBLE RULES (ABSOLUTE — never violate):
1. CEILING area = FLOORING area (dry ceiling = dry floor area, wet ceiling = wet floor area)
2. Balcony waterproofing is ALWAYS a SEPARATE item from balcony flooring
3. L.S (Lump Sum) items: qty = 1, NEVER skip
4. Road Base: area = excavation area × thickness, DEDUCTED from backfill
5. Ground-floor-only projects: NO 2nd floor items (no upper slab, no upper block, no upper plaster/tiles)

## STRUCTURAL EQUATIONS (extract data from structural drawings and calculate):
- Foundation Volume m³ = Σ(L × W × D × qty) for each footing type in foundation schedule
- PCC m³ = Σ((L+0.2) × (W+0.2) × 0.10 × qty)
- Neck Columns m³ = Σ(L × W × neck_height × qty) — neck_height from section drawing
- Tie Beams m³ = Σ(total_length × width × depth) — widths/depths from beam schedule
- Strap Beams m³ = Σ(total_length × width × depth)
- Columns m³ = Σ(L × W × clear_height × total_count_all_floors) — from column schedule
- Beams m³ = Σ(span × width × (depth − slab_thickness))
- Slab m³ = net_slab_area × thickness — from slab layout
- Staircase m³ = estimate from waist slab + steps in staircase detail
- Excavation m³ = excavation_area × excavation_depth
- Backfill m³ = excavation_volume − (foundations + PCC + tie_beams + neck_columns + slab_on_grade)

## ARCHITECTURAL EQUATIONS (verify from floor plans):
- External Block m² = external_perimeter × Σ(floor_heights) − (window_areas + main_door_area)
- Internal Block 20cm m² = internal_wall_lengths_20cm × floor_height − door_areas
- Internal Block 10cm m² = internal_wall_lengths_10cm × floor_height − door_areas
- Parapet m² = roof_perimeter × parapet_height (typically 1.0–1.2m, measured in m²)
- Internal Plaster m² = (all_wall_lengths × 2 + ext_perimeter) × floor_height − openings
- External Finish m² = ext_block_area + parapet_area (both sides)
- Waterproofing m² = sum of wet room areas + balcony areas
- Combo Roof m² = roof_slab_area + parapet_upstand (perimeter × parapet_height)

## FINISHING EQUATIONS (verify room areas from floor plans):
- Dry Flooring m² = sum of dry room floor areas (bedrooms, living, corridor, kitchen, etc.)
- Wet Flooring m² = sum of wet room floor areas (bathrooms, laundry, WC)
- Dry Ceiling m² = EXACTLY SAME as Dry Flooring (BIBLE RULE #1)
- Wet Ceiling m² = EXACTLY SAME as Wet Flooring (BIBLE RULE #1)
- Paint m² = dry_room_wall_areas + wet_room_walls_above_tile_height − openings
- Skirting LM = dry room perimeters − door widths (ground floor only)
- Wall Tiles m² = wet_room_perimeters × (floor_height − 0.3m)
- Balcony Tiles m² = balcony floor area
- Marble Threshold LM = number_of_doors × threshold_length_per_door

## SCALED REFERENCE AVERAGES for {نوع} projects (area {مساحة_المشروع:.0f} م²):
From {_إحصاء_مجموعة_التدريب['total_analyzed']}-project UAE villa training corpus ({_عدد_الكوربوس} {نوع} projects):
{سطور_avg}
## REAL TRAINING DATA RANGES (from {_إحصاء_مجموعة_التدريب['total_analyzed']} analyzed UAE villa BOQ files, 2020–2025):
These are ACTUAL extracted quantities — use as sanity-check bounds on your corrections:
{سطور_تدريب}
If your corrected value falls OUTSIDE these real-project ranges, double-check your reading from the drawings.

## TRAINING CORPUS SUMMARY:
- Total villa projects analyzed: {_إحصاء_مجموعة_التدريب['total_analyzed']}
- Ground-only (G): {_إحصاء_مجموعة_التدريب['ground_only']} projects in full corpus
- G+1: {_إحصاء_مجموعة_التدريب['g_plus_1']} | G+1+Service: {_إحصاء_مجموعة_التدريب['g_plus_1_service']} | G+2: {_إحصاء_مجموعة_التدريب['g_plus_2']}
- Unknown-but-villa-like folders with BOQ+PDF: {_إحصاء_مجموعة_التدريب['unknown_villa_like']}
- Source: {_إحصاء_مجموعة_التدريب['data_source']}

## PROJECT INFO:
{معلومات}

## ENGINE OUTPUT (review and correct):
{ملخص_نص}

## YOUR TASK:
1. Look at EVERY drawing page — structural AND architectural
2. For STRUCTURAL items: Find foundation schedule (table with F1, F2...), column schedule (C1, C2...), beam schedule → calculate REAL volumes using equations above
3. For ARCHITECTURAL items: Verify wall lengths, perimeters, room areas from floor plans
4. For FINISHING items: Verify room areas match floor plan measurements
5. Compare EVERY item with scaled reference averages AND real training ranges — correct if > 2× or < 0.3× expected
6. ENFORCE Bible Rule #1: Dry Ceiling MUST = Dry Flooring, Wet Ceiling MUST = Wet Flooring
7. If you find data in drawings that engine missed → provide the CORRECTED quantity
8. If you CANNOT determine a quantity → DO NOT include it (engine value stays)

## OUTPUT FORMAT:
Return ONLY valid JSON. No markdown, no backticks, no explanation text.
Only include items you want to CORRECT. Keep the exact Arabic item names.
Example: {{"إجمالي حجم الأسس": 42.5, "بلوك 20 سم خارجي": 380.0, "سقف المناطق الجافة": 280.0}}
"""

    أجزاء = []
    for صورة in صور_base64:
        أجزاء.append({
            "inline_data": {"mime_type": "image/png", "data": صورة}
        })
    أجزاء.append({"text": prompt})

    الطلب_json = {
        "contents": [{"parts": أجزاء}],
        "generationConfig": {
            "temperature": 0.1,
            "maxOutputTokens": 8192,
        }
    }

    # ── 5. إرسال لـ Gemini 2.5 Flash ─────────────────────────────
    رابط = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={مفتاح_api}"

    try:
        مسجل.info(f"مشرف Gemini المتكامل: إرسال {len(صور_base64)} صورة + Bible + Averages + Training + Equations...")
        استجابة = requests.post(رابط, json=الطلب_json, timeout=180)
        استجابة.raise_for_status()
        نص_رد = استجابة.json()

        محتوى = نص_رد.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "")

        if not محتوى:
            مسجل.warning("مشرف Gemini المتكامل: استجابة فارغة")
            return نتائج

        # تنظيف markdown
        محتوى = re.sub(r'```json\s*|```\s*', '', محتوى).strip()

        بداية = محتوى.find('{')
        نهاية = محتوى.rfind('}')
        if بداية == -1 or نهاية == -1:
            مسجل.warning(f"مشرف Gemini المتكامل: لا يوجد JSON في الاستجابة")
            return نتائج

        تصحيحات = json.loads(محتوى[بداية:نهاية + 1])
        مسجل.info(f"مشرف Gemini المتكامل: {len(تصحيحات)} تصحيح مستلم")

    except requests.exceptions.Timeout:
        مسجل.warning("مشرف Gemini المتكامل: timeout بعد 180 ثانية")
        return نتائج
    except json.JSONDecodeError as خطأ:
        مسجل.warning(f"مشرف Gemini المتكامل: فشل تحليل JSON: {خطأ}")
        return نتائج
    except Exception as خطأ:
        مسجل.warning(f"مشرف Gemini المتكامل: خطأ: {خطأ}")
        return نتائج

    # ── 6. تطبيق التصحيحات مع حماية ──────────────────────────────
    عداد = 0
    for قسم, بنود in نتائج.items():
        for بند in بنود:
            اسم = بند.get("البند", "")
            if اسم in تصحيحات:
                قيمة_جديدة = تصحيحات[اسم]
                try:
                    قيمة_رقمية = round(float(قيمة_جديدة), 1)
                    if قيمة_رقمية > 0:
                        بند["الكمية"] = قيمة_رقمية
                        بند["الحالة"] = "GEMINI_VERIFIED"
                        عداد += 1
                except (TypeError, ValueError):
                    pass

    # ── 7. إعادة فرض Bible بعد Gemini (ceiling=flooring) ─────────
    _فرض_قواعد_الكتاب(نتائج, بيانات_فضاء)

    مسجل.info(f"مشرف Gemini المتكامل: تم تصحيح {عداد} بند وأُعيد فرض Bible Rules")
    return نتائج


# ─────────────────────────────────────────────────────────────────────────────
# 4/4 — النظام المتكامل (Unified System)
# V15 Engine + Bible Rules + Smart Averages + Gemini Supervisor = ONE SYSTEM
# ─────────────────────────────────────────────────────────────────────────────

def _نظام_متكامل(نتائج: dict, بيانات_فضاء: dict,
                  مسار_المجلد: str = "", مفتاح_api: str = "") -> dict:
    """
    النظام المتكامل: كل المكونات تعمل كنظام واحد.

    التسلسل:
    ┌──────────────────────────────────────────────────────────┐
    │ 1. Bible Rules: فرض القواعد المطلقة                      │
    │ 2. Smart Averages: شبكة أمان + fallback للبنود الناقصة   │
    │ 3. Gemini Supervisor: مراجعة بالرسومات + كل المعرفة      │
    │ 4. Bible Rules مرة أخرى: ضمان عدم مخالفة Gemini للقواعد │
    └──────────────────────────────────────────────────────────┘

    كل مكون يبني على نتائج المكون السابق.
    الكل يعمل معاً — ليس engine ثم fix.
    """
    # === الخطوة 1: فرض قواعد الكتاب ===
    _فرض_قواعد_الكتاب(نتائج, بيانات_فضاء)

    # === الخطوة 2: المتوسطات كشبكة أمان ===
    _تصحيح_بالمتوسطات_الذكي(نتائج, بيانات_فضاء)

    # === الخطوة 3: مراجعة Gemini بكل المعرفة ===
    if مفتاح_api and مسار_المجلد:
        _مراجعة_gemini_المتكاملة(نتائج, بيانات_فضاء, مسار_المجلد, مفتاح_api)

    # === الخطوة 4: إعادة فرض Bible (Gemini قد يغير ceiling ≠ flooring) ===
    _فرض_قواعد_الكتاب(نتائج, بيانات_فضاء)

    return نتائج


# ─────────────────────────────────────────────────────────────────────────────
# 1. طبقة التحقق من المنطق
# ─────────────────────────────────────────────────────────────────────────────

class مُدقِّق_المنطق:
    """
    يتحقق من منطقية نتائج الحساب ويُصدر تحذيرات واضحة للمهندس.
    يصنف التحذيرات: خطأ 🔴 / تحذير 🟡 / ملاحظة 🔵
    """

    def __init__(self, نتائج: Dict, فضاء: Dict, ثوابت: ثوابت_المشروع):
        self.نتائج  = نتائج
        self.فضاء   = فضاء
        self.ثوابت  = ثوابت
        self.تنبيهات: List[Dict] = []

    def _تنبيه(self, مستوى: str, البند: str, الرسالة: str, القيمة=None):
        رمز = {"خطأ": "🔴", "تحذير": "🟡", "ملاحظة": "🔵"}.get(مستوى, "⚪")
        self.تنبيهات.append({
            "المستوى": مستوى,
            "الرمز":   رمز,
            "البند":   البند,
            "الرسالة": الرسالة,
            "القيمة":  القيمة,
        })

    def _قيمة_البند(self, اسم: str) -> Optional[float]:
        كل_البنود = [ب for ق in self.نتائج.values() for ب in ق]
        for ب in كل_البنود:
            if ب["البند"] == اسم:
                return ب["الكمية"]
        return None

    def تحقق(self) -> List[Dict]:
        م_كلية  = self.فضاء.get("المساحة_الكلية", 0)
        م_حفر   = self.فضاء.get("مساحة_الحفر", 0)
        ث = self.ثوابت

        # ── تحقق من الحفر ─────────────────────────────────────────────────────
        if م_حفر > م_كلية * 3:
            self._تنبيه("خطأ", "مساحة الحفر",
                f"مساحة الحفر ({م_حفر:.1f}م²) أكبر بكثير من مساحة الفيلا ({م_كلية:.1f}م²) — تحقق من الرسم",
                م_حفر)

        if م_حفر < م_كلية * 0.8:
            self._تنبيه("تحذير", "مساحة الحفر",
                f"مساحة الحفر ({م_حفر:.1f}م²) أصغر من مساحة الفيلا ({م_كلية:.1f}م²) — قد يكون الكشف ناقصاً",
                م_حفر)

        # ── تحقق من الردم ─────────────────────────────────────────────────────
        ردم = self._قيمة_البند("حجم الردم")
        حفر_ح = self._قيمة_البند("حجم الحفر")
        if ردم is not None and حفر_ح is not None:
            if ردم == 0 and حفر_ح > 0:
                self._تنبيه("تحذير", "حجم الردم",
                    "الردم = صفر — قد يعني إن حجم العناصر الإنشائية أكبر من حجم الحفر، تحقق من الأبعاد")
            نسبة_الردم = ردم / حفر_ح if حفر_ح > 0 else 0
            if نسبة_الردم > 0.85:
                self._تنبيه("ملاحظة", "حجم الردم",
                    f"الردم {نسبة_الردم*100:.0f}% من الحفر — نسبة مرتفعة، تحقق من أبعاد الأسس",
                    ردم)

        # ── تحقق من المحيط الخارجي ───────────────────────────────────────────
        محيط = self.فضاء.get("المحيط_الخارجي", 0)
        if م_كلية > 0:
            # دائرة مثالية: محيط = 2π√(مساحة/π) = 2√(πA)
            محيط_أدنى = 4 * math.sqrt(م_كلية)  # مربع
            if محيط < محيط_أدنى * 0.5:
                self._تنبيه("خطأ", "المحيط الخارجي",
                    f"المحيط ({محيط:.1f}م) صغير جداً لمساحة {م_كلية:.1f}م² — تحقق من اكتشاف الحدود",
                    محيط)
            if محيط > محيط_أدنى * 5:
                self._تنبيه("تحذير", "المحيط الخارجي",
                    f"المحيط ({محيط:.1f}م) كبير جداً — قد يشمل إطار الورقة أو جداول",
                    محيط)

        # ── تحقق من الأعمدة ──────────────────────────────────────────────────
        ح_أع = self._قيمة_البند("إجمالي حجم الأعمدة")
        if ح_أع is not None and م_كلية > 0:
            ارتفاع_مرجعي_للتحقق = max(
                float(getattr(ث, "ارتفاع_الدور_الأرضي", 0.0) or 0.0),
                float(getattr(ث, "ارتفاع_الدور_الأول", 0.0) or 0.0),
                float(getattr(ث, "ارتفاع_الدور_الثاني", 0.0) or 0.0),
                float(getattr(ث, "ارتفاع_دور_السطح", 0.0) or 0.0),
                float(getattr(ث, "ارتفاع_مبنى_الخدمة", 0.0) or 0.0),
                float(getattr(ث, "ارتفاع_الدور", 0.0) or 0.0),
            )
            if ارتفاع_مرجعي_للتحقق > 0 and ح_أع > م_كلية * ارتفاع_مرجعي_للتحقق * 0.10:
                self._تنبيه("تحذير", "حجم الأعمدة",
                    f"حجم الأعمدة ({ح_أع:.2f}م³) يبدو مرتفعاً — تحقق من عدد الأعمدة أو أبعادها",
                    ح_أع)

        # ── تحقق من البلاطة ──────────────────────────────────────────────────
        ح_بلاطة = self._قيمة_البند("حجم بلاطة الدور الأول")
        if ح_بلاطة is not None and م_كلية > 0 and ث.سماكة_البلاطة > 0:
            سماكة_محسوبة = ح_بلاطة / م_كلية if م_كلية > 0 else 0
            if abs(سماكة_محسوبة - ث.سماكة_البلاطة) > 0.05:
                self._تنبيه("ملاحظة", "سماكة البلاطة",
                    f"السماكة المحسوبة {سماكة_محسوبة*100:.0f}سم تختلف عن المدخلة {ث.سماكة_البلاطة*100:.0f}سم")

        # ── تحقق من الكمرات ──────────────────────────────────────────────────
        ح_كمرات = self._قيمة_البند("إجمالي حجم كمرات الربط")
        if ح_كمرات == 0:
            self._تنبيه("تحذير", "كمرات الربط",
                "لم يُعثر على كمرات ربط — تحقق من أسماء الطبقات أو أدخلها يدوياً")

        # ── تحقق من الفتحات ──────────────────────────────────────────────────
        عدد_الفتحات = len(self.نتائج.get("الفتحات", []))
        if عدد_الفتحات == 0:
            self._تنبيه("تحذير", "الفتحات",
                "لم يُعثر على فتحات (أبواب/نوافذ) — تحقق من الجدول أو أدخلها يدوياً")
        elif all((ب.get("الطول", 0) or 0) == 0 and (ب.get("الارتفاع", 0) or 0) == 0 for ب in self.نتائج.get("الفتحات", [])):
            self._تنبيه("ملاحظة", "الفتحات",
                "تم استخراج عدد الفتحات فقط بدون أبعادها — بنود الخصم المساحي والعتبات تظل جزئية")

        # ── تحقق من كمرات الستراب ─────────────────────────────────────────────
        ح_ست = self._قيمة_البند("إجمالي حجم كمرات الستراب")
        if False and ح_ست == 0:
            self._تنبيه("ملاحظة", "كمرات الستراب",
                "لم تُستخرج كمرات ستراب من الرسم في هذا التشغيل — قد تكون غير مسماة أو غير مرتبطة هندسياً")

        # ── تحقق من المناطق المبللة ───────────────────────────────────────────
        م_جاف = self._قيمة_البند("بلاط المناطق الجافة")
        if م_جاف is not None and م_جاف == 0:
            self._تنبيه("تحذير", "المناطق الجافة",
                "مساحة المناطق الجافة = صفر — قد يعني إن كل الغرف صُنِّفت مناطق مبللة")

        # ── تحقق من ثوابت المشروع ────────────────────────────────────────────
        for اسم_ارتفاع, قيمة_ارتفاع in [
            ("ارتفاع الدور الأرضي", float(getattr(ث, "ارتفاع_الدور_الأرضي", 0.0) or 0.0)),
            ("ارتفاع الدور الأول", float(getattr(ث, "ارتفاع_الدور_الأول", 0.0) or 0.0)),
            ("ارتفاع دور السطح", float(getattr(ث, "ارتفاع_دور_السطح", 0.0) or 0.0)),
            ("ارتفاع مبنى الخدمة", float(getattr(ث, "ارتفاع_مبنى_الخدمة", 0.0) or 0.0)),
            ("ارتفاع الدور العام", float(getattr(ث, "ارتفاع_الدور", 0.0) or 0.0)),
        ]:
            if قيمة_ارتفاع > 0 and (قيمة_ارتفاع < 2.5 or قيمة_ارتفاع > 6.0):
                self._تنبيه("تحذير", اسم_ارتفاع,
                    f"{اسم_ارتفاع} {قيمة_ارتفاع}م خارج النطاق المعتاد (2.5م–6م)",
                    قيمة_ارتفاع)

        if ث.عمق_الحفر < 0.5 or ث.عمق_الحفر > 4.0:
            self._تنبيه("تحذير", "عمق الحفر",
                f"عمق الحفر {ث.عمق_الحفر}م خارج النطاق المعتاد (0.5م–4م)",
                ث.عمق_الحفر)

        # ── ملخص ─────────────────────────────────────────────────────────────
        عدد_الأخطاء   = sum(1 for ت in self.تنبيهات if ت["المستوى"] == "خطأ")
        عدد_التحذيرات = sum(1 for ت in self.تنبيهات if ت["المستوى"] == "تحذير")

        if عدد_الأخطاء == 0 and عدد_التحذيرات == 0:
            self._تنبيه("ملاحظة", "التحقق العام", "✅ جميع القيم ضمن النطاق المنطقي المتوقع")

        return self.تنبيهات


# ─────────────────────────────────────────────────────────────────────────────
# 2. عدّاد الفتحات من المسقط
# ─────────────────────────────────────────────────────────────────────────────

class عدّاد_الفتحات:
    """
    يعدّ كمية كل فتحة (D1، W1، MD1...) من المسقط تلقائياً.
    يبحث عن نصوص تطابق الرمز في المسقط ويحسب تكرارها.
    """

    def __init__(self, msp, مقياس: float):
        self.msp    = msp
        self.مقياس  = مقياس
        self.العداد: Dict[str, int] = defaultdict(int)
        self.العداد_حسب_المستوى: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        self.صناديق_المخططات: List[Dict] = []

    @staticmethod
    def _نص_موحد(نص: str) -> str:
        قيمة = re.sub(r'%%[A-Z]', '', str(نص).upper())
        return re.sub(r'[\s\-_]+', '', قيمة)

    @staticmethod
    def _داخل_صندوق(x: float, y: float, صندوق: Dict) -> bool:
        return صندوق["xmin"] <= x <= صندوق["xmax"] and صندوق["ymin"] <= y <= صندوق["ymax"]

    def _اكتشف_صناديق_المخططات(self) -> List[Dict]:
        صناديق: List[Dict] = []
        for نص in self.msp.query('TEXT MTEXT'):
            try:
                خام = (نص.dxf.text if نص.dxftype() == 'TEXT' else نص.plain_text()).strip().upper()
                x = نص.dxf.insert.x * self.مقياس
                y = نص.dxf.insert.y * self.مقياس
                if "GROUND FLOOR PLAN" in خام:
                    مستوى = "GF"
                elif "FIRST FLOOR PLAN" in خام:
                    مستوى = "1F"
                elif "SECOND FLOOR PLAN" in خام or "2ND FLOOR PLAN" in خام:
                    مستوى = "2F"
                elif "ROOF PLAN" in خام and "TOP" not in خام:
                    مستوى = "RF"
                else:
                    continue
                صندوق = {
                    "المستوى": مستوى,
                    "xmin": x - 32.0,
                    "xmax": x + 32.0,
                    "ymin": y - 2.0,
                    "ymax": y + 34.0,
                }
                if any(
                    abs(((ق["xmin"] + ق["xmax"]) / 2.0) - x) <= 5.0 and
                    abs(((ق["ymin"] + ق["ymax"]) / 2.0) - (y + 16.0)) <= 5.0
                    for ق in صناديق
                ):
                    continue
                صناديق.append(صندوق)
            except Exception:
                pass
        return صناديق

    # طبقات تحوي رموز الفتحات — موسّعة لتغطية تسميات CAD المختلفة
    _طبقات_فتحة = frozenset([
        "FURN", "DOOR", "GLAZ", "WINDOW", "TEXT",
        "A-DOOR", "A-GLAZ", "A-WIND", "A-ANNO", "ANNO",
        "DIMENSION", "DIM", "LABEL", "LABELS", "TAGS",
        "OPENING", "WIN", "DR", "ARCH", "SCHEDULE",
        "0",  # طبقة 0 الافتراضية شائعة في الرسومات القديمة
    ])

    def عدّ(self) -> Dict[str, int]:
        """يبحث عن رموز الفتحات في المسقط ويحسب عددها."""
        # نمط موسّع: D, W, FD (fire door), BD, WD, WIN, DR, MD
        نمط_الفتحة = re.compile(
            r'^(F?D|W|MD|WD|WIN|DR|BD|FD)\s*\d+[A-Z]{0,2}$',
            re.IGNORECASE
        )
        self.صناديق_المخططات = self._اكتشف_صناديق_المخططات()
        مستويات_معتمدة = {"GF", "1F", "2F", "RF", "KB"}
        # عند غياب بكسات المخططات نقبل جميع النصوص
        بدون_صناديق = not self.صناديق_المخططات
        for نص in self.msp.query('TEXT MTEXT'):
            try:
                قيمة = (نص.dxf.text if نص.dxftype()=='TEXT'
                        else نص.plain_text()).strip().upper()
                قيمة = self._نص_موحد(قيمة)
                x = نص.dxf.insert.x * self.مقياس
                y = نص.dxf.insert.y * self.مقياس
                طبقة = نص.dxf.layer.upper()
                صناديق_مطابقة = []
                if not بدون_صناديق:
                    صناديق_مطابقة = [
                        ص for ص in self.صناديق_المخططات
                        if self._داخل_صندوق(x, y, ص) and ص["المستوى"] in مستويات_معتمدة
                    ]
                    if not صناديق_مطابقة:
                        continue
                # قبول أي طبقة تحتوي إحدى الكلمات المفتاحية الموسّعة
                # أو عند غياب الصناديق نقبل بشرط مطابقة النمط فقط
                طبقة_معتمدة = any(ك in طبقة for ك in self._طبقات_فتحة)
                if not طبقة_معتمدة and not بدون_صناديق:
                    continue
                if نمط_الفتحة.match(قيمة):
                    رمز = رمز_فتحة_موحد(قيمة)
                    self.العداد[رمز] += 1
                    if صناديق_مطابقة:
                        مستوى = صناديق_مطابقة[0]["المستوى"]
                        self.العداد_حسب_المستوى[مستوى][رمز] += 1
            except Exception:
                pass

        مسجل.info(f"عدّاد الفتحات: {dict(self.العداد)}")
        return dict(self.العداد)

    def تحديث_الفتحات(self, الفتحات: List[بيانات_فتحة]) -> List[بيانات_فتحة]:
        """يحدّث كمية كل فتحة من المسقط فقط، ويُبقي الأبعاد من الجدول."""
        if not self.العداد:
            return [
                بيانات_فتحة(
                    الرمز=ف.الرمز,
                    الطول=ف.الطول,
                    الارتفاع=ف.الارتفاع,
                    الكمية=0,
                    مصدر_الكمية="PLANS",
                    مصدر_الأبعاد=(ف.مصدر_الأبعاد or ("DETAIL" if (ف.الطول or ف.الارتفاع) else "")),
                    ملف_الأبعاد=getattr(ف, "ملف_الأبعاد", ""),
                )
                for ف in الفتحات
            ]
        محدّثة = []
        for ف in الفتحات:
            رمز = رمز_فتحة_موحد(ف.الرمز)
            كمية_مسقط = int(self.العداد.get(رمز, 0) or 0)
            محدّثة.append(بيانات_فتحة(
                الرمز=ف.الرمز,
                الطول=ف.الطول,
                الارتفاع=ف.الارتفاع,
                الكمية=كمية_مسقط,
                مصدر_الكمية="PLANS",
                مصدر_الأبعاد=(ف.مصدر_الأبعاد or ("DETAIL" if (ف.الطول or ف.الارتفاع) else "")),
                ملف_الأبعاد=getattr(ف, "ملف_الأبعاد", ""),
            ))
            مسجل.info(f"فتحة {ف.الرمز}: كمية المسقط = {كمية_مسقط}")
        return محدّثة

    def إنشاء_فتحات_من_العد(self) -> List[بيانات_فتحة]:
        """يبني قائمة فتحات من العدّ فقط عند غياب جدول الفتحات."""
        نتائج: List[بيانات_فتحة] = []
        for رمز in sorted(self.العداد):
            كمية = int(self.العداد.get(رمز, 0) or 0)
            if كمية <= 0:
                continue
            نتائج.append(بيانات_فتحة(
                الرمز=رمز,
                الطول=0.0,
                الارتفاع=0.0,
                الكمية=كمية,
                مصدر_الكمية="PLANS",
            ))
        return نتائج


# ─────────────────────────────────────────────────────────────────────────────
# 3. معالج المجلد (مخططات متعددة)
# ─────────────────────────────────────────────────────────────────────────────

class معالج_المجلد:
    """
    يقبل مجلداً يحتوي مخططات متعددة ويصنفها:
    - مخطط إنشائي (Structural): يحتوي FOUNDATION أو BEAM أو COLUMN
    - مخطط معماري (Architectural): يحتوي A-WALL أو DOOR أو ROOM

    يدمج النتائج من كل الملفات في بيانات موحدة.
    """

    # كلمات تدل على المخطط الإنشائي
    كلمات_إنشائية = ["FOUNDATION", "FOOTING", "BEAM", "COLUMN", "STRUCTURAL",
                     "STRUCT", "STR", "S-PLAN", "BASEMENT", "GROUND FL STR"]

    # كلمات تدل على المخطط المعماري
    كلمات_معمارية = ["ARCHITECTURAL", "ARCH", "A-WALL", "FLOOR PLAN",
                     "GF PLAN", "F1 PLAN", "FIRST FLOOR", "GROUND FLOOR"]

    def __init__(self, مسار_المجلد: str, وحدة: str, مفتاح_gemini: str = ""):
        self.مسار_المجلد = مسار_المجلد
        self.وحدة        = وحدة
        self.مفتاح_gemini = مفتاح_gemini
        self.الملفات_الإنشائية: List[str] = []
        self.الملفات_المعمارية: List[str] = []
        self.ملفات_غير_مصنفة:  List[str] = []

    @staticmethod
    def _ادمج_فتحات_الحزمة(
        فتحات_الأبعاد: Dict[str, Dict],
        عدادات_الخطط: Dict[str, int]
    ) -> List[Dict]:
        رموز = sorted(set(فتحات_الأبعاد) | set(عدادات_الخطط))
        نتائج: List[Dict] = []
        for رمز in رموز:
            بيانات = فتحات_الأبعاد.get(رمز, {})
            نتائج.append({
                "الرمز": بيانات.get("الرمز", رمز),
                "الطول": float(بيانات.get("الطول", 0.0) or 0.0),
                "الارتفاع": float(بيانات.get("الارتفاع", 0.0) or 0.0),
                "الكمية": int(عدادات_الخطط.get(رمز, 0) or 0),
                "مصدر_الكمية": "PLANS",
                "مصدر_الأبعاد": str(بيانات.get("مصدر_الأبعاد", "PACKAGE-DETAIL") or "PACKAGE-DETAIL"),
                "ملف_الأبعاد": str(بيانات.get("ملف_الأبعاد", "") or ""),
            })
        return نتائج

    def _نوع_الملف(self, مسار: str) -> str:
        """يصنف الملف بناءً على اسمه."""
        اسم = os.path.basename(مسار).upper()
        if اسم.startswith("ARC-") or اسم.startswith("ARCH-"):
            return "معماري"
        # A + رقم (مثل A101, A-01) = معماري
        if re.match(r'^A\d', اسم) or re.match(r'^A-\d', اسم):
            return "معماري"
        نتائج_إنشائي = any(ك in اسم for ك in self.كلمات_إنشائية)
        نتائج_معماري = any(ك in اسم for ك in self.كلمات_معمارية)
        # لو الاثنين موجودين — نختار حسب بداية الاسم (S = إنشائي، A = معماري)
        if نتائج_إنشائي and نتائج_معماري:
            if اسم[0] == 'S':
                return "إنشائي"
            return "معماري"
        if نتائج_إنشائي:
            return "إنشائي"
        if نتائج_معماري:
            return "معماري"
        # محاولة القراءة السريعة لأول 200 طبقة
        try:
            if مسار.endswith('.dxf'):
                وثيقة = ezdxf.readfile(مسار)
                الطبقات = [ط.dxf.name.upper() for ط in وثيقة.layers][:200]
                if any(any(ك in ط for ط in الطبقات) for ك in ["FOOTING","FOUNDATION","S-BEAM"]):
                    return "إنشائي"
                if any(any(ك in ط for ط in الطبقات) for ك in ["A-WALL","A-DOOR","A-FLOR"]):
                    return "معماري"
        except Exception:
            pass
        return "غير_محدد"

    def مسح(self):
        """يمسح المجلد ويصنف الملفات."""
        امتدادات = ['*.dxf', '*.pdf', '*.DXF', '*.PDF']
        كل_الملفات = []
        for امتداد in امتدادات:
            كل_الملفات.extend(glob.glob(os.path.join(self.مسار_المجلد, امتداد)))
        كل_الملفات = sorted(set(os.path.abspath(ملف) for ملف in كل_الملفات))

        for ملف in كل_الملفات:
            نوع = self._نوع_الملف(ملف)
            if نوع == "إنشائي":
                self.الملفات_الإنشائية.append(ملف)
            elif نوع == "معماري":
                self.الملفات_المعمارية.append(ملف)
            else:
                self.ملفات_غير_مصنفة.append(ملف)

        مسجل.info(
            f"المجلد: {len(self.الملفات_الإنشائية)} إنشائي | "
            f"{len(self.الملفات_المعمارية)} معماري | "
            f"{len(self.ملفات_غير_مصنفة)} غير محدد"
        )

        # لو ما في تصنيف واضح، عامل الكل كـ إنشائي+معماري
        if not self.الملفات_الإنشائية and not self.الملفات_المعمارية:
            self.الملفات_الإنشائية = [str(f) for f in كل_الملفات]
            self.الملفات_المعمارية = [str(f) for f in كل_الملفات]

    def معالجة(self) -> Dict:
        """يعالج كل الملفات ويدمج النتائج."""
        self.مسح()

        بيانات_مدمجة = {
            "المساحة_الكلية": 0.0,
            "المحيط_الخارجي": 0.0,
            "مساحة_الحفر": 0.0,
            "مساحات_البلاطات": [],
            "طول_جدران_بلوك_20": 0.0,
            "طول_جدران_بلوك_10": 0.0,
            "الغرف": [],
            "الجداول": {
                "الأسس": [], "أعمدة_العنق": [],
                "الأعمدة": [], "الفتحات": []
            },
            "الأعمدة_من_الرسم": {
                "الأعمدة": [],
                "إحصاء_حسب_المستوى": {},
            },
            "الكمرات_من_الرسم": {
                "كمرات_الربط": [], "كمرات_الستراب": [], "الكمرات": []
            },
            "بيانات_gemini": {},
            "درجة_الثقة": {"النتيجة": 0, "الدرجة": "غير_محدد"},
            "ملفات_معالجة": [],
        }
        فتحات_أبعاد_مدمجة: Dict[str, Dict] = {}
        عدادات_فتحات_مدمجة: Dict[str, int] = {}
        عدادات_فتحات_حسب_المستوى_مدمجة: Dict[str, Dict[str, int]] = defaultdict(dict)
        كاش_نتائج_الملفات: Dict[str, Dict] = {}

        # معالجة كل ملف
        كل_الملفات = list(set(self.الملفات_الإنشائية + self.الملفات_المعمارية))
        for ملف in كل_الملفات:
            try:
                مسجل.info(f"معالجة: {os.path.basename(ملف)}")
                محرك = محرك_الفضاء(ملف, self.وحدة, self.مفتاح_gemini)
                نتيجة = محرك.تنفيذ()
                كاش_نتائج_الملفات[os.path.abspath(ملف)] = نتيجة

                # دمج المساحات (أكبر قيمة)
                if نتيجة["المساحة_الكلية"] > بيانات_مدمجة["المساحة_الكلية"]:
                    بيانات_مدمجة["المساحة_الكلية"] = نتيجة["المساحة_الكلية"]
                if نتيجة["المحيط_الخارجي"] > بيانات_مدمجة["المحيط_الخارجي"]:
                    بيانات_مدمجة["المحيط_الخارجي"] = نتيجة["المحيط_الخارجي"]
                if نتيجة["مساحة_الحفر"] > بيانات_مدمجة["مساحة_الحفر"]:
                    بيانات_مدمجة["مساحة_الحفر"] = نتيجة["مساحة_الحفر"]

                # البلاطات: نحتفظ بأغنى قائمة مقاسة من الرسم
                if len(نتيجة.get("مساحات_البلاطات", [])) > len(بيانات_مدمجة["مساحات_البلاطات"]):
                    بيانات_مدمجة["مساحات_البلاطات"] = نتيجة.get("مساحات_البلاطات", [])

                # دمج الجدران (جمع)
                بيانات_مدمجة["طول_جدران_بلوك_20"] += نتيجة.get("طول_جدران_بلوك_20", 0)
                بيانات_مدمجة["طول_جدران_بلوك_10"] += نتيجة.get("طول_جدران_بلوك_10", 0)

                # دمج الغرف بتصفية التكرار عبر الملفات — O(n) بدل O(n²)
                # مفتاح: (اسم, مستوى) → احتفظ بالغرفة الأكبر مساحةً
                if "_غرف_مؤقتة" not in بيانات_مدمجة:
                    بيانات_مدمجة["_غرف_مؤقتة"] = {
                        (غ.get("الاسم", ""), غ.get("المستوى", "")): غ
                        for غ in بيانات_مدمجة["الغرف"]
                    }
                for غرفة_جديدة in نتيجة.get("الغرف", []):
                    _مفتاح = (غرفة_جديدة.get("الاسم", ""), غرفة_جديدة.get("المستوى", ""))
                    _مساحة = float(غرفة_جديدة.get("المساحة", 0) or 0)
                    موجودة = بيانات_مدمجة["_غرف_مؤقتة"].get(_مفتاح)
                    if موجودة is None:
                        بيانات_مدمجة["_غرف_مؤقتة"][_مفتاح] = غرفة_جديدة
                    else:
                        _مساحة_موجودة = float(موجودة.get("المساحة", 0) or 0)
                        if (_مساحة_موجودة > 0
                                and abs(_مساحة - _مساحة_موجودة) / _مساحة_موجودة < 0.25
                                and _مساحة > _مساحة_موجودة):
                            بيانات_مدمجة["_غرف_مؤقتة"][_مفتاح] = غرفة_جديدة
                        elif _مساحة_موجودة == 0:
                            بيانات_مدمجة["_غرف_مؤقتة"][_مفتاح] = غرفة_جديدة

                # دمج الجداول (بدون تكرار)
                for مفتاح in ["الأسس", "أعمدة_العنق", "الأعمدة", "الفتحات"]:
                    موجود = {ع.get("الرمز") for ع in بيانات_مدمجة["الجداول"][مفتاح]}
                    for عنصر in نتيجة.get("الجداول", {}).get(مفتاح, []):
                        if عنصر.get("الرمز") not in موجود:
                            بيانات_مدمجة["الجداول"][مفتاح].append(عنصر)

                # دمج أعمدة الرسم حسب المستوى
                for عنصر in نتيجة.get("الأعمدة_من_الرسم", {}).get("الأعمدة", []):
                    موجود = {
                        (ع.get("الرمز"), ع.get("المستوى"))
                        for ع in بيانات_مدمجة["الأعمدة_من_الرسم"]["الأعمدة"]
                    }
                    مفتاح_عنصر = (عنصر.get("الرمز"), عنصر.get("المستوى"))
                    if مفتاح_عنصر not in موجود:
                        بيانات_مدمجة["الأعمدة_من_الرسم"]["الأعمدة"].append(عنصر)
                for مستوى, عدادات in نتيجة.get("الأعمدة_من_الرسم", {}).get("إحصاء_حسب_المستوى", {}).items():
                    مجمعة = بيانات_مدمجة["الأعمدة_من_الرسم"]["إحصاء_حسب_المستوى"].setdefault(مستوى, {})
                    for رمز, كمية in عدادات.items():
                        مجمعة[رمز] = مجمعة.get(رمز, 0) + كمية

                # دمج الكمرات
                for مفتاح in ["كمرات_الربط", "كمرات_الستراب", "الكمرات"]:
                    موجود = {ك.get("الرمز") for ك in بيانات_مدمجة["الكمرات_من_الرسم"][مفتاح]}
                    for ك in نتيجة.get("الكمرات_من_الرسم", {}).get(مفتاح, []):
                        if ك.get("الرمز") not in موجود:
                            بيانات_مدمجة["الكمرات_من_الرسم"][مفتاح].append(ك)
                        else:
                            # لو الرمز موجود، اجمع الطول
                            for ك2 in بيانات_مدمجة["الكمرات_من_الرسم"][مفتاح]:
                                if ك2.get("الرمز") == ك.get("الرمز"):
                                    ك2["الطول"] = round(ك2.get("الطول",0) + ك.get("الطول",0), 3)

                # درجة الثقة (أعلى درجة)
                if نتيجة["درجة_الثقة"]["النتيجة"] > بيانات_مدمجة["درجة_الثقة"]["النتيجة"]:
                    بيانات_مدمجة["درجة_الثقة"] = نتيجة["درجة_الثقة"]

                بيانات_مدمجة["ملفات_معالجة"].append(os.path.basename(ملف))

            except Exception as خطأ:
                مسجل.error(f"فشل معالجة {ملف}: {خطأ}")

        # الفتحات: العد من جميع ملفات المساقط في الحزمة مع منع تكرار نفس المستوى، والأبعاد من أي schedule/detail file
        كل_ملفات_الفتحان = sorted(set(
            os.path.abspath(ملف)
            for ملف in (self.الملفات_الإنشائية + self.الملفات_المعمارية + self.ملفات_غير_مصنفة)
        ))
        for ملف in كل_ملفات_الفتحان:
            try:
                نتيجة = كاش_نتائج_الملفات.get(ملف)
                if نتيجة is None:
                    مسجل.info(f"معالجة فتحات الحزمة: {os.path.basename(ملف)}")
                    محرك = محرك_الفضاء(ملف, self.وحدة, self.مفتاح_gemini)
                    نتيجة = محرك.تنفيذ()
                    كاش_نتائج_الملفات[ملف] = نتيجة

                عدادات_حسب_المستوى = نتيجة.get("عداد_الفتحات_حسب_المستوى", {}) or {}
                if عدادات_حسب_المستوى:
                    for مستوى, عدادات in عدادات_حسب_المستوى.items():
                        مجمعة_المستوى = عدادات_فتحات_حسب_المستوى_مدمجة.setdefault(مستوى, {})
                        for رمز, كمية in (عدادات or {}).items():
                            رمز_موحد = رمز_فتحة_موحد(رمز)
                            if not رمز_موحد:
                                continue
                            كمية_صحيحة = int(كمية or 0)
                            if كمية_صحيحة > int(مجمعة_المستوى.get(رمز_موحد, 0) or 0):
                                مجمعة_المستوى[رمز_موحد] = كمية_صحيحة
                else:
                    for رمز, كمية in (نتيجة.get("عداد_الفتحات", {}) or {}).items():
                        رمز_موحد = رمز_فتحة_موحد(رمز)
                        if not رمز_موحد:
                            continue
                        if int(كمية or 0) > int(عدادات_فتحات_مدمجة.get(رمز_موحد, 0) or 0):
                            عدادات_فتحات_مدمجة[رمز_موحد] = int(كمية or 0)

                for فتحة in نتيجة.get("الجداول", {}).get("الفتحات", []):
                    رمز_موحد = رمز_فتحة_موحد(فتحة.get("الرمز", ""))
                    if not رمز_موحد:
                        continue
                    طول = float(فتحة.get("الطول", 0.0) or 0.0)
                    ارتفاع = float(فتحة.get("الارتفاع", 0.0) or 0.0)
                    حالي = فتحات_أبعاد_مدمجة.get(رمز_موحد)
                    if طول > 0 and ارتفاع > 0:
                        if (not حالي) or ((طول * ارتفاع) > (float(حالي.get("الطول", 0.0) or 0.0) * float(حالي.get("الارتفاع", 0.0) or 0.0))):
                            فتحات_أبعاد_مدمجة[رمز_موحد] = {
                                "الرمز": فتحة.get("الرمز", رمز_موحد),
                                "الطول": طول,
                                "الارتفاع": ارتفاع,
                                "مصدر_الأبعاد": str(فتحة.get("مصدر_الأبعاد", "PACKAGE-DETAIL") or "PACKAGE-DETAIL"),
                                "ملف_الأبعاد": os.path.basename(ملف),
                            }
            except Exception as خطأ:
                مسجل.error(f"فشل دمج فتحات الحزمة من {ملف}: {خطأ}")

        if عدادات_فتحات_حسب_المستوى_مدمجة:
            عدادات_فتحات_مدمجة = {}
            for عدادات in عدادات_فتحات_حسب_المستوى_مدمجة.values():
                for رمز, كمية in عدادات.items():
                    عدادات_فتحات_مدمجة[رمز] = عدادات_فتحات_مدمجة.get(رمز, 0) + int(كمية or 0)

        if فتحات_أبعاد_مدمجة or عدادات_فتحات_مدمجة:
            بيانات_مدمجة["الجداول"]["الفتحات"] = self._ادمج_فتحات_الحزمة(
                فتحات_أبعاد_مدمجة,
                عدادات_فتحات_مدمجة,
            )
            بيانات_مدمجة["عداد_الفتحات"] = dict(sorted(عدادات_فتحات_مدمجة.items()))
            بيانات_مدمجة["عداد_الفتحات_حسب_المستوى"] = {
                مستوى: dict(sorted(عدادات.items()))
                for مستوى, عدادات in sorted(عدادات_فتحات_حسب_المستوى_مدمجة.items())
            }

        # تفريغ dict الغرف المؤقت إلى قائمة وإعادة حساب المساحة_الكلية
        if "_غرف_مؤقتة" in بيانات_مدمجة:
            بيانات_مدمجة["الغرف"] = list(بيانات_مدمجة.pop("_غرف_مؤقتة").values())

        # إعادة حساب المساحة_الكلية من الغرف بعد تصفية التكرار
        if بيانات_مدمجة["الغرف"]:
            _مساحة_من_غرف = round(
                sum(float(غ.get("المساحة", 0) or 0) for غ in بيانات_مدمجة["الغرف"]
                    if "FLOOR" not in str(غ.get("الاسم", ""))),
                2
            )
            if _مساحة_من_غرف > 0:
                بيانات_مدمجة["المساحة_الكلية"] = _مساحة_من_غرف

        return بيانات_مدمجة
# ─────────────────────────────────────────────────────────────────────────────

class مُصدِّر_Excel:
    """
    يصدّر نتائج الكميات إلى ملف Excel منسّق احترافياً.
    """

    # ألوان الأقسام
    ألوان = {
        "البنية_التحتية": "1F4E79",
        "البنية_الفوقية": "2E75B6",
        "المعمارية":      "375623",
        "التشطيبات":      "7030A0",
        "الفتحات":        "833C00",
    }

    أسماء_الأقسام = {
        "البنية_التحتية": "أ — البنية التحتية (Substructure)",
        "البنية_الفوقية": "ب — البنية الفوقية (Superstructure)",
        "المعمارية":      "ج — الأعمال المعمارية (Architectural)",
        "التشطيبات":      "د — التشطيبات (Finishes)",
        "الفتحات":        "هـ — الفتحات (Openings)",
    }

    def __init__(self, نتائج: Dict, بيانات_مشروع: Dict, تنبيهات: List[Dict]):
        self.نتائج         = نتائج
        self.بيانات_مشروع  = بيانات_مشروع
        self.تنبيهات       = تنبيهات

    @staticmethod
    def _خلية_رأس(ورقة, صف, عمود, نص, لون_خلفية="1F4E79", لون_نص="FFFFFF", حجم=12):
        خلية = ورقة.cell(row=صف, column=عمود, value=نص)
        خلية.font       = Font(bold=True, color=لون_نص, size=حجم)
        خلية.fill       = PatternFill("solid", fgColor=لون_خلفية)
        خلية.alignment  = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True, reading_order=2)
        خلية.border     = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        return خلية

    @staticmethod
    def _خلية_بيانات(ورقة, صف, عمود, قيمة, محاذاة="right"):
        خلية = ورقة.cell(row=صف, column=عمود, value=قيمة)
        خلية.alignment = Alignment(horizontal=محاذاة, reading_order=2)
        خلية.border    = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        return خلية

    def تصدير(self, مسار_الإخراج: str) -> str:
        if not EXCEL_متاح:
            raise RuntimeError("openpyxl غير مثبت. نفذ: pip install openpyxl")

        مصنف = openpyxl.Workbook()

        # ── ورقة 1: ملخص المشروع ───────────────────────────────────────────────
        ورقة_ملخص = مصنف.active
        ورقة_ملخص.title = "ملخص المشروع"
        ورقة_ملخص.sheet_view.rightToLeft = True
        ورقة_ملخص.column_dimensions['A'].width = 30
        ورقة_ملخص.column_dimensions['B'].width = 30

        self._خلية_رأس(ورقة_ملخص, 1, 1, "محرك حساب الكميات الإماراتي v15", حجم=14)
        ورقة_ملخص.merge_cells('A1:B1')

        بيانات_ملخص = [
            ("رقم المشروع",   self.بيانات_مشروع.get("رقم_المشروع", "")),
            ("التاريخ",        self.بيانات_مشروع.get("التاريخ", "")),
            ("المساحة الكلية", f"{self.بيانات_مشروع.get('المساحة_الكلية', 0):.2f} م²"),
            ("المحيط الخارجي", f"{self.بيانات_مشروع.get('المحيط_الخارجي', 0):.2f} م"),
            ("ارتفاع الأرضي GF", f"{self.بيانات_مشروع.get('ارتفاع_الدور_الأرضي', 0):.2f} م"),
            ("ارتفاع الأول 1F", f"{self.بيانات_مشروع.get('ارتفاع_الدور_الأول', 0):.2f} م"),
            ("ارتفاع الثاني 2F", f"{self.بيانات_مشروع.get('ارتفاع_الدور_الثاني', 0):.2f} م"),
            ("ارتفاع السطح RF", f"{self.بيانات_مشروع.get('ارتفاع_دور_السطح', 0):.2f} م"),
            ("ارتفاع الخدمة KB", f"{self.بيانات_مشروع.get('ارتفاع_مبنى_الخدمة', 0):.2f} م") if self.بيانات_مشروع.get("يوجد_مبنى_خدمة", False) else None,
            ("ارتفاع عام احتياطي", f"{self.بيانات_مشروع.get('ارتفاع_الدور', 0):.2f} م"),
            ("عمق الحفر",      f"{self.بيانات_مشروع.get('عمق_الحفر', 0):.2f} م"),
            ("عدد الأدوار",    self.بيانات_مشروع.get("عدد_الأدوار", "")),
            ("نوع المشروع المستخرج", self.بيانات_مشروع.get("نوع_المشروع", "")),
            ("المستويات الصريحة", ", ".join(self.بيانات_مشروع.get("المستويات_الصريحة", []) or [])),
            ("درجة الثقة",     f"{self.بيانات_مشروع.get('درجة_الثقة', {}).get('النتيجة', 0)}% — {self.بيانات_مشروع.get('درجة_الثقة', {}).get('الدرجة', '')}"),
        ]
        بيانات_ملخص = [س for س in بيانات_ملخص if س]
        for i, (مفتاح, قيمة) in enumerate(بيانات_ملخص, start=3):
            self._خلية_بيانات(ورقة_ملخص, i, 1, مفتاح)
            self._خلية_بيانات(ورقة_ملخص, i, 2, قيمة, "center")

        # ── ورقة 2: جدول الكميات الكامل ──────────────────────────────────────
        ورقة_BOQ = مصنف.create_sheet("جدول الكميات")
        ورقة_BOQ.sheet_view.rightToLeft = True
        ورقة_BOQ.column_dimensions['A'].width = 5
        ورقة_BOQ.column_dimensions['B'].width = 45
        ورقة_BOQ.column_dimensions['C'].width = 15
        ورقة_BOQ.column_dimensions['D'].width = 10

        # رأس الجدول
        self._خلية_رأس(ورقة_BOQ, 1, 1, "م", حجم=11)
        self._خلية_رأس(ورقة_BOQ, 1, 2, "البند", حجم=11)
        self._خلية_رأس(ورقة_BOQ, 1, 3, "الكمية", حجم=11)
        self._خلية_رأس(ورقة_BOQ, 1, 4, "الوحدة", حجم=11)
        ورقة_BOQ.row_dimensions[1].height = 25

        صف_حالي = 2
        رقم_بند  = 1

        for قسم, بنود in self.نتائج.items():
            لون = self.ألوان.get(قسم, "404040")
            اسم_قسم = self.أسماء_الأقسام.get(قسم, قسم)

            # رأس القسم
            self._خلية_رأس(ورقة_BOQ, صف_حالي, 1, "", لون)
            self._خلية_رأس(ورقة_BOQ, صف_حالي, 2, اسم_قسم, لون, حجم=11)
            self._خلية_رأس(ورقة_BOQ, صف_حالي, 3, "", لون)
            self._خلية_رأس(ورقة_BOQ, صف_حالي, 4, "", لون)
            ورقة_BOQ.merge_cells(f'B{صف_حالي}:D{صف_حالي}')
            صف_حالي += 1

            for بند in بنود:
                self._خلية_بيانات(ورقة_BOQ, صف_حالي, 1, رقم_بند, "center")
                self._خلية_بيانات(ورقة_BOQ, صف_حالي, 2, بند["البند"])
                كمية_خلية = ورقة_BOQ.cell(row=صف_حالي, column=3, value=بند["الكمية"])
                كمية_خلية.number_format = '#,##0.000'
                كمية_خلية.alignment = Alignment(horizontal="center")
                كمية_خلية.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
                self._خلية_بيانات(ورقة_BOQ, صف_حالي, 4, بند["الوحدة"], "center")
                صف_حالي += 1
                رقم_بند  += 1

            صف_حالي += 1  # فراغ بين الأقسام

        # ── ورقة 3: التحذيرات ─────────────────────────────────────────────────
        if self.تنبيهات:
            ورقة_تنبيهات = مصنف.create_sheet("التحذيرات والتنبيهات")
            ورقة_تنبيهات.sheet_view.rightToLeft = True
            ورقة_تنبيهات.column_dimensions['A'].width = 10
            ورقة_تنبيهات.column_dimensions['B'].width = 20
            ورقة_تنبيهات.column_dimensions['C'].width = 60

            self._خلية_رأس(ورقة_تنبيهات, 1, 1, "المستوى")
            self._خلية_رأس(ورقة_تنبيهات, 1, 2, "البند")
            self._خلية_رأس(ورقة_تنبيهات, 1, 3, "الرسالة")

            ألوان_تنبيه = {"خطأ": "FF0000", "تحذير": "FFC000", "ملاحظة": "92D050"}
            for i, ت in enumerate(self.تنبيهات, start=2):
                لون = ألوان_تنبيه.get(ت["المستوى"], "FFFFFF")
                for عمود, قيمة in [(1, f"{ت['الرمز']} {ت['المستوى']}"),
                                    (2, ت["البند"]),
                                    (3, ت["الرسالة"])]:
                    خ = ورقة_تنبيهات.cell(row=i, column=عمود, value=قيمة)
                    خ.fill      = PatternFill("solid", fgColor=لون)
                    خ.alignment = Alignment(reading_order=2, wrap_text=True)
                    خ.border    = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        مصنف.save(مسار_الإخراج)
        مسجل.info(f"Excel حُفظ: {مسار_الإخراج}")
        return مسار_الإخراج


# ─────────────────────────────────────────────────────────────────────────────
# واجهة المستخدم HTML
# ─────────────────────────────────────────────────────────────────────────────

HTML_الواجهة = """<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>محرك حساب الكميات الإماراتي v15</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Arial, sans-serif; }
  body { background: #0f1117; color: #e0e0e0; min-height: 100vh; }

  .header {
    background: linear-gradient(135deg, #1a237e 0%, #283593 50%, #1565c0 100%);
    padding: 24px 32px; border-bottom: 2px solid #3949ab;
  }
  .header h1 { font-size: 1.8rem; color: #fff; }
  .header p  { color: #90caf9; font-size: 0.9rem; margin-top: 4px; }

  .container { max-width: 1200px; margin: 0 auto; padding: 24px; }

  .card {
    background: #1a1d2e; border: 1px solid #2a2d3e;
    border-radius: 12px; padding: 24px; margin-bottom: 20px;
  }
  .card h2 { color: #90caf9; font-size: 1.1rem; margin-bottom: 16px;
             border-bottom: 1px solid #2a2d3e; padding-bottom: 8px; }

  .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
  .grid-3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; }

  label { display: block; color: #90caf9; font-size: 0.85rem; margin-bottom: 4px; }
  input, select {
    width: 100%; padding: 10px 12px; background: #0f1117;
    border: 1px solid #3949ab; border-radius: 8px; color: #e0e0e0;
    font-size: 0.9rem; transition: border 0.2s;
  }
  input:focus, select:focus { outline: none; border-color: #90caf9; }

  .upload-area {
    border: 2px dashed #3949ab; border-radius: 12px;
    padding: 32px; text-align: center; cursor: pointer;
    transition: all 0.3s; background: #0f1117;
  }
  .upload-area:hover { border-color: #90caf9; background: #1a1d2e; }
  .upload-area .icon { font-size: 2.5rem; margin-bottom: 8px; }
  .upload-area p { color: #90caf9; }
  .upload-area small { color: #5c6bc0; }

  .btn {
    padding: 12px 32px; border: none; border-radius: 8px;
    font-size: 1rem; font-weight: bold; cursor: pointer;
    transition: all 0.2s; width: 100%;
  }
  .btn-primary {
    background: linear-gradient(135deg, #1565c0, #1976d2);
    color: white;
  }
  .btn-primary:hover { background: linear-gradient(135deg, #1976d2, #42a5f5); }
  .btn-primary:disabled { opacity: 0.5; cursor: not-allowed; }

  .btn-excel {
    background: linear-gradient(135deg, #1b5e20, #2e7d32);
    color: white; margin-top: 12px;
  }
  .btn-excel:hover { background: linear-gradient(135deg, #2e7d32, #43a047); }

  .results { display: none; }
  .results.show { display: block; }

  .confidence-badge {
    display: inline-block; padding: 6px 16px; border-radius: 20px;
    font-weight: bold; font-size: 0.9rem;
  }
  .conf-A { background: #1b5e20; color: #a5d6a7; }
  .conf-B { background: #1a237e; color: #90caf9; }
  .conf-C { background: #e65100; color: #ffcc80; }
  .conf-D { background: #b71c1c; color: #ef9a9a; }

  .section-header {
    background: #1a237e; color: #90caf9;
    padding: 10px 16px; border-radius: 6px;
    font-weight: bold; margin: 16px 0 8px; font-size: 0.95rem;
  }

  table { width: 100%; border-collapse: collapse; font-size: 0.88rem; }
  th {
    background: #1a237e; color: #90caf9;
    padding: 10px 12px; text-align: right; font-weight: 600;
  }
  td { padding: 8px 12px; border-bottom: 1px solid #2a2d3e; }
  tr:hover td { background: #1e2237; }
  .qty { color: #a5d6a7; font-weight: bold; font-family: monospace; }
  .unit { color: #ef9a9a; font-size: 0.85rem; }

  .warnings { margin-top: 16px; }
  .warn-خطأ   { background: #311; border-right: 4px solid #f44336; }
  .warn-تحذير { background: #221500; border-right: 4px solid #ff9800; }
  .warn-ملاحظة{ background: #1a2a1a; border-right: 4px solid #4caf50; }
  .warn-item  { padding: 10px 14px; border-radius: 6px; margin-bottom: 8px; font-size: 0.88rem; }

  .loader { display: none; text-align: center; padding: 40px; }
  .loader.show { display: block; }
  .spinner {
    width: 48px; height: 48px; border: 4px solid #3949ab;
    border-top-color: #90caf9; border-radius: 50%;
    animation: spin 1s linear infinite; margin: 0 auto 16px;
  }
  @keyframes spin { to { transform: rotate(360deg); } }

  .tag { display: inline-block; background: #1a237e; color: #90caf9;
         padding: 2px 8px; border-radius: 4px; font-size: 0.8rem; margin: 2px; }
  .stat-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; }
  .stat-card { background: #0f1117; border: 1px solid #2a2d3e; border-radius: 8px;
               padding: 14px; text-align: center; }
  .stat-card .val { font-size: 1.4rem; font-weight: bold; color: #90caf9; }
  .stat-card .lbl { font-size: 0.78rem; color: #7986cb; margin-top: 4px; }
</style>
</head>
<body>

<div class="header">
  <h1>⚙️ محرك حساب الكميات الإماراتي</h1>
  <p>الإصدار 15 — تحليل تلقائي لمخططات DXF وPDF مع 31 بنداً حسابياً</p>
</div>

<div class="container">

  <!-- رفع الملف -->
  <div class="card">
    <h2>📁 رفع المخططات</h2>
    <div class="upload-area" onclick="document.getElementById('fileInput').click()">
      <div class="icon">📐</div>
      <p>اسحب ملفات DXF أو PDF هنا</p>
      <small>يدعم: ملف واحد أو مجلد كامل (.dxf, .pdf)</small>
      <input type="file" id="fileInput" multiple accept=".dxf,.pdf" style="display:none" onchange="onFilesSelected(this)">
    </div>
    <div id="filesList" style="margin-top:12px"></div>
    <div id="profileInfo" class="warn-item warn-ملاحظة" style="display:none; margin-top:12px"></div>
  </div>

  <!-- ثوابت المشروع -->
  <div class="card">
    <h2>📊 ثوابت المشروع</h2>
    <div class="grid-3">
      <div>
        <label>رقم المشروع</label>
        <input type="text" id="projectId" placeholder="PRJ-001" value="PRJ-001">
      </div>
      <div>
        <label>وحدة القياس في الرسم</label>
        <select id="unit">
          <option value="mm" selected>ملليمتر (mm)</option>
          <option value="cm">سنتيمتر (cm)</option>
          <option value="m">متر (m)</option>
        </select>
      </div>
      <div>
        <label>عدد الأدوار - احتياطي فقط عند فشل اكتشاف المشروع</label>
        <input type="number" id="numFloors" value="0" min="0" max="10">
      </div>
      <div id="gfHeightWrap">
        <label>ارتفاع الأرضي GF (م) - عند فشل الاستخراج</label>
        <input type="number" id="gfHeight" value="0" step="0.1">
      </div>
      <div id="firstHeightWrap">
        <label>ارتفاع الأول 1F (م) - عند فشل الاستخراج</label>
        <input type="number" id="firstHeight" value="0" step="0.1">
      </div>
      <div id="secondHeightWrap">
        <label>ارتفاع الثاني 2F (م) - عند فشل الاستخراج</label>
        <input type="number" id="secondHeight" value="0" step="0.1">
      </div>
      <div id="roofHeightWrap">
        <label>ارتفاع السطح RF (م) - عند فشل الاستخراج</label>
        <input type="number" id="roofHeight" value="0" step="0.1">
      </div>
      <div id="serviceHeightWrap">
        <label>ارتفاع مبنى الخدمة KB (م) - اختياري</label>
        <input type="number" id="serviceHeight" value="0" step="0.1">
      </div>
      <div>
        <label>عمق الحفر (م) - إدخال يدوي</label>
        <input type="number" id="excavationDepth" value="0" step="0.05">
      </div>
      <div>
        <label>مستوى بلاطة الأرضي (م) - اختياري عند فشل الاستخراج</label>
        <input type="number" id="gfslLevel" value="0" step="0.05">
      </div>
      <div>
        <label>عمق كمرة الربط (م) - اختياري عند فشل الاستخراج</label>
        <input type="number" id="tbDepth" value="0" step="0.05">
      </div>
      <div>
        <label>سماكة البلاطة (م) - اختياري عند فشل الاستخراج</label>
        <input type="number" id="slabThickness" value="0" step="0.01">
      </div>
      <div>
        <label>خرسانة الدرج (م³) - اختياري</label>
        <input type="number" id="staircaseConcrete" value="0" step="0.1">
      </div>
    </div>
    <div class="grid-2" style="margin-top:16px">
      <div>
        <label>مفتاح Gemini API (لـ PDF المسح الضوئي)</label>
        <input type="password" id="geminiKey" placeholder="AIza...">
      </div>
      <div>
        <label>يوجد رصيف طرق؟</label>
        <select id="roadBase">
          <option value="false">لا</option>
          <option value="true">نعم</option>
        </select>
      </div>
    </div>
  </div>

  <!-- زر التشغيل -->
  <button class="btn btn-primary" onclick="runQTO()" id="runBtn">
    🚀 تشغيل حساب الكميات
  </button>

  <!-- مؤشر التحميل -->
  <div class="loader" id="loader">
    <div class="spinner"></div>
    <p style="color:#90caf9">جاري تحليل المخططات...</p>
  </div>

  <!-- النتائج -->
  <div class="results" id="results">

    <!-- إحصائيات سريعة -->
    <div class="card" style="margin-top:20px">
      <h2>📈 ملخص التحليل</h2>
      <div class="stat-grid" id="statsGrid"></div>
    </div>

    <!-- التنبيهات -->
    <div class="card" id="warningsCard" style="display:none">
      <h2>⚠️ التحذيرات والتنبيهات</h2>
      <div class="warnings" id="warningsList"></div>
    </div>

    <!-- جدول الكميات -->
    <div class="card">
      <h2>📋 جدول الكميات</h2>
      <div id="boqTable"></div>
      <button class="btn btn-excel" onclick="downloadExcel()" id="excelBtn" style="display:none">
        📥 تحميل Excel
      </button>
    </div>

  </div>

</div>

<script>
let uploadedFiles = [];
let lastResults = null;
let lastProjectId = '';

async function onFilesSelected(input) {
  uploadedFiles = Array.from(input.files);
  const list = document.getElementById('filesList');
  list.innerHTML = uploadedFiles.map(f =>
    `<span class="tag">📄 ${f.name}</span>`
  ).join('');
  await detectProjectProfile();
}

async function runQTO() {
  if (!uploadedFiles.length) {
    alert('⚠️ يرجى رفع ملف DXF أو PDF أولاً');
    return;
  }

  document.getElementById('runBtn').disabled = true;
  document.getElementById('loader').classList.add('show');
  document.getElementById('results').classList.remove('show');

  const formData = new FormData();
  uploadedFiles.forEach(f => formData.append('files', f));
  formData.append('project_id',        document.getElementById('projectId').value);
  formData.append('unit',              document.getElementById('unit').value);
  formData.append('gf_height',         document.getElementById('gfHeight').value);
  formData.append('first_floor_height',document.getElementById('firstHeight').value);
  formData.append('second_floor_height',document.getElementById('secondHeight').value);
  formData.append('roof_floor_height', document.getElementById('roofHeight').value);
  formData.append('service_height',    document.getElementById('serviceHeight').value);
  formData.append('excavation_depth',  document.getElementById('excavationDepth').value);
  formData.append('gfsl_level',        document.getElementById('gfslLevel').value);
  formData.append('tb_depth',          document.getElementById('tbDepth').value);
  formData.append('slab_thickness',    document.getElementById('slabThickness').value);
  formData.append('num_floors',        document.getElementById('numFloors').value);
  formData.append('staircase',         document.getElementById('staircaseConcrete').value);
  formData.append('road_base',         document.getElementById('roadBase').value);
  formData.append('gemini_key',        document.getElementById('geminiKey').value);

  try {
    const resp = await fetch('/رفع-وحساب', { method: 'POST', body: formData });
    const data = await resp.json();

    if (data.الحالة !== 'نجاح') throw new Error(data.detail || 'خطأ غير معروف');

    lastResults  = data;
    lastProjectId = document.getElementById('projectId').value;
    renderResults(data);

  } catch (err) {
    alert('❌ ' + err.message);
  } finally {
    document.getElementById('runBtn').disabled = false;
    document.getElementById('loader').classList.remove('show');
  }
}

function toggleHeightField(wrapperId, visible) {
  const wrap = document.getElementById(wrapperId);
  if (!wrap) return;
  wrap.style.display = visible ? '' : 'none';
  const input = wrap.querySelector('input');
  if (input && !visible) input.value = '0';
}

function applyDetectedProfile(profile) {
  const levels = profile.explicit_levels || [];
  const requested = new Set(profile.requested_height_fields || []);

  toggleHeightField('gfHeightWrap', requested.has('GF') || levels.length === 0);
  toggleHeightField('firstHeightWrap', requested.has('1F'));
  toggleHeightField('secondHeightWrap', requested.has('2F'));
  toggleHeightField('roofHeightWrap', requested.has('RF'));
  toggleHeightField('serviceHeightWrap', requested.has('KB'));

  const structuralLevels = profile.structural_levels || [];
  if (structuralLevels.length) {
    document.getElementById('numFloors').value = structuralLevels.length;
  }
}

async function detectProjectProfile() {
  const info = document.getElementById('profileInfo');
  if (!uploadedFiles.length) {
    info.style.display = 'none';
    return;
  }

  const formData = new FormData();
  uploadedFiles.forEach(f => formData.append('files', f));
  formData.append('unit', document.getElementById('unit').value);
  formData.append('gemini_key', document.getElementById('geminiKey').value);

  info.style.display = 'block';
  info.className = 'warn-item warn-ملاحظة';
  info.textContent = 'جارٍ اكتشاف نوع المشروع والمستويات من الرسم...';

  try {
    const response = await fetch('/فحص-المشروع', { method: 'POST', body: formData });
    const data = await response.json();
    if (!response.ok || data.الحالة !== 'نجاح') {
      throw new Error(data.detail || data.message || 'تعذر اكتشاف ملف المشروع');
    }

    const profile = data.project_profile || {};
    applyDetectedProfile(profile);

    const explicitLevels = (profile.explicit_levels || []).join(', ') || 'غير واضحة';
    const projectType = profile.project_type || 'UNKNOWN';
    const hasService = profile.has_service_block ? 'نعم' : 'لا';
    info.innerHTML = `نوع المشروع: <strong>${projectType}</strong> | المستويات: <strong>${explicitLevels}</strong> | مبنى خدمة: <strong>${hasService}</strong>`;
  } catch (error) {
    toggleHeightField('gfHeightWrap', true);
    toggleHeightField('firstHeightWrap', true);
    toggleHeightField('secondHeightWrap', true);
    toggleHeightField('roofHeightWrap', true);
    toggleHeightField('serviceHeightWrap', true);
    info.className = 'warn-item warn-تحذير';
    info.textContent = `تعذر اكتشاف ملف المشروع تلقائيًا: ${error.message}`;
  }
}

document.addEventListener('DOMContentLoaded', () => {
  const unit = document.getElementById('unit');
  if (unit) {
    unit.addEventListener('change', () => {
      if (uploadedFiles.length) detectProjectProfile();
    });
  }
});

function renderResults(data) {
  // إحصائيات
  const conf = data.درجة_الثقة || {};
  const مكتشف = data.مكتشف_تلقائياً || {};
  const confClass = {'ممتاز':'A','جيد جداً':'B','جيد':'C','ضعيف':'D'}[conf.الدرجة] || 'D';

  document.getElementById('statsGrid').innerHTML = `
    <div class="stat-card">
      <div class="val">${(مكتشف.المحيط_الخارجي||0).toFixed(1)} م</div>
      <div class="lbl">المحيط الخارجي</div>
    </div>
    <div class="stat-card">
      <div class="val">${(مكتشف.مساحة_الحفر||0).toFixed(1)} م²</div>
      <div class="lbl">مساحة الحفر</div>
    </div>
    <div class="stat-card">
      <div class="val">${(مكتشف.طول_جدران_بلوك_20||0).toFixed(1)} م</div>
      <div class="lbl">جدران بلوك 20</div>
    </div>
    <div class="stat-card">
      <div class="val">
        <span class="confidence-badge conf-${confClass}">${conf.النتيجة||0}% — ${conf.الدرجة||''}</span>
      </div>
      <div class="lbl">درجة الثقة</div>
    </div>
  `;

  // التحذيرات
  const تنبيهات = data.التنبيهات || [];
  if (تنبيهات.length) {
    document.getElementById('warningsCard').style.display = 'block';
    document.getElementById('warningsList').innerHTML = تنبيهات.map(ت =>
      `<div class="warn-item warn-${ت.المستوى}">
        ${ت.الرمز} <strong>${ت.البند}</strong> — ${ت.الرسالة}
       </div>`
    ).join('');
  }

  // جدول الكميات
  const أقسام = data.النتائج_بالأقسام || {};
  const أسماء = {
    'البنية_التحتية': '🏗️ أ — البنية التحتية',
    'البنية_الفوقية': '🏛️ ب — البنية الفوقية',
    'المعمارية':      '🧱 ج — الأعمال المعمارية',
    'التشطيبات':      '🎨 د — التشطيبات',
    'الفتحات':        '🚪 هـ — الفتحات',
  };

  let html = '';
  let رقم = 1;
  for (const [قسم, بنود] of Object.entries(أقسام)) {
    if (!بنود.length) continue;
    html += `<div class="section-header">${أسماء[قسم]||قسم}</div>`;
    html += `<table><thead><tr><th>#</th><th>البند</th><th>الكمية</th><th>الوحدة</th></tr></thead><tbody>`;
    for (const ب of بنود) {
      html += `<tr>
        <td style="text-align:center;color:#5c6bc0">${رقم++}</td>
        <td>${ب.البند}</td>
        <td class="qty" style="text-align:center">${ب.الكمية.toLocaleString('ar-AE',{minimumFractionDigits:2,maximumFractionDigits:3})}</td>
        <td class="unit" style="text-align:center">${ب.الوحدة}</td>
      </tr>`;
    }
    html += `</tbody></table>`;
  }
  document.getElementById('boqTable').innerHTML = html;
  document.getElementById('excelBtn').style.display = 'block';
  document.getElementById('results').classList.add('show');
}

async function downloadExcel() {
  if (!lastResults) return;
  const resp = await fetch('/تحميل-excel', {
    method: 'POST',
    headers: {'Content-Type':'application/json'},
    body: JSON.stringify({ project_id: lastProjectId, results: lastResults })
  });
  const blob = await resp.blob();
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href     = url;
  a.download = `QTO_${lastProjectId}.xlsx`;
  a.click();
}
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# نقطة نهاية API
# ─────────────────────────────────────────────────────────────────────────────

import tempfile
import shutil
from datetime import datetime

# ── الواجهة الرئيسية ─────────────────────────────────────────────────────────

@التطبيق.get("/", response_class=HTMLResponse)
async def الصفحة_الرئيسية():
    return HTML_الواجهة


@التطبيق.post("/فحص-المشروع")
async def فحص_المشروع(
    files: List[UploadFile] = File(...),
    unit: str = Form("mm"),
    gemini_key: str = Form(""),
):
    مجلد_مؤقت = tempfile.mkdtemp()
    try:
        مسارات_الملفات = []
        for ملف in files:
            مسار = os.path.join(مجلد_مؤقت, ملف.filename)
            with open(مسار, "wb") as f:
                f.write(await ملف.read())
            مسارات_الملفات.append(مسار)

        if not مسارات_الملفات:
            raise HTTPException(status_code=400, detail="لم يتم رفع أي ملف")

        مسار_مرجعي = next((م for م in مسارات_الملفات if م.lower().endswith(".dxf")), مسارات_الملفات[0])
        محرك = محرك_الفضاء(مسار_مرجعي, unit, gemini_key)
        بيانات_الفضاء = محرك.تنفيذ()
        ملف_المشروع = بيانات_الفضاء.get("ملف_المشروع_المستخرج", {}) or {}

        مستويات_صريحة = list(ملف_المشروع.get("المستويات_الصريحة", []) or [])
        مستويات_إنشائية = list(ملف_المشروع.get("المستويات_الإنشائية", []) or [])
        يوجد_مبنى_خدمة = bool(ملف_المشروع.get("يوجد_مبنى_خدمة", False))
        مستويات_تحتاج_ارتفاع = [م for م in ["GF", "1F", "2F", "RF"] if م in مستويات_صريحة]
        if يوجد_مبنى_خدمة:
            مستويات_تحتاج_ارتفاع.append("KB")

        return {
            "الحالة": "نجاح",
            "الملف_المرجعي": os.path.basename(مسار_مرجعي),
            "project_profile": {
                "project_type": ملف_المشروع.get("نوع_المشروع", "UNKNOWN"),
                "explicit_levels": مستويات_صريحة,
                "structural_levels": مستويات_إنشائية,
                "has_service_block": يوجد_مبنى_خدمة,
                "requested_height_fields": مستويات_تحتاج_ارتفاع,
            },
        }
    except HTTPException:
        raise
    except Exception as خطأ:
        raise HTTPException(status_code=500, detail=str(خطأ))
    finally:
        shutil.rmtree(مجلد_مؤقت, ignore_errors=True)


# ── نقطة الرفع والحساب (للواجهة) ────────────────────────────────────────────

@التطبيق.post("/رفع-وحساب")
async def رفع_وحساب(
    files:             List[UploadFile] = File(...),
    project_id:        str  = Form("PRJ-001"),
    unit:              str  = Form("mm"),
    floor_height:      float = Form(0.0),
    gf_height:         float = Form(0.0),
    first_floor_height: float = Form(0.0),
    second_floor_height: float = Form(0.0),
    roof_floor_height: float = Form(0.0),
    service_height:    float = Form(0.0),
    excavation_depth:  float = Form(0.0),
    gfsl_level:        float = Form(0.0),
    tb_depth:          float = Form(0.0),
    slab_thickness:    float = Form(0.0),
    num_floors:        int   = Form(0),
    staircase:         float = Form(0.0),
    road_base:         str   = Form("false"),
    gemini_key:        str   = Form(""),
):
    # حفظ الملفات المرفوعة في مجلد مؤقت
    مجلد_مؤقت = tempfile.mkdtemp()
    try:
        مسارات_الملفات = []
        for ملف in files:
            مسار = os.path.join(مجلد_مؤقت, ملف.filename)
            with open(مسار, "wb") as f:
                f.write(await ملف.read())
            مسارات_الملفات.append(مسار)

        # بناء ثوابت المشروع
        الثوابت = ثوابت_المشروع(
            ارتفاع_الدور          = floor_height,
            ارتفاع_الدور_الأرضي   = gf_height,
            ارتفاع_الدور_الأول    = first_floor_height,
            ارتفاع_الدور_الثاني   = second_floor_height,
            ارتفاع_دور_السطح      = roof_floor_height,
            ارتفاع_مبنى_الخدمة    = service_height,
            عمق_الحفر             = excavation_depth,
            مستوى_بلاطة_الأرضي   = gfsl_level,
            عمق_كمرة_الربط        = tb_depth,
            عدد_الأدوار           = num_floors,
            سماكة_البلاطة         = slab_thickness,
            خرسانة_الدرج          = staircase,
            يوجد_رصيف_طرق         = road_base.lower() == "true",
            مفتاح_gemini          = gemini_key,
        )

        # معالجة: ملف واحد أو متعدد
        if len(مسارات_الملفات) == 1:
            محرك = محرك_الفضاء(مسارات_الملفات[0], unit, gemini_key)
            بيانات_الفضاء = محرك.تنفيذ()

            # عدّ الفتحات من المسقط (DXF فقط)
            if مسارات_الملفات[0].endswith('.dxf'):
                وثيقة_dxf = ezdxf.readfile(مسارات_الملفات[0])
                مقياس = {"m":1.0,"cm":0.01,"mm":0.001}.get(unit.lower(), 0.001)
                عداد = عدّاد_الفتحات(وثيقة_dxf.modelspace(), مقياس)
                عداد.عدّ()
                # تحديث الفتحات في الجداول
                فتحات_محدثة = عداد.تحديث_الفتحات(
                    [بيانات_فتحة(**ف) for ف in بيانات_الفضاء["الجداول"].get("الفتحات", [])]
                )
                بيانات_الفضاء["الجداول"]["الفتحات"] = [ف.__dict__ for ف in فتحات_محدثة]
                بيانات_الفضاء["عداد_الفتحات"] = dict(عداد.العداد)
                بيانات_الفضاء["عداد_الفتحات_حسب_المستوى"] = {
                    مستوى: dict(sorted(عدادات.items()))
                    for مستوى, عدادات in sorted(عداد.العداد_حسب_المستوى.items())
                }
        else:
            معالج = معالج_المجلد(مجلد_مؤقت, unit, gemini_key)
            بيانات_الفضاء = معالج.معالجة()

        # بناء الطلب
        الطلب = طلب_حساب_الكميات(
            مسار_الملف   = مسارات_الملفات[0] if مسارات_الملفات else "",
            رقم_المشروع  = project_id,
            وحدة_القياس  = unit,
            الثوابت      = الثوابت,
        )

        # حساب الكميات
        النتائج = كتاب_الكميات.احسب(
            بيانات_الغرف = بيانات_الفضاء["الغرف"],
            الفضاء       = بيانات_الفضاء,
            الطلب        = الطلب,
        )

        # التحقق من المنطق
        مدقق  = مُدقِّق_المنطق(النتائج, بيانات_الفضاء, الثوابت)
        تنبيهات = مدقق.تحقق()

        مسطح = [ب for قسم in النتائج.values() for ب in قسم]
        درجة_ثقة_نهائية = _حساب_درجة_ثقة_محسّنة(بيانات_الفضاء["درجة_الثقة"], مسطح)
        _نظام_متكامل(النتائج, بيانات_الفضاء, مجلد_مؤقت, gemini_key)

        return {
            "الحالة":          "نجاح",
            "رقم_المشروع":     project_id,
            "التاريخ":         datetime.now().strftime("%Y-%m-%d %H:%M"),
            "درجة_الثقة":      درجة_ثقة_نهائية,
            "التنبيهات":       تنبيهات,
            "مكتشف_تلقائياً": {
                "المحيط_الخارجي":    بيانات_الفضاء["المحيط_الخارجي"],
                "طول_جدران_بلوك_20": بيانات_الفضاء.get("طول_جدران_بلوك_20", 0),
                "طول_جدران_بلوك_10": بيانات_الفضاء.get("طول_جدران_بلوك_10", 0),
                "مساحة_الحفر":       بيانات_الفضاء["مساحة_الحفر"],
                "المساحة_الكلية":    بيانات_الفضاء["المساحة_الكلية"],
                "الجداول":           بيانات_الفضاء["الجداول"],
                "الكمرات_من_الرسم":  بيانات_الفضاء.get("الكمرات_من_الرسم", {}),
                "عداد_الفتحات":      بيانات_الفضاء.get("عداد_الفتحات", {}),
                "ملف_المشروع_المستخرج": بيانات_الفضاء.get("ملف_المشروع_المستخرج", {}),
                "ملفات_معالجة":      بيانات_الفضاء.get("ملفات_معالجة", []),
            },
            "النتائج_بالأقسام": النتائج,
            "النتائج_مسطحة":    مسطح,
            "الغرف":            بيانات_الفضاء["الغرف"],
        }

    except Exception as خطأ:
        مسجل.error(f"فشل: {خطأ}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(خطأ))
    finally:
        shutil.rmtree(مجلد_مؤقت, ignore_errors=True)


# ── تحميل Excel ──────────────────────────────────────────────────────────────

@التطبيق.post("/تحميل-excel")
async def تحميل_excel(بيانات: Dict):
    try:
        نتائج      = بيانات.get("results", {})
        project_id = بيانات.get("project_id", "QTO")

        بيانات_مشروع = {
            "رقم_المشروع":    project_id,
            "التاريخ":        datetime.now().strftime("%Y-%m-%d"),
            "المساحة_الكلية":  نتائج.get("مكتشف_تلقائياً", {}).get("المساحة_الكلية", 0),
            "المحيط_الخارجي":  نتائج.get("مكتشف_تلقائياً", {}).get("المحيط_الخارجي", 0),
            "درجة_الثقة":     نتائج.get("درجة_الثقة", {}),
        }

        مسار_excel = os.path.join(tempfile.gettempdir(), f"QTO_{project_id}.xlsx")
        مُصدِّر = مُصدِّر_Excel(
            نتائج.get("النتائج_بالأقسام", {}),
            بيانات_مشروع,
            نتائج.get("التنبيهات", [])
        )
        مُصدِّر.تصدير(مسار_excel)

        return FileResponse(
            path=مسار_excel,
            filename=f"QTO_{project_id}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as خطأ:
        raise HTTPException(status_code=500, detail=str(خطأ))


# ── API التقليدي (JSON) ───────────────────────────────────────────────────────

@التطبيق.post("/تشغيل-حساب-الكميات")
async def تشغيل_الكميات(الطلب: طلب_حساب_الكميات):
    """نقطة نهاية JSON للتكامل مع أنظمة أخرى."""
    try:
        مسار = الطلب.مسار_الملف or (
            list(glob.glob(os.path.join(الطلب.مسار_المجلد, "*.dxf")) +
                 glob.glob(os.path.join(الطلب.مسار_المجلد, "*.pdf")))[0]
            if الطلب.مسار_المجلد else ""
        )

        if الطلب.مسار_المجلد:
            معالج = معالج_المجلد(
                الطلب.مسار_المجلد,
                الطلب.وحدة_القياس,
                الطلب.الثوابت.مفتاح_gemini
            )
            بيانات_الفضاء = معالج.معالجة()
        else:
            محرك = محرك_الفضاء(مسار, الطلب.وحدة_القياس, الطلب.الثوابت.مفتاح_gemini)
            بيانات_الفضاء = محرك.تنفيذ()

        النتائج = كتاب_الكميات.احسب(
            بيانات_الغرف = بيانات_الفضاء["الغرف"],
            الفضاء       = بيانات_الفضاء,
            الطلب        = الطلب,
        )

        مدقق    = مُدقِّق_المنطق(النتائج, بيانات_الفضاء, الطلب.الثوابت)
        تنبيهات = مدقق.تحقق()
        مسطح    = [ب for ق in النتائج.values() for ب in ق]
        درجة_ثقة_نهائية = _حساب_درجة_ثقة_محسّنة(بيانات_الفضاء["درجة_الثقة"], مسطح)
        _نظام_متكامل(النتائج, بيانات_الفضاء, الطلب.مسار_المجلد or os.path.dirname(مسار), الطلب.الثوابت.مفتاح_gemini)

        return {
            "الحالة":          "نجاح",
            "رقم_المشروع":     الطلب.رقم_المشروع,
            "درجة_الثقة":      درجة_ثقة_نهائية,
            "التنبيهات":       تنبيهات,
            "ملخص_التنبيهات": {
                "أخطاء":   sum(1 for ت in تنبيهات if ت["المستوى"]=="خطأ"),
                "تحذيرات": sum(1 for ت in تنبيهات if ت["المستوى"]=="تحذير"),
                "ملاحظات": sum(1 for ت in تنبيهات if ت["المستوى"]=="ملاحظة"),
            },
            "مكتشف_تلقائياً": {
                "المحيط_الخارجي":    بيانات_الفضاء["المحيط_الخارجي"],
                "طول_جدران_بلوك_20": بيانات_الفضاء.get("طول_جدران_بلوك_20", 0),
                "طول_جدران_بلوك_10": بيانات_الفضاء.get("طول_جدران_بلوك_10", 0),
                "مساحة_الحفر":       بيانات_الفضاء["مساحة_الحفر"],
                "الجداول":           بيانات_الفضاء["الجداول"],
                "الكمرات_من_الرسم":  بيانات_الفضاء.get("الكمرات_من_الرسم", {}),
            },
            "النتائج_بالأقسام": النتائج,
            "النتائج_مسطحة":    مسطح,
            "الغرف":            بيانات_الفضاء["الغرف"],
        }

    except Exception as خطأ:
        مسجل.error(f"فشل: {خطأ}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(خطأ))


# ── التكامل مع QS Hub (Legacy Endpoints) ──────────────────────────────────────────

@التطبيق.post("/run-master-qto")
async def run_master_qto(req: QTORequest):
    """
    نقطة النهاية الأساسية التي يتوقعها QS Hub.
    تقوم بتحويل طلب Hub إلى منطق V15 الداخلي.
    """
    try:
        # 1. بناء الثوابت الداخلية من طلب Hub
        الثوابت = ثوابت_المشروع(
            ارتفاع_الدور          = req.constants.floor_height,
            ارتفاع_الدور_الأرضي   = req.constants.floor_height, # Hub uses single height usually
            عمق_الحفر             = req.constants.excavation_depth,
            مستوى_بلاطة_الأرضي   = req.constants.gfsl_level,
            عمق_كمرة_الربط        = req.constants.tb_depth,
            عدد_الأدوار           = req.constants.no_of_floors,
            سماكة_البلاطة         = req.constants.slab_thickness,
            خرسانة_الدرج          = req.constants.staircase_concrete,
            يوجد_رصيف_طرق         = req.constants.road_base_exists,
            سماكة_رصيف_الطرق      = req.constants.road_base_thickness,
            مفتاح_gemini          = req.gemini_api_key,
        )

        # 2. تشغيل معالجة الفضاء
        محرك = محرك_الفضاء(req.file_path, req.unit, req.gemini_api_key or "",
                          pdf_drawing_scale=float(req.pdf_drawing_scale or 0.0))
        بيانات_الفضاء = محرك.تنفيذ()

        # 3. بناء الطلب الداخلي
        الطلب_الداخلي = طلب_حساب_الكميات(
            مسار_الملف   = req.file_path,
            رقم_المشروع  = req.project_id,
            وحدة_القياس  = req.unit,
            الثوابت      = الثوابت,
        )

        # 4. حساب الكميات باستخدام منطق V15
        النتائج = كتاب_الكميات.احسب(
            بيانات_الغرف = بيانات_الفضاء["الغرف"],
            الفضاء       = بيانات_الفضاء,
            الطلب        = الطلب_الداخلي,
        )

        # 5. تدقيق المنطق وبناء التنبيهات
        مدقق    = مُدقِّق_المنطق(النتائج, بيانات_الفضاء, الثوابت)
        تنبيهات = مدقق.تحقق()

        # 6. تحويل البنود إلى الأسماء الإنجليزية الدقيقة التي يتوقعها Hub
        قائمة_بأسماء_hub = {
            # ── Legacy keys (backward compat) ────────────────────────────────
            "حفر": "Excavation",
            "مساحة الحفر": "Excavation Area",
            "رديم": "Back Filling Volume",
            "PCC للمؤسسات": "Total Foundation PCC",
            "RCC للمؤسسات": "Total Foundation Volume",
            "رقاب أعمدة": "Total Neck Columns Volume",
            "ميدات": "Total Tie Beams Volume",
            "ستراب بيم": "Total Strap Beams Volume",
            "أعمدة سوبر": "Total Columns Volume",
            "كمرات السقف": "Total Beams Volume",
            "بلاستر داخلي": "Internal Plaster",
            "بلاستر خارجي": "External Wall Finish",
            "بلاط أرضيات": "Dry Area Flooring",
            "بلاط حمامات": "Wet Area Flooring (Ceramic)",
            "بلاط بلكونات": "Balcony Flooring",
            "وزرات": "Skirting",
            "عزل السطح": "Combo Roof System",
            # ── Modern V15 exact totals ───────────────────────────────────────
            "حجم الحفر": "Excavation",
            "إجمالي حجم الأسس": "Total Foundation Volume",
            "إجمالي حجم أعمدة العنق": "Total Neck Columns Volume",
            "إجمالي حجم كمرات الربط": "Total Tie Beams Volume",
            "إجمالي حجم كمرات الستراب": "Total Strap Beams Volume",
            "حجم بلاطة على الأرض": "Slab on Grade Volume",
            "مساحة بلاطة على الأرض": "Slab on Grade Area",
            "حجم الردم": "Back Filling Volume",
            "مبيد النمل الأبيض": "Anti-Termite Treatment",
            "نايلون أسود (بولي إيثيلين)": "Polyethylene Sheet",
            "نايلون أسود": "Polyethylene Sheet",
            "حجم بلاطة الدور الأول": "Slab Volume (Floor 1)",
            "حجم البلاطة (الدور 1)": "Slab Volume (Floor 1)",
            "حجم بلاطة الدور الثاني": "Slab Volume (Floor 2)",
            "حجم البلاطة (الدور 2)": "Slab Volume (Floor 2)",
            "حجم بلاطة السقف": "Slab Volume (Floor 2)",
            "حجم بلاطات أعلى السطح": "Slab Volume (Floor 2)",
            "إجمالي حجم الأعمدة": "Total Columns Volume",
            "إجمالي حجم الكمرات": "Total Beams Volume",
            "خرسانة الدرج": "Staircase Concrete",
            "بلوك 20 سم خارجي": "Block 20cm External",
            "بلوك 20 سم تصوينة السطح": "Block 20cm External",
            "بلوك 20 سم داخلي": "Block 20cm Internal",
            "بلوك 10 سم داخلي": "Block 10cm Internal",
            "لياسة داخلية": "Internal Plaster",
            "تشطيب الواجهة الخارجية": "External Wall Finish",
            "عزل مائي": "Waterproofing",
            "نظام السقف المركب": "Combo Roof System",
            "بلاط المناطق الجافة": "Dry Area Flooring",
            "سيراميك المناطق المبللة": "Wet Area Flooring (Ceramic)",
            "بلاط البلكونة": "Balcony Flooring",
            "سكرتة": "Skirting",
            "دهان": "Paint",
            "سقف المناطق الجافة": "Dry Area Ceiling",
            "بلاط الجدران": "Wall Tiles",
            "سقف المناطق المبللة": "Wet Area Ceiling",
            "عتبات رخام": "Marble Threshold",
            "مساحة رصيف الطرق": "Road Base Area",
            "حجم رصيف الطرق": "Road Base Volume",
            # ── Modern V15 per-element PREFIXES (matched with " (" suffix) ────
            "حجم الأساس": "Foundation Volume",
            "مساحة الأساس": "Foundation Area",
            "PCC الأساس": "Foundation PCC",
            "بيتومين الأساس": "Foundation Bitumen",
            "حجم عمود العنق": "Neck Column Volume",
            "بيتومين عمود العنق": "Neck Column Bitumen",
            "حجم كمرة الربط": "Tie Beam Volume",
            "PCC كمرة الربط": "Tie Beam PCC",
            "بيتومين كمرة الربط": "Tie Beam Bitumen",
            "حجم كمرة الستراب": "Strap Beam Volume",
            "مساحة بلوك تحت الأرض": "Underground Block Area",
            "بيتومين بلوك تحت الأرض": "Underground Block Bitumen",
            "حجم العمود": "Column Volume",
            "حجم الكمرة": "Beam Volume",
        }

        def _ترجمة(اسم: str) -> str:
            if اسم in قائمة_بأسماء_hub:
                return قائمة_بأسماء_hub[اسم]
            for عربي, إنجليزي in قائمة_بأسماء_hub.items():
                if اسم.startswith(عربي + " ("):
                    return إنجليزي + اسم[len(عربي):]
            return اسم

        نتائج_Hub_بأقسام = {}
        نتائج_Hub_مسطحة = []
        مسطح_عربي = [ب for قسم in النتائج.values() for ب in قسم]
        درجة_ثقة_نهائية = _حساب_درجة_ثقة_محسّنة(بيانات_الفضاء["درجة_الثقة"], مسطح_عربي)
        _نظام_متكامل(النتائج, بيانات_الفضاء, os.path.dirname(req.file_path), req.gemini_api_key or "")

        for قسم, بنود in النتائج.items():
            بنود_جديدة = []
            for بند in بنود:
                بند_معدل = {
                    "item":   _ترجمة(بند["البند"]),
                    "qty":    بند["الكمية"],
                    "unit":   بند["الوحدة"],
                    "status": بند.get("الحالة", ""),
                    "description": بند.get("الوصف", "")
                }
                بنود_جديدة.append(بند_معدل)
                نتائج_Hub_مسطحة.append(بند_معدل)
            نتائج_Hub_بأقسام[قسم] = بنود_جديدة

        return {
            "status": "Success",
            "project_id": req.project_id,
            "confidence": درجة_ثقة_نهائية,
            "auto_detected": {
                "external_perimeter": بيانات_الفضاء["المحيط_الخارجي"],
                "internal_walls_20_length": بيانات_الفضاء.get("طول_جدران_بلوك_20", 0),
                "internal_walls_10_length": بيانات_الفضاء.get("طول_جدران_بلوك_10", 0),
                "excavation_area": بيانات_الفضاء["مساحة_الحفر"],
                "schedule": بيانات_الفضاء["الجداول"],
            },
            "results_by_section": نتائج_Hub_بأقسام,
            "results_flat": نتائج_Hub_مسطحة,
            "rooms": بيانات_الفضاء["الغرف"],
            "alerts": تنبيهات
        }

    except Exception as e:
        logger.error(f"Hub Integration Failure: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@التطبيق.post("/run-pdf-qto")
async def run_pdf_qto(req: QTORequest):
    """نقطة نهاية إضافية لمعالجة PDF بشكل مستقل إذا لزم الأمر."""
    return await run_master_qto(req)


if __name__ == "__main__":
    import sys
    import json
    import argparse
    from pydantic import ValidationError

    # إعداد معالج الأوامر
    parser = argparse.ArgumentParser(description="محرك حساب الكميات V15 - تشغيل مباشر")
    parser.add_argument("--config", type=str, help="مسار ملف JSON يحتوي على بيانات QTORequest")
    parser.add_argument("--server", action="store_true", help="تشغيل كخادم FastAPI (الوضع القديم)")
    
    args, unknown = parser.parse_known_args()

    if args.server:
        import uvicorn
        مسجل.info("🚀 محرك حساب الكميات v15 يعمل بنظام التكامل مع QS Hub على http://localhost:8000")
        uvicorn.run(التطبيق, host="0.0.0.0", port=8000)
    
    elif args.config:
        try:
            with open(args.config, 'r', encoding='utf-8-sig') as f:
                config_data = json.load(f)
            
            # تحويل البيانات إلى Model
            req = QTORequest(**config_data)
            
            # تنفيذ الحساب (نفس منطق الـ Endpoint)
            # 1. بناء الثوابت
            الثوابت = ثوابت_المشروع(
                ارتفاع_الدور          = req.constants.floor_height,
                ارتفاع_الدور_الأرضي   = req.constants.floor_height,
                عمق_الحفر             = req.constants.excavation_depth,
                مستوى_بلاطة_الأرضي   = req.constants.gfsl_level,
                عمق_كمرة_الربط        = req.constants.tb_depth,
                عدد_الأدوار           = req.constants.no_of_floors,
                سماكة_البلاطة         = req.constants.slab_thickness,
                خرسانة_الدرج          = req.constants.staircase_concrete,
                يوجد_رصيف_طرق         = req.constants.road_base_exists,
                سماكة_رصيف_الطرق      = req.constants.road_base_thickness,
                مفتاح_gemini          = req.gemini_api_key or "",
            )

            # 2. تشغيل معالجة الفضاء
            محرك = محرك_الفضاء(req.file_path, req.unit, req.gemini_api_key or "",
                              pdf_drawing_scale=float(req.pdf_drawing_scale or 0.0))
            بيانات_الفضاء = محرك.تنفيذ()

            # 3. بناء الطلب الداخلي
            الطلب_الداخلي = طلب_حساب_الكميات(
                مسار_الملف   = req.file_path,
                رقم_المشروع  = req.project_id,
                وحدة_القياس  = req.unit,
                الثوابت      = الثوابت,
            )

            # 4. حساب الكميات
            النتائج = كتاب_الكميات.احسب(
                بيانات_الغرف = بيانات_الفضاء["الغرف"],
                الفضاء       = بيانات_الفضاء,
                الطلب        = الطلب_الداخلي,
            )

            # 5. تدقيق المنطق
            مدقق    = مُدقِّق_المنطق(النتائج, بيانات_الفضاء, الثوابت)
            تنبيهات = مدقق.تحقق()

            # 6. تحويل البنود لأسماء Hub
            قائمة_بأسماء_hub = {
                # ── Legacy keys (backward compat) ────────────────────────────
                "حفر": "Excavation",
                "مساحة الحفر": "Excavation Area",
                "رديم": "Back Filling Volume",
                "PCC للمؤسسات": "Total Foundation PCC",
                "RCC للمؤسسات": "Total Foundation Volume",
                "رقاب أعمدة": "Total Neck Columns Volume",
                "ميدات": "Total Tie Beams Volume",
                "ستراب بيم": "Total Strap Beams Volume",
                "أعمدة سوبر": "Total Columns Volume",
                "كمرات السقف": "Total Beams Volume",
                "بلاستر داخلي": "Internal Plaster",
                "بلاستر خارجي": "External Wall Finish",
                "بلاط أرضيات": "Dry Area Flooring",
                "بلاط حمامات": "Wet Area Flooring (Ceramic)",
                "بلاط بلكونات": "Balcony Flooring",
                "وزرات": "Skirting",
                "عزل السطح": "Combo Roof System",
                "عزل مائي (Wet Areas)": "Waterproofing",
                "رخام عتبات": "Marble Threshold",
                "أرضيات جافة": "Dry Area Flooring",
                "أرضيات مبللة": "Wet Area Flooring (Ceramic)",
                # ── Modern V15 exact totals ───────────────────────────────────
                "حجم الحفر": "Excavation",
                "إجمالي حجم الأسس": "Total Foundation Volume",
                "إجمالي حجم أعمدة العنق": "Total Neck Columns Volume",
                "إجمالي حجم كمرات الربط": "Total Tie Beams Volume",
                "إجمالي حجم كمرات الستراب": "Total Strap Beams Volume",
                "حجم بلاطة على الأرض": "Slab on Grade Volume",
                "مساحة بلاطة على الأرض": "Slab on Grade Area",
                "حجم الردم": "Back Filling Volume",
                "مبيد النمل الأبيض": "Anti-Termite Treatment",
                "نايلون أسود (بولي إيثيلين)": "Polyethylene Sheet",
                "نايلون أسود": "Polyethylene Sheet",
                "حجم بلاطة الدور الأول": "Slab Volume (Floor 1)",
                "حجم البلاطة (الدور 1)": "Slab Volume (Floor 1)",
                "حجم بلاطة الدور الثاني": "Slab Volume (Floor 2)",
                "حجم البلاطة (الدور 2)": "Slab Volume (Floor 2)",
                "حجم بلاطة السقف": "Slab Volume (Floor 2)",
                "حجم بلاطات أعلى السطح": "Slab Volume (Floor 2)",
                "إجمالي حجم الأعمدة": "Total Columns Volume",
                "إجمالي حجم الكمرات": "Total Beams Volume",
                "خرسانة الدرج": "Staircase Concrete",
                "بلوك 20 سم خارجي": "Block 20cm External",
                "بلوك 20 سم تصوينة السطح": "Block 20cm External",
                "بلوك 20 سم داخلي": "Block 20cm Internal",
                "بلوك 10 سم داخلي": "Block 10cm Internal",
                "لياسة داخلية": "Internal Plaster",
                "تشطيب الواجهة الخارجية": "External Wall Finish",
                "عزل مائي": "Waterproofing",
                "نظام السقف المركب": "Combo Roof System",
                "بلاط المناطق الجافة": "Dry Area Flooring",
                "سيراميك المناطق المبللة": "Wet Area Flooring (Ceramic)",
                "بلاط البلكونة": "Balcony Flooring",
                "سكرتة": "Skirting",
                "دهان": "Paint",
                "سقف المناطق الجافة": "Dry Area Ceiling",
                "بلاط الجدران": "Wall Tiles",
                "سقف المناطق المبللة": "Wet Area Ceiling",
                "عتبات رخام": "Marble Threshold",
                "مساحة رصيف الطرق": "Road Base Area",
                "حجم رصيف الطرق": "Road Base Volume",
                # ── Modern V15 per-element PREFIXES (matched with " (" suffix) ─
                "حجم الأساس": "Foundation Volume",
                "مساحة الأساس": "Foundation Area",
                "PCC الأساس": "Foundation PCC",
                "بيتومين الأساس": "Foundation Bitumen",
                "حجم عمود العنق": "Neck Column Volume",
                "بيتومين عمود العنق": "Neck Column Bitumen",
                "حجم كمرة الربط": "Tie Beam Volume",
                "PCC كمرة الربط": "Tie Beam PCC",
                "بيتومين كمرة الربط": "Tie Beam Bitumen",
                "حجم كمرة الستراب": "Strap Beam Volume",
                "مساحة بلوك تحت الأرض": "Underground Block Area",
                "بيتومين بلوك تحت الأرض": "Underground Block Bitumen",
                "حجم العمود": "Column Volume",
                "حجم الكمرة": "Beam Volume",
            }

            def _ترجمة(اسم: str) -> str:
                if اسم in قائمة_بأسماء_hub:
                    return قائمة_بأسماء_hub[اسم]
                for عربي, إنجليزي in قائمة_بأسماء_hub.items():
                    if اسم.startswith(عربي + " ("):
                        return إنجليزي + اسم[len(عربي):]
                return اسم

            نتائج_Hub_بأقسام = {}
            نتائج_Hub_مسطحة = []
            مسطح_عربي = [ب for قسم in النتائج.values() for ب in قسم]
            درجة_ثقة_نهائية = _حساب_درجة_ثقة_محسّنة(بيانات_الفضاء["درجة_الثقة"], مسطح_عربي)
            _نظام_متكامل(النتائج, بيانات_الفضاء, os.path.dirname(req.file_path), req.gemini_api_key or "")
            for قسم, بنود in النتائج.items():
                بنود_جديدة = []
                for بند in بنود:
                    بند_معدل = {
                        "item":   _ترجمة(بند["البند"]),
                        "qty":    بند["الكمية"],
                        "unit":   بند["الوحدة"],
                        "status": بند.get("الحالة", ""),
                        "description": بند.get("الوصف", "")
                    }
                    بنود_جديدة.append(بند_معدل)
                    نتائج_Hub_مسطحة.append(بند_معدل)
                نتائج_Hub_بأقسام[قسم] = بنود_جديدة

            # مخرجات JSON النهائية لـ Node.js
            output = {
                "status": "Success",
                "results_by_section": نتائج_Hub_بأقسام,
                "results_flat": نتائج_Hub_مسطحة,
                "confidence": درجة_ثقة_نهائية,
                "rooms": بيانات_الفضاء["الغرف"],
                "layers": بيانات_الفضاء.get("الطبقات_المكتشفة", []),
                "raw_spatial_evidence": بيانات_الفضاء,
                "alerts": تنبيهات
            }
            print(json.dumps(output, ensure_ascii=False))

        except Exception as e:
            print(json.dumps({"status": "Error", "detail": str(e)}))
            sys.exit(1)
    
    else:
        parser.print_help()
